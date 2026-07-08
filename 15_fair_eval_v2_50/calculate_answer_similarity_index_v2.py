from __future__ import annotations

import difflib
import itertools
import statistics
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("오류: openpyxl이 필요합니다. 프로젝트 가상환경에서 'pip install openpyxl' 후 다시 실행하세요.")
    sys.exit(1)

BASE_DIR = Path(__file__).resolve().parent
INPUT_PATH = BASE_DIR / "answer_similarity_input_template_50_v2.xlsx"
OUTPUT_PATH = BASE_DIR / "answer_similarity_result_50_v2.xlsx"


def normalize(value):
    return "" if value is None else str(value).strip()


def header_map(ws):
    return {normalize(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}


def stability_level(score):
    if score >= 0.85:
        return "높음"
    if score >= 0.70:
        return "보통"
    return "낮음"


def difflib_similarity(texts):
    return [
        difflib.SequenceMatcher(None, a, b).ratio()
        for a, b in itertools.combinations(texts, 2)
    ]


def tfidf_similarity(texts):
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.metrics.pairwise import cosine_similarity
    except ImportError:
        return None
    matrix = TfidfVectorizer().fit_transform(texts)
    sim = cosine_similarity(matrix)
    values = []
    for i in range(len(texts)):
        for j in range(i + 1, len(texts)):
            values.append(float(sim[i, j]))
    return values


def main():
    if not INPUT_PATH.exists():
        print(f"오류: 입력 파일을 찾을 수 없습니다: {INPUT_PATH}")
        sys.exit(1)

    wb = load_workbook(INPUT_PATH, data_only=True)
    ws = wb.active
    h = header_map(ws)
    required = ["question_id", "original_model", "run_id", "answer_text"]
    missing = [c for c in required if c not in h]
    if missing:
        print("오류: 입력 파일에 필요한 컬럼이 없습니다: " + ", ".join(missing))
        sys.exit(1)

    groups = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        qid = normalize(ws.cell(row, h["question_id"]).value)
        model = normalize(ws.cell(row, h["original_model"]).value)
        answer = normalize(ws.cell(row, h["answer_text"]).value)
        if qid and model and answer:
            groups[(qid, model)].append(answer)

    out = Workbook()
    result = out.active
    result.title = "similarity_result"
    result.append(["question_id", "original_model", "similarity_mean", "similarity_min", "similarity_max", "stability_level"])

    method_used = "TF-IDF cosine similarity if available, otherwise difflib SequenceMatcher"
    for (qid, model), texts in sorted(groups.items()):
        if len(texts) < 2:
            continue
        values = tfidf_similarity(texts)
        if values is None:
            values = difflib_similarity(texts)
            method_used = "difflib SequenceMatcher"
        else:
            method_used = "TF-IDF cosine similarity"
        mean_v = statistics.mean(values)
        result.append([qid, model, round(mean_v, 4), round(min(values), 4), round(max(values), 4), stability_level(mean_v)])

    note = out.create_sheet("method_note")
    note.append(["item", "value"])
    note.append(["method_used", method_used])
    note.append(["stability_high", "0.85 이상"])
    note.append(["stability_medium", "0.70 이상 0.85 미만"])
    note.append(["stability_low", "0.70 미만"])
    out.save(OUTPUT_PATH)
    print(f"유사성 지수 결과 파일 생성 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
