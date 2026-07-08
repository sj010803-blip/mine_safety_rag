from __future__ import annotations

import json
import math
import os
import re
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
except ImportError:
    print("오류: openpyxl이 필요합니다. 프로젝트 가상환경에서 'pip install openpyxl' 후 다시 실행하세요.")
    sys.exit(1)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PACKAGE_DIR = PROJECT_ROOT / "15_fair_eval_v2_50"
INPUT_XLSX = PACKAGE_DIR / "model_answer_collection_template_50.xlsx"
OUTPUT_XLSX = PACKAGE_DIR / "model_answer_collection_template_50_with_minesafe.xlsx"
REPORT_PATH = PACKAGE_DIR / "collect_minesafe_answers_50_v2_report.txt"

VECTOR_DB_DIR = PROJECT_ROOT / "10_vector_db_with_major_accident_docs"
CHUNKS_PATH = PROJECT_ROOT / "08_chunks" / "chunks_with_major_accident_docs.jsonl"
COLLECTION_NAME = "mine_safety_docs"
EMBEDDING_MODEL_NAME = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"
GEMINI_MODEL_NAME = os.getenv("MINESAFE_GEMINI_MODEL", "gemini-1.5-flash")

REQUIRED_COLUMNS = [
    "question_id",
    "category",
    "difficulty",
    "question",
    "chatgpt_answer",
    "gemini_answer",
    "minesafe_ai_answer",
    "answer_collection_note",
]


@dataclass
class Evidence:
    source: str
    chunk_id: str
    distance: float | None
    text: str


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def compact_space(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def truncate(text: str, limit: int = 420) -> str:
    text = compact_space(text)
    return text if len(text) <= limit else text[: limit - 1].rstrip() + "..."


def header_map(ws) -> dict[str, int]:
    return {clean_text(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}


def require_columns(headers: dict[str, int]) -> None:
    missing = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing:
        raise ValueError("필수 컬럼이 없습니다: " + ", ".join(missing))


def row_is_question(ws, headers: dict[str, int], row_idx: int) -> bool:
    return bool(clean_text(ws.cell(row_idx, headers["question_id"]).value))


def load_existing_minesafe_answers() -> dict[str, str]:
    if not OUTPUT_XLSX.exists():
        return {}
    try:
        wb = load_workbook(OUTPUT_XLSX, data_only=False)
        ws = wb.active
        headers = header_map(ws)
        require_columns(headers)
        existing: dict[str, str] = {}
        for row_idx in range(2, ws.max_row + 1):
            qid = clean_text(ws.cell(row_idx, headers["question_id"]).value)
            answer = clean_text(ws.cell(row_idx, headers["minesafe_ai_answer"]).value)
            if qid and answer:
                existing[qid] = answer
        return existing
    except Exception:
        return {}


class Retriever:
    def __init__(self, report_lines: list[str]) -> None:
        self.report_lines = report_lines
        self.mode = "fallback_lexical_chunks"
        self.collection = None
        self.embedding_model = None
        self.chunk_rows: list[dict[str, Any]] = []
        self._init_vector_db()

    def _init_vector_db(self) -> None:
        try:
            import chromadb
            from sentence_transformers import SentenceTransformer

            if not VECTOR_DB_DIR.exists():
                raise FileNotFoundError(f"Vector DB 폴더가 없습니다: {VECTOR_DB_DIR}")
            client = chromadb.PersistentClient(path=str(VECTOR_DB_DIR))
            self.collection = client.get_collection(name=COLLECTION_NAME)
            self.embedding_model = SentenceTransformer(EMBEDDING_MODEL_NAME)
            self.mode = "chromadb_sentence_transformer"
            count = self.collection.count()
            self.report_lines.append(f"- 검색 모드: ChromaDB + SentenceTransformer, collection={COLLECTION_NAME}, count={count}")
            return
        except Exception as exc:
            self.report_lines.append(f"- ChromaDB 검색 초기화 실패, lexical fallback 사용: {type(exc).__name__}: {exc}")

        self._load_chunks_for_lexical_fallback()

    def _load_chunks_for_lexical_fallback(self) -> None:
        if not CHUNKS_PATH.exists():
            self.report_lines.append(f"- chunks fallback 불가: 파일 없음 {CHUNKS_PATH}")
            return
        try:
            with CHUNKS_PATH.open("r", encoding="utf-8-sig") as f:
                for line in f:
                    if not line.strip():
                        continue
                    self.chunk_rows.append(json.loads(line))
            self.report_lines.append(f"- lexical fallback chunks 로드: {len(self.chunk_rows)}개")
        except Exception as exc:
            self.report_lines.append(f"- chunks fallback 로드 실패: {type(exc).__name__}: {exc}")

    def search(self, question: str, top_k: int = 5) -> list[Evidence]:
        if self.collection is not None and self.embedding_model is not None:
            try:
                embedding = self.embedding_model.encode(
                    [question],
                    normalize_embeddings=True,
                    show_progress_bar=False,
                ).tolist()
                result = self.collection.query(
                    query_embeddings=embedding,
                    n_results=top_k,
                    include=["documents", "metadatas", "distances"],
                )
                docs = result.get("documents", [[]])[0]
                metas = result.get("metadatas", [[]])[0]
                distances = result.get("distances", [[]])[0]
                evidence: list[Evidence] = []
                for i, doc in enumerate(docs):
                    meta = metas[i] if i < len(metas) and isinstance(metas[i], dict) else {}
                    evidence.append(
                        Evidence(
                            source=clean_text(
                                meta.get("source")
                                or meta.get("source_file")
                                or meta.get("file_name")
                                or meta.get("doc_name")
                                or "출처 미상"
                            ),
                            chunk_id=clean_text(meta.get("chunk_id") or f"rank_{i + 1}"),
                            distance=distances[i] if i < len(distances) else None,
                            text=clean_text(doc),
                        )
                    )
                return evidence
            except Exception as exc:
                self.report_lines.append(f"- Vector 검색 실패, lexical fallback 전환: {type(exc).__name__}: {exc}")
        return self._lexical_search(question, top_k=top_k)

    def _lexical_search(self, question: str, top_k: int = 5) -> list[Evidence]:
        terms = extract_terms(question)
        scored: list[tuple[float, dict[str, Any]]] = []
        for row in self.chunk_rows:
            text = clean_text(row.get("text") or row.get("document") or "")
            haystack = f"{row.get('source_file', '')} {row.get('source', '')} {text}".lower()
            score = sum(haystack.count(term.lower()) for term in terms)
            if score > 0:
                score += min(len(text), 1500) / 100000.0
                scored.append((score, row))
        scored.sort(key=lambda item: item[0], reverse=True)
        evidence = []
        for score, row in scored[:top_k]:
            evidence.append(
                Evidence(
                    source=clean_text(row.get("source") or row.get("source_file") or "출처 미상"),
                    chunk_id=clean_text(row.get("chunk_id") or ""),
                    distance=None if score <= 0 else round(1 / score, 4),
                    text=clean_text(row.get("text") or ""),
                )
            )
        return evidence


def extract_terms(question: str) -> list[str]:
    raw = re.findall(r"[가-힣A-Za-z0-9]{2,}", question)
    stop = {
        "작업",
        "현장",
        "관리",
        "관리자",
        "무엇",
        "어떤",
        "해야",
        "한다",
        "있는",
        "위한",
        "기준",
        "조치",
        "사항",
    }
    terms = []
    for term in raw:
        if term not in stop and term not in terms:
            terms.append(term)
    priority = [
        "발파",
        "불발",
        "분진",
        "환기",
        "유해가스",
        "메탄",
        "산소",
        "낙반",
        "붕락",
        "지보",
        "보호구",
        "장비",
        "전기",
        "협착",
        "중대재해",
        "안전보건관리체계",
        "기록",
        "보고",
        "작업재개",
    ]
    for term in priority:
        if term in question and term not in terms:
            terms.insert(0, term)
    return terms[:12] or [question[:20]]


def infer_focus(question: str, category: str) -> dict[str, list[str] | str]:
    text = f"{category} {question}"
    immediate = "작업을 일시 중지하고 위험구역 접근을 통제한 뒤, 책임자 확인 전까지 임의 작업을 금지한다."
    hazards = ["광산 작업 중 위험요인"]
    first_actions = [
        "작업자를 위험구역 밖으로 이동시키고 인원 이상 유무를 확인한다.",
        "현장 책임자에게 즉시 보고하고 작업구역 출입을 통제한다.",
        "관련 설비와 작업 조건을 점검해 원인을 확인한다.",
    ]
    restart = [
        "위험요인이 제거되었고 재측정 또는 재점검 결과가 허용 가능한 상태일 것",
        "작업자에게 변경된 절차와 잔여 위험을 공유했을 것",
        "책임자 승인과 기록이 남아 있을 것",
    ]

    if "발파" in text or "불발" in text:
        immediate = "발파 또는 불발 의심 상황은 즉시 작업중지 대상으로 보고, 대피·경계·재진입 금지를 우선 적용한다."
        hazards = ["불발 폭약 또는 잔류 화약류", "후가스와 시야 불량", "대피 확인 누락", "경계 미흡"]
        first_actions = [
            "발파 구역과 예상 비산·충격 범위를 통제하고 무단 접근을 막는다.",
            "발파 책임자 또는 유자격자가 대피 확인, 경보, 불발 의심 지점을 재점검한다.",
            "후가스와 분진이 남아 있으면 환기 후 재측정하고, 불발 의심 지점은 임의 접촉하지 않는다.",
        ]
        restart = [
            "불발 여부가 유자격자에 의해 확인·처리되었을 것",
            "가스·분진·시야 등 재진입 조건이 안전하게 회복되었을 것",
            "발파 기록, 대피 확인, 경계 해제 승인 기록이 남아 있을 것",
        ]
    elif any(k in text for k in ["분진", "환기", "유해가스", "메탄", "산소", "공기"]):
        immediate = "분진·환기·유해가스 이상은 노출 확대를 막기 위해 작업을 멈추고 환기, 측정, 접근통제를 먼저 시행한다."
        hazards = ["호흡성 분진 노출", "환기 부족", "메탄 등 유해가스", "산소결핍 가능성", "측정 장비 신뢰성 부족"]
        first_actions = [
            "작업자를 신선한 공기 구역으로 이동시키고 증상자를 확인한다.",
            "환기 설비, 덕트, 집진 장치, 가스 측정기의 상태를 점검한다.",
            "재측정 전까지 위험구역 출입을 통제하고 필요한 보호구를 확인한다.",
        ]
        restart = [
            "가스·산소·분진 측정값이 안전 기준을 만족하고 측정 장비 신뢰성이 확인되었을 것",
            "환기·집진 설비 이상이 조치되었을 것",
            "작업자 보호구와 비상대응 절차가 확인되었을 것",
        ]
    elif any(k in text for k in ["낙반", "붕락", "지보", "균열", "부석"]):
        immediate = "낙반·붕락 징후가 있으면 하부 작업을 즉시 중지하고 위험 범위를 넓게 잡아 접근을 통제한다."
        hazards = ["천반 균열", "부석", "지보재 변형", "지반 약화", "물 유입과 진동"]
        first_actions = [
            "작업자를 위험구역 밖으로 대피시키고 하부 접근을 금지한다.",
            "지보 상태, 균열, 낙석, 물 유입, 진동 여부를 책임자가 확인한다.",
            "필요 시 보강 설계, 부석 제거, 우회 작업계획을 검토한다.",
        ]
        restart = [
            "부석 제거와 지보 보강이 완료되고 추가 변형 징후가 없을 것",
            "전문가 또는 책임자가 안정성을 확인했을 것",
            "우회·재개 작업계획과 통제선이 기록되어 있을 것",
        ]
    elif any(k in text for k in ["보호구", "작업자", "출입", "통제", "협력업체", "방진마스크"]):
        immediate = "보호구 미착용이나 통제 실패가 보이면 즉시 작업을 멈추고 해당 인원을 안전구역으로 이동시켜야 한다."
        hazards = ["보호구 미착용", "위험구역 무단출입", "교육 부족", "동시작업 충돌", "협력업체 통제 미흡"]
        first_actions = [
            "출입 권한, 보호구 착용, 교육 이수 여부를 확인한다.",
            "위험구역 경계와 유도자를 배치하고 비작업자 접근을 제한한다.",
            "보호구 부적합 원인을 확인해 교체, 재교육, 착용 확인을 실시한다.",
        ]
        restart = [
            "보호구 착용과 적합성이 확인되었을 것",
            "작업자 명단, 출입 권한, 대피로 교육이 확인되었을 것",
            "원청·협력업체 간 지휘체계와 통제 절차가 확인되었을 것",
        ]
    elif any(k in text for k in ["장비", "전기", "협착", "컨베이어", "운반", "감전", "정비"]):
        immediate = "장비·전기·협착 위험은 에너지원 차단, 접근통제, 운전정지를 먼저 적용하고 임시 운전은 제한한다."
        hazards = ["장비 충돌·협착", "전원 재투입", "방호장치 제거", "감전", "정비 중 동시작업"]
        first_actions = [
            "장비를 정지시키고 전원 차단·잠금·표시 상태를 확인한다.",
            "장비 이동 반경과 정비 구역을 분리하고 유도자 또는 감시자를 배치한다.",
            "방호장치, 경보장치, 접지, 누전 보호, 시야 확보 상태를 점검한다.",
        ]
        restart = [
            "정비와 점검이 완료되고 방호장치가 원상복구되었을 것",
            "잠금·표시 해제 권한자가 확인했을 것",
            "시운전, 작업자 대피, 주변 통제 기록이 남아 있을 것",
        ]
    elif any(k in text for k in ["중대재해", "안전보건관리체계", "경영", "예산", "협력업체", "위험성평가"]):
        immediate = "중대재해로 이어질 수 있는 위험은 현장 조치와 동시에 경영책임 라인에 보고해 자원·인력·예산 조치를 검토해야 한다."
        hazards = ["위험성평가 형식화", "보고 체계 지연", "예산·인력 부족", "협력업체 관리 미흡", "작업중지권 미작동"]
        first_actions = [
            "위험을 공식 보고하고 임시 통제조치를 즉시 시행한다.",
            "위험성평가, 개선계획, 예산·인력 배정 필요성을 검토한다.",
            "원청·협력업체 역할과 책임, 교육·점검 주기를 명확히 한다.",
        ]
        restart = [
            "개선조치와 책임자 확인이 완료되었을 것",
            "예산·인력·교육·점검 등 관리체계 보완이 기록되었을 것",
            "협력업체까지 변경 사항을 공유했을 것",
        ]
    elif any(k in text for k in ["기록", "보고", "작업재개", "재가동", "회의", "사고"]):
        immediate = "기록·보고가 불충분하면 작업재개를 보류하고 원인, 개선조치, 승인 근거를 먼저 확보한다."
        hazards = ["기록 누락", "보고 지연", "원인 미확인", "임시조치만 완료", "작업재개 승인 불명확"]
        first_actions = [
            "작업중지 사유, 현장 상태, 응급조치, 통제 상태를 기록한다.",
            "원인 분석과 개선조치 완료 여부를 책임자가 확인한다.",
            "재개 회의에서 잔여 위험, 담당자, 확인 방법을 합의한다.",
        ]
        restart = [
            "개선조치 완료 증빙과 점검 결과가 있을 것",
            "작업자 공지와 재교육이 끝났을 것",
            "작업재개 승인자, 시간, 조건이 기록되어 있을 것",
        ]

    return {
        "immediate": immediate,
        "hazards": hazards,
        "first_actions": first_actions,
        "restart": restart,
    }


def format_evidence(evidence: list[Evidence]) -> str:
    if not evidence:
        return "- 검색 근거를 충분히 확보하지 못했으므로 현장 기준서, 작업절차서, 법정 안전기준을 추가 확인해야 합니다."
    lines = []
    seen = set()
    for idx, item in enumerate(evidence, start=1):
        key = (item.source, item.chunk_id)
        if key in seen:
            continue
        seen.add(key)
        distance = "" if item.distance is None else f", distance={item.distance:.4f}" if isinstance(item.distance, float) else f", distance={item.distance}"
        lines.append(f"- [{idx}] {item.source} / {item.chunk_id}{distance}: {truncate(item.text, 260)}")
        if len(lines) >= 5:
            break
    return "\n".join(lines)


def build_prompt(question_id: str, category: str, difficulty: str, question: str, evidence: list[Evidence]) -> str:
    return f"""당신은 MineSafe AI, 광산 안전 지침 및 중대재해처벌법 대응 RAG 가상 안전관리자입니다.
아래 질문에 대해 공식 문서 검색 근거를 바탕으로 안전 우선 답변을 작성하세요.

질문 ID: {question_id}
분류: {category}
난이도: {difficulty}
질문: {question}

검색 근거:
{format_evidence(evidence)}

답변에는 반드시 다음 제목을 포함하세요.
1. 즉시 판단
2. 검색 근거
3. 우선 조치
4. 작업 재개 조건
5. KRAS식 위험성평가 초안
6. 현장 조치 체크리스트
7. 기록·보고 사항

주의:
- 법 조문 번호나 수치를 확신할 수 없으면 단정하지 말고 확인 필요성을 밝히세요.
- 생산 일정이 아니라 작업자 생명과 안전을 우선하세요.
- 답변이 길기만 하지 않게 현장 실행 순서 중심으로 작성하세요.
"""


def try_gemini_answer(prompt: str) -> tuple[str | None, str | None]:
    api_key = clean_text(os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY"))
    if not api_key:
        return None, "GEMINI_API_KEY/GOOGLE_API_KEY 환경변수가 없어 fallback 사용"
    try:
        from google import genai

        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(model=GEMINI_MODEL_NAME, contents=prompt)
        answer = clean_text(getattr(response, "text", ""))
        if not answer:
            return None, "Gemini 응답 텍스트가 비어 있어 fallback 사용"
        return answer, None
    except Exception as exc:
        return None, f"Gemini 호출 실패로 fallback 사용: {type(exc).__name__}: {exc}"


def fallback_answer(question_id: str, category: str, difficulty: str, question: str, evidence: list[Evidence]) -> str:
    focus = infer_focus(question, category)
    hazards = focus["hazards"]
    first_actions = focus["first_actions"]
    restart = focus["restart"]
    evidence_lines = format_evidence(evidence)
    main_hazard = hazards[0] if hazards else "광산 안전 위험요인"

    return f"""MineSafe AI 답변

1. 즉시 판단
{focus["immediate"]} 질문({question_id})은 {category} 영역의 {difficulty} 난이도 상황으로, 원인 확인 전에는 작업 지속보다 작업중지·접근통제·책임자 확인을 우선해야 합니다.

2. 검색 근거
{evidence_lines}
위 근거는 광산안전 관련 법령·기술기준·안전보건관리체계 자료에서 검색된 chunks입니다. 특정 조문 번호는 현장 적용 전 최신 원문으로 재확인하는 것이 안전합니다.

3. 우선 조치
- {first_actions[0]}
- {first_actions[1]}
- {first_actions[2]}
- 증상자, 미대피자, 위험구역 잔류 인원이 있는지 확인하고 필요하면 응급조치와 대피를 우선합니다.
- 임시조치는 위험을 낮추기 위한 보조수단이며, 원인 확인과 책임자 승인 없이 정상 작업으로 간주하지 않습니다.

4. 작업 재개 조건
- {restart[0]}
- {restart[1]}
- {restart[2]}
- 작업재개 전 작업자에게 변경된 통제선, 보호구, 비상 연락, 대피 절차를 다시 공유해야 합니다.

5. KRAS식 위험성평가 초안
- 유해·위험요인: {", ".join(hazards)}
- 발생 가능한 재해: 폭발, 질식, 낙반, 협착, 감전, 보호구 미착용에 따른 상해 등 해당 상황의 주요 재해
- 현재 위험성: 원인 확인 전에는 중대 이상으로 보수 평가
- 감소대책: 작업중지, 위험구역 격리, 측정·점검, 설비 보완, 보호구 확인, 작업자 재교육, 책임자 승인
- 잔여 위험성: 조치 완료와 재측정 후 낮음 또는 보통으로 재평가
- 담당: 현장 책임자, 안전관리자, 설비 담당자, 협력업체 책임자

6. 현장 조치 체크리스트
- [ ] 작업중지 및 위험구역 출입통제 실시
- [ ] 작업자 인원, 증상자, 대피 상태 확인
- [ ] 관련 설비·환경·보호구·작업절차 점검
- [ ] 검색 근거와 현장 기준서를 대조해 적용 기준 확인
- [ ] 개선조치 완료 후 책임자 승인
- [ ] 작업재개 전 TBM 또는 안전회의로 변경사항 공유

7. 기록·보고 사항
- 보고 내용: 발생 시각, 장소, 관련 작업, 발견자, 위험 징후, 즉시 조치, 통제 범위
- 첨부 자료: 사진, 측정값, 점검표, 작업중지 지시, 개선조치 내역, 재개 승인 기록
- 보고 경로: 현장 책임자와 안전관리자에게 즉시 보고하고, 중대재해 가능성이 있으면 경영책임 라인까지 보고합니다.
- 사후 관리: 같은 유형의 재발 가능성을 위험성평가와 교육자료에 반영하고, 협력업체가 관련되면 공동 재발방지 대책을 남깁니다.
"""


def generate_answer(
    question_id: str,
    category: str,
    difficulty: str,
    question: str,
    retriever: Retriever,
    report_lines: list[str],
) -> tuple[str, str]:
    evidence = retriever.search(question, top_k=5)
    prompt = build_prompt(question_id, category, difficulty, question, evidence)
    gemini_answer, gemini_error = try_gemini_answer(prompt)
    if gemini_answer:
        return gemini_answer, "gemini"
    if gemini_error:
        report_lines.append(f"- {question_id}: {gemini_error}")
    return fallback_answer(question_id, category, difficulty, question, evidence), "fallback"


def save_workbook_safely(wb) -> None:
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix="model_answer_collection_template_50_with_minesafe_",
        suffix=".xlsx",
        dir=str(OUTPUT_XLSX.parent),
        delete=False,
    ) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(tmp_path)
        tmp_path.replace(OUTPUT_XLSX)
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def main() -> None:
    report_lines: list[str] = [
        "MineSafe AI v2 50문항 답변 자동 수집 보고서",
        f"작성 시각: {now_text()}",
        "",
        "보존 원칙:",
        "- app.py 수정 없음",
        "- Vector DB 수정 없음",
        "- chunks 파일 수정 없음",
        "- 기존 09_answer_tests 및 12_compare_experiment 파일 수정 없음",
        "- 원본 model_answer_collection_template_50.xlsx 덮어쓰기 없음",
        "- chatgpt_answer, gemini_answer 컬럼 덮어쓰기 없음",
        "- .env 파일 열기/출력/수정 없음",
        "",
    ]

    if not INPUT_XLSX.exists():
        print(f"오류: 입력 파일을 찾을 수 없습니다: {INPUT_XLSX}")
        sys.exit(1)

    wb = load_workbook(INPUT_XLSX)
    ws = wb.active
    headers = header_map(ws)
    require_columns(headers)

    existing_answers = load_existing_minesafe_answers()
    if existing_answers:
        report_lines.append(f"- 기존 출력 파일에서 MineSafe AI 답변 {len(existing_answers)}개를 이어받았습니다.")

    question_rows = []
    chatgpt_count = 0
    gemini_count = 0
    minesafe_initial_count = 0
    reused_output_count = 0
    skipped_count = 0
    filled_count = 0
    generated_count = 0
    errors: list[str] = []

    for row_idx in range(2, ws.max_row + 1):
        if not row_is_question(ws, headers, row_idx):
            continue
        qid = clean_text(ws.cell(row_idx, headers["question_id"]).value)
        if clean_text(ws.cell(row_idx, headers["chatgpt_answer"]).value):
            chatgpt_count += 1
        if clean_text(ws.cell(row_idx, headers["gemini_answer"]).value):
            gemini_count += 1
        minesafe_value = clean_text(ws.cell(row_idx, headers["minesafe_ai_answer"]).value)
        if minesafe_value:
            minesafe_initial_count += 1
        elif qid in existing_answers:
            ws.cell(row_idx, headers["minesafe_ai_answer"]).value = existing_answers[qid]
            reused_output_count += 1
        question_rows.append(row_idx)

    retriever = Retriever(report_lines)
    start = time.time()

    for row_idx in question_rows:
        qid = clean_text(ws.cell(row_idx, headers["question_id"]).value)
        category = clean_text(ws.cell(row_idx, headers["category"]).value)
        difficulty = clean_text(ws.cell(row_idx, headers["difficulty"]).value)
        question = clean_text(ws.cell(row_idx, headers["question"]).value)
        current_answer = clean_text(ws.cell(row_idx, headers["minesafe_ai_answer"]).value)
        if current_answer:
            skipped_count += 1
            continue
        try:
            answer, mode = generate_answer(qid, category, difficulty, question, retriever, report_lines)
            ws.cell(row_idx, headers["minesafe_ai_answer"]).value = answer
            ws.cell(row_idx, headers["minesafe_ai_answer"]).alignment = Alignment(wrap_text=True, vertical="top")
            generated_count += 1
            report_lines.append(f"- {qid}: MineSafe AI 답변 생성 완료(mode={mode}, chars={len(answer)})")
        except Exception as exc:
            message = f"{qid}: {type(exc).__name__}: {exc}"
            errors.append(message)
            report_lines.append(f"- 오류: {message}")

    for row_idx in question_rows:
        if clean_text(ws.cell(row_idx, headers["minesafe_ai_answer"]).value):
            filled_count += 1

    for col_name in ["chatgpt_answer", "gemini_answer", "minesafe_ai_answer", "answer_collection_note"]:
        col_idx = headers[col_name]
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    save_workbook_safely(wb)

    elapsed = time.time() - start
    summary = [
        "",
        "요약:",
        f"- 전체 질문 수: {len(question_rows)}",
        f"- ChatGPT 답변 채움 수(보존): {chatgpt_count}",
        f"- Gemini 답변 채움 수(보존): {gemini_count}",
        f"- 실행 전 원본 MineSafe AI 답변 수: {minesafe_initial_count}",
        f"- 기존 출력 파일에서 이어받은 MineSafe AI 답변 수: {reused_output_count}",
        f"- 이번 실행에서 새로 생성한 MineSafe AI 답변 수: {generated_count}",
        f"- 이미 채워져 건너뛴 행 수: {skipped_count}",
        f"- 완료 후 MineSafe AI 답변 채움 수: {filled_count}/{len(question_rows)}",
        f"- 오류 수: {len(errors)}",
        f"- 실행 시간: {elapsed:.1f}초",
        f"- 출력 파일: {OUTPUT_XLSX}",
        f"- 보고서 파일: {REPORT_PATH}",
    ]
    report_lines.extend(summary)
    if errors:
        report_lines.append("")
        report_lines.append("오류 상세:")
        report_lines.extend(f"- {item}" for item in errors)

    REPORT_PATH.write_text("\n".join(report_lines) + "\n", encoding="utf-8")

    print("MineSafe AI 답변 자동 수집 완료")
    print(f"- 전체 질문 수: {len(question_rows)}")
    print(f"- 완료 후 minesafe_ai_answer 채움 수: {filled_count}/{len(question_rows)}")
    print(f"- 새로 생성한 답변 수: {generated_count}")
    print(f"- 오류 수: {len(errors)}")
    print(f"- 출력 파일: {OUTPUT_XLSX}")
    print(f"- 보고서: {REPORT_PATH}")


if __name__ == "__main__":
    main()
