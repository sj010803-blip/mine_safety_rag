from __future__ import annotations

import random
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("오류: openpyxl이 필요합니다. 프로젝트 가상환경에서 'pip install openpyxl' 후 다시 실행하세요.")
    sys.exit(1)

SEED = 20260707
BASE_DIR = Path(__file__).resolve().parent
INPUT_PATH = BASE_DIR / "model_answer_collection_template_50_with_minesafe_fixed.xlsx"
BLIND_OUTPUT_PATH = BASE_DIR / "blind_eval_questions_50_v2.xlsx"
KEY_OUTPUT_PATH = BASE_DIR / "blind_eval_answer_key_50_v2.xlsx"
REPORT_PATH = BASE_DIR / "create_blind_eval_50_v2_report.txt"

MODEL_COLUMNS = [
    ("ChatGPT", "chatgpt_answer"),
    ("Gemini", "gemini_answer"),
    ("MineSafe AI", "minesafe_ai_answer"),
]
OUTPUT_LABELS = ["Answer A", "Answer B", "Answer C"]
MODEL_NAME_PATTERNS = [
    (re.compile(r"MineSafe\s*AI\s*답변", re.IGNORECASE), "블라인드 답변"),
    (re.compile(r"MineSafe\s*AI", re.IGNORECASE), "블라인드 평가 대상"),
    (re.compile(r"ChatGPT", re.IGNORECASE), "블라인드 평가 대상"),
    (re.compile(r"Gemini", re.IGNORECASE), "블라인드 평가 대상"),
]


def normalize(value):
    return "" if value is None else str(value).strip()


def get_header_map(ws):
    return {normalize(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}


def require_columns(header_map, columns):
    missing = [col for col in columns if col not in header_map]
    if missing:
        raise ValueError("입력 파일에 필요한 컬럼이 없습니다: " + ", ".join(missing))


def redact_model_names(text: str) -> tuple[str, int]:
    redacted = normalize(text)
    count = 0
    for pattern, replacement in MODEL_NAME_PATTERNS:
        redacted, n = pattern.subn(replacement, redacted)
        count += n
    return redacted, count


def style_sheet(ws, widths: dict[str, int]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def count_answer_cells(ws) -> dict[str, int]:
    headers = get_header_map(ws)
    counts = {"rows": 0, "answer_A": 0, "answer_B": 0, "answer_C": 0}
    for row_idx in range(2, ws.max_row + 1):
        qid = normalize(ws.cell(row_idx, headers["question_id"]).value)
        if not qid:
            continue
        counts["rows"] += 1
        for col in ["answer_A", "answer_B", "answer_C"]:
            if normalize(ws.cell(row_idx, headers[col]).value):
                counts[col] += 1
    return counts


def find_model_name_leaks(ws) -> list[tuple[str, str, str]]:
    headers = get_header_map(ws)
    leaks: list[tuple[str, str, str]] = []
    model_terms = ["ChatGPT", "Gemini", "MineSafe AI"]
    for row_idx in range(2, ws.max_row + 1):
        qid = normalize(ws.cell(row_idx, headers["question_id"]).value)
        if not qid:
            continue
        for col in ["answer_A", "answer_B", "answer_C"]:
            value = normalize(ws.cell(row_idx, headers[col]).value)
            for term in model_terms:
                if term.lower() in value.lower():
                    leaks.append((qid, col, term))
    return leaks


def main():
    if not INPUT_PATH.exists():
        print(f"오류: 입력 파일을 찾을 수 없습니다: {INPUT_PATH}")
        sys.exit(1)

    wb = load_workbook(INPUT_PATH, data_only=False)
    ws = wb.active
    header_map = get_header_map(ws)
    base_columns = ["question_id", "category", "difficulty", "question"]
    answer_columns = [col for _, col in MODEL_COLUMNS]
    require_columns(header_map, base_columns + answer_columns)

    rows = []
    empty_cells = []
    redaction_count = 0
    source_answer_counts = Counter()

    for row_idx in range(2, ws.max_row + 1):
        question_id = normalize(ws.cell(row_idx, header_map["question_id"]).value)
        if not question_id:
            continue
        base = {
            "question_id": question_id,
            "category": normalize(ws.cell(row_idx, header_map["category"]).value),
            "difficulty": normalize(ws.cell(row_idx, header_map["difficulty"]).value),
            "question": normalize(ws.cell(row_idx, header_map["question"]).value),
        }
        answers = []
        for model_name, col_name in MODEL_COLUMNS:
            answer_text = normalize(ws.cell(row_idx, header_map[col_name]).value)
            if not answer_text:
                empty_cells.append(f"{question_id} / {col_name}")
            else:
                source_answer_counts[model_name] += 1
            redacted_text, redacted_n = redact_model_names(answer_text)
            redaction_count += redacted_n
            answers.append({"original_model": model_name, "answer_text": redacted_text})
        rows.append((base, answers))

    if empty_cells:
        print("오류: 아직 비어 있는 답변 칸이 있습니다.")
        print(f"{INPUT_PATH.name}에서 세 모델 답변을 모두 채운 뒤 다시 실행하세요.")
        for item in empty_cells[:20]:
            print(" - " + item)
        if len(empty_cells) > 20:
            print(f" ... 외 {len(empty_cells) - 20}개")
        sys.exit(1)

    rng = random.Random(SEED)
    blind_wb = Workbook()
    blind_ws = blind_wb.active
    blind_ws.title = "blind_eval"
    blind_ws.append(["question_id", "category", "difficulty", "question", "answer_A", "answer_B", "answer_C"])

    key_wb = Workbook()
    key_ws = key_wb.active
    key_ws.title = "answer_key"
    key_ws.append(["question_id", "answer_label", "original_model"])

    order_counter = Counter()
    for base, answers in rows:
        shuffled = answers[:]
        rng.shuffle(shuffled)
        order_signature = " | ".join(item["original_model"] for item in shuffled)
        order_counter[order_signature] += 1
        label_to_answer = dict(zip(OUTPUT_LABELS, shuffled))
        blind_ws.append([
            base["question_id"],
            base["category"],
            base["difficulty"],
            base["question"],
            label_to_answer["Answer A"]["answer_text"],
            label_to_answer["Answer B"]["answer_text"],
            label_to_answer["Answer C"]["answer_text"],
        ])
        for label in OUTPUT_LABELS:
            key_ws.append([base["question_id"], label, label_to_answer[label]["original_model"]])

    style_sheet(
        blind_ws,
        {"A": 12, "B": 28, "C": 12, "D": 70, "E": 55, "F": 55, "G": 55},
    )
    style_sheet(key_ws, {"A": 12, "B": 16, "C": 18})

    blind_wb.save(BLIND_OUTPUT_PATH)
    key_wb.save(KEY_OUTPUT_PATH)

    verify_blind = load_workbook(BLIND_OUTPUT_PATH, data_only=True)
    verify_ws = verify_blind.active
    answer_counts = count_answer_cells(verify_ws)
    leaks = find_model_name_leaks(verify_ws)
    randomized = len(order_counter) > 1

    report_lines = [
        "MineSafe AI v2 50문항 블라인드 평가 파일 생성 보고서",
        f"작성 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        f"입력 파일: {INPUT_PATH}",
        f"랜덤 seed: {SEED}",
        f"평가자 제공 파일: {BLIND_OUTPUT_PATH}",
        f"정답키 파일: {KEY_OUTPUT_PATH}",
        "주의: blind_eval_answer_key_50_v2.xlsx는 Answer A/B/C와 실제 모델명 매칭표이므로 평가자에게 보여주면 안 됩니다.",
        "",
        "입력 답변 개수:",
        f"- ChatGPT: {source_answer_counts['ChatGPT']}",
        f"- Gemini: {source_answer_counts['Gemini']}",
        f"- MineSafe AI: {source_answer_counts['MineSafe AI']}",
        "",
        "블라인드 파일 검증:",
        f"- 질문 수: {answer_counts['rows']}",
        f"- answer_A 채움 수: {answer_counts['answer_A']}",
        f"- answer_B 채움 수: {answer_counts['answer_B']}",
        f"- answer_C 채움 수: {answer_counts['answer_C']}",
        f"- 실제 모델명 redaction 건수: {redaction_count}",
        f"- 실제 모델명 노출 여부: {'노출 없음' if not leaks else '확인 필요'}",
        f"- 질문별 A/B/C 순서 랜덤화 여부: {'랜덤 섞임 확인' if randomized else '확인 필요'}",
        "",
        "A/B/C 실제 모델 순서 분포:",
    ]
    for order, count in sorted(order_counter.items()):
        report_lines.append(f"- {order}: {count}문항")
    if leaks:
        report_lines.append("")
        report_lines.append("모델명 노출 의심 셀:")
        for qid, col, term in leaks:
            report_lines.append(f"- {qid} / {col} / {term}")
    report_lines.extend([
        "",
        "보존 확인:",
        "- 최종 답변 수집 파일은 읽기만 했고 수정하지 않음",
        "- app.py 수정 없음",
        "- Vector DB 수정 없음",
        "- chunks 파일 수정 없음",
        "- 기존 09_answer_tests 평가 파일 수정 없음",
        "- 기존 12_compare_experiment 비교 실험 파일 수정 없음",
        "- 기존 94.90점 재계산 없음",
        "- .env 파일 열기/출력/수정 없음",
        "- API Key 출력 없음",
    ])
    REPORT_PATH.write_text("\n".join(report_lines) + "\n", encoding="utf-8")

    print("블라인드 평가 파일 생성 완료")
    print(f"- 입력 파일: {INPUT_PATH}")
    print(f"- 질문 수: {answer_counts['rows']}")
    print(f"- Answer A/B/C 채움: {answer_counts['answer_A']}/{answer_counts['answer_B']}/{answer_counts['answer_C']}")
    print(f"- 모델명 노출: {'없음' if not leaks else '확인 필요'}")
    print(f"- Answer key: {KEY_OUTPUT_PATH}")
    print("주의: blind_eval_answer_key_50_v2.xlsx는 평가자에게 보여주면 안 됩니다.")


if __name__ == "__main__":
    main()
