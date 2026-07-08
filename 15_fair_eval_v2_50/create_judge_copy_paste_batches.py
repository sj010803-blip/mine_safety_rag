from __future__ import annotations

import re
import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("오류: openpyxl이 필요합니다. 프로젝트 가상환경에서 'pip install openpyxl' 후 다시 실행하세요.")
    sys.exit(1)


PROJECT_ROOT = Path(r"C:\Users\USER\Desktop\mine_safety_rag")
PACKAGE_DIR = PROJECT_ROOT / "15_fair_eval_v2_50"
INPUT_XLSX = PACKAGE_DIR / "blind_eval_questions_50_v2.xlsx"
OUTPUT_DIR = PACKAGE_DIR / "judge_copy_paste_batches"
REPORT_PATH = OUTPUT_DIR / "create_judge_batches_report.txt"

JUDGES = ["GPT", "Gemini", "Claude"]
ANSWER_COLUMNS = ["answer_A", "answer_B", "answer_C"]
REQUIRED_COLUMNS = ["question_id", "category", "difficulty", "question", *ANSWER_COLUMNS]
ACTUAL_MODEL_TERMS = ["ChatGPT", "MineSafe AI"]


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def plain_text(value) -> str:
    text = clean(value)
    text = text.replace("\t", " ")
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text


def get_header_map(ws) -> dict[str, int]:
    return {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}


def load_blind_rows() -> list[dict[str, str]]:
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {INPUT_XLSX}")
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = get_header_map(ws)
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise ValueError("필수 컬럼이 없습니다: " + ", ".join(missing))

    rows: list[dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        qid = clean(ws.cell(row_idx, headers["question_id"]).value)
        if not qid:
            continue
        item = {col: plain_text(ws.cell(row_idx, headers[col]).value) for col in REQUIRED_COLUMNS}
        rows.append(item)
    rows.sort(key=lambda item: item["question_id"])
    return rows


def make_prompt(judge: str, batch_no: int, batch_rows: list[dict[str, str]]) -> str:
    first_qid = batch_rows[0]["question_id"]
    last_qid = batch_rows[-1]["question_id"]
    expected_lines = len(batch_rows) * 3
    example_line = f"{first_qid}\tAnswer A\t{judge}\t20\t18\t21\t19\t78\t근거는 있으나 현장 조치가 다소 일반적임"

    parts = [
        f"광산 안전관리 v2 블라인드 평가 - {judge} 평가자용 batch {batch_no:02d} ({first_qid}~{last_qid})",
        "",
        "역할",
        "- 당신은 광산 안전관리 답변 평가자이다.",
        "- 실제 답변 생성 모델명은 숨겨져 있다.",
        "- Answer A, Answer B, Answer C를 동일한 기준으로 평가하라.",
        "- 특정 답변이 어떤 모델인지 추정하지 마라.",
        "- 문장력보다 광산 안전관리 업무 적합성을 우선 평가하라.",
        "",
        "평가 기준",
        "- 검색 적합성: 25점",
        "- 근거 기반성: 25점",
        "- 안전·법령 판단 정확성: 25점",
        "- 실무성: 25점",
        "- 총점: 100점",
        "",
        "평가 시 중요하게 볼 내용",
        "- 공식 문서 근거, 출처 추적성, 현장 조치, 작업중지, 접근통제, 기록·보고, 작업재개 조건이 있으면 높게 평가할 수 있다.",
        "- 답변이 길기만 하고 조치가 불명확하면 높은 점수를 주지 마라.",
        "- 법령 조문을 확정할 수 없는 상황에서 무리하게 단정하지 않은 답변은 긍정적으로 평가할 수 있다.",
        "- 없는 법령이나 근거를 지어낸 답변은 낮게 평가하라.",
        "",
        "출력 지시",
        "- 반드시 TSV 형식으로만 출력하라.",
        "- 마크다운 표를 쓰지 마라.",
        "- 탭 문자로 컬럼을 구분하라.",
        "- 추가 설명 문장 없이 header 1줄 + 평가 결과만 출력하라.",
        f"- 이 batch는 {len(batch_rows)}개 질문 x Answer A/B/C = 총 {expected_lines}개 평가 행을 출력해야 한다.",
        f"- judge_model 컬럼은 모든 행에 반드시 {judge}로 입력하라.",
        "- total_score_100은 네 항목 점수의 합으로 입력하라.",
        "",
        "TSV 출력 header",
        "question_id\tanswer_label\tjudge_model\tsearch_relevance_25\tevidence_grounding_25\tsafety_legal_accuracy_25\tpracticality_25\ttotal_score_100\tcomment",
        "",
        "TSV 출력 예시",
        example_line,
        "",
        "평가 대상",
    ]

    for item in batch_rows:
        parts.extend(
            [
                "",
                f"## {item['question_id']}",
                f"category: {item['category']}",
                f"difficulty: {item['difficulty']}",
                f"question: {item['question']}",
                "",
                "[Answer A]",
                item["answer_A"],
                "",
                "[Answer B]",
                item["answer_B"],
                "",
                "[Answer C]",
                item["answer_C"],
            ]
        )
    parts.append("")
    return "\n".join(parts)


def write_batches(rows: list[dict[str, str]]) -> tuple[list[Path], dict[str, list[str]]]:
    if len(rows) != 50:
        raise ValueError(f"질문 수가 50개가 아닙니다: {len(rows)}")
    ranges: dict[str, list[str]] = {}
    written: list[Path] = []
    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
    for judge in JUDGES:
        (OUTPUT_DIR / judge).mkdir(parents=True, exist_ok=True)

    for batch_idx in range(10):
        batch_no = batch_idx + 1
        batch_rows = rows[batch_idx * 5 : (batch_idx + 1) * 5]
        first_qid = batch_rows[0]["question_id"]
        last_qid = batch_rows[-1]["question_id"]
        ranges[f"batch {batch_no:02d}"] = [first_qid, last_qid]
        for judge in JUDGES:
            prompt = make_prompt(judge, batch_no, batch_rows)
            output_path = OUTPUT_DIR / judge / f"{judge}_batch_{batch_no:02d}_{first_qid}_{last_qid}.txt"
            output_path.write_text(prompt, encoding="utf-8")
            written.append(output_path)
    return written, ranges


def check_files(written: list[Path]) -> dict[str, object]:
    counts_by_judge = Counter(path.parent.name for path in written)
    actual_model_leaks = []
    gemini_mentions = []
    for path in written:
        text = path.read_text(encoding="utf-8")
        for term in ACTUAL_MODEL_TERMS:
            if term.lower() in text.lower():
                actual_model_leaks.append((str(path), term))
        if "gemini" in text.lower():
            gemini_mentions.append(str(path))
    return {
        "counts_by_judge": counts_by_judge,
        "actual_model_leaks": actual_model_leaks,
        "gemini_mentions": gemini_mentions,
    }


def write_report(written: list[Path], ranges: dict[str, list[str]], checks: dict[str, object]) -> None:
    counts_by_judge: Counter = checks["counts_by_judge"]
    actual_model_leaks: list[tuple[str, str]] = checks["actual_model_leaks"]
    gemini_mentions: list[str] = checks["gemini_mentions"]

    lines = [
        "MineSafe AI v2 평가자 복붙용 batch txt 생성 보고서",
        f"작성 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        f"입력 파일: {INPUT_XLSX}",
        f"출력 폴더: {OUTPUT_DIR}",
        "",
        "1. 생성한 batch 파일 개수",
        f"- GPT: {counts_by_judge.get('GPT', 0)}개",
        f"- Gemini: {counts_by_judge.get('Gemini', 0)}개",
        f"- Claude: {counts_by_judge.get('Claude', 0)}개",
        "",
        "2. GPT/Gemini/Claude 각각 10개씩 생성되었는지",
        f"- {'예' if all(counts_by_judge.get(j, 0) == 10 for j in JUDGES) else '확인 필요'}",
        "",
        "3. 총 txt 파일 수",
        f"- {len(written)}개",
        "",
        "4. 각 batch의 질문 범위",
    ]
    for batch_name, (first_qid, last_qid) in ranges.items():
        lines.append(f"- {batch_name}: {first_qid}~{last_qid}")

    lines.extend(
        [
            "",
            "5. 실제 모델명 노출 여부 검사 결과",
            f"- 평가용 batch txt 내 ChatGPT/MineSafe AI 노출: {'없음' if not actual_model_leaks else '확인 필요'}",
            "- Gemini 문자열은 Gemini 평가자 파일의 judge_model 및 평가자명으로만 필요하게 포함될 수 있음",
            f"- Gemini 문자열 포함 파일 수: {len(gemini_mentions)}개",
        ]
    )
    if actual_model_leaks:
        lines.append("- 노출 의심 상세:")
        for path, term in actual_model_leaks:
            lines.append(f"  - {path}: {term}")

    lines.extend(
        [
            "",
            "6. 사용자가 이후 해야 할 일",
            "- GPT 폴더의 batch txt 10개를 차례로 열어 GPT 평가자에게 붙여넣고 TSV 결과를 저장한다.",
            "- Gemini 폴더의 batch txt 10개를 차례로 열어 Gemini 평가자에게 붙여넣고 TSV 결과를 저장한다.",
            "- Claude 폴더의 batch txt 10개를 차례로 열어 Claude 평가자에게 붙여넣고 TSV 결과를 저장한다.",
            "- 각 평가자가 출력한 TSV를 multi_judge_score_input_50_v2.xlsx에 붙여넣는다.",
            "- blind_eval_answer_key_50_v2.xlsx는 평가자에게 절대 보여주지 않는다.",
            "",
            "보존 확인",
            "- app.py 수정 없음",
            "- Vector DB 수정 없음",
            "- chunks 파일 수정 없음",
            "- 기존 평가 파일 수정 없음",
            "- 기존 비교 실험 파일 수정 없음",
            "- .env 파일 열기/출력/수정 없음",
            "- API Key 출력 없음",
        ]
    )
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = load_blind_rows()
    written, ranges = write_batches(rows)
    checks = check_files(written)
    write_report(written, ranges, checks)
    print("평가자 복붙용 batch txt 생성 완료")
    print(f"- 총 txt 파일 수: {len(written)}")
    for judge in JUDGES:
        print(f"- {judge}: {checks['counts_by_judge'].get(judge, 0)}개")
    print(f"- 보고서: {REPORT_PATH}")


if __name__ == "__main__":
    main()
