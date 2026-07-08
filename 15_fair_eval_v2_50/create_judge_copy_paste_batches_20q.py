from __future__ import annotations

import re
import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("오류: openpyxl이 필요합니다. 프로젝트 가상환경에서 'pip install openpyxl' 후 다시 실행하세요.")
    sys.exit(1)


PROJECT_ROOT = Path(r"C:\Users\USER\Desktop\mine_safety_rag")
PACKAGE_DIR = PROJECT_ROOT / "15_fair_eval_v2_50"
INPUT_XLSX = PACKAGE_DIR / "blind_eval_questions_50_v2.xlsx"
OUTPUT_DIR = PACKAGE_DIR / "judge_copy_paste_batches_20q"
REPORT_PATH = OUTPUT_DIR / "create_judge_batches_20q_report.txt"

JUDGES = ["GPT", "Gemini", "Claude"]
ANSWER_COLUMNS = ["answer_A", "answer_B", "answer_C"]
REQUIRED_COLUMNS = ["question_id", "category", "difficulty", "question", *ANSWER_COLUMNS]
ACTUAL_MODEL_TERMS = ["ChatGPT", "MineSafe AI"]
BATCH_RANGES = [
    ("batch 01", "FQ001", "FQ020"),
    ("batch 02", "FQ021", "FQ040"),
    ("batch 03", "FQ041", "FQ050"),
]


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def plain_text(value) -> str:
    text = clean(value)
    text = text.replace("\t", " ")
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text


def get_header_map(ws) -> dict[str, int]:
    return {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}


def load_rows() -> list[dict[str, str]]:
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {INPUT_XLSX}")
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = get_header_map(ws)
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise ValueError("필수 컬럼이 없습니다: " + ", ".join(missing))
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        qid = clean(ws.cell(row_idx, headers["question_id"]).value)
        if not qid:
            continue
        rows.append({col: plain_text(ws.cell(row_idx, headers[col]).value) for col in REQUIRED_COLUMNS})
    rows.sort(key=lambda item: item["question_id"])
    return rows


def row_number(qid: str) -> int:
    return int(qid.replace("FQ", ""))


def select_batch(rows: list[dict[str, str]], start_qid: str, end_qid: str) -> list[dict[str, str]]:
    start = row_number(start_qid)
    end = row_number(end_qid)
    return [row for row in rows if start <= row_number(row["question_id"]) <= end]


def make_prompt(judge: str, batch_no: int, start_qid: str, end_qid: str, batch_rows: list[dict[str, str]]) -> str:
    expected_lines = len(batch_rows) * 3
    example_line = f"{start_qid}\tAnswer A\t{judge}\t20\t18\t21\t19\t78\t근거는 있으나 현장 조치가 다소 일반적임"
    parts = [
        f"광산 안전관리 v2 블라인드 평가 - {judge} 평가자용 20문항 batch {batch_no:02d} ({start_qid}~{end_qid})",
        "",
        "역할",
        "- 당신은 광산 안전관리 답변 평가자이다.",
        "- 실제 답변 생성 모델명은 숨겨져 있다.",
        "- Answer A, Answer B, Answer C를 동일한 기준으로 평가하라.",
        "- 특정 답변이 어떤 모델인지 추정하지 마라.",
        "- 문장력보다 광산 안전관리 업무 적합성을 우선 평가하라.",
        "",
        "평가 기준",
        "- 검색 적합성: 25점",
        "- 근거 기반성: 25점",
        "- 안전·법령 판단 정확성: 25점",
        "- 실무성: 25점",
        "- 총점: 100점",
        "",
        "평가 시 중요하게 볼 내용",
        "- 공식 문서 근거, 출처 추적성, 현장 조치, 작업중지, 접근통제, 기록·보고, 작업재개 조건이 있으면 높게 평가할 수 있다.",
        "- 답변이 길기만 하고 조치가 불명확하면 높은 점수를 주지 마라.",
        "- 법령 조문을 확정할 수 없는 상황에서 무리하게 단정하지 않은 답변은 긍정적으로 평가할 수 있다.",
        "- 없는 법령이나 근거를 지어낸 답변은 낮게 평가하라.",
        "",
        "출력 지시",
        "- 반드시 TSV 형식으로만 출력하라.",
        "- 마크다운 표를 쓰지 마라.",
        "- 탭 문자로 컬럼을 구분하라.",
        "- 설명 문장 없이 TSV header 1줄 + TSV 결과만 출력하라.",
        "- comment는 1문장으로 짧게 작성하라.",
        f"- 이 batch는 {len(batch_rows)}개 질문 x Answer A/B/C = 총 {expected_lines}개 평가 행을 출력해야 한다.",
        f"- judge_model 컬럼은 모든 행에 반드시 {judge}로 입력하라.",
        "- total_score_100은 네 항목 점수의 합으로 입력하라.",
        "",
        "TSV 출력 header",
        "question_id\tanswer_label\tjudge_model\tsearch_relevance_25\tevidence_grounding_25\tsafety_legal_accuracy_25\tpracticality_25\ttotal_score_100\tcomment",
        "",
        "TSV 출력 예시",
        example_line,
        "",
        "평가 대상",
    ]
    for item in batch_rows:
        parts.extend(
            [
                "",
                f"## {item['question_id']}",
                f"category: {item['category']}",
                f"difficulty: {item['difficulty']}",
                f"question: {item['question']}",
                "",
                "[Answer A]",
                item["answer_A"],
                "",
                "[Answer B]",
                item["answer_B"],
                "",
                "[Answer C]",
                item["answer_C"],
            ]
        )
    parts.append("")
    return "\n".join(parts)


def write_batches(rows: list[dict[str, str]]) -> tuple[list[Path], dict[str, dict[str, int | str]]]:
    if len(rows) != 50:
        raise ValueError(f"질문 수가 50개가 아닙니다: {len(rows)}")
    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
    for judge in JUDGES:
        (OUTPUT_DIR / judge).mkdir(parents=True, exist_ok=True)

    written = []
    batch_info: dict[str, dict[str, int | str]] = {}
    for batch_index, (_label, start_qid, end_qid) in enumerate(BATCH_RANGES, start=1):
        batch_rows = select_batch(rows, start_qid, end_qid)
        batch_info[f"batch {batch_index:02d}"] = {
            "start": start_qid,
            "end": end_qid,
            "question_count": len(batch_rows),
        }
        for judge in JUDGES:
            text = make_prompt(judge, batch_index, start_qid, end_qid, batch_rows)
            output_path = OUTPUT_DIR / judge / f"{judge}_batch_{batch_index:02d}_{start_qid}_{end_qid}.txt"
            output_path.write_text(text, encoding="utf-8")
            written.append(output_path)
    return written, batch_info


def check_written_files(written: list[Path]) -> dict[str, object]:
    counts_by_judge = Counter(path.parent.name for path in written)
    actual_model_leaks = []
    question_counts = {}
    answer_block_issues = []
    header = "question_id\tanswer_label\tjudge_model\tsearch_relevance_25\tevidence_grounding_25\tsafety_legal_accuracy_25\tpracticality_25\ttotal_score_100\tcomment"
    missing_header = []
    for path in written:
        text = path.read_text(encoding="utf-8")
        for term in ACTUAL_MODEL_TERMS:
            if term.lower() in text.lower():
                actual_model_leaks.append((str(path), term))
        if header not in text:
            missing_header.append(str(path))
        qids = re.findall(r"^## (FQ\d{3})$", text, flags=re.MULTILINE)
        question_counts[str(path)] = len(qids)
        for qid in qids:
            section_match = re.search(rf"^## {qid}\n(.*?)(?=^## FQ\d{{3}}|\Z)", text, flags=re.MULTILINE | re.DOTALL)
            section = section_match.group(1) if section_match else ""
            for label in ["[Answer A]", "[Answer B]", "[Answer C]"]:
                if label not in section:
                    answer_block_issues.append((str(path), qid, label))
    return {
        "counts_by_judge": counts_by_judge,
        "actual_model_leaks": actual_model_leaks,
        "question_counts": question_counts,
        "answer_block_issues": answer_block_issues,
        "missing_header": missing_header,
    }


def write_report(written: list[Path], batch_info: dict[str, dict[str, int | str]], checks: dict[str, object]) -> None:
    counts_by_judge: Counter = checks["counts_by_judge"]
    actual_model_leaks: list[tuple[str, str]] = checks["actual_model_leaks"]
    question_counts: dict[str, int] = checks["question_counts"]
    answer_block_issues: list[tuple[str, str, str]] = checks["answer_block_issues"]
    missing_header: list[str] = checks["missing_header"]

    lines = [
        "MineSafe AI v2 20문항 평가자 복붙용 batch txt 생성 보고서",
        f"작성 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "1. 생성한 파일 목록",
    ]
    for path in sorted(written):
        lines.append(f"- {path}")
    lines.extend(
        [
            "",
            "2. 총 batch 파일 수",
            f"- {len(written)}개",
            "",
            "3. GPT/Gemini/Claude별 파일 수",
            f"- GPT: {counts_by_judge.get('GPT', 0)}개",
            f"- Gemini: {counts_by_judge.get('Gemini', 0)}개",
            f"- Claude: {counts_by_judge.get('Claude', 0)}개",
            "",
            "4. batch별 질문 범위",
        ]
    )
    for name, info in batch_info.items():
        lines.append(f"- {name}: {info['start']}~{info['end']}")
    lines.append("")
    lines.append("5. 각 batch 질문 수")
    for name, info in batch_info.items():
        lines.append(f"- {name}: {info['question_count']}개")
    lines.extend(
        [
            "",
            "6. 실제 모델명 노출 검사 결과",
            f"- 평가용 txt 내 ChatGPT/MineSafe AI 노출: {'없음' if not actual_model_leaks else '확인 필요'}",
            "- Gemini 문자열은 Gemini 평가자 파일의 judge_model 및 평가자명으로만 필요한 경우 허용",
            f"- TSV header 누락 파일: {len(missing_header)}개",
            f"- Answer A/B/C 누락 이슈: {len(answer_block_issues)}개",
        ]
    )
    if actual_model_leaks:
        lines.append("- 노출 의심 상세:")
        for path, term in actual_model_leaks:
            lines.append(f"  - {path}: {term}")
    if missing_header:
        lines.append("- TSV header 누락 상세:")
        for path in missing_header:
            lines.append(f"  - {path}")
    if answer_block_issues:
        lines.append("- Answer block 누락 상세:")
        for path, qid, label in answer_block_issues:
            lines.append(f"  - {path}: {qid} {label}")
    lines.extend(
        [
            "",
            "7. 기존 파일 보존 여부",
            "- 기존 5문항 batch 폴더 삭제/수정 없음",
            "- 기존 10문항 batch가 있더라도 삭제/수정 없음",
            "- app.py 수정 없음",
            "- Vector DB 수정 없음",
            "- chunks 파일 수정 없음",
            "- 기존 평가 파일 수정 없음",
            "- 기존 비교 실험 파일 수정 없음",
            "- .env 파일 열기/출력/수정 없음",
            "- API Key 출력 없음",
            "",
            "8. 사용자가 이후 해야 할 일",
            "- GPT/Gemini/Claude 폴더에서 각 평가자용 20문항 batch txt를 연다.",
            "- 파일 전체를 Ctrl+A, Ctrl+C로 복사해 해당 평가 모델에 붙여넣는다.",
            "- 평가자가 출력한 TSV를 multi_judge_score_input_50_v2.xlsx에 붙여넣는다.",
            "- blind_eval_answer_key_50_v2.xlsx는 평가자에게 절대 보여주지 않는다.",
        ]
    )
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = load_rows()
    written, batch_info = write_batches(rows)
    checks = check_written_files(written)
    write_report(written, batch_info, checks)
    print("20문항 평가자 복붙용 batch txt 생성 완료")
    print(f"- 총 평가용 txt 파일 수: {len(written)}")
    for judge in JUDGES:
        print(f"- {judge}: {checks['counts_by_judge'].get(judge, 0)}개")
    print(f"- 보고서: {REPORT_PATH}")


if __name__ == "__main__":
    main()
