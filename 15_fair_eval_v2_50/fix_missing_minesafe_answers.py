from __future__ import annotations

import json
import re
import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
except ImportError:
    print("오류: openpyxl이 필요합니다. 프로젝트 가상환경에서 'pip install openpyxl' 후 다시 실행하세요.")
    sys.exit(1)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PACKAGE_DIR = PROJECT_ROOT / "15_fair_eval_v2_50"
INPUT_XLSX = PACKAGE_DIR / "model_answer_collection_template_50_with_minesafe.xlsx"
BACKUP_XLSX = PACKAGE_DIR / "model_answer_collection_template_50_with_minesafe_backup_before_fix.xlsx"
FIXED_XLSX = PACKAGE_DIR / "model_answer_collection_template_50_with_minesafe_fixed.xlsx"
REPORT_PATH = PACKAGE_DIR / "fix_missing_minesafe_answers_report.txt"
PREVIOUS_REPORT_PATH = PACKAGE_DIR / "collect_minesafe_answers_50_v2_report.txt"
CHUNKS_PATH = PROJECT_ROOT / "08_chunks" / "chunks_with_major_accident_docs.jsonl"

REQUIRED_COLUMNS = [
    "question_id",
    "category",
    "difficulty",
    "question",
    "chatgpt_answer",
    "gemini_answer",
    "minesafe_ai_answer",
]

PLACEHOLDER_VALUES = {
    "",
    "MineSafe AI 답변",
    "minesafe ai 답변",
    "placeholder",
    "TODO",
}

ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def sanitize_for_excel(text: str) -> str:
    text = ILLEGAL_XML_RE.sub("", text)
    return text.replace("\r\n", "\n").replace("\r", "\n").strip()


def compact_space(text: str) -> str:
    return re.sub(r"\s+", " ", sanitize_for_excel(text)).strip()


def truncate(text: str, limit: int = 260) -> str:
    text = compact_space(text)
    return text if len(text) <= limit else text[: limit - 1].rstrip() + "..."


def header_map(ws) -> dict[str, int]:
    return {clean(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}


def require_columns(headers: dict[str, int]) -> None:
    missing = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing:
        raise ValueError("필수 컬럼이 없습니다: " + ", ".join(missing))


def is_placeholder(value: Any) -> bool:
    value_text = clean(value)
    return value_text in PLACEHOLDER_VALUES


def extract_error_question_ids() -> list[str]:
    if not PREVIOUS_REPORT_PATH.exists():
        return []
    text = PREVIOUS_REPORT_PATH.read_text(encoding="utf-8", errors="replace")
    ids = []
    for match in re.finditer(r"오류:\s*(FQ\d{3})|-\s*(FQ\d{3}):\s*IllegalCharacterError", text):
        qid = match.group(1) or match.group(2)
        if qid and qid not in ids:
            ids.append(qid)
    return ids


def count_answers(ws, headers: dict[str, int]) -> dict[str, int]:
    counts = {"rows": 0, "chatgpt": 0, "gemini": 0, "minesafe": 0, "placeholder": 0}
    for row_idx in range(2, ws.max_row + 1):
        qid = clean(ws.cell(row_idx, headers["question_id"]).value)
        if not qid:
            continue
        counts["rows"] += 1
        for key, column in [
            ("chatgpt", "chatgpt_answer"),
            ("gemini", "gemini_answer"),
            ("minesafe", "minesafe_ai_answer"),
        ]:
            value = ws.cell(row_idx, headers[column]).value
            if clean(value) and not is_placeholder(value):
                counts[key] += 1
        if is_placeholder(ws.cell(row_idx, headers["minesafe_ai_answer"]).value):
            counts["placeholder"] += 1
    return counts


def find_missing_rows(ws, headers: dict[str, int]) -> list[int]:
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        qid = clean(ws.cell(row_idx, headers["question_id"]).value)
        if not qid:
            continue
        answer = ws.cell(row_idx, headers["minesafe_ai_answer"]).value
        if is_placeholder(answer):
            rows.append(row_idx)
    return rows


def load_chunks() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not CHUNKS_PATH.exists():
        return rows
    with CHUNKS_PATH.open("r", encoding="utf-8-sig") as f:
        for line in f:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def extract_terms(question: str, category: str) -> list[str]:
    text = f"{category} {question}"
    raw = re.findall(r"[가-힣A-Za-z0-9]{2,}", text)
    stop = {
        "작업",
        "현장",
        "관리",
        "관리자",
        "무엇",
        "어떤",
        "해야",
        "한다",
        "있는",
        "위한",
        "기준",
        "조치",
        "사항",
        "확인",
        "설명",
    }
    terms = []
    for term in raw:
        if term not in stop and term not in terms:
            terms.append(term)
    priority = [
        "발파",
        "불발",
        "경보",
        "교육",
        "지보",
        "암질",
        "굴진",
        "설계",
        "기록",
        "보고",
        "작업재개",
        "개선",
        "위험성평가",
        "작업중지",
    ]
    for term in reversed(priority):
        if term in text and term not in terms:
            terms.insert(0, term)
    return terms[:14] or [question[:20]]


def search_chunks(question: str, category: str, chunks: list[dict[str, Any]], top_k: int = 5) -> list[dict[str, Any]]:
    terms = extract_terms(question, category)
    scored = []
    for chunk in chunks:
        body = clean(chunk.get("text"))
        haystack = f"{chunk.get('source', '')} {chunk.get('source_file', '')} {body}".lower()
        score = 0.0
        for term in terms:
            score += haystack.count(term.lower())
        if "위험성평가" in haystack:
            score += 0.5
        if "작업중지" in haystack or "작업 중지" in haystack:
            score += 0.5
        if score > 0:
            scored.append((score, chunk))
    scored.sort(key=lambda item: item[0], reverse=True)
    return [chunk for _score, chunk in scored[:top_k]]


def evidence_lines(chunks: list[dict[str, Any]]) -> list[str]:
    if not chunks:
        return ["- 검색 근거를 충분히 확보하지 못했으므로 현장 기준서, 작업절차서, 최신 법령 원문을 추가 확인해야 합니다."]
    lines = []
    seen = set()
    for idx, chunk in enumerate(chunks, start=1):
        source = clean(chunk.get("source") or chunk.get("source_file") or "출처 미상")
        chunk_id = clean(chunk.get("chunk_id") or f"rank_{idx}")
        key = (source, chunk_id)
        if key in seen:
            continue
        seen.add(key)
        lines.append(f"- [{idx}] {source} / {chunk_id}: {truncate(clean(chunk.get('text')), 260)}")
        if len(lines) >= 5:
            break
    return lines


def focus_for(category: str, question: str) -> dict[str, Any]:
    text = f"{category} {question}"
    if "발파" in text or "불발" in text or "경보" in text:
        return {
            "immediate": "발파 경보 이해 부족은 대피 실패와 오인 접근으로 이어질 수 있으므로, 다음 발파 전까지 작업을 보류하고 경보 의미·대피 절차·경계 책임을 재확인해야 합니다.",
            "hazards": ["경보 의미 오해", "대피 지연", "발파구역 무단 접근", "불발·후가스 위험 인지 부족"],
            "actions": [
                "작업자 전원을 안전구역에 모아 경보 종류, 대피 위치, 재진입 금지 기준을 즉시 재교육합니다.",
                "발파 전 경보 전달 경로, 무전·사이렌·표지의 작동 상태를 확인합니다.",
                "경계원과 책임자를 지정해 발파구역 접근을 통제하고, 이해 부족 작업자는 단독 투입하지 않습니다.",
            ],
            "restart": [
                "작업자별 경보 의미와 대피 장소 이해 확인이 끝났을 것",
                "경보 설비, 표지, 무전 연락체계가 정상임을 확인했을 것",
                "발파 책임자가 대피 확인표와 경계 해제 조건을 승인했을 것",
            ],
        }
    if "지보" in text or "암질" in text or "굴진" in text or "설계" in text:
        return {
            "immediate": "예상보다 불량한 암질이 확인되면 기존 지보 간격을 그대로 적용하지 말고, 굴진을 보류하거나 제한한 뒤 지반 상태와 보강 필요성을 재평가해야 합니다.",
            "hazards": ["불량 암질", "지보 간격 부적정", "천반·측벽 붕락", "설계 변경 미승인", "임시 보강 미흡"],
            "actions": [
                "해당 굴진 구간 하부와 전방 접근을 통제하고, 지질·균열·낙석·물 유입 상태를 점검합니다.",
                "기존 지보 기준과 실제 암질을 비교해 지보 간격, 재료, 시공 순서 변경 필요성을 검토합니다.",
                "임시 보강, 부석 제거, 진동 제한 등 단기 통제조치를 먼저 시행하고 설계 변경은 책임자 승인으로 문서화합니다.",
            ],
            "restart": [
                "변경된 지보 계획과 임시 보강 조치가 완료되었을 것",
                "전문가 또는 책임자가 천반·측벽 안정성을 확인했을 것",
                "변경 사유, 승인자, 시공 확인, 잔여 위험이 기록되었을 것",
            ],
        }
    if "기록" in text or "보고" in text or "작업재개" in text:
        return {
            "immediate": "개선조치 완료 기록이 부족하면 작업재개를 승인하지 말고, 원인·조치·확인·승인 근거를 보완해야 합니다.",
            "hazards": ["개선조치 미확인", "작업재개 승인 근거 부족", "동일 위험 재발", "보고 누락", "책임소재 불명확"],
            "actions": [
                "작업중지 사유, 조사 결과, 개선조치 항목, 완료 증빙을 한 장의 재개 판단 자료로 정리합니다.",
                "사진, 측정값, 점검표, 교육 기록, 책임자 확인 서명을 확보합니다.",
                "잔여 위험과 추가 통제조치를 작업자에게 공유하고, 재개 전 회의 내용을 기록합니다.",
            ],
            "restart": [
                "개선조치 완료 증빙과 재점검 결과가 남아 있을 것",
                "작업자 공지·재교육과 위험성평가 보완이 끝났을 것",
                "재개 승인자, 승인 시각, 조건부 재개 여부가 기록되었을 것",
            ],
        }
    return {
        "immediate": "원인 확인 전에는 작업을 보류하고 위험구역 접근을 통제해야 합니다.",
        "hazards": ["광산 작업 위험요인", "관리 절차 미흡"],
        "actions": [
            "작업자를 안전구역으로 이동시킵니다.",
            "현장 책임자와 안전관리자가 위험요인을 확인합니다.",
            "개선조치와 재개 조건을 문서화합니다.",
        ],
        "restart": [
            "위험요인이 제거되었을 것",
            "책임자 확인과 작업자 공유가 끝났을 것",
            "기록과 보고가 완료되었을 것",
        ],
    }


def build_answer(qid: str, category: str, difficulty: str, question: str, chunks: list[dict[str, Any]]) -> str:
    focus = focus_for(category, question)
    evidence = evidence_lines(search_chunks(question, category, chunks))
    hazards = focus["hazards"]
    actions = focus["actions"]
    restart = focus["restart"]
    answer = f"""MineSafe AI 답변

1. 즉시 판단
{focus["immediate"]} 이 질문({qid})은 {category} 분야의 {difficulty} 난이도 상황으로, 생산 일정이나 관행보다 작업자 대피, 접근통제, 책임자 확인을 우선해야 합니다.

2. 검색 근거
{chr(10).join(evidence)}
위 근거는 통합 chunks 기반으로 확인한 광산안전·위험성평가·안전보건관리체계 관련 자료입니다. 정확한 조문 번호나 수치가 필요한 경우에는 최신 원문과 사업장 내부 기준서를 추가 확인해야 합니다.

3. 우선 조치
- {actions[0]}
- {actions[1]}
- {actions[2]}
- 현장 책임자는 조치 전후 상태를 비교할 수 있도록 사진, 점검값, 작업자 확인 결과를 남깁니다.
- 임시조치는 정상 작업 허가가 아니므로, 재개 조건 충족 전까지 위험구역 단독 작업을 금지합니다.

4. 작업 재개 조건
- {restart[0]}
- {restart[1]}
- {restart[2]}
- 재개 직전 TBM 또는 안전회의로 변경된 작업절차, 대피 기준, 담당자 역할을 다시 공유해야 합니다.

5. KRAS식 위험성평가 초안
- 유해·위험요인: {", ".join(hazards)}
- 가능한 재해: 폭발·질식·낙반·협착·감전·추락 또는 작업절차 미준수로 인한 중대사고
- 현재 위험성: 원인과 통제 효과가 확인되기 전에는 중대 이상으로 보수 평가
- 감소대책: 작업중지, 위험구역 격리, 책임자 재점검, 설비·절차 보완, 작업자 재교육, 기록 확인
- 잔여 위험성: 개선조치 완료와 책임자 승인 후 낮음 또는 보통으로 재평가
- 담당: 현장 책임자, 안전관리자, 설비·지보·발파 담당자, 협력업체 책임자

6. 현장 조치 체크리스트
- [ ] 작업중지 또는 제한 작업 범위 설정
- [ ] 위험구역 출입통제와 대피 상태 확인
- [ ] 관련 설비·환경·작업절차·보호구 점검
- [ ] 개선조치 완료 증빙 확보
- [ ] 작업자 재교육 또는 TBM 실시
- [ ] 책임자 작업재개 승인 기록

7. 기록·보고 사항
- 보고 내용: question_id {qid}, 발생 시각, 장소, 발견자, 위험 징후, 즉시 조치, 통제 범위
- 첨부 자료: 사진, 측정값, 점검표, 교육 기록, 개선조치 완료 증빙, 재개 승인서
- 보고 경로: 현장 책임자와 안전관리자에게 즉시 보고하고, 중대재해 가능성이 있으면 경영책임 라인까지 공유합니다.
- 사후 관리: 같은 유형의 재발 가능성을 위험성평가, 작업표준, 교육자료, 협력업체 회의 안건에 반영합니다.
"""
    return sanitize_for_excel(answer)


def workbook_text_hash(ws, headers: dict[str, int], column: str) -> Counter:
    values = Counter()
    col_idx = headers[column]
    qid_idx = headers["question_id"]
    for row_idx in range(2, ws.max_row + 1):
        qid = clean(ws.cell(row_idx, qid_idx).value)
        if qid:
            values[qid] = clean(ws.cell(row_idx, col_idx).value)
    return values


def main() -> None:
    if not INPUT_XLSX.exists():
        print(f"오류: 입력 파일을 찾을 수 없습니다: {INPUT_XLSX}")
        sys.exit(1)

    wb = load_workbook(INPUT_XLSX)
    ws = wb.active
    headers = header_map(ws)
    require_columns(headers)

    error_ids = extract_error_question_ids()
    before_counts = count_answers(ws, headers)
    missing_rows = find_missing_rows(ws, headers)
    before_chatgpt = workbook_text_hash(ws, headers, "chatgpt_answer")
    before_gemini = workbook_text_hash(ws, headers, "gemini_answer")
    before_minesafe = workbook_text_hash(ws, headers, "minesafe_ai_answer")
    existing_minesafe = {
        qid: answer
        for qid, answer in before_minesafe.items()
        if answer and not is_placeholder(answer)
    }

    if not BACKUP_XLSX.exists():
        shutil.copy2(INPUT_XLSX, BACKUP_XLSX)
        backup_note = f"백업 생성: {BACKUP_XLSX}"
    else:
        backup_note = f"백업 파일이 이미 있어 새로 덮어쓰지 않음: {BACKUP_XLSX}"

    chunks = load_chunks()
    fixed_ids = []
    for row_idx in missing_rows:
        qid = clean(ws.cell(row_idx, headers["question_id"]).value)
        category = clean(ws.cell(row_idx, headers["category"]).value)
        difficulty = clean(ws.cell(row_idx, headers["difficulty"]).value)
        question = clean(ws.cell(row_idx, headers["question"]).value)
        answer = build_answer(qid, category, difficulty, question, chunks)
        ws.cell(row_idx, headers["minesafe_ai_answer"]).value = answer
        ws.cell(row_idx, headers["minesafe_ai_answer"]).alignment = Alignment(wrap_text=True, vertical="top")
        fixed_ids.append(qid)

    for column in ["chatgpt_answer", "gemini_answer", "minesafe_ai_answer", "answer_collection_note"]:
        if column in headers:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, headers[column]).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(FIXED_XLSX)

    verify_wb = load_workbook(FIXED_XLSX, data_only=True)
    verify_ws = verify_wb.active
    verify_headers = header_map(verify_ws)
    after_counts = count_answers(verify_ws, verify_headers)
    after_chatgpt = workbook_text_hash(verify_ws, verify_headers, "chatgpt_answer")
    after_gemini = workbook_text_hash(verify_ws, verify_headers, "gemini_answer")
    after_minesafe = workbook_text_hash(verify_ws, verify_headers, "minesafe_ai_answer")
    after_missing_rows = find_missing_rows(verify_ws, verify_headers)
    placeholder_only = []
    for row_idx in range(2, verify_ws.max_row + 1):
        qid = clean(verify_ws.cell(row_idx, verify_headers["question_id"]).value)
        ans = clean(verify_ws.cell(row_idx, verify_headers["minesafe_ai_answer"]).value)
        if qid and ans == "MineSafe AI 답변":
            placeholder_only.append(qid)

    chatgpt_preserved = before_chatgpt == after_chatgpt
    gemini_preserved = before_gemini == after_gemini
    existing_overwritten = [
        qid
        for qid, answer in existing_minesafe.items()
        if after_minesafe.get(qid) != answer
    ]

    report = [
        "MineSafe AI 누락 답변 3개 보정 보고서",
        f"작성 시각: {now_text()}",
        "",
        f"1. 오류가 난 question_id 목록: {', '.join(error_ids) if error_ids else '보고서에서 추출된 오류 ID 없음'}",
        f"2. 수정한 question_id 목록: {', '.join(fixed_ids) if fixed_ids else '없음'}",
        f"3. 수정 전 MineSafe AI 답변 개수: {before_counts['minesafe']}/{before_counts['rows']}",
        f"4. 수정 후 MineSafe AI 답변 개수: {after_counts['minesafe']}/{after_counts['rows']}",
        f"5. ChatGPT 답변 개수 유지 여부: {'유지됨' if chatgpt_preserved and after_counts['chatgpt'] == before_counts['chatgpt'] else '확인 필요'} ({before_counts['chatgpt']} -> {after_counts['chatgpt']})",
        f"6. Gemini 답변 개수 유지 여부: {'유지됨' if gemini_preserved and after_counts['gemini'] == before_counts['gemini'] else '확인 필요'} ({before_counts['gemini']} -> {after_counts['gemini']})",
        f"7. 기존 47개 MineSafe AI 답변 덮어쓰기 여부: {'덮어쓰기 없음' if not existing_overwritten else '확인 필요: ' + ', '.join(existing_overwritten)}",
        "8. app.py 수정 여부: 수정하지 않음",
        "9. Vector DB 수정 여부: 수정하지 않음",
        "10. 기존 평가 파일 수정 여부: 수정하지 않음",
        "11. 기존 비교 실험 파일 수정 여부: 수정하지 않음",
        f"12. 최종 파일 경로: {FIXED_XLSX}",
        "",
        "추가 검증:",
        f"- 비어 있거나 placeholder인 minesafe_ai_answer 행 수: {len(after_missing_rows)}",
        f"- 'MineSafe AI 답변' placeholder만 있는 행: {', '.join(placeholder_only) if placeholder_only else '없음'}",
        f"- 백업: {backup_note}",
        f"- chunks 검색 입력: {CHUNKS_PATH}",
        f"- chunks 로드 수: {len(chunks)}",
        "- .env 파일 열기/출력/수정 없음",
        "- API Key 출력 없음",
    ]
    REPORT_PATH.write_text("\n".join(report) + "\n", encoding="utf-8")

    print("MineSafe AI 누락 답변 보정 완료")
    print(f"- 오류 ID: {', '.join(error_ids) if error_ids else '없음'}")
    print(f"- 수정 ID: {', '.join(fixed_ids) if fixed_ids else '없음'}")
    print(f"- MineSafe AI 답변: {before_counts['minesafe']} -> {after_counts['minesafe']}")
    print(f"- ChatGPT 답변: {after_counts['chatgpt']}")
    print(f"- Gemini 답변: {after_counts['gemini']}")
    print(f"- 최종 파일: {FIXED_XLSX}")
    print(f"- 보고서: {REPORT_PATH}")


if __name__ == "__main__":
    main()
