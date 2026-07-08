from __future__ import annotations

import math
import statistics
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("오류: openpyxl이 필요합니다. 프로젝트 가상환경에서 'pip install openpyxl' 후 다시 실행하세요.")
    sys.exit(1)

BASE_DIR = Path(__file__).resolve().parent
SCORE_INPUT_PATH = BASE_DIR / "multi_judge_score_input_50_v2.xlsx"
ANSWER_KEY_PATH = BASE_DIR / "blind_eval_answer_key_50_v2.xlsx"
OUTPUT_PATH = BASE_DIR / "multi_judge_score_summary_50_v2.xlsx"
JUDGES = ["GPT", "Gemini", "Claude"]


def normalize(value):
    return "" if value is None else str(value).strip()


def to_float(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(value)
    except ValueError:
        return None


def header_map(ws):
    return {normalize(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}


def require_columns(found, required, path):
    missing = [c for c in required if c not in found]
    if missing:
        raise ValueError(f"{path.name}에 필요한 컬럼이 없습니다: {', '.join(missing)}")


def main():
    if not SCORE_INPUT_PATH.exists():
        print(f"오류: 점수 입력 파일을 찾을 수 없습니다: {SCORE_INPUT_PATH}")
        sys.exit(1)
    if not ANSWER_KEY_PATH.exists():
        print(f"오류: answer key 파일을 찾을 수 없습니다: {ANSWER_KEY_PATH}")
        sys.exit(1)

    key_wb = load_workbook(ANSWER_KEY_PATH, data_only=True)
    key_ws = key_wb.active
    key_h = header_map(key_ws)
    require_columns(key_h, ["question_id", "answer_label", "original_model"], ANSWER_KEY_PATH)
    answer_key = {}
    for row in range(2, key_ws.max_row + 1):
        qid = normalize(key_ws.cell(row, key_h["question_id"]).value)
        label = normalize(key_ws.cell(row, key_h["answer_label"]).value)
        model = normalize(key_ws.cell(row, key_h["original_model"]).value)
        if qid and label and model:
            answer_key[(qid, label)] = model

    score_wb = load_workbook(SCORE_INPUT_PATH, data_only=True)
    score_ws = score_wb.active
    score_h = header_map(score_ws)
    required = [
        "question_id", "answer_label", "judge_model",
        "search_relevance_25", "evidence_grounding_25",
        "safety_legal_accuracy_25", "practicality_25", "total_score_100",
    ]
    require_columns(score_h, required, SCORE_INPUT_PATH)

    scores = defaultdict(dict)
    warnings = []
    for row in range(2, score_ws.max_row + 1):
        qid = normalize(score_ws.cell(row, score_h["question_id"]).value)
        label = normalize(score_ws.cell(row, score_h["answer_label"]).value)
        judge = normalize(score_ws.cell(row, score_h["judge_model"]).value)
        if not qid and not label and not judge:
            continue
        model = answer_key.get((qid, label))
        if not model:
            warnings.append(f"answer key 없음: {qid} / {label}")
            continue
        total = to_float(score_ws.cell(row, score_h["total_score_100"]).value)
        if total is None:
            parts = [
                to_float(score_ws.cell(row, score_h["search_relevance_25"]).value),
                to_float(score_ws.cell(row, score_h["evidence_grounding_25"]).value),
                to_float(score_ws.cell(row, score_h["safety_legal_accuracy_25"]).value),
                to_float(score_ws.cell(row, score_h["practicality_25"]).value),
            ]
            if all(v is not None for v in parts):
                total = sum(parts)
        if total is None:
            warnings.append(f"점수 없음: {qid} / {label} / {judge}")
            continue
        if total < 0 or total > 100:
            warnings.append(f"총점 범위 확인 필요: {qid} / {label} / {judge} = {total}")
        scores[(qid, model)][judge] = total

    out_wb = Workbook()
    ws = out_wb.active
    ws.title = "summary_by_question_model"
    ws.append(["question_id", "original_model", "gpt_judge_score", "gemini_judge_score", "claude_judge_score", "average_score", "std_score", "final_rank"])

    grouped_by_question = defaultdict(list)
    for key, judge_scores in scores.items():
        qid, model = key
        vals = [judge_scores[j] for j in JUDGES if j in judge_scores]
        avg = statistics.mean(vals) if vals else None
        std = statistics.stdev(vals) if len(vals) >= 2 else 0 if len(vals) == 1 else None
        grouped_by_question[qid].append((model, judge_scores, avg, std))

    for qid in sorted(grouped_by_question):
        rows = grouped_by_question[qid]
        ranked = sorted(rows, key=lambda item: (-math.inf if item[2] is None else -item[2], item[0]))
        rank_by_model = {model: idx + 1 for idx, (model, _, _, _) in enumerate(ranked)}
        for model, judge_scores, avg, std in sorted(rows, key=lambda item: item[0]):
            ws.append([
                qid,
                model,
                judge_scores.get("GPT", ""),
                judge_scores.get("Gemini", ""),
                judge_scores.get("Claude", ""),
                round(avg, 2) if avg is not None else "",
                round(std, 2) if std is not None else "",
                rank_by_model[model],
            ])

    overall = out_wb.create_sheet("overall_model_summary")
    overall.append(["original_model", "question_count", "overall_average_score", "overall_std_score"])
    by_model = defaultdict(list)
    for (_qid, model), judge_scores in scores.items():
        vals = [judge_scores[j] for j in JUDGES if j in judge_scores]
        if vals:
            by_model[model].append(statistics.mean(vals))
    for model in sorted(by_model):
        vals = by_model[model]
        overall.append([model, len(vals), round(statistics.mean(vals), 2), round(statistics.stdev(vals), 2) if len(vals) >= 2 else 0])

    notes = out_wb.create_sheet("warnings")
    notes.append(["warning"])
    for w in warnings:
        notes.append([w])

    out_wb.save(OUTPUT_PATH)
    print(f"요약 파일 생성 완료: {OUTPUT_PATH}")
    if warnings:
        print(f"주의: 확인 메시지 {len(warnings)}개가 warnings 시트에 저장되었습니다.")


if __name__ == "__main__":
    main()
