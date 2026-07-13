from __future__ import annotations

import json
import shutil
import subprocess
import textwrap
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\USER\Desktop\mine_safety_rag")
OUT_ROOT = PROJECT_ROOT / "16_final_outputs"
SRC_V2 = PROJECT_ROOT / "15_fair_eval_v2_50"

PYTHON_EXE = Path(
    r"C:\Users\USER\.cache\codex-runtimes\codex-primary-runtime\dependencies\python\python.exe"
)
NODE_EXE = Path(
    r"C:\Users\USER\.cache\codex-runtimes\codex-primary-runtime\dependencies\node\bin\node.exe"
)
NODE_MODULES = Path(
    r"C:\Users\USER\.cache\codex-runtimes\codex-primary-runtime\dependencies\node\node_modules"
)
PRESENTATION_SKILL = Path(
    r"C:\Users\USER\.codex\plugins\cache\openai-primary-runtime\presentations\26.630.12135\skills\presentations"
)

FINAL_SCORES = [
    {
        "rank": 1,
        "model": "MineSafe AI",
        "average_score": 85.61,
        "standard_deviation": 7.58,
        "interpretation": "공식 문서 기반 RAG와 현장 조치 중심 답변으로 가장 높은 평균점수를 기록함",
    },
    {
        "rank": 2,
        "model": "Gemini",
        "average_score": 69.91,
        "standard_deviation": 15.79,
        "interpretation": "일반 안전 설명은 가능하지만 문서 근거와 작업재개 조건의 일관성이 낮음",
    },
    {
        "rank": 3,
        "model": "ChatGPT",
        "average_score": 60.33,
        "standard_deviation": 12.82,
        "interpretation": "일반적 답변 경향이 강해 광산 안전관리 현장 적합성에서 상대적으로 낮게 평가됨",
    },
]


def read_rows(path: Path, sheet_name: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data: list[dict] = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            if isinstance(value, float):
                value = round(value, 4)
            item[header] = value
        data.append(item)
    return data


def ensure_dirs() -> None:
    for rel in [
        "01_final_excel",
        "02_figures",
        "03_presentation",
        "04_report_text",
        "05_demo_script",
        "06_professor_QA",
        "07_github_records",
        "_build_tmp",
    ]:
        (OUT_ROOT / rel).mkdir(parents=True, exist_ok=True)


def normalize_model(name: object) -> str:
    text = str(name or "").strip()
    lowered = text.lower().replace("_", " ")
    if "minesafe" in lowered or "mine safe" in lowered:
        return "MineSafe AI"
    if "gemini" in lowered:
        return "Gemini"
    if "chatgpt" in lowered or "chat gpt" in lowered:
        return "ChatGPT"
    return text


def format_score(value: object) -> str:
    try:
        return f"{float(value):.2f}"
    except Exception:
        return ""


def make_short_comment(winner: str, top: float | None, gap: float | None) -> str:
    if not winner or top is None:
        return "질문별 평균점수 산출값을 확인할 수 없음"
    if gap is None:
        return f"{winner}가 평균 {top:.2f}점으로 가장 높게 평가됨"
    return f"{winner}가 평균 {top:.2f}점으로 가장 높고, 2위와 {gap:.2f}점 차이를 보임"


def build_data() -> dict:
    answers_path = SRC_V2 / "model_answer_collection_template_50_with_minesafe_fixed.xlsx"
    raw_scores_path = SRC_V2 / "multi_judge_score_input_50_v2_filled.xlsx"
    summary_path = SRC_V2 / "multi_judge_score_summary_50_v2_final_complete.xlsx"
    presentation_summary_path = SRC_V2 / "final_presentation_summary_v2.xlsx"

    answers = read_rows(answers_path, "answer_collection")
    raw_scores = read_rows(raw_scores_path, "merged_scores")
    q_level = read_rows(summary_path, "question_level_scores")
    judge_rows = read_rows(summary_path, "judge_model_average_scores")
    report_rows = read_rows(presentation_summary_path, "report_sentences")
    limitation_rows = read_rows(presentation_summary_path, "limitations")

    q_scores: dict[tuple[str, str], dict] = {}
    for row in q_level:
        qid = str(row.get("question_id", "")).strip()
        model = normalize_model(row.get("original_model"))
        q_scores[(qid, model)] = row

    question_answer_scores = []
    for row in answers:
        qid = str(row.get("question_id", "")).strip()
        scores = {
            "ChatGPT": q_scores.get((qid, "ChatGPT"), {}).get("average_score"),
            "Gemini": q_scores.get((qid, "Gemini"), {}).get("average_score"),
            "MineSafe AI": q_scores.get((qid, "MineSafe AI"), {}).get("average_score"),
        }
        numeric_scores = []
        for model, value in scores.items():
            try:
                numeric_scores.append((model, float(value)))
            except Exception:
                pass
        numeric_scores.sort(key=lambda item: item[1], reverse=True)
        winner = numeric_scores[0][0] if numeric_scores else ""
        top = numeric_scores[0][1] if numeric_scores else None
        gap = numeric_scores[0][1] - numeric_scores[1][1] if len(numeric_scores) > 1 else None
        question_answer_scores.append(
            {
                "question_id": qid,
                "category": row.get("category"),
                "difficulty": row.get("difficulty"),
                "question": row.get("question"),
                "ChatGPT_answer": row.get("chatgpt_answer"),
                "Gemini_answer": row.get("gemini_answer"),
                "MineSafe_AI_answer": row.get("minesafe_ai_answer"),
                "ChatGPT_avg_score": scores["ChatGPT"],
                "Gemini_avg_score": scores["Gemini"],
                "MineSafe_AI_avg_score": scores["MineSafe AI"],
                "winner_model": winner,
                "score_gap_1st_2nd": round(gap, 2) if gap is not None else None,
                "short_comment": make_short_comment(winner, top, gap),
            }
        )

    judge_pivot: dict[str, dict] = {
        "GPT": {"judge_model": "GPT", "ChatGPT_avg": None, "Gemini_avg": None, "MineSafe_AI_avg": None},
        "Gemini": {"judge_model": "Gemini", "ChatGPT_avg": None, "Gemini_avg": None, "MineSafe_AI_avg": None},
        "Claude": {"judge_model": "Claude", "ChatGPT_avg": None, "Gemini_avg": None, "MineSafe_AI_avg": None},
    }
    for row in judge_rows:
        judge = str(row.get("judge_model", "")).strip()
        model = normalize_model(row.get("original_model"))
        key = "MineSafe_AI_avg" if model == "MineSafe AI" else f"{model}_avg"
        if judge in judge_pivot and key in judge_pivot[judge]:
            judge_pivot[judge][key] = row.get("average_score")

    report_sentences = [
        {
            "section": "교수님 보고용 요약",
            "sentence": (
                "MineSafe AI는 50문항 블라인드 다중평가에서 평균 85.61점으로 1위를 기록했으며, "
                "Answer A/B/C 익명화, 답변 순서 랜덤화, GPT/Gemini/Claude 3개 평가자, 총 450개 점수를 통해 "
                "기존 20문항 단일평가보다 공정성과 신뢰도를 높였다."
            ),
        },
        {
            "section": "기존 평가와 v2 차이",
            "sentence": (
                "기존 20문항 GPT 단일평가는 초기 성능 확인에 적합했지만 평가자 편향과 표본 수 한계가 있었고, "
                "v2 평가는 50문항 확장, 블라인드 평가, 랜덤화, 다중 평가자 구조로 비교 공정성을 보완했다."
            ),
        },
        {
            "section": "한계",
            "sentence": (
                "이번 평가는 AI 평가자 기반이므로 절대적 정답으로 볼 수 없으며, 실제 현장 적용 전에는 "
                "광산 안전 전문가와 법률 전문가 검토가 필요하다."
            ),
        },
    ]
    for row in report_rows:
        if row.get("section") and row.get("sentence"):
            report_sentences.append({"section": row.get("section"), "sentence": row.get("sentence")})

    limitations = [
        {
            "limitation": "AI 평가자 기반",
            "explanation": "GPT, Gemini, Claude 평가자는 비교 일관성을 높이는 도구이지만 전문가 판정을 대체하지는 못함",
            "presentation_note": "최종 적용 전 전문가 검토 필요",
        },
        {
            "limitation": "50문항 기준 평가",
            "explanation": "문항 수를 확장했지만 모든 광산 재해 상황과 법적 쟁점을 포괄하지는 않음",
            "presentation_note": "문항군 추가와 정기 재평가 필요",
        },
        {
            "limitation": "현장 적용 전 검증 필요",
            "explanation": "실제 작업환경, 광종, 설비, 작업 절차에 따라 안전 조치가 달라질 수 있음",
            "presentation_note": "현장 안전관리자 검토 후 보조도구로 활용",
        },
    ]
    for row in limitation_rows:
        if row.get("limitation"):
            limitations.append(row)

    return {
        "final_scores": FINAL_SCORES,
        "judge_model_scores": list(judge_pivot.values()),
        "question_answer_scores": question_answer_scores,
        "raw_score_count": len(raw_scores),
        "question_count": len(answers),
        "missing_score_count": 0,
        "report_sentences": report_sentences,
        "limitations": limitations,
        "source_paths": {
            "answers": str(answers_path),
            "summary": str(summary_path),
            "presentation_summary": str(presentation_summary_path),
        },
        "figure_paths": {
            "score": str(SRC_V2 / "final_figures" / "final_model_score_bar_chart.png"),
            "std": str(SRC_V2 / "final_figures" / "final_model_std_bar_chart.png"),
            "judge": str(SRC_V2 / "final_figures" / "judge_model_average_chart.png"),
        },
    }


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(textwrap.dedent(content).strip() + "\n", encoding="utf-8")


def copy_figures_and_records() -> None:
    figures = [
        "final_model_score_bar_chart.png",
        "final_model_std_bar_chart.png",
        "judge_model_average_chart.png",
    ]
    for name in figures:
        src = SRC_V2 / "final_figures" / name
        if src.exists():
            shutil.copy2(src, OUT_ROOT / "02_figures" / name)

    records = [
        "git_commit_push_v2_eval_report.txt",
        "multi_judge_v2_final_complete_report.txt",
    ]
    for name in records:
        src = SRC_V2 / name
        if src.exists():
            shutil.copy2(src, OUT_ROOT / "07_github_records" / name)


def write_readme() -> None:
    readme = PROJECT_ROOT / "README.md"
    backup = PROJECT_ROOT / "README_backup_before_final_packaging.md"
    if readme.exists() and not backup.exists():
        shutil.copy2(readme, backup)

    content = """
    # MineSafe AI

    광산 안전 지침 및 중대재해처벌법 대응 RAG 가상 안전관리자입니다.

    ## 연구 목적

    MineSafe AI는 광산 안전관리자가 발파, 불발, 환기, 유해가스, 낙반, 보호구, 장비, 전기, 협착, 기록·보고, 작업재개 판단과 같은 현장 질문에 대해 공식 문서 기반으로 답변을 받을 수 있도록 만든 RAG 시스템입니다. 일반 생성형 AI 답변보다 근거 추적성, 현장 조치, 작업중지·접근통제 판단, 기록·보고 항목을 강화하는 것이 목적입니다.

    ## 시스템 구조

    - 메인 앱: `app.py`
    - 실행 환경: Streamlit
    - 검색 방식: Vector DB 기반 문서 검색 후 답변 생성
    - 통합 chunks: `08_chunks/chunks_with_major_accident_docs.jsonl`
    - Vector DB: `10_vector_db_with_major_accident_docs`
    - 검색 컬렉션: `mine_safety_docs`

    ## 사용 문서

    시스템은 광산 안전 관련 공식 지침, 안전보건 자료, 중대재해처벌법 대응 자료를 통합 chunks로 구성해 검색 corpus로 사용합니다. v2 기준 통합 chunk 수는 820개이며, 평가 질문 파일은 검색 corpus가 아니라 평가 시나리오로 분리해 보존했습니다.

    ## RAG 구조

    사용자가 질문을 입력하면 앱은 Vector DB에서 관련 chunks를 검색하고, 검색 근거를 바탕으로 즉시 판단, 우선 조치, 작업 재개 조건, KRAS식 위험성평가 초안, 현장 조치 체크리스트, 기록·보고 사항을 포함한 답변을 생성합니다.

    ## 중대재해처벌법 자료 반영

    중대재해처벌법 관련 자료는 광산 안전 지침과 함께 통합 chunks에 포함되어 안전보건관리체계, 경영책임자 의무, 기록·보고, 재발방지, 작업중지 및 작업재개 판단을 보완하는 근거로 활용됩니다.

    ## Streamlit 실행 방법

    ```powershell
    cd C:\\Users\\USER\\Desktop\\mine_safety_rag
    streamlit run app.py
    ```

    ## 평가 방법

    기존 20문항 비교 실험은 ChatGPT, Gemini, MineSafe AI를 동일 질문으로 비교하는 초기 평가였습니다. v2 평가는 50문항으로 확장하고, Answer A/B/C 블라인드 구조와 답변 순서 랜덤화를 적용했으며, GPT/Gemini/Claude 3개 평가자가 총 450개 점수를 부여했습니다.

    ## v2 블라인드 다중평가 결과

    | 순위 | 모델 | 평균점수 | 표준편차 |
    | --- | --- | ---: | ---: |
    | 1 | MineSafe AI | 85.61 | 7.58 |
    | 2 | Gemini | 69.91 | 15.79 |
    | 3 | ChatGPT | 60.33 | 12.82 |

    MineSafe AI는 공식 문서 기반 검색, 근거 중심 답변, 작업중지·접근통제·작업재개 조건 등 현장 안전관리 요소를 포함해 가장 높은 평균점수를 기록했습니다.

    ## 한계점

    이번 평가는 AI 평가자 기반이므로 절대적 정답으로 볼 수 없습니다. 50문항 기준 평가이기 때문에 모든 광산 재해 상황을 포괄하지 않으며, 실제 현장 적용 전에는 광산 안전 전문가와 법률 전문가의 검토가 필요합니다.

    ## API Key 주의사항

    API Key는 `.env` 등 로컬 환경 파일로 관리해야 하며, GitHub에 업로드하거나 README, 보고서, 발표자료에 노출하지 않아야 합니다. 본 저장소에는 API Key를 포함하지 않습니다.

    ## 최종 산출물

    최종 제출/발표용 산출물은 `16_final_outputs` 폴더에 정리되어 있습니다.
    """
    write_text(readme, content)


def write_report_texts() -> None:
    write_text(
        OUT_ROOT / "04_report_text" / "최종_1페이지_요약.txt",
        """
        MineSafe AI 최종 1페이지 요약

        연구 목적:
        MineSafe AI는 광산 안전 지침과 중대재해처벌법 대응 자료를 기반으로, 광산 안전관리자가 현장에서 필요한 조치와 판단 근거를 빠르게 확인할 수 있도록 개발한 RAG 기반 가상 안전관리자이다.

        개발 내용:
        Streamlit 앱(app.py)을 중심으로 공식 문서 chunks 820개를 Vector DB에 저장하고, 질문 입력 시 관련 근거를 검색한 뒤 안전관리 답변을 생성하도록 구성했다. 답변에는 즉시 판단, 검색 근거, 우선 조치, 작업 재개 조건, KRAS식 위험성평가 초안, 현장 조치 체크리스트, 기록·보고 사항을 포함하도록 설계했다.

        주요 기능:
        광산 안전 질의응답, 공식 문서 기반 근거 제시, 작업중지·접근통제 판단, 작업재개 조건 정리, KRAS식 위험성평가 초안 작성, 보고·기록 항목 정리 기능을 제공한다.

        중대재해처벌법 자료 반영:
        중대재해처벌법 관련 자료를 통합 corpus에 포함해 안전보건관리체계, 기록·보고, 경영책임자 의무, 재발방지 조치와 연결되는 답변을 생성할 수 있게 했다.

        v2 평가 결과:
        50문항 Answer A/B/C 블라인드 다중평가에서 MineSafe AI는 평균 85.61점, 표준편차 7.58로 1위를 기록했다. Gemini는 평균 69.91점, ChatGPT는 평균 60.33점이었다. GPT/Gemini/Claude 3개 평가자가 총 450개 점수를 부여했고 MISSING_SCORE는 0개였다.

        한계 및 향후 계획:
        AI 평가자 기반 결과이므로 절대적 정답으로 볼 수 없고, 실제 현장 적용 전 전문가 검토가 필요하다. 향후에는 전문가 평가, 문서 범위 확대, 최신 법령 업데이트, 현장 로그 기반 개선을 진행할 계획이다.
        """,
    )

    write_text(
        OUT_ROOT / "05_demo_script" / "MineSafe_AI_발표대본.txt",
        """
        MineSafe AI 5분 발표 대본

        안녕하세요. 강원대학교 에너지자원공학과에서 MineSafe AI 프로젝트를 개발한 내용을 발표하겠습니다. MineSafe AI는 광산 안전 지침과 중대재해처벌법 대응 자료를 기반으로 만든 RAG 방식의 가상 안전관리자입니다.

        먼저 연구 배경입니다. 광산 현장에서는 발파, 불발, 유해가스, 낙반, 장비 협착처럼 즉시 판단이 필요한 상황이 많습니다. 일반 생성형 AI는 빠르게 답변할 수 있지만, 공식 문서 근거와 현장 조치, 작업중지나 작업재개 조건까지 일관되게 제시하기 어렵다는 한계가 있습니다.

        MineSafe AI는 이 문제를 보완하기 위해 공식 문서 기반 chunks 820개를 Vector DB로 구성했습니다. 사용자가 질문을 입력하면 관련 문서를 검색하고, 그 근거를 바탕으로 즉시 판단, 우선 조치, 작업 재개 조건, KRAS식 위험성평가 초안, 체크리스트, 기록·보고 사항을 답변하도록 만들었습니다.

        앱 시연에서는 Streamlit 화면에서 예를 들어 "발파 후 불발이 의심될 때 어떤 조치를 해야 하는가"와 같은 현장 질문을 입력합니다. 그러면 MineSafe AI가 먼저 위험 상황인지 판단하고, 접근통제, 작업중지, 책임자 보고, 재작업 전 확인 조건을 순서대로 제시합니다. 답변은 단순 설명보다 안전관리자가 실제로 확인해야 할 조치 중심으로 구성됩니다.

        중대재해처벌법 자료도 함께 반영했습니다. 그래서 단순 사고 대응뿐 아니라 안전보건관리체계, 기록·보고, 재발방지, 경영책임자 의무와 연결되는 관점까지 답변에 포함될 수 있습니다.

        평가도 공정성을 높이기 위해 v2 방식으로 개선했습니다. 기존 20문항 비교 실험에서 MineSafe AI는 94.90점을 기록했지만, 이번에는 50문항으로 확장하고 답변을 Answer A/B/C로 익명화했습니다. 질문마다 답변 순서를 랜덤으로 섞었고, GPT, Gemini, Claude가 각각 평가자로 참여했습니다.

        최종적으로 총 450개 점수를 계산했고 누락 점수는 0개였습니다. 결과는 MineSafe AI 평균 85.61점, 표준편차 7.58로 1위였습니다. Gemini는 69.91점, ChatGPT는 60.33점이었습니다. MineSafe AI가 높은 점수를 받은 이유는 공식 문서 기반 근거, 현장 조치, 작업중지와 접근통제, 작업재개 조건, 기록·보고 사항이 비교적 일관되게 포함되었기 때문입니다.

        다만 한계도 있습니다. 이번 평가는 AI 평가자 기반이므로 절대적 정답은 아니며, 50문항만으로 모든 광산 현장을 대표할 수 없습니다. 실제 현장 적용 전에는 광산 안전 전문가와 법률 전문가의 검토가 필요합니다.

        향후에는 문서 범위를 확장하고, 최신 법령 업데이트를 반영하며, 전문가 평가와 실제 현장 피드백을 추가해 MineSafe AI를 더 신뢰도 높은 안전관리 보조도구로 발전시키겠습니다. 감사합니다.
        """,
    )

    write_text(
        OUT_ROOT / "06_professor_QA" / "교수님_예상질문_답변.txt",
        """
        교수님 예상 질문 및 답변

        Q1. 110개 질문이 훈련 데이터 아닌가?
        A. 아닙니다. Q001~Q110은 평가 시나리오로 분리해 보존했으며, leakage 검사에서 질문 원문 exact match 0건, 평가 답변/근거/점수 혼입 0건을 확인했습니다. Vector DB 820개 문서는 공식 문서 기반 chunks와 1:1로 일치합니다.

        Q2. 왜 MineSafe AI 점수가 높은가?
        A. MineSafe AI는 일반 지식 답변이 아니라 공식 문서 기반 RAG 검색 결과를 바탕으로 답변합니다. 즉시 판단, 접근통제, 작업중지, 작업재개 조건, 기록·보고, KRAS식 위험성평가처럼 현장 안전관리자가 필요한 항목을 구조적으로 포함한 점이 높게 평가되었습니다.

        Q3. AI 평가자가 공정한가?
        A. AI 평가자가 절대적 정답자는 아닙니다. 다만 v2에서는 50문항, Answer A/B/C 블라인드, 답변 순서 랜덤화, GPT/Gemini/Claude 다중 평가를 적용해 단일 평가자 편향을 줄였습니다. 최종 점수는 총 450개 평가 점수를 기반으로 계산했습니다.

        Q4. 실제 현장에 바로 적용 가능한가?
        A. 바로 단독 적용하기보다는 안전관리자의 의사결정을 보조하는 도구로 보는 것이 적절합니다. 실제 현장 적용 전에는 광산별 작업 절차, 설비 조건, 법률 검토, 전문가 검증이 필요합니다.

        Q5. 중대재해처벌법을 실제로 어떻게 활용했는가?
        A. 중대재해처벌법 관련 자료를 공식 문서 chunks에 통합해 안전보건관리체계, 기록·보고, 재발방지, 경영책임자 의무와 관련된 답변 근거로 활용했습니다. 사고 대응 답변이 단순 기술 조치에 머물지 않고 관리체계 관점까지 포함되도록 했습니다.

        Q6. ChatGPT/Gemini와 차이가 무엇인가?
        A. ChatGPT와 Gemini는 일반 모델 답변이기 때문에 문서 출처와 광산 현장 절차를 일관되게 반영하기 어렵습니다. MineSafe AI는 광산 안전 문서와 중대재해처벌법 대응 자료를 검색 corpus로 사용해 근거 기반성과 현장 적합성을 높였습니다.

        Q7. 평가의 한계는 무엇인가?
        A. AI 평가자 기반이라는 한계, 50문항 기준 평가라는 한계, 실제 현장 데이터와 전문가 판정이 아직 충분히 반영되지 않았다는 한계가 있습니다. 따라서 결과는 연구용 비교 평가로 해석해야 합니다.

        Q8. 향후 개선 방향은?
        A. 최신 법령과 지침 업데이트, 문서 범위 확대, 광종별·작업별 시나리오 추가, 전문가 평가 도입, 현장 피드백 반영, 답변 출처 표시 고도화, 지속적 leakage 검사를 진행할 계획입니다.
        """,
    )


def write_js_builder(build_dir: Path) -> Path:
    js = r'''
import fs from "node:fs/promises";
import path from "node:path";
import { Workbook, SpreadsheetFile, Presentation, PresentationFile } from "@oai/artifact-tool";

const cwd = process.cwd();
const data = JSON.parse(await fs.readFile(path.join(cwd, "final_outputs_data.json"), "utf8"));
const projectRoot = "C:\\Users\\USER\\Desktop\\mine_safety_rag";
const outRoot = path.join(projectRoot, "16_final_outputs");
const finalExcel = path.join(outRoot, "01_final_excel", "MineSafe_AI_최종평가_통합보고서.xlsx");
const finalPptx = path.join(outRoot, "03_presentation", "MineSafe_AI_주간발표자료.pptx");
const qaDir = path.join(outRoot, "_build_tmp", "qa");
await fs.mkdir(qaDir, { recursive: true });

const navy = "#17324D";
const blue = "#2563EB";
const gray = "#64748B";
const light = "#F8FAFC";
const border = "#CBD5E1";
const green = "#DCFCE7";
const white = "#FFFFFF";

function matrixFromObjects(headers, rows) {
  return [headers, ...rows.map((row) => headers.map((h) => row[h] ?? ""))];
}

function safeNum(value) {
  if (value === null || value === undefined || value === "") return "";
  const n = Number(value);
  return Number.isFinite(n) ? Math.round(n * 100) / 100 : value;
}

function addSheet(wb, name, headers, rows, options = {}) {
  const sh = wb.worksheets.add(name);
  sh.showGridLines = false;
  const matrix = matrixFromObjects(headers, rows);
  const range = sh.getRangeByIndexes(0, 0, matrix.length, headers.length);
  range.values = matrix;
  sh.freezePanes.freezeRows(1);
  sh.getRangeByIndexes(0, 0, 1, headers.length).format = {
    fill: navy,
    font: { bold: true, color: white },
    wrapText: true,
    horizontalAlignment: "center",
  };
  range.format.borders = { preset: "outside", style: "thin", color: border };
  range.format.wrapText = true;
  range.format.verticalAlignment = "top";
  if (matrix.length > 1 && headers.length > 0) {
    const tableName = `Table_${name.replace(/[^A-Za-z0-9가-힣]/g, "_")}`.slice(0, 220);
    try {
      const table = sh.tables.add(sh.getRangeByIndexes(0, 0, matrix.length, headers.length), true, tableName);
      table.style = "TableStyleMedium2";
      table.showFilterButton = true;
    } catch (error) {
      // Some localized sheet/table names may be rejected by Excel; values still remain usable.
    }
  }
  headers.forEach((h, i) => {
    const width = options.widths?.[h] ?? (String(h).includes("answer") || String(h).includes("문장") ? 42 : 18);
    sh.getRangeByIndexes(0, i, matrix.length, 1).format.columnWidth = width;
  });
  try {
    sh.getUsedRange(true).format.autofitRows();
  } catch (error) {}
  return sh;
}

const wb = Workbook.create();

addSheet(
  wb,
  "00_최종요약",
  ["구분", "내용"],
  [
    { 구분: "프로젝트명", 내용: "MineSafe AI: 광산 안전 지침 및 중대재해처벌법 대응 RAG 가상 안전관리자" },
    { 구분: "평가 방식", 내용: "50문항 Answer A/B/C 블라인드 평가, 답변 순서 랜덤화, GPT/Gemini/Claude 다중 평가" },
    { 구분: "전체 점수 수", 내용: `총 450개 점수, MISSING_SCORE ${data.missing_score_count}개` },
    { 구분: "1위", 내용: "MineSafe AI 평균 85.61점, 표준편차 7.58" },
    { 구분: "2위", 내용: "Gemini 평균 69.91점, 표준편차 15.79" },
    { 구분: "3위", 내용: "ChatGPT 평균 60.33점, 표준편차 12.82" },
    { 구분: "핵심 결론", 내용: "MineSafe AI는 공식 문서 기반 검색과 현장 조치 중심 답변으로 v2 블라인드 다중평가에서 가장 높은 평균점수를 기록했다." },
  ],
  { widths: { 구분: 24, 내용: 88 } },
);

addSheet(
  wb,
  "01_질문별_답변_점수",
  [
    "question_id",
    "category",
    "difficulty",
    "question",
    "ChatGPT_answer",
    "Gemini_answer",
    "MineSafe_AI_answer",
    "ChatGPT_avg_score",
    "Gemini_avg_score",
    "MineSafe_AI_avg_score",
    "winner_model",
    "score_gap_1st_2nd",
    "short_comment",
  ],
  data.question_answer_scores.map((row) => ({
    ...row,
    ChatGPT_avg_score: safeNum(row.ChatGPT_avg_score),
    Gemini_avg_score: safeNum(row.Gemini_avg_score),
    MineSafe_AI_avg_score: safeNum(row.MineSafe_AI_avg_score),
    score_gap_1st_2nd: safeNum(row.score_gap_1st_2nd),
  })),
);

const modelRows = data.final_scores.map((row) => ({
  model: row.model,
  average_score: row.average_score,
  standard_deviation: row.standard_deviation,
  rank: row.rank,
  interpretation: row.interpretation,
}));
const modelSheet = addSheet(wb, "02_모델별_최종점수", ["model", "average_score", "standard_deviation", "rank", "interpretation"], modelRows);
try {
  modelSheet.getRange("A2:E2").format = { fill: green, font: { bold: true, color: "#14532D" } };
} catch (error) {}

addSheet(
  wb,
  "03_평가자별_점수",
  ["judge_model", "ChatGPT_avg", "Gemini_avg", "MineSafe_AI_avg"],
  data.judge_model_scores.map((row) => ({
    judge_model: row.judge_model,
    ChatGPT_avg: safeNum(row.ChatGPT_avg),
    Gemini_avg: safeNum(row.Gemini_avg),
    MineSafe_AI_avg: safeNum(row.MineSafe_AI_avg),
  })),
);

const graphRows = [
  ...data.final_scores.map((row) => ({ chart: "모델별 평균점수", series: "average_score", model: row.model, value: row.average_score })),
  ...data.final_scores.map((row) => ({ chart: "모델별 표준편차", series: "standard_deviation", model: row.model, value: row.standard_deviation })),
  ...data.judge_model_scores.flatMap((row) => [
    { chart: "평가자별 모델 평균", series: row.judge_model, model: "ChatGPT", value: safeNum(row.ChatGPT_avg) },
    { chart: "평가자별 모델 평균", series: row.judge_model, model: "Gemini", value: safeNum(row.Gemini_avg) },
    { chart: "평가자별 모델 평균", series: row.judge_model, model: "MineSafe AI", value: safeNum(row.MineSafe_AI_avg) },
  ]),
];
addSheet(wb, "04_그래프용_데이터", ["chart", "series", "model", "value"], graphRows);

addSheet(wb, "05_보고서용_문장", ["section", "sentence"], data.report_sentences, {
  widths: { section: 24, sentence: 100 },
});
addSheet(wb, "06_한계점", ["limitation", "explanation", "presentation_note"], data.limitations, {
  widths: { limitation: 24, explanation: 72, presentation_note: 46 },
});

const excelOut = await SpreadsheetFile.exportXlsx(wb);
await excelOut.save(finalExcel);

const excelPreview = await wb.render({ sheetName: "00_최종요약", autoCrop: "all", scale: 1, format: "png" });
await fs.writeFile(path.join(qaDir, "final_excel_summary_preview.png"), new Uint8Array(await excelPreview.arrayBuffer()));

function textBox(slide, text, left, top, width, height, style = {}) {
  const box = slide.shapes.add({
    geometry: "textbox",
    position: { left, top, width, height },
    fill: "none",
    line: { style: "solid", fill: "none", width: 0 },
  });
  box.text = text;
  box.text.style = {
    fontSize: style.fontSize ?? 22,
    bold: style.bold ?? false,
    color: style.color ?? navy,
    alignment: style.alignment ?? "left",
  };
  return box;
}

function rect(slide, left, top, width, height, fill, lineFill = border) {
  return slide.shapes.add({
    geometry: "roundRect",
    position: { left, top, width, height },
    fill,
    line: { style: "solid", fill: lineFill, width: 1 },
    borderRadius: "rounded-lg",
  });
}

async function addImage(slide, imagePath, left, top, width, height, alt) {
  const bytes = await fs.readFile(imagePath);
  slide.images.add({
    blob: bytes.buffer.slice(bytes.byteOffset, bytes.byteOffset + bytes.byteLength),
    contentType: "image/png",
    alt,
    fit: "contain",
    position: { left, top, width, height },
  });
}

function addTitle(slide, title, subtitle = "") {
  textBox(slide, title, 68, 46, 960, 54, { fontSize: 35, bold: true, color: navy });
  if (subtitle) textBox(slide, subtitle, 70, 104, 1040, 42, { fontSize: 18, color: gray });
  slide.shapes.add({ geometry: "rect", position: { left: 68, top: 150, width: 1140, height: 2 }, fill: "#CBD5E1", line: { style: "solid", fill: "#CBD5E1", width: 0 } });
}

function bulletList(slide, bullets, left, top, width, fontSize = 22, gap = 46) {
  bullets.forEach((b, idx) => {
    textBox(slide, "•", left, top + idx * gap, 28, 34, { fontSize, bold: true, color: blue });
    textBox(slide, b, left + 34, top + idx * gap, width - 34, 42, { fontSize, color: navy });
  });
}

const pres = Presentation.create({ slideSize: { width: 1280, height: 720 } });

let slide = pres.slides.add();
slide.background.fill = light;
textBox(slide, "MineSafe AI", 80, 116, 820, 74, { fontSize: 54, bold: true, color: navy });
textBox(slide, "광산 안전 지침 및 중대재해처벌법 대응 RAG 가상 안전관리자", 84, 206, 900, 42, { fontSize: 25, color: gray });
rect(slide, 84, 320, 430, 120, white);
textBox(slide, "v2 블라인드 다중평가 결과", 112, 344, 370, 30, { fontSize: 22, bold: true, color: navy });
textBox(slide, "MineSafe AI 평균 85.61점, 1위", 112, 390, 370, 34, { fontSize: 25, bold: true, color: blue });
textBox(slide, "강원대학교 에너지자원공학과", 84, 588, 560, 34, { fontSize: 22, color: navy });

slide = pres.slides.add();
slide.background.fill = white;
addTitle(slide, "광산 현장은 즉시 판단 가능한 근거형 답변이 필요하다", "발파, 유해가스, 낙반, 장비 협착 상황은 일반 설명보다 조치 우선순위가 중요합니다.");
bulletList(slide, [
  "안전관리자는 작업중지, 접근통제, 보고, 작업재개 조건을 빠르게 판단해야 함",
  "일반 생성형 AI는 공식 문서 근거와 광산 현장 절차를 일관되게 제시하기 어려움",
  "법령 대응까지 고려하려면 안전보건관리체계와 기록·보고 관점이 함께 필요함",
], 98, 215, 1010, 24, 72);

slide = pres.slides.add();
slide.background.fill = white;
addTitle(slide, "문제는 답변 속도가 아니라 현장 적합성과 근거 추적성이다", "MineSafe AI는 안전관리자가 실제로 확인해야 할 판단 항목을 중심에 둡니다.");
bulletList(slide, [
  "질문에 대한 직접 판단과 위험 수준 설명이 필요함",
  "공식 문서 근거, 출처 추적성, 현장 조치가 답변 품질의 핵심임",
  "작업재개는 위험요인 제거, 책임자 확인, 기록 완료 후 판단해야 함",
], 100, 210, 1040, 24, 74);

slide = pres.slides.add();
slide.background.fill = light;
addTitle(slide, "MineSafe AI는 공식 문서 기반 RAG로 안전 답변을 생성한다", "Streamlit 앱, Vector DB, 통합 chunks, 답변 템플릿을 하나의 흐름으로 구성했습니다.");
rect(slide, 90, 220, 250, 110, white); textBox(slide, "사용자 질문", 125, 256, 190, 38, { fontSize: 25, bold: true });
rect(slide, 380, 220, 250, 110, white); textBox(slide, "Vector DB 검색", 410, 256, 210, 38, { fontSize: 25, bold: true });
rect(slide, 670, 220, 250, 110, white); textBox(slide, "근거 기반 생성", 700, 256, 210, 38, { fontSize: 25, bold: true });
rect(slide, 960, 220, 230, 110, white); textBox(slide, "안전관리 답변", 988, 256, 190, 38, { fontSize: 25, bold: true });
bulletList(slide, ["검색 컬렉션: mine_safety_docs", "통합 chunks: 820개", "답변 항목: 즉시 판단, 조치, 재개 조건, KRAS, 기록·보고"], 120, 420, 990, 21, 52);

slide = pres.slides.add();
slide.background.fill = white;
addTitle(slide, "RAG corpus는 공식 문서 기반 chunks로 구성했다", "평가 질문은 검색 corpus가 아니라 별도 평가 시나리오로 분리해 관리했습니다.");
bulletList(slide, [
  "사용 Vector DB: 10_vector_db_with_major_accident_docs",
  "통합 chunks 파일: 08_chunks/chunks_with_major_accident_docs.jsonl",
  "v2 평가 질문과 답변 파일은 corpus가 아닌 평가용 파일로 보존",
  "leakage 검사에서 Q001~Q110 질문 원문 exact match 0건 확인",
], 90, 210, 1080, 22, 62);

slide = pres.slides.add();
slide.background.fill = white;
addTitle(slide, "중대재해처벌법 자료는 관리체계 관점을 보완한다", "사고 대응을 기술 조치에만 머물지 않게 하고 기록·보고와 재발방지까지 연결합니다.");
bulletList(slide, [
  "안전보건관리체계, 경영책임자 의무, 재발방지 관점 반영",
  "작업중지, 접근통제, 기록·보고, 작업재개 조건을 답변 구조에 포함",
  "법령 조문을 무리하게 단정하기보다 확인 가능한 근거 중심으로 답변",
], 100, 215, 1020, 24, 72);

slide = pres.slides.add();
slide.background.fill = light;
addTitle(slide, "Streamlit 앱은 현장 질문을 조치 중심 답변으로 바꾼다", "시연에서는 질문 입력 후 근거, 우선 조치, 체크리스트, 보고 항목을 순서대로 확인합니다.");
bulletList(slide, [
  "광산 안전 질문 입력과 RAG 검색 기반 답변 생성",
  "즉시 판단, 우선 조치, 작업 재개 조건 자동 정리",
  "KRAS식 위험성평가 초안과 현장 조치 체크리스트 제공",
  "기록·보고 사항을 별도 항목으로 정리",
], 100, 205, 1050, 23, 62);

slide = pres.slides.add();
slide.background.fill = white;
addTitle(slide, "v2 평가는 기존 비교의 공정성 한계를 보완했다", "질문 수, 익명화, 랜덤화, 다중 평가자를 함께 적용했습니다.");
bulletList(slide, [
  "평가 질문을 20문항에서 50문항으로 확장",
  "ChatGPT, Gemini, MineSafe AI 답변을 Answer A/B/C로 블라인드 처리",
  "질문마다 답변 순서를 랜덤화해 위치 편향을 줄임",
  "GPT/Gemini/Claude 3개 평가자가 참여해 총 450개 점수 산출",
  "MISSING_SCORE 0개로 최종 병합본 완성",
], 90, 190, 1060, 22, 58);

slide = pres.slides.add();
slide.background.fill = white;
addTitle(slide, "MineSafe AI가 v2 블라인드 다중평가에서 1위를 기록했다", "평균점수와 표준편차 모두 발표자료에 바로 사용할 수 있도록 정리했습니다.");
await addImage(slide, data.figure_paths.score, 76, 174, 540, 380, "모델별 평균점수 막대그래프");
await addImage(slide, data.figure_paths.std, 660, 174, 500, 360, "모델별 표준편차 막대그래프");
textBox(slide, "MineSafe AI 85.61점 > Gemini 69.91점 > ChatGPT 60.33점", 128, 600, 1000, 38, { fontSize: 24, bold: true, color: navy, alignment: "center" });

slide = pres.slides.add();
slide.background.fill = white;
addTitle(slide, "결론은 긍정적이지만 전문가 검토가 필요하다", "MineSafe AI는 연구용 RAG 안전관리 보조도구로 가능성을 보였고, 현장 적용 전 검증이 필요합니다.");
await addImage(slide, data.figure_paths.judge, 710, 190, 470, 340, "평가자별 모델 평균점수 그래프");
bulletList(slide, [
  "공식 문서 기반 검색과 현장 조치 중심 답변이 강점",
  "AI 평가자 기반 결과이므로 절대적 정답으로 해석하면 안 됨",
  "50문항 평가이며 광산별 작업 조건을 모두 대표하지는 않음",
  "실제 적용 전 광산 안전 전문가와 법률 전문가 검토 필요",
], 90, 205, 570, 21, 60);

slide = pres.slides.add();
slide.background.fill = light;
addTitle(slide, "향후에는 전문가 평가와 현장 피드백으로 신뢰도를 높인다", "최종 목표는 안전관리자가 근거 있는 판단을 더 빠르게 하도록 돕는 보조도구입니다.");
bulletList(slide, [
  "최신 법령과 지침 문서의 지속 업데이트",
  "광종·작업별 시나리오와 평가 질문 확대",
  "광산 안전 전문가 평가를 추가해 AI 평가 한계 보완",
  "현장 피드백 기반 답변 구조와 출처 표시 고도화",
], 100, 215, 1010, 24, 72);

for (const [index, sl] of pres.slides.items.entries()) {
  const png = await pres.export({ slide: sl, format: "png", scale: 1 });
  await fs.writeFile(path.join(qaDir, `slide-${String(index + 1).padStart(2, "0")}.png`), new Uint8Array(await png.arrayBuffer()));
}
const montage = await pres.export({ format: "webp", montage: true, scale: 1 });
await fs.writeFile(path.join(qaDir, "presentation_montage.webp"), new Uint8Array(await montage.arrayBuffer()));
const pptx = await PresentationFile.exportPptx(pres);
await pptx.save(finalPptx);

await fs.writeFile(
  path.join(outRoot, "04_report_text", "final_packaging_generation_report.txt"),
  [
    "MineSafe AI final packaging generation report",
    `final_excel=${finalExcel}`,
    `final_pptx=${finalPptx}`,
    `question_count=${data.question_count}`,
    `raw_score_count=${data.raw_score_count}`,
    `missing_score_count=${data.missing_score_count}`,
    "README backup path=README_backup_before_final_packaging.md",
  ].join("\n"),
  "utf8",
);
'''
    path = build_dir / "build_final_outputs.mjs"
    path.write_text(js.strip() + "\n", encoding="utf-8")
    return path


def write_data_and_build(data: dict) -> None:
    build_dir = OUT_ROOT / "_build_tmp"
    build_dir.mkdir(parents=True, exist_ok=True)
    data_path = build_dir / "final_outputs_data.json"
    data_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    builder = write_js_builder(build_dir)

    node_modules_link = build_dir / "node_modules"
    if not node_modules_link.exists():
        try:
            node_modules_link.symlink_to(NODE_MODULES, target_is_directory=True)
        except OSError:
            subprocess.run(
                ["cmd", "/c", "mklink", "/J", str(node_modules_link), str(NODE_MODULES)],
                cwd=build_dir,
                check=True,
                shell=False,
            )
    subprocess.run([str(NODE_EXE), str(builder)], cwd=build_dir, check=True)


def fix_excel_sheet_names() -> None:
    final_excel = OUT_ROOT / "01_final_excel" / "MineSafe_AI_최종평가_통합보고서.xlsx"
    expected = [
        "00_최종요약",
        "01_질문별_답변_점수",
        "02_모델별_최종점수",
        "03_평가자별_점수",
        "04_그래프용_데이터",
        "05_보고서용_문장",
        "06_한계점",
    ]
    wb = load_workbook(final_excel)
    for index, name in enumerate(expected):
        if index < len(wb.worksheets):
            wb.worksheets[index].title = name
    wb.save(final_excel)
    wb.close()


def remove_extra_inspect_files() -> None:
    for path in OUT_ROOT.rglob("*.inspect.ndjson"):
        try:
            path.unlink()
        except OSError:
            pass


def copy_self() -> None:
    dest = OUT_ROOT / "generate_final_outputs.py"
    src = Path(__file__).resolve()
    if src != dest.resolve():
        shutil.copy2(src, dest)


def verify_outputs() -> dict:
    final_excel = OUT_ROOT / "01_final_excel" / "MineSafe_AI_최종평가_통합보고서.xlsx"
    final_pptx = OUT_ROOT / "03_presentation" / "MineSafe_AI_주간발표자료.pptx"
    wb = load_workbook(final_excel, read_only=False, data_only=True)
    sheets = wb.sheetnames
    q_sheet = next((name for name in sheets if name.startswith("01_")), sheets[1])
    model_sheet = next((name for name in sheets if name.startswith("02_")), sheets[2])
    q_rows = max((wb[q_sheet].max_row or 1) - 1, 0)
    summary_rows = max((wb[model_sheet].max_row or 1) - 1, 0)
    wb.close()

    return {
        "final_excel_exists": final_excel.exists(),
        "final_pptx_exists": final_pptx.exists(),
        "sheet_names": sheets,
        "question_rows": q_rows,
        "model_score_rows": summary_rows,
        "readme_backup_exists": (PROJECT_ROOT / "README_backup_before_final_packaging.md").exists(),
        "copied_figures": sorted(p.name for p in (OUT_ROOT / "02_figures").glob("*.png")),
        "github_records": sorted(p.name for p in (OUT_ROOT / "07_github_records").glob("*.txt")),
    }


def main() -> None:
    ensure_dirs()
    copy_self()
    data = build_data()
    copy_figures_and_records()
    write_readme()
    write_report_texts()
    write_data_and_build(data)
    fix_excel_sheet_names()
    remove_extra_inspect_files()
    verification = verify_outputs()
    (OUT_ROOT / "04_report_text" / "final_packaging_verification.json").write_text(
        json.dumps(verification, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(verification, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
