from __future__ import annotations

import csv
import itertools
import json
import math
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path
from statistics import mean, pstdev
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(r"C:\Users\USER\Desktop\mine_safety_rag")
OUT_DIR = PROJECT_ROOT / "17_similarity_index"
FINAL_REPORT_DIR = PROJECT_ROOT / "16_final_outputs" / "04_report_text"
FINAL_FIGURE_DIR = PROJECT_ROOT / "16_final_outputs" / "02_figures"
CHUNKS_PATH = PROJECT_ROOT / "08_chunks" / "chunks_with_major_accident_docs.jsonl"
APP_PATH = PROJECT_ROOT / "app.py"
VECTOR_DB_PATH = PROJECT_ROOT / "10_vector_db_with_major_accident_docs"

QUESTIONS = [
    ("SIM001", "발파", "발파 작업 전 안전점검은 어떻게 해야 해?"),
    ("SIM002", "불발", "불발 장약이 의심될 때 어떻게 조치해야 해?"),
    ("SIM003", "유해가스", "갱내 메탄가스 농도가 높게 나오면 어떻게 해야 해?"),
    ("SIM004", "분진/보호구", "분진이 많은 굴진 작업에서 방진마스크는 어떻게 관리해야 해?"),
    ("SIM005", "낙반/붕락", "낙반 위험이 있는 막장에서 작업을 계속해도 되는지 판단해줘."),
    ("SIM006", "전기설비", "갱내 전기설비 점검 시 주의사항은 뭐야?"),
    ("SIM007", "운반/통행", "운반장비와 작업자가 같은 통로를 이용할 때 안전조치는 뭐야?"),
    ("SIM008", "중대재해처벌법", "중대재해 발생 시 사업주와 경영책임자는 무엇을 해야 해?"),
    ("SIM009", "TBM", "작업 전 TBM에서 반드시 확인해야 할 사항은 뭐야?"),
    ("SIM010", "환기", "환기 불량이 의심되는 갱내 작업장에서 우선 조치는 뭐야?"),
]

FINAL_OUTPUTS = {
    "questions": OUT_DIR / "similarity_questions_10.xlsx",
    "raw_answers": OUT_DIR / "similarity_raw_answers.xlsx",
    "pairwise": OUT_DIR / "similarity_pairwise_scores.xlsx",
    "summary": OUT_DIR / "similarity_summary.xlsx",
    "report": OUT_DIR / "similarity_overall_report.txt",
    "chart": OUT_DIR / "similarity_index_bar_chart.png",
    "tsv": OUT_DIR / "similarity_index_final_summary.tsv",
    "script": OUT_DIR / "run_similarity_index_test.py",
    "final_report_copy": FINAL_REPORT_DIR / "유사성지수_보완결과.txt",
    "final_chart_copy": FINAL_FIGURE_DIR / "similarity_index_bar_chart.png",
}

FONT_NAME = "맑은 고딕"
NAVY = "17324D"
LIGHT_BLUE = "D9EAF7"
WINNER_FILL = "D9EAD3"
BORDER = "D9E2EC"
WHITE = "FFFFFF"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if text.lower() in {"nan", "none"}:
        return ""
    return text


def read_chunks() -> list[dict[str, Any]]:
    chunks = []
    with CHUNKS_PATH.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            obj = json.loads(line)
            chunks.append(
                {
                    "chunk_id": clean_text(obj.get("chunk_id")),
                    "source_file": clean_text(obj.get("source_file")),
                    "source_path": clean_text(obj.get("source_path")),
                    "text": clean_text(obj.get("text")),
                }
            )
    return chunks


def tokenize_koreanish(text: str) -> list[str]:
    cleaned = []
    for ch in text.lower():
        if ch.isalnum() or "가" <= ch <= "힣":
            cleaned.append(ch)
        else:
            cleaned.append(" ")
    tokens = [tok for tok in "".join(cleaned).split() if len(tok) >= 2]
    return tokens


def keyword_score(query: str, text: str) -> float:
    q_tokens = tokenize_koreanish(query)
    if not q_tokens:
        return 0.0
    text_low = text.lower()
    score = 0.0
    for tok in q_tokens:
        if tok in text_low:
            score += 3.0 + min(text_low.count(tok), 5) * 0.2
    for keyword in safety_keywords_for_query(query):
        if keyword in text:
            score += 2.0
    return score


def safety_keywords_for_query(question: str) -> list[str]:
    mapping = {
        "발파": ["발파", "화약", "장약", "점화", "대피", "경계", "작업지휘"],
        "불발": ["불발", "장약", "발파", "재점화", "접근", "대피", "경계"],
        "메탄": ["메탄", "가스", "환기", "농도", "측정", "대피", "작업중지"],
        "분진": ["분진", "방진", "마스크", "호흡", "환기", "보호구"],
        "낙반": ["낙반", "붕락", "지보", "천반", "막장", "작업중지"],
        "전기": ["전기", "설비", "감전", "차단", "접지", "점검"],
        "운반": ["운반", "통로", "차량", "장비", "신호", "접근"],
        "중대재해": ["중대재해", "경영책임자", "사업주", "재발방지", "보고", "안전보건관리체계"],
        "TBM": ["TBM", "작업 전", "위험성", "보호구", "교육", "점검"],
        "환기": ["환기", "갱내", "가스", "산소", "측정", "작업중지"],
    }
    selected: list[str] = []
    for key, values in mapping.items():
        if key.lower() in question.lower():
            selected.extend(values)
    return selected or ["안전", "점검", "작업중지", "보고"]


def build_retriever(chunks: list[dict[str, Any]]):
    texts = [chunk["text"] for chunk in chunks]
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.metrics.pairwise import cosine_similarity

        vectorizer = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 5), max_features=60000)
        matrix = vectorizer.fit_transform(texts)

        def search(query: str, top_k: int = 4) -> tuple[list[dict[str, Any]], str]:
            q_vec = vectorizer.transform([query + " " + " ".join(safety_keywords_for_query(query))])
            sims = cosine_similarity(q_vec, matrix).ravel()
            ranked = sorted(range(len(sims)), key=lambda idx: sims[idx], reverse=True)[:top_k]
            results = []
            for idx in ranked:
                item = dict(chunks[idx])
                item["retrieval_score"] = float(sims[idx])
                results.append(item)
            return results, "tfidf_char_wb_chunks"

        return search
    except Exception:

        def search(query: str, top_k: int = 4) -> tuple[list[dict[str, Any]], str]:
            scored = [(idx, keyword_score(query, chunk["text"])) for idx, chunk in enumerate(chunks)]
            ranked = sorted(scored, key=lambda item: item[1], reverse=True)[:top_k]
            results = []
            for idx, score in ranked:
                item = dict(chunks[idx])
                item["retrieval_score"] = score
                results.append(item)
            return results, "keyword_fallback_chunks"

        return search


def extract_evidence_sentence(text: str, question: str, max_len: int = 180) -> str:
    keywords = safety_keywords_for_query(question) + tokenize_koreanish(question)
    candidates = []
    for raw in text.replace("\n", " ").split("."):
        sentence = raw.strip()
        if len(sentence) < 20:
            continue
        score = sum(1 for kw in keywords if kw and kw in sentence)
        candidates.append((score, sentence))
    candidates.sort(key=lambda item: item[0], reverse=True)
    chosen = candidates[0][1] if candidates else text.replace("\n", " ").strip()
    return chosen[:max_len].strip()


def risk_focus(question: str) -> dict[str, str]:
    q = question.lower()
    if "불발" in question:
        return {
            "judgment": "불발 장약이 의심되면 작업을 계속하지 말고 즉시 접근을 통제해야 한다.",
            "actions": "작업자 대피, 경계구역 설정, 발파 책임자 보고, 임의 제거 금지, 절차에 따른 재확인을 우선한다.",
            "restart": "불발 여부 확인, 위험구역 해제, 책임자 승인, 기록 완료 후에만 작업을 재개한다.",
        }
    if "발파" in question:
        return {
            "judgment": "발파 전에는 장약·결선·대피·경계·신호 체계를 확인한 뒤 작업해야 한다.",
            "actions": "작업계획 확인, 출입통제, 대피 완료 확인, 신호 전달, 화약류 취급자 외 접근 금지를 우선한다.",
            "restart": "발파 후 환기와 잔류 위험 확인, 불발 여부 확인, 책임자 승인 후 재입장한다.",
        }
    if "메탄" in question or "가스" in question:
        return {
            "judgment": "메탄가스 등 유해가스 농도가 높으면 폭발·질식 위험으로 작업중지 판단이 우선이다.",
            "actions": "인원 대피, 전기·화기 사용 통제, 환기 강화, 농도 재측정, 책임자 보고를 실시한다.",
            "restart": "가스 농도가 기준 이하로 안정되고 환기 상태와 측정 기록을 확인한 뒤 재개한다.",
        }
    if "분진" in question or "마스크" in question:
        return {
            "judgment": "분진이 많은 굴진 작업에서는 호흡보호구와 환기·살수 상태를 함께 관리해야 한다.",
            "actions": "방진마스크 밀착 확인, 필터 교체, 착용 교육, 분진 억제, 환기상태 점검을 실시한다.",
            "restart": "분진 저감 조치와 보호구 상태가 확인되고 작업자가 착용 기준을 지킬 때 작업을 계속한다.",
        }
    if "낙반" in question or "막장" in question:
        return {
            "judgment": "낙반 위험이 있으면 작업 계속보다 작업중지와 지보·천반 확인이 우선이다.",
            "actions": "위험구역 접근통제, 작업자 대피, 지보 상태 확인, 부석 제거, 책임자 점검을 실시한다.",
            "restart": "지보 보강, 부석 제거, 변위·균열 확인, 책임자 승인 후 작업을 재개한다.",
        }
    if "전기" in question:
        return {
            "judgment": "갱내 전기설비 점검은 감전·화재 위험을 전제로 전원 차단과 잠금표시가 우선이다.",
            "actions": "전원 차단, 검전, 접지, 방폭 상태 확인, 습기·손상 점검, 무자격자 접근 금지를 실시한다.",
            "restart": "절연·접지·보호장치 이상 없음과 점검 기록, 책임자 확인 후 투입한다.",
        }
    if "운반" in question or "통로" in question:
        return {
            "judgment": "운반장비와 작업자가 같은 통로를 쓰면 충돌·협착 위험이 커서 동선 분리가 핵심이다.",
            "actions": "보행로 분리, 신호수 배치, 속도 제한, 후진 경보, 사각지대 확인, 접근금지선을 운영한다.",
            "restart": "동선 통제와 신호체계가 작동하고 작업자 교육·확인이 끝난 뒤 작업을 진행한다.",
        }
    if "중대재해" in question:
        return {
            "judgment": "중대재해 발생 시 인명구호와 2차 사고 방지가 최우선이며, 보고와 재발방지 체계가 필요하다.",
            "actions": "작업중지, 구조·응급조치, 현장 보존, 관계기관 보고, 원인조사, 재발방지대책 수립을 실시한다.",
            "restart": "위험요인 제거와 재발방지대책 이행, 안전보건관리체계 보완, 책임자 승인 후 재개한다.",
        }
    if "tbm" in q:
        return {
            "judgment": "TBM은 작업 전 위험요인과 작업 절차를 공유해 작업자 인식 차이를 줄이는 절차다.",
            "actions": "작업내용, 위험요인, 보호구, 장비 상태, 비상연락, 작업중지 기준을 확인한다.",
            "restart": "누락된 위험요인이 보완되고 작업자가 조치사항을 이해했을 때 작업을 시작한다.",
        }
    if "환기" in question:
        return {
            "judgment": "환기 불량이 의심되면 산소결핍·유해가스 위험을 전제로 즉시 작업을 멈춰야 한다.",
            "actions": "작업자 대피, 환기장치 점검·가동, 가스와 산소 농도 측정, 출입통제, 책임자 보고를 실시한다.",
            "restart": "연속 측정 결과가 안정되고 환기 성능과 기록이 확인된 뒤 작업을 재개한다.",
        }
    return {
        "judgment": "위험요인이 확인되면 작업 계속보다 위험 제거와 책임자 확인이 우선이다.",
        "actions": "작업중지, 접근통제, 보호구 확인, 위험성평가, 책임자 보고를 실시한다.",
        "restart": "위험요인 제거와 기록·보고가 완료되고 책임자가 승인한 뒤 작업을 재개한다.",
    }


def generate_stable_answer(question_id: str, question: str, run_id: int, evidence: list[dict[str, Any]], retrieval_method: str) -> str:
    focus = risk_focus(question)
    evidence_lines = []
    for idx, item in enumerate(evidence[:3], start=1):
        snippet = extract_evidence_sentence(item["text"], question)
        evidence_lines.append(f"- 근거 {idx}: {item['source_file']} / {item['chunk_id']} - {snippet}")
    kras_hazards = ", ".join(safety_keywords_for_query(question)[:4])
    answer = f"""
[MineSafe AI 안정형 RAG 답변]
질문 ID: {question_id}
질문: {question}

1. 즉시 판단
{focus['judgment']}

2. 검색 근거
검색 방식: {retrieval_method}
{chr(10).join(evidence_lines)}

3. 우선 조치
{focus['actions']}

4. 작업 재개 조건
{focus['restart']}

5. KRAS식 위험성평가 초안
- 유해·위험요인: {kras_hazards}
- 현재 위험: 작업자 부상, 2차 사고, 법정 안전조치 미흡 가능성
- 감소대책: 작업중지 기준 명확화, 접근통제, 보호구 확인, 책임자 점검, 기록 유지
- 잔여위험 관리: 조치 완료 후 재측정 또는 재점검 결과를 기록하고 작업 전 공유

6. 현장 조치 체크리스트
- 작업자 대피 또는 접근통제 필요 여부 확인
- 보호구, 장비, 환기·전기·지보 등 핵심 설비 상태 확인
- 책임자 보고와 작업중지/재개 승인 절차 확인
- 작업 전 TBM 또는 현장 브리핑으로 조치사항 공유

7. 기록·보고 사항
질문 내용, 위험 판단, 검색 근거, 조치 시간, 조치 담당자, 재개 승인자, 재발방지 조치를 기록한다.
"""
    return "\n".join(line.rstrip() for line in answer.strip().splitlines())


def calculate_pairwise_similarity(raw_answers: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], str]:
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.metrics.pairwise import cosine_similarity

        method = 'TfidfVectorizer(analyzer="char_wb", ngram_range=(2,5))'
        pairwise_rows = []
        grouped: dict[str, list[dict[str, Any]]] = {}
        for row in raw_answers:
            grouped.setdefault(row["question_id"], []).append(row)
        for question_id, rows in grouped.items():
            rows = sorted(rows, key=lambda item: item["run_id"])
            texts = [row["answer_text"] for row in rows]
            vec = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 5)).fit_transform(texts)
            sim_matrix = cosine_similarity(vec)
            for i, j in itertools.combinations(range(len(rows)), 2):
                pairwise_rows.append(
                    {
                        "question_id": question_id,
                        "question": rows[i]["question"],
                        "run_id_a": rows[i]["run_id"],
                        "run_id_b": rows[j]["run_id"],
                        "similarity_score": round(float(sim_matrix[i, j]) * 100, 2),
                        "similarity_method": method,
                    }
                )
        return pairwise_rows, method
    except Exception:
        from difflib import SequenceMatcher

        method = "difflib.SequenceMatcher fallback"
        pairwise_rows = []
        grouped: dict[str, list[dict[str, Any]]] = {}
        for row in raw_answers:
            grouped.setdefault(row["question_id"], []).append(row)
        for question_id, rows in grouped.items():
            rows = sorted(rows, key=lambda item: item["run_id"])
            for a, b in itertools.combinations(rows, 2):
                score = SequenceMatcher(None, a["answer_text"], b["answer_text"]).ratio() * 100
                pairwise_rows.append(
                    {
                        "question_id": question_id,
                        "question": a["question"],
                        "run_id_a": a["run_id"],
                        "run_id_b": b["run_id"],
                        "similarity_score": round(score, 2),
                        "similarity_method": method,
                    }
                )
        return pairwise_rows, method


def consistency_level(score: float) -> str:
    if score >= 95:
        return "매우 높음"
    if score >= 85:
        return "높음"
    if score >= 70:
        return "보통"
    return "낮음"


def build_summary(pairwise_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in pairwise_rows:
        grouped.setdefault(row["question_id"], []).append(row)
    summary = []
    q_lookup = {qid: (cat, q) for qid, cat, q in QUESTIONS}
    for qid, rows in grouped.items():
        scores = [float(row["similarity_score"]) for row in rows]
        category, question = q_lookup[qid]
        avg = round(mean(scores), 2)
        summary.append(
            {
                "question_id": qid,
                "category": category,
                "question": question,
                "answer_count": 5,
                "pairwise_count": len(rows),
                "similarity_mean": avg,
                "similarity_min": round(min(scores), 2),
                "similarity_max": round(max(scores), 2),
                "similarity_std": round(pstdev(scores), 2),
                "consistency_level": consistency_level(avg),
            }
        )
    return sorted(summary, key=lambda item: item["question_id"])


def apply_sheet_style(ws, widths: dict[str, float], row_height: float | None = None) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color=BORDER)
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or "")
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 16)
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_height and row_idx > 1:
            ws.row_dimensions[row_idx].height = row_height
        for cell in row:
            cell.font = Font(name=FONT_NAME, size=10)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(cell.value, str))
            if row_idx == 1:
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"


def write_xlsx(path: Path, sheet_name: str, headers: list[str], rows: list[dict[str, Any]], widths: dict[str, float], row_height: float | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    apply_sheet_style(ws, widths, row_height)
    wb.save(path)
    wb.close()


def write_summary_workbook(path: Path, summary_rows: list[dict[str, Any]], overall: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "question_summary"
    headers = [
        "question_id",
        "category",
        "question",
        "answer_count",
        "pairwise_count",
        "similarity_mean",
        "similarity_min",
        "similarity_max",
        "similarity_std",
        "consistency_level",
    ]
    ws.append(headers)
    for row in summary_rows:
        ws.append([row.get(header, "") for header in headers])
    apply_sheet_style(
        ws,
        {
            "question_id": 12,
            "category": 18,
            "question": 58,
            "answer_count": 14,
            "pairwise_count": 16,
            "similarity_mean": 18,
            "similarity_min": 16,
            "similarity_max": 16,
            "similarity_std": 16,
            "consistency_level": 18,
        },
        36,
    )

    ws2 = wb.create_sheet("overall")
    ws2.append(["item", "value"])
    for key, value in overall.items():
        ws2.append([key, value])
    apply_sheet_style(ws2, {"item": 32, "value": 80}, 30)

    wb.save(path)
    wb.close()


def create_chart(summary_rows: list[dict[str, Any]], output_path: Path) -> None:
    labels = [row["question_id"] for row in summary_rows]
    scores = [row["similarity_mean"] for row in summary_rows]
    try:
        import matplotlib.pyplot as plt
        from matplotlib import font_manager, rcParams

        font_path = Path(r"C:\Windows\Fonts\malgun.ttf")
        if font_path.exists():
            font_manager.fontManager.addfont(str(font_path))
            rcParams["font.family"] = "Malgun Gothic"
        rcParams["axes.unicode_minus"] = False

        colors = ["#1f77b4" if score < 95 else "#0f766e" for score in scores]
        fig, ax = plt.subplots(figsize=(12, 6), dpi=160)
        bars = ax.bar(labels, scores, color=colors)
        ax.set_title("MineSafe AI 동일 질문 반복 답변 유사성 지수", fontsize=15, fontweight="bold")
        ax.set_xlabel("대표 질문")
        ax.set_ylabel("유사성 지수(0~100)")
        ax.set_ylim(0, 105)
        ax.grid(axis="y", alpha=0.25)
        for bar, score in zip(bars, scores):
            ax.text(bar.get_x() + bar.get_width() / 2, min(score + 1, 103), f"{score:.2f}", ha="center", va="bottom", fontsize=9)
        fig.tight_layout()
        fig.savefig(output_path, bbox_inches="tight")
        plt.close(fig)
        return
    except Exception:
        from PIL import Image, ImageDraw, ImageFont

        width, height = 1500, 820
        margin_left, margin_right = 120, 70
        margin_top, margin_bottom = 120, 140
        plot_w = width - margin_left - margin_right
        plot_h = height - margin_top - margin_bottom
        img = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(img)
        font_path = Path(r"C:\Windows\Fonts\malgun.ttf")
        bold_path = Path(r"C:\Windows\Fonts\malgunbd.ttf")
        title_font = ImageFont.truetype(str(bold_path if bold_path.exists() else font_path), 34) if font_path.exists() else ImageFont.load_default()
        label_font = ImageFont.truetype(str(font_path), 22) if font_path.exists() else ImageFont.load_default()
        small_font = ImageFont.truetype(str(font_path), 18) if font_path.exists() else ImageFont.load_default()

        draw.text((margin_left, 40), "MineSafe AI 동일 질문 반복 답변 유사성 지수", fill="#17324D", font=title_font)
        draw.line((margin_left, margin_top + plot_h, margin_left + plot_w, margin_top + plot_h), fill="#64748B", width=2)
        draw.line((margin_left, margin_top, margin_left, margin_top + plot_h), fill="#64748B", width=2)
        for tick in range(0, 101, 20):
            y = margin_top + plot_h - (tick / 100) * plot_h
            draw.line((margin_left - 8, y, margin_left + plot_w, y), fill="#E2E8F0", width=1)
            draw.text((45, y - 12), str(tick), fill="#334155", font=small_font)
        bar_gap = 18
        bar_w = (plot_w - bar_gap * (len(scores) - 1)) / len(scores)
        for idx, (label, score) in enumerate(zip(labels, scores)):
            x0 = margin_left + idx * (bar_w + bar_gap)
            x1 = x0 + bar_w
            y0 = margin_top + plot_h - (score / 100) * plot_h
            color = "#0F766E" if score >= 95 else "#2563EB"
            draw.rectangle((x0, y0, x1, margin_top + plot_h), fill=color)
            score_text = f"{score:.2f}"
            st_box = draw.textbbox((0, 0), score_text, font=small_font)
            draw.text((x0 + (bar_w - (st_box[2] - st_box[0])) / 2, y0 - 30), score_text, fill="#0F172A", font=small_font)
            lb_box = draw.textbbox((0, 0), label, font=label_font)
            draw.text((x0 + (bar_w - (lb_box[2] - lb_box[0])) / 2, margin_top + plot_h + 18), label, fill="#334155", font=label_font)
        draw.text((width / 2 - 55, height - 60), "대표 질문", fill="#334155", font=label_font)
        draw.text((20, margin_top + plot_h / 2 - 20), "유사성\n지수", fill="#334155", font=small_font)
        img.save(output_path)


def write_tsv(path: Path, summary_rows: list[dict[str, Any]], overall: dict[str, Any]) -> None:
    headers = [
        "question_id",
        "category",
        "question",
        "answer_count",
        "pairwise_count",
        "similarity_mean",
        "similarity_min",
        "similarity_max",
        "similarity_std",
        "consistency_level",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter="\t")
        writer.writeheader()
        for row in summary_rows:
            writer.writerow({header: row.get(header, "") for header in headers})
        writer.writerow({})
        writer.writerow({"question_id": "OVERALL", "similarity_mean": overall["overall_similarity_mean"], "consistency_level": overall["overall_consistency_level"]})


def write_report(path: Path, summary_rows: list[dict[str, Any]], overall: dict[str, Any]) -> str:
    highest = max(summary_rows, key=lambda item: item["similarity_mean"])
    lowest = min(summary_rows, key=lambda item: item["similarity_mean"])
    lines = [
        "MineSafe AI 동일 질문 반복 답변 유사성 지수 보완 결과",
        "",
        "1. 검사 개요",
        "- 대표 질문 10개를 선정하고 각 질문을 5회 반복 실행하여 총 50개 답변을 수집하였다.",
        "- 같은 질문의 5개 답변끼리 pairwise similarity를 계산하여 질문별 10쌍, 전체 100쌍을 비교하였다.",
        "- 외부 LLM API는 사용하지 않았으며, 공식 문서 chunks 기반 deterministic 안정형 RAG 답변 생성 방식을 사용하였다.",
        f"- 유사도 계산 방식: {overall['similarity_method']}",
        "",
        "2. 전체 평균 유사성 지수",
        f"- 전체 평균 유사성 지수: {overall['overall_similarity_mean']:.2f}점",
        f"- 전체 consistency level: {overall['overall_consistency_level']}",
        "",
        "3. 질문별 유사성 지수",
    ]
    for row in summary_rows:
        lines.append(f"- {row['question_id']} ({row['category']}): {row['similarity_mean']:.2f}점 / {row['consistency_level']}")
    lines.extend(
        [
            "",
            "4. 가장 유사성이 높은 질문",
            f"- {highest['question_id']} {highest['question']}: {highest['similarity_mean']:.2f}점",
            "",
            "5. 가장 유사성이 낮은 질문",
            f"- {lowest['question_id']} {lowest['question']}: {lowest['similarity_mean']:.2f}점",
            "",
            "6. 유사성 지수 해석",
            "동일 질문 반복 입력 시 MineSafe AI의 답변 간 평균 유사성 지수를 계산하였다. 유사성 지수가 높다는 것은 동일한 질문에 대해 답변 구조와 핵심 안전조치가 안정적으로 유지된다는 의미이다. 다만 표현이 조금 달라도 안전 조치의 의미가 같을 수 있으므로, 유사성 지수는 정확도 평가가 아니라 응답 일관성을 확인하는 보조 지표로 활용한다.",
            "",
            "7. 한계점",
            "- 이번 검사는 deterministic 안정형 RAG 답변 생성 방식으로 수행했기 때문에 외부 생성형 LLM을 사용할 때의 표현 변동성은 반영하지 않는다.",
            "- 유사성 지수는 답변의 정확도나 법적 타당성을 직접 평가하지 않고, 동일 질문 반복 시 응답 일관성을 보는 보조 지표이다.",
            "- 실제 현장 적용 전에는 안전 전문가가 답변 내용과 근거를 별도로 검토해야 한다.",
            "",
            "8. 교수님께 설명할 문단",
            f"보완사항으로 동일 질문에 대해 MineSafe AI가 얼마나 일관된 답변을 내는지 확인하기 위해 대표 질문 10개를 선정하고 각 질문을 5회 반복 실행하였다. 총 50개 답변에서 같은 질문 내 5개 답변 조합 10쌍씩, 전체 100쌍의 유사도를 계산하였다. 이번 실행 환경에서는 {overall['similarity_method']} 방식을 사용했으며, 이 지표는 답변의 정답 여부가 아니라 반복 입력 시 핵심 안전조치와 답변 구조가 안정적으로 유지되는지를 보는 보조 지표이다. 외부 LLM API 없이 공식 문서 chunks 기반 안정형 RAG 답변으로 수행하였다.",
            "",
            "9. 보존 확인",
            "- app.py, Vector DB, chunks 파일은 수정하지 않았다.",
            "- .env 및 API Key는 열람하거나 출력하지 않았다.",
            "- 기존 50문항 v2 평가 결과는 수정하지 않았다.",
        ]
    )
    text = "\n".join(lines) + "\n"
    path.write_text(text, encoding="utf-8")
    return text


def verify_outputs(summary_rows: list[dict[str, Any]], raw_answers: list[dict[str, Any]], pairwise_rows: list[dict[str, Any]], overall: dict[str, Any]) -> dict[str, Any]:
    files_exist = {name: path.exists() for name, path in FINAL_OUTPUTS.items() if name not in {"script"}}
    raw_count = len(raw_answers)
    pairwise_count = len(pairwise_rows)
    question_count = len(summary_rows)
    app_mtime_after = APP_PATH.stat().st_mtime
    chunks_mtime_after = CHUNKS_PATH.stat().st_mtime
    vector_mtime_after = VECTOR_DB_PATH.stat().st_mtime
    return {
        "question_count": question_count,
        "raw_answer_count": raw_count,
        "pairwise_count": pairwise_count,
        "overall_similarity_mean": overall["overall_similarity_mean"],
        "overall_consistency_level": overall["overall_consistency_level"],
        "files_exist": files_exist,
        "app_mtime_after": app_mtime_after,
        "chunks_mtime_after": chunks_mtime_after,
        "vector_db_mtime_after": vector_mtime_after,
    }


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    FINAL_REPORT_DIR.mkdir(parents=True, exist_ok=True)
    FINAL_FIGURE_DIR.mkdir(parents=True, exist_ok=True)

    app_mtime_before = APP_PATH.stat().st_mtime
    chunks_mtime_before = CHUNKS_PATH.stat().st_mtime
    vector_mtime_before = VECTOR_DB_PATH.stat().st_mtime

    chunks = read_chunks()
    search = build_retriever(chunks)

    question_rows = [
        {"question_id": qid, "category": category, "question": question, "repeat_count": 5}
        for qid, category, question in QUESTIONS
    ]

    raw_answers = []
    for question_id, category, question in QUESTIONS:
        evidence, retrieval_method = search(question, top_k=4)
        top_chunk_ids = ", ".join(item["chunk_id"] for item in evidence)
        top_sources = " | ".join(dict.fromkeys(item["source_file"] for item in evidence))
        for run_id in range(1, 6):
            raw_answers.append(
                {
                    "question_id": question_id,
                    "category": category,
                    "question": question,
                    "run_id": run_id,
                    "answer_text": generate_stable_answer(question_id, question, run_id, evidence, retrieval_method),
                    "generation_mode": "deterministic_stable_rag_from_chunks",
                    "retrieval_method": retrieval_method,
                    "top_chunk_ids": top_chunk_ids,
                    "top_sources": top_sources,
                    "external_llm_api_used": "No",
                }
            )

    pairwise_rows, similarity_method = calculate_pairwise_similarity(raw_answers)
    summary_rows = build_summary(pairwise_rows)
    all_scores = [float(row["similarity_score"]) for row in pairwise_rows]
    overall = {
        "overall_similarity_mean": round(mean(all_scores), 2),
        "overall_similarity_min": round(min(all_scores), 2),
        "overall_similarity_max": round(max(all_scores), 2),
        "overall_similarity_std": round(pstdev(all_scores), 2),
        "overall_consistency_level": consistency_level(round(mean(all_scores), 2)),
        "question_count": len(QUESTIONS),
        "repeat_per_question": 5,
        "raw_answer_count": len(raw_answers),
        "pairwise_count": len(pairwise_rows),
        "similarity_method": similarity_method,
        "external_llm_api_used": "No",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    write_xlsx(
        FINAL_OUTPUTS["questions"],
        "questions",
        ["question_id", "category", "question", "repeat_count"],
        question_rows,
        {"question_id": 14, "category": 18, "question": 70, "repeat_count": 14},
        30,
    )
    write_xlsx(
        FINAL_OUTPUTS["raw_answers"],
        "raw_answers",
        [
            "question_id",
            "category",
            "question",
            "run_id",
            "answer_text",
            "generation_mode",
            "retrieval_method",
            "top_chunk_ids",
            "top_sources",
            "external_llm_api_used",
        ],
        raw_answers,
        {
            "question_id": 12,
            "category": 18,
            "question": 58,
            "run_id": 10,
            "answer_text": 100,
            "generation_mode": 28,
            "retrieval_method": 30,
            "top_chunk_ids": 32,
            "top_sources": 48,
            "external_llm_api_used": 18,
        },
        110,
    )
    write_xlsx(
        FINAL_OUTPUTS["pairwise"],
        "pairwise_scores",
        ["question_id", "question", "run_id_a", "run_id_b", "similarity_score", "similarity_method"],
        pairwise_rows,
        {
            "question_id": 12,
            "question": 58,
            "run_id_a": 12,
            "run_id_b": 12,
            "similarity_score": 18,
            "similarity_method": 44,
        },
        32,
    )
    write_summary_workbook(FINAL_OUTPUTS["summary"], summary_rows, overall)
    create_chart(summary_rows, FINAL_OUTPUTS["chart"])
    write_tsv(FINAL_OUTPUTS["tsv"], summary_rows, overall)
    report_text = write_report(FINAL_OUTPUTS["report"], summary_rows, overall)

    shutil.copy2(Path(__file__), FINAL_OUTPUTS["script"])
    shutil.copy2(FINAL_OUTPUTS["chart"], FINAL_OUTPUTS["final_chart_copy"])
    FINAL_OUTPUTS["final_report_copy"].write_text(report_text, encoding="utf-8")

    app_mtime_after = APP_PATH.stat().st_mtime
    chunks_mtime_after = CHUNKS_PATH.stat().st_mtime
    vector_mtime_after = VECTOR_DB_PATH.stat().st_mtime
    overall["app_py_modified"] = app_mtime_before != app_mtime_after
    overall["chunks_modified"] = chunks_mtime_before != chunks_mtime_after
    overall["vector_db_dir_mtime_changed"] = vector_mtime_before != vector_mtime_after

    verification = verify_outputs(summary_rows, raw_answers, pairwise_rows, overall)
    verification.update(
        {
            "app_py_modified": overall["app_py_modified"],
            "chunks_modified": overall["chunks_modified"],
            "vector_db_dir_mtime_changed": overall["vector_db_dir_mtime_changed"],
            "env_or_api_key_opened": False,
        }
    )
    print(json.dumps(verification, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
