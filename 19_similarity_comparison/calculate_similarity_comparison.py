from __future__ import annotations

from collections import defaultdict
from difflib import SequenceMatcher
from itertools import combinations
from pathlib import Path
import math
import statistics
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
SIM_DIR = ROOT / "19_similarity_comparison"
DEFAULT_INPUT = SIM_DIR / "similarity_answer_collection_template.xlsx"
MODELS = ["ChatGPT", "Gemini", "MineSafe AI"]


def style_sheet(ws) -> None:
    thin = Side(style="thin", color="D9E2EC")
    header_fill = PatternFill("solid", fgColor="17324D")
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx > 1:
            ws.row_dimensions[row_idx].height = 30
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if row_idx == 1:
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, idx).value or "")
        width = 18
        if "question" in header:
            width = 42
        elif "answer" in header or "comment" in header:
            width = 44
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def save_xlsx(path: Path, sheet_name: str, headers: list[str], rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    wb.save(path)
    wb.close()


def font(size: int, bold: bool = False):
    candidates = [
        Path(r"C:\Windows\Fonts\malgunbd.ttf" if bold else r"C:\Windows\Fonts\malgun.ttf"),
        Path(r"C:\Windows\Fonts\NanumGothicBold.ttf" if bold else r"C:\Windows\Fonts\NanumGothic.ttf"),
    ]
    for path in candidates:
        if path.exists():
            return ImageFont.truetype(str(path), size)
    return ImageFont.load_default()


def similarity(a: str, b: str) -> float:
    return round(SequenceMatcher(None, a, b).ratio() * 100, 2)


def level(score: float | None) -> str:
    if score is None or math.isnan(score):
        return "NEED_INPUT"
    if score >= 95:
        return "매우 높음"
    if score >= 85:
        return "높음"
    if score >= 70:
        return "보통"
    return "낮음"


def read_input(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    required = ["model", "question_id", "repeat_no", "question", "answer"]
    missing = [header for header in required if header not in headers]
    if missing:
        raise RuntimeError(f"필수 컬럼 누락: {missing}")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        rows.append(
            {
                "model": str(data.get("model", "")).strip(),
                "question_id": str(data.get("question_id", "")).strip(),
                "repeat_no": int(data.get("repeat_no", 0) or 0),
                "question": str(data.get("question", "") or "").strip(),
                "answer": str(data.get("answer", "") or "").strip(),
            }
        )
    wb.close()
    return rows


def draw_bar_chart(model_rows: list[list]) -> None:
    path = SIM_DIR / "similarity_comparison_final_bar_chart.png"
    img = Image.new("RGB", (1200, 720), "#ffffff")
    draw = ImageDraw.Draw(img)
    title_font = font(30, True)
    label_font = font(20, True)
    text_font = font(18)
    small_font = font(15)
    left, top, right, bottom = 120, 110, 1130, 600
    draw.text((left, 25), "동일 질문 반복 답변 유사성 지수 비교", fill="#17324d", font=title_font)
    draw.rectangle([left, top, right, bottom], outline="#cbd5e1", width=2)
    for score in range(0, 101, 20):
        y = bottom - int((score / 100) * (bottom - top))
        draw.line([left, y, right, y], fill="#e2e8f0", width=1)
        draw.text((left - 45, y - 10), str(score), fill="#475569", font=small_font)
    models = [row[0] for row in model_rows]
    scores = [float(row[4]) if row[4] != "NEED_INPUT" else 0 for row in model_rows]
    statuses = [row[8] for row in model_rows]
    colors = ["#94a3b8" if status == "NEED_INPUT" else "#2563eb" for status in statuses]
    bar_w = 150
    gap = (right - left - bar_w * len(models)) // (len(models) + 1)
    for idx, (model, score, status, color) in enumerate(zip(models, scores, statuses, colors)):
        x1 = left + gap * (idx + 1) + bar_w * idx
        x2 = x1 + bar_w
        y = bottom - int((score / 100) * (bottom - top))
        draw.rounded_rectangle([x1, y, x2, bottom], radius=8, fill=color)
        label = "NEED_INPUT" if status == "NEED_INPUT" else f"{score:.2f}"
        draw.text((x1 + bar_w / 2, max(y - 26, top + 20)), label, fill="#0f172a", font=text_font, anchor="mm")
        draw.text((x1 + bar_w / 2, bottom + 30), model, fill="#17324d", font=label_font, anchor="mm")
    img.save(path)


def main() -> int:
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    rows = read_input(input_path)
    grouped = defaultdict(list)
    question_text = {}
    for row in rows:
        if row["model"] and row["question_id"] and row["repeat_no"]:
            grouped[(row["model"], row["question_id"])].append(row)
            question_text[row["question_id"]] = row["question"]

    pairwise_rows = []
    question_rows = []
    model_scores = defaultdict(list)
    model_expected_groups = defaultdict(int)
    model_complete_groups = defaultdict(int)
    model_ready_answers = defaultdict(int)
    question_ids = sorted({row["question_id"] for row in rows if row["question_id"]})

    for model in MODELS:
        for qid in question_ids:
            group = sorted(grouped.get((model, qid), []), key=lambda x: x["repeat_no"])
            model_expected_groups[model] += 1
            answers_by_repeat = {row["repeat_no"]: row["answer"] for row in group if row["answer"]}
            model_ready_answers[model] += len(answers_by_repeat)
            missing_repeats = [str(i) for i in range(1, 6) if not answers_by_repeat.get(i)]
            if missing_repeats:
                question_rows.append([model, qid, question_text.get(qid, ""), len(answers_by_repeat), "NEED_INPUT", "", "", "", "NEED_INPUT", f"missing repeat_no: {', '.join(missing_repeats)}"])
                continue
            scores = []
            for a, b in combinations(range(1, 6), 2):
                score = similarity(answers_by_repeat[a], answers_by_repeat[b])
                scores.append(score)
                pairwise_rows.append([model, qid, a, b, score])
            q_mean = round(statistics.mean(scores), 2)
            q_std = round(statistics.stdev(scores), 2) if len(scores) > 1 else 0
            question_rows.append([model, qid, question_text.get(qid, ""), 5, 10, q_mean, q_std, min(scores), max(scores), level(q_mean)])
            model_scores[model].extend(scores)
            model_complete_groups[model] += 1

    model_rows = []
    for model in MODELS:
        scores = model_scores.get(model, [])
        if scores:
            mean_score = round(statistics.mean(scores), 2)
            std_score = round(statistics.stdev(scores), 2) if len(scores) > 1 else 0
            min_score = round(min(scores), 2)
            max_score = round(max(scores), 2)
            status = "COMPLETE" if model_complete_groups[model] == len(question_ids) else "PARTIAL"
            level_text = level(mean_score)
        else:
            mean_score = std_score = min_score = max_score = "NEED_INPUT"
            status = "NEED_INPUT"
            level_text = "NEED_INPUT"
        model_rows.append([model, len(question_ids) * 5, model_ready_answers[model], model_complete_groups[model], len(scores), mean_score, std_score, min_score, max_score, level_text, status])

    save_xlsx(
        SIM_DIR / "similarity_comparison_pairwise_scores.xlsx",
        "pairwise_scores",
        ["model", "question_id", "repeat_a", "repeat_b", "similarity_score"],
        pairwise_rows if pairwise_rows else [["", "", "", "", "NEED_INPUT"]],
    )
    save_xlsx(
        SIM_DIR / "similarity_comparison_model_summary.xlsx",
        "model_summary",
        ["model", "expected_answers", "ready_answers", "complete_question_count", "pairwise_score_count", "similarity_mean", "similarity_std", "similarity_min", "similarity_max", "consistency_level", "calculation_status"],
        model_rows,
    )
    save_xlsx(
        SIM_DIR / "similarity_comparison_question_summary.xlsx",
        "question_summary",
        ["model", "question_id", "question", "ready_answer_count", "pairwise_count", "similarity_mean", "similarity_std", "similarity_min", "similarity_max", "calculation_status_or_level"],
        question_rows,
    )
    draw_bar_chart(model_rows)

    report = ["ChatGPT/Gemini/MineSafe AI 유사성 지수 비교 평가 보고서", ""]
    report.append(f"- 입력 파일: {input_path}")
    report.append(f"- 대표 질문 수: {len(question_ids)}개")
    report.append("- 반복 기준: 각 질문당 5회 답변")
    report.append("- pairwise 기준: 질문별 10개, 모델별 최대 100개 유사도 점수")
    report.append("- 계산 방식: difflib.SequenceMatcher")
    report.append("")
    report.append("1. 모델별 계산 상태")
    for row in model_rows:
        report.append(f"- {row[0]}: ready_answers {row[2]}/{row[1]}, pairwise {row[4]}, mean {row[5]}, status {row[10]}")
    report.append("")
    report.append("2. 해석")
    report.append("- 유사성 지수는 정확도 평가가 아니라 동일 질문 반복 입력 시 답변 구조와 핵심 조치가 얼마나 일관되게 유지되는지 확인하는 보조 지표이다.")
    report.append("- MineSafe AI는 기존 반복 답변 기준 계산 완료 상태이다.")
    report.append("- ChatGPT/Gemini는 사용자가 동일 질문을 각 5회씩 입력해 답변을 붙여넣은 뒤 계산해야 한다.")
    report.append("- 답변이 비어 있는 모델은 최종 비교 완료로 표시하지 않고 NEED_INPUT으로 표시한다.")
    report.append("")
    report.append("3. 생성 파일")
    report.append("- similarity_comparison_pairwise_scores.xlsx")
    report.append("- similarity_comparison_model_summary.xlsx")
    report.append("- similarity_comparison_question_summary.xlsx")
    report.append("- similarity_comparison_final_report.txt")
    report.append("- similarity_comparison_final_bar_chart.png")
    (SIM_DIR / "similarity_comparison_final_report.txt").write_text("\n".join(report) + "\n", encoding="utf-8")
    print("CALCULATION_COMPLETE")
    for row in model_rows:
        print(row)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
