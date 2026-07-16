from __future__ import annotations

from collections import Counter
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
import csv
import gc
import importlib
from io import BytesIO
import json
import uuid
from datetime import date, datetime
from html import escape
from pathlib import Path
import os
import sys
import time
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    __import__("pysqlite3")
    sys.modules["sqlite3"] = sys.modules.pop("pysqlite3")
except Exception:
    pass

import chromadb
import streamlit as st
from dotenv import load_dotenv
from google import genai
from google.genai import types
from sentence_transformers import SentenceTransformer

import verified_case_review as verified_review


VERIFIED_REVIEW_REQUIRED_ATTRIBUTES = (
    "AUTO_SCREENED_PUBLIC_TIER",
    "AUTO_SCREENED_STATUS",
    "HIDDEN_PUBLIC_TIER",
    "ReviewWorkflowBlocked",
    "TEXT_SAFE_FALLBACK_TIER",
    "VERIFIED_PUBLIC_TIER",
    "classify_public_case_relation",
    "detect_corrupted_ocr_text",
    "effective_public_case_tier",
    "generate_priority_card_images",
    "initialize_review_store",
    "is_public_display_safe_case",
    "load_review_records",
    "priority_review_candidates",
    "rank_public_official_cases",
    "rebuild_auto_screened_case_db",
    "rebuild_text_safe_case_db",
    "rebuild_verified_case_db",
    "review_status_counts",
    "sanitize_display_case",
    "save_review_update",
)


def ensure_verified_review_contract(module: Any) -> Any:
    """Validate and refresh the local review module after a Streamlit hot reload."""
    expected_path = Path(__file__).resolve().with_name("verified_case_review.py")
    module_file = getattr(module, "__file__", "")
    loaded_path = Path(module_file).resolve() if module_file else None
    if loaded_path != expected_path:
        raise ImportError("로컬 verified_case_review 모듈 경로를 확인할 수 없습니다.")

    missing = [
        name for name in VERIFIED_REVIEW_REQUIRED_ATTRIBUTES if not hasattr(module, name)
    ]
    if missing:
        importlib.invalidate_caches()
        module = importlib.reload(module)
        missing = [
            name for name in VERIFIED_REVIEW_REQUIRED_ATTRIBUTES if not hasattr(module, name)
        ]
    if missing:
        raise ImportError(
            "verified_case_review 공개 계약 누락: " + ", ".join(sorted(missing))
        )
    return module


verified_review = ensure_verified_review_contract(verified_review)


# ==============================
# 기본 설정
# ==============================
ROOT_DIR = Path(__file__).resolve().parent
load_dotenv(ROOT_DIR / ".env", override=True)

VECTOR_DB_DIR = ROOT_DIR / "10_vector_db_with_major_accident_docs"
UNVERIFIED_OFFICIAL_CASE_VECTOR_DB_DIR = ROOT_DIR / "23_official_accident_case_vector_db"
OFFICIAL_CASE_VECTOR_DB_DIR = ROOT_DIR / "23_verified_official_accident_case_vector_db"
AUTO_SCREENED_OFFICIAL_CASE_VECTOR_DB_DIR = (
    ROOT_DIR / "23_auto_screened_official_accident_case_vector_db"
)
TEXT_SAFE_OFFICIAL_CASE_VECTOR_DB_DIR = (
    ROOT_DIR / "23_text_safe_official_accident_case_vector_db"
)
SCENARIO_PATH = ROOT_DIR / "02_질문시나리오" / "question_scenarios_30.tsv"
SCENARIO_PATH_65 = ROOT_DIR / "02_질문시나리오" / "question_scenarios_65.tsv"
SCENARIO_PATH_100 = ROOT_DIR / "02_질문시나리오" / "question_scenarios_100.tsv"
SCENARIO_PATH_110 = ROOT_DIR / "02_질문시나리오" / "question_scenarios_110.tsv"
SCENARIO_SET_OPTIONS = {
    "30개 기본 평가 세트": SCENARIO_PATH,
    "65개 1차 확장 평가 세트": SCENARIO_PATH_65,
    "100개 전체 확장 평가 세트": SCENARIO_PATH_100,
    "110개 분진·보호구 보강 평가 세트": SCENARIO_PATH_110,
}
SCENARIO_SET_DESCRIPTIONS = {
    "30개 기본 평가 세트": "기존 30개 질문 시나리오를 기준으로 한 기본 평가 세트입니다.",
    "65개 1차 확장 평가 세트": "기존 30개 질문에 추가 35개 질문을 더한 확장 세트입니다.",
    "100개 전체 확장 평가 세트": "기존 30개 질문에 작업 전 상황 중심 질문 70개를 추가하여 총 100개 질문 시나리오로 구성했습니다.",
    "110개 분진·보호구 보강 평가 세트": "기존 100개 질문에 분진 관리 5개와 보호구/PPE 5개를 추가하여 총 110개 질문 시나리오로 구성했습니다.",
}
EVALUATION_PATH = ROOT_DIR / "09_answer_tests" / "evaluation_template.tsv"
EVALUATION_CRITERIA_PATH = ROOT_DIR / "09_answer_tests" / "evaluation_criteria.md"
EVALUATION_EXAMPLE_Q01_PATH = ROOT_DIR / "09_answer_tests" / "evaluation_example_Q01.md"
AUTO_EVAL_SUMMARY_PATH = ROOT_DIR / "09_answer_tests" / "auto_eval_110_full_summary.tsv"
AUTO_EVAL_DETAIL_PATH = ROOT_DIR / "09_answer_tests" / "auto_eval_Q001_Q110.tsv"
AUTO_EVAL_BATCH_PATHS = [
    ROOT_DIR / "09_answer_tests" / "auto_eval_Q001_Q030.tsv",
    ROOT_DIR / "09_answer_tests" / "auto_eval_Q031_Q110.tsv",
]
COLLECTION_NAME = "mine_safety_docs"
UNVERIFIED_OFFICIAL_CASE_COLLECTION_NAME = "mine_official_accident_cases"
OFFICIAL_CASE_COLLECTION_NAME = "mine_verified_official_accident_cases"
AUTO_SCREENED_OFFICIAL_CASE_COLLECTION_NAME = "mine_auto_screened_official_accident_cases"
TEXT_SAFE_OFFICIAL_CASE_COLLECTION_NAME = "mine_text_safe_official_accident_cases"
OFFICIAL_CASE_TOP_K = 3
OFFICIAL_CASE_INTERNAL_SEARCH_COUNT = 75
EMBEDDING_MODEL_NAME = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"
GEMINI_MODEL_NAME = "gemini-2.5-flash-lite"
GEMINI_MODEL_OPTIONS = [
    "gemini-2.5-flash-lite",
    "gemini-2.5-flash",
    "gemini-2.5-pro",
]

MAJOR_ACCIDENT_DOC_TOTAL_CHUNKS = 820
MAJOR_ACCIDENT_DOC_ADDED_CHUNKS = 333
MAJOR_ACCIDENT_LAW_DOCS = [
    ("FAQ", "중대재해처벌법령 FAQ 중대산업재해"),
    ("질의회시", "중대재해처벌법 질의회시집"),
    ("해설서", "중대재해처벌법 해설서"),
    ("따라하기", "중대재해처벌법 따라하기 안내서"),
]

# 발표/평가 중에는 외부 API 때문에 화면이 오래 멈추지 않도록 짧게 제한합니다.
GEMINI_HTTP_TIMEOUT_MS = 60_000
GEMINI_RESPONSE_TIMEOUT_SECONDS = 60
GEMINI_MAX_ATTEMPTS = 3
GEMINI_RETRY_SLEEP_SECONDS = 2.0
GEMINI_MAX_OUTPUT_TOKENS = 800
GEMINI_CONTEXT_TOP_K = 3
CONTEXT_CHUNK_CHAR_LIMIT = 800

# 공식 RAG 근거의 구조적 충분성만 보수적으로 확인합니다.
# ChromaDB distance의 의미는 collection 설정에 따라 달라질 수 있어 판정에 사용하지 않습니다.
EVIDENCE_MIN_OFFICIAL_CHUNKS = 3
EVIDENCE_MIN_UNIQUE_DOCUMENTS = 2
EVIDENCE_MIN_NON_EMPTY_CHUNKS = 3
EVIDENCE_MAX_DUPLICATE_RATIO = 0.0
EVIDENCE_MAX_SINGLE_DOCUMENT_RATIO = 0.67
EVIDENCE_STATUS_SUFFICIENT = "sufficient"
EVIDENCE_STATUS_NEEDS_REVIEW = "needs_review"
EVIDENCE_STATUS_INSUFFICIENT = "insufficient"
EVIDENCE_STATUS_LABELS = {
    EVIDENCE_STATUS_SUFFICIENT: "공식 근거 충분",
    EVIDENCE_STATUS_NEEDS_REVIEW: "공식 근거 보완 필요",
    EVIDENCE_STATUS_INSUFFICIENT: "공식 근거 부족",
}
EVIDENCE_STATUS_REASONS = {
    EVIDENCE_STATUS_SUFFICIENT: "검색된 공식 문서가 현재 질문의 주요 안전조치를 뒷받침합니다.",
    EVIDENCE_STATUS_NEEDS_REVIEW: "관련 공식 문서가 검색되었지만 일부 세부 기준은 추가 확인이 필요합니다.",
    EVIDENCE_STATUS_INSUFFICIENT: "검색된 공식 문서만으로는 구체적인 법령·수치·작업 재개 조건을 단정하기 어렵습니다.",
}
EVIDENCE_INSUFFICIENT_GUIDANCE = (
    "검색된 공식 문서만으로는 세부 기준을 충분히 확인하기 어렵습니다. "
    "우선 보수적인 안전조치를 적용하고 담당 안전관리자, 관계기관 또는 전문가를 통해 "
    "현장 조건과 최신 법령을 추가 확인해야 합니다."
)

STABLE_MODE = "안정성 모드: 검색 근거 기반 체크리스트형 답변"
GEMINI_MODE = "자연어 설명 모드: 검색 근거 기반 설명형 답변"
HYBRID_MODE = "하이브리드 모드: 검색 근거 답변을 먼저 표시하고 Gemini 답변도 추가 시도"
WORKER_EASY_MODE_LABEL = "근로자 쉬운 설명 모드"
STABLE_MODE_HELP = "공식 문서 기반의 핵심 안전조치를 일정한 체크리스트 형식으로 제공합니다."
WORKER_EASY_MODE_HELP = (
    "신규 근로자와 비전문 작업자가 이해하기 쉽도록 전문용어를 줄이고, "
    "지금 해야 할 행동과 위험한 이유를 짧게 설명합니다."
)
HYBRID_MODE_HELP = "현장 대응 체크리스트와 함께 조치 이유, 관리상 주의사항을 종합적으로 설명합니다."
WORKER_EASY_TERM_EXPLANATIONS = (
    "격리조치 → 다른 사람이 가까이 가지 못하게 막기",
    "에너지 차단 → 전원과 움직이는 힘을 완전히 끊기",
    "재가동 방지 → 다른 사람이 기계를 다시 켜지 못하게 하기",
    "산소결핍 → 숨 쉴 산소가 부족한 상태",
    "유해가스 → 마시면 몸에 해로운 가스",
    "비산분진 → 공기 중에 날리는 미세한 먼지",
    "불발공 → 폭약이 터지지 않고 남아 있을 수 있는 구멍",
    "지보 상태 → 천장과 벽을 받치는 시설의 상태",
    "작업 재개 승인 → 관리자가 안전을 확인한 뒤 다시 일하도록 허용하는 것",
)

DATA_DIR = ROOT_DIR / "data"
FEATURE_OUTPUT_DIR = ROOT_DIR / "18_legal_evidence_features"
LEGAL_CHECKLIST_STATUS_PATH = DATA_DIR / "legal_checklist_status.json"
CONVERSATION_HISTORY_PATH = DATA_DIR / "conversation_history.jsonl"
LEGAL_CHECKLIST_EXPORT_PATH = FEATURE_OUTPUT_DIR / "legal_checklist_export.xlsx"
RISK_ASSESSMENT_EXPORT_PATH = FEATURE_OUTPUT_DIR / "risk_assessment_draft_export.xlsx"
CONVERSATION_HISTORY_EXPORT_PATH = FEATURE_OUTPUT_DIR / "conversation_history_export.xlsx"
FEATURE_REPORT_PATH = FEATURE_OUTPUT_DIR / "legal_evidence_history_feature_report.txt"
LATEST_REFERENCE_CASES_PATH = FEATURE_OUTPUT_DIR / "latest_reference_cases.json"
NAVER_NEWS_API_URL = "https://openapi.naver.com/v1/search/news.json"
LIVE_CASE_SEARCH_TTL_SECONDS = 1800

BLASTING_QUESTION_TYPE = "발파/불발"
BLASTING_KEYWORDS = [
    "발파",
    "불발",
    "화약",
    "폭약",
    "장약",
    "점화",
    "발파 후",
    "굴진",
    "대피",
    "위험구역",
    "출입통제",
]
BLASTING_CORE_KEYWORDS = [
    "발파",
    "불발",
    "화약",
    "폭약",
    "장약",
    "점화",
    "발파 후",
]
BLASTING_PREFERRED_SOURCE_FILES = [
    "06_광산안전법.txt",
    "08_광산안전법_시행규칙.txt",
    "09_광산안전기술기준_제10차개정_2024-12-19.txt",
    "10_광산안전업무_처리지침_2025-35_latest.txt",
]
BLASTING_PREFERRED_SOURCE_MARKERS = [
    "광산안전기술기준",
    "광산안전업무",
    "광산안전업무_처리지침",
    "광산안전법",
    "광산안전법_시행규칙",
]
BLASTING_GENERAL_SOURCE_MARKERS = [
    "안전보건관리체계",
    "산업안전보건법",
    "산업안전보건법_시행령",
    "산업안전보건법_시행규칙",
]
BLASTING_INTERNAL_SEARCH_COUNT = 20

ELECTRICAL_QUESTION_TYPE = "전기안전"
ELECTRICAL_KEYWORDS = [
    "전기설비",
    "전기",
    "누전",
    "감전",
    "접지",
    "절연",
    "누전차단기",
    "전원 차단",
    "방폭",
    "케이블",
    "배선",
    "습기",
    "물기",
    "전기안전",
]
ELECTRICAL_PREFERRED_SOURCE_FILES = [
    "05_산업안전보건법_시행규칙.txt",
    "06_광산안전법.txt",
    "07_광산안전법_시행령.txt",
    "08_광산안전법_시행규칙.txt",
    "09_광산안전기술기준_제10차개정_2024-12-19.txt",
    "10_광산안전업무_처리지침_2025-35_latest.txt",
]
ELECTRICAL_PREFERRED_SOURCE_MARKERS = [
    "광산안전기술기준",
    "광산안전법",
    "광산안전법_시행규칙",
    "광산안전업무",
    "광산안전업무_처리지침",
    "산업안전보건법_시행규칙",
]
ELECTRICAL_GENERAL_SOURCE_MARKERS = [
    "중대재해처벌법",
    "중대재해처벌법_시행령",
    "중대재해_처벌",
    "안전보건관리체계",
    "위험성평가",
]
ELECTRICAL_INTERNAL_SEARCH_COUNT = 20

PPE_DUST_QUESTION_TYPE = "보호구/PPE/분진"
PPE_DUST_KEYWORDS = [
    "안전모",
    "방진마스크",
    "보호구",
    "호흡보호구",
    "분진",
    "먼지",
    "굴진",
    "파쇄",
    "작업환경측정",
    "노출농도",
    "유해인자",
    "착용",
    "지급",
]
PPE_DUST_CORE_KEYWORDS = [
    "안전모",
    "방진마스크",
    "보호구",
    "호흡보호구",
    "분진",
    "먼지",
    "작업환경측정",
    "노출농도",
    "유해인자",
]

RISK_ASSESSMENT_QUESTION_TYPE = "위험성평가"
RISK_ASSESSMENT_KEYWORDS = [
    "위험성평가",
    "유해위험요인",
    "유해·위험요인",
    "감소대책",
    "추가 평가",
    "재검토",
    "새 장비",
    "작업 방식 변경",
    "신규 작업",
    "새로운 굴진 구간",
    "작업 전 평가",
]

ACCIDENT_RESPONSE_QUESTION_TYPE = "사고보고/응급조치/중대재해 대응"
ACCIDENT_RESPONSE_KEYWORDS = [
    "사고",
    "부상",
    "다침",
    "접촉사고",
    "사망",
    "중대부상",
    "중대재해",
    "응급조치",
    "119",
    "구조",
    "구급",
    "보고",
    "현장 보존",
    "재해조사",
    "사고조사",
    "재발방지대책",
]
ACCIDENT_RESPONSE_CORE_KEYWORDS = [
    keyword for keyword in ACCIDENT_RESPONSE_KEYWORDS
    if keyword != "보고"
]

DAILY_INSPECTION_QUESTION_TYPE = "일상 안전점검/기록"
DAILY_INSPECTION_KEYWORDS = [
    "일상 점검",
    "매일 점검",
    "안전일지",
    "기록",
    "안전관리 사항",
    "광산 관리자",
    "점검하고 기록",
]
DAILY_INSPECTION_CORE_KEYWORDS = [
    "일상 점검",
    "매일 점검",
    "안전일지",
    "안전관리 사항",
    "광산 관리자",
    "점검하고 기록",
]

PASSAGE_SAFETY_QUESTION_TYPE = "통로/조명/표지/대피로 점검"
PASSAGE_SAFETY_KEYWORDS = [
    "통로",
    "조명",
    "표지판",
    "안내표지",
    "위험표시",
    "대피로",
    "운반갱도",
    "주요 통행장소",
]

TBM_QUESTION_TYPE = "작업 전 안전회의/TBM"
TBM_KEYWORDS = [
    "작업 전 안전회의",
    "안전점검회의",
    "tbm",
    "매일 작업 시작 전",
    "관리감독자",
    "회의 내용",
    "작업 전 회의",
]

COMPLEX_RISK_QUESTION_TYPE = "복합위험 통제"
COMPLEX_RISK_KEYWORDS = [
    "복합 작업",
    "복합위험",
    "발파 후 환기",
    "환기가 충분하지",
    "운반차량 투입",
    "순서로 위험 통제",
]

RERANK_TYPE_CONFIGS = {
    BLASTING_QUESTION_TYPE: {
        "keywords": BLASTING_KEYWORDS,
        "preferred_files": BLASTING_PREFERRED_SOURCE_FILES,
        "preferred_markers": BLASTING_PREFERRED_SOURCE_MARKERS,
        "general_markers": BLASTING_GENERAL_SOURCE_MARKERS,
        "internal_count": BLASTING_INTERNAL_SEARCH_COUNT,
        "label": "광산 특화 문서 우선 정렬 적용",
        "keyword_weight": 0.35,
        "general_penalty": 0.35,
    },
    ELECTRICAL_QUESTION_TYPE: {
        "keywords": ELECTRICAL_KEYWORDS,
        "preferred_files": ELECTRICAL_PREFERRED_SOURCE_FILES,
        "preferred_markers": ELECTRICAL_PREFERRED_SOURCE_MARKERS,
        "general_markers": ELECTRICAL_GENERAL_SOURCE_MARKERS,
        "internal_count": ELECTRICAL_INTERNAL_SEARCH_COUNT,
        "label": "전기안전 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.30,
        "general_penalty": 0.45,
    },
    PPE_DUST_QUESTION_TYPE: {
        "keywords": PPE_DUST_KEYWORDS,
        "preferred_files": [
            "05_산업안전보건법_시행규칙.txt",
            "08_광산안전법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "10_광산안전업무_처리지침_2025-35_latest.txt",
            "13_2023_새로운_위험성평가_안내서.txt",
        ],
        "preferred_markers": [
            "광산안전법_시행규칙",
            "광산안전업무",
            "광산안전기술기준",
            "산업안전보건법_시행규칙",
            "위험성평가_안내서",
            "새로운_위험성평가",
        ],
        "general_markers": ["중대재해처벌법", "중대재해_처벌"],
        "internal_count": 20,
        "label": "보호구·분진 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.32,
        "general_penalty": 0.40,
    },
    RISK_ASSESSMENT_QUESTION_TYPE: {
        "keywords": RISK_ASSESSMENT_KEYWORDS,
        "preferred_files": [
            "03_산업안전보건법.txt",
            "05_산업안전보건법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "11_안전보건관리체계_구축_가이드북.txt",
            "13_2023_새로운_위험성평가_안내서.txt",
        ],
        "preferred_markers": [
            "새로운_위험성평가",
            "위험성평가_안내서",
            "안전보건관리체계_구축",
            "산업안전보건법",
            "산업안전보건법_시행규칙",
            "광산안전기술기준",
        ],
        "general_markers": ["중대재해처벌법", "중대재해_처벌"],
        "internal_count": 20,
        "label": "위험성평가 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.35,
        "general_penalty": 0.35,
    },
    ACCIDENT_RESPONSE_QUESTION_TYPE: {
        "keywords": ACCIDENT_RESPONSE_KEYWORDS,
        "preferred_files": [
            "01_중대재해_처벌_등에_관한_법률.txt",
            "02_중대재해처벌법_시행령.txt",
            "03_산업안전보건법.txt",
            "04_산업안전보건법_시행령.txt",
            "10_광산안전업무_처리지침_2025-35_latest.txt",
            "11_안전보건관리체계_구축_가이드북.txt",
        ],
        "preferred_markers": [
            "중대재해_처벌",
            "중대재해처벌법",
            "산업안전보건법",
            "산업안전보건법_시행령",
            "광산안전업무",
            "안전보건관리체계_구축",
        ],
        "general_markers": ["위험성평가_안내서", "새로운_위험성평가"],
        "internal_count": 20,
        "label": "사고보고·응급조치 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.32,
        "general_penalty": 0.30,
    },
    DAILY_INSPECTION_QUESTION_TYPE: {
        "keywords": DAILY_INSPECTION_KEYWORDS,
        "preferred_files": [
            "08_광산안전법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "10_광산안전업무_처리지침_2025-35_latest.txt",
        ],
        "preferred_markers": [
            "광산안전기술기준",
            "광산안전법_시행규칙",
            "광산안전업무",
        ],
        "general_markers": [
            "중대재해처벌법",
            "중대재해_처벌",
            "안전보건관리체계",
            "위험성평가",
        ],
        "internal_count": 20,
        "label": "광산 일상점검·기록 문서 우선 정렬 적용",
        "keyword_weight": 0.30,
        "general_penalty": 0.40,
    },
    PASSAGE_SAFETY_QUESTION_TYPE: {
        "keywords": PASSAGE_SAFETY_KEYWORDS,
        "preferred_files": [
            "08_광산안전법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "10_광산안전업무_처리지침_2025-35_latest.txt",
        ],
        "preferred_markers": [
            "광산안전기술기준",
            "광산안전법_시행규칙",
            "광산안전업무",
        ],
        "general_markers": [
            "중대재해처벌법",
            "중대재해_처벌",
            "안전보건관리체계",
            "위험성평가",
        ],
        "internal_count": 20,
        "label": "통로·조명·대피로 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.34,
        "general_penalty": 0.40,
    },
    TBM_QUESTION_TYPE: {
        "keywords": TBM_KEYWORDS,
        "preferred_files": [
            "08_광산안전법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "11_안전보건관리체계_구축_가이드북.txt",
            "13_2023_새로운_위험성평가_안내서.txt",
        ],
        "preferred_markers": [
            "새로운_위험성평가",
            "위험성평가_안내서",
            "안전보건관리체계_구축",
            "광산안전법_시행규칙",
            "광산안전기술기준",
        ],
        "general_markers": ["중대재해처벌법", "중대재해_처벌"],
        "internal_count": 20,
        "label": "작업 전 안전회의·TBM 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.38,
        "general_penalty": 0.35,
    },
    COMPLEX_RISK_QUESTION_TYPE: {
        "keywords": COMPLEX_RISK_KEYWORDS + [
            "불발",
            "잔류화약류",
            "후가스",
            "메탄",
            "산소",
            "일산화탄소",
            "환기",
            "낙반",
            "부석",
            "운반차량",
        ],
        "preferred_files": [
            "08_광산안전법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "10_광산안전업무_처리지침_2025-35_latest.txt",
        ],
        "preferred_markers": [
            "광산안전기술기준",
            "광산안전법_시행규칙",
            "광산안전업무",
        ],
        "general_markers": [
            "중대재해처벌법",
            "중대재해_처벌",
            "안전보건관리체계",
            "위험성평가",
        ],
        "internal_count": 24,
        "label": "복합위험 통제 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.30,
        "general_penalty": 0.40,
    },
}


# ==============================
# 페이지 설정
# ==============================
st.set_page_config(
    page_title="MineSafe AI",
    page_icon="⛑️",
    layout="wide",
)


def inject_dashboard_css() -> None:
    st.markdown(
        """
        <style>
        :root {
            --ms-bg: #0f172a;
            --ms-panel: #182033;
            --ms-panel-2: #1f2937;
            --ms-border: #334155;
            --ms-text: #f8fafc;
            --ms-muted: #cbd5e1;
            --ms-accent: #b45309;
            --ms-accent-hover: #92400e;
            --ms-success: #10b981;
        }

        .stApp {
            background: var(--ms-bg);
            color: var(--ms-text);
        }

        [data-testid="stHeader"] {
            background: rgba(15, 23, 42, 0.96);
            border-bottom: 1px solid var(--ms-border);
        }

        [data-testid="stSidebar"] {
            background: #111827;
            border-right: 1px solid var(--ms-border);
        }

        [data-testid="stSidebar"] * {
            color: var(--ms-text);
        }

        .block-container {
            max-width: 1500px;
            padding-top: 2rem;
            padding-bottom: 2rem;
        }

        h1, h2, h3, h4, p, li, label, [data-testid="stMarkdownContainer"] {
            color: var(--ms-text);
        }

        h1, h2, h3 {
            letter-spacing: 0;
        }

        h2 {
            font-size: 1.5rem;
            line-height: 1.3;
        }

        h3 {
            font-size: 1.3rem;
            line-height: 1.35;
        }

        h2, h3 {
            border-left: 0;
            border-bottom: 1px solid var(--ms-border);
            padding-left: 0;
            padding-bottom: 0.45rem;
            margin-top: 1.4rem;
            margin-bottom: 0.75rem;
        }

        .mscc-header {
            border-bottom: 1px solid var(--ms-border);
            padding: 0.8rem 0 1.1rem 0;
            margin-bottom: 1.1rem;
        }

        .mscc-kicker {
            color: #d97706;
            font-size: 0.76rem;
            font-weight: 650;
            margin-bottom: 0.35rem;
        }

        .mscc-title {
            color: #ffffff;
            font-size: 2.15rem;
            line-height: 1.18;
            font-weight: 760;
            margin: 0;
        }

        .mscc-title-ko {
            color: #d7dde4;
            font-size: 1.05rem;
            font-weight: 600;
            margin-top: 0.35rem;
        }

        .mscc-subtitle {
            color: var(--ms-muted);
            font-size: 0.92rem;
            margin-top: 0.55rem;
        }

        .mscc-status-card {
            min-height: 112px;
            background: var(--ms-panel);
            border: 1px solid var(--ms-border);
            border-radius: 6px;
            padding: 0.85rem 0.95rem;
            margin-bottom: 0.4rem;
            box-shadow: none;
        }

        .mscc-status-label {
            color: var(--ms-muted);
            font-size: 0.8rem;
            font-weight: 600;
            text-transform: none;
        }

        .mscc-status-value {
            color: var(--ms-text);
            font-size: 1.28rem;
            font-weight: 750;
            line-height: 1.25;
            margin-top: 0.35rem;
            overflow-wrap: anywhere;
        }

        .mscc-status-description {
            color: var(--ms-muted);
            font-size: 0.78rem;
            margin-top: 0.35rem;
        }

        .mscc-priority-strip {
            background: var(--ms-panel-2);
            border: 1px solid var(--ms-border);
            border-left: 3px solid var(--ms-accent);
            border-radius: 6px;
            padding: 0.8rem 0.95rem;
            margin: 0.35rem 0 1rem 0;
        }

        .mscc-priority-strip strong {
            color: #f1c38f;
        }

        .mscc-section-label {
            color: #d9a66f;
            font-size: 0.78rem;
            font-weight: 650;
            margin-bottom: 0.45rem;
        }

        .mscc-evidence-title {
            color: var(--ms-text);
            font-size: 1rem;
            font-weight: 720;
            margin-bottom: 0.25rem;
        }

        .mscc-evidence-meta {
            color: #a7b4c5;
            font-size: 0.78rem;
            margin-bottom: 0.65rem;
        }

        .mscc-evidence-preview {
            color: #d7dde4;
            font-size: 0.88rem;
            line-height: 1.58;
        }

        [data-testid="stMetric"] {
            background: var(--ms-panel);
            border: 1px solid var(--ms-border);
            border-radius: 6px;
            padding: 0.75rem 0.85rem;
            box-shadow: none;
        }

        [data-testid="stMetricLabel"] {
            color: var(--ms-muted);
        }

        [data-testid="stMetricValue"] {
            color: var(--ms-text);
        }

        [data-testid="stVerticalBlockBorderWrapper"] {
            background: var(--ms-panel);
            border-color: var(--ms-border);
            border-radius: 6px;
        }

        [data-testid="stExpander"] {
            background: var(--ms-panel-2);
            border: 1px solid var(--ms-border);
            border-radius: 6px;
        }

        [data-testid="stDataFrame"] {
            border: 1px solid var(--ms-border);
            border-radius: 6px;
            overflow: hidden;
        }

        [data-testid="stTabs"] [role="tablist"] {
            border-bottom: 1px solid var(--ms-border);
            gap: 0.25rem;
        }

        [data-testid="stTabs"] button[role="tab"] {
            color: var(--ms-muted);
            border-radius: 4px 4px 0 0;
            padding: 0.65rem 1rem;
        }

        [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
            color: #e5e7eb;
            background: var(--ms-panel);
            border-bottom-color: var(--ms-accent);
        }

        [data-baseweb="tab-highlight"] {
            background-color: var(--ms-accent) !important;
        }

        .stButton > button {
            border-radius: 5px;
            border: 1px solid var(--ms-accent);
            background: var(--ms-accent);
            color: #ffffff;
            font-weight: 650;
            box-shadow: none;
        }

        .stButton > button:hover {
            background: var(--ms-accent-hover);
            border-color: var(--ms-accent-hover);
            color: #ffffff;
        }

        [data-testid="stTextArea"] textarea,
        [data-testid="stTextInput"] input,
        [data-baseweb="select"] > div {
            background: var(--ms-panel);
            color: var(--ms-text);
            border-color: var(--ms-border);
            box-shadow: none;
        }

        [data-testid="stTextArea"] textarea:focus,
        [data-testid="stTextInput"] input:focus,
        [data-baseweb="select"] > div:focus-within {
            border-color: #64748b !important;
            box-shadow: 0 0 0 1px #64748b !important;
        }

        [data-testid="stAlert"] {
            background-color: var(--ms-panel-2) !important;
            border-color: var(--ms-border);
            color: var(--ms-text);
            box-shadow: none;
        }

        [data-testid="stAlert"] > div,
        div[role="alert"] {
            background-color: var(--ms-panel-2) !important;
        }

        code {
            color: #e2e8f0;
            background: #111827;
        }

        hr {
            border-color: var(--ms-border);
        }

        .mscc-sidebar-brand {
            border-bottom: 1px solid var(--ms-border);
            padding-bottom: 0.8rem;
            margin-bottom: 0.8rem;
        }

        .mscc-sidebar-brand strong {
            color: var(--ms-text);
            font-size: 1rem;
        }

        .mscc-sidebar-note {
            background: var(--ms-panel);
            border: 1px solid var(--ms-border);
            border-left: 2px solid #64748b;
            border-radius: 5px;
            color: var(--ms-muted);
            font-size: 0.78rem;
            padding: 0.65rem 0.75rem;
            margin-top: 0.75rem;
        }

        .mscc-footer {
            color: #7f8b97;
            font-size: 0.76rem;
            text-align: center;
            padding-top: 0.7rem;
        }

        @media (max-width: 800px) {
            .block-container {
                padding-left: 1rem;
                padding-right: 1rem;
            }

            .mscc-title {
                font-size: 1.65rem;
            }

            .mscc-status-card {
                min-height: auto;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <style>
        :root {
            --portal-bg: #f4f7fb;
            --portal-card: #ffffff;
            --portal-border: #dce3ec;
            --portal-text: #1f2937;
            --portal-muted: #64748b;
            --portal-blue: #1d4ed8;
            --portal-blue-dark: #173a6b;
            --portal-sidebar: #10213d;
            --portal-sidebar-soft: #172c4d;
            --portal-success: #16845b;
            --portal-warning: #b45309;
        }

        .stApp {
            background: var(--portal-bg);
            color: var(--portal-text);
        }

        [data-testid="stHeader"] {
            background: rgba(244, 247, 251, 0.96);
            border-bottom: 1px solid var(--portal-border);
        }

        [data-testid="stSidebar"] {
            background: var(--portal-sidebar);
            border-right: 1px solid #253b5d;
        }

        [data-testid="stSidebar"] * {
            color: #e8eef7;
        }

        [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] .stCaption {
            color: #c8d5e7;
        }

        [data-testid="stSidebar"] [data-baseweb="select"] > div,
        [data-testid="stSidebar"] [data-testid="stTextInput"] input {
            background: var(--portal-sidebar-soft);
            border-color: #365072;
            color: #f8fafc;
        }

        [data-testid="stSidebar"] [data-testid="stAlert"],
        [data-testid="stSidebar"] [data-testid="stExpander"] {
            background: var(--portal-sidebar-soft) !important;
            border-color: #365072;
        }

        .block-container {
            box-sizing: border-box;
            width: 100%;
            max-width: 1650px;
            padding-top: 1.75rem;
            padding-bottom: 2.5rem;
            padding-left: 2rem;
            padding-right: 2rem;
            overflow: visible;
        }

        [data-testid="stMain"] h1,
        [data-testid="stMain"] h2,
        [data-testid="stMain"] h3,
        [data-testid="stMain"] h4,
        [data-testid="stMain"] p,
        [data-testid="stMain"] li,
        [data-testid="stMain"] label,
        [data-testid="stMain"] [data-testid="stMarkdownContainer"] {
            color: var(--portal-text);
        }

        [data-testid="stMain"] h2,
        [data-testid="stMain"] h3 {
            border-bottom: 0;
            padding-bottom: 0;
        }

        .portal-page-header {
            margin-bottom: 0.5rem;
        }

        .portal-breadcrumb {
            color: #64748b;
            font-size: 0.82rem;
            margin-bottom: 0.45rem;
        }

        .portal-page-title {
            color: #172033;
            font-size: 1.8rem;
            font-weight: 760;
            line-height: 1.25;
            margin: 0;
        }

        .portal-page-subtitle {
            color: #64748b;
            font-size: 0.9rem;
            margin-top: 0.35rem;
        }

        .mscc-header {
            display: none;
        }

        .mscc-status-card {
            min-height: 112px;
            height: auto;
            background: var(--portal-card);
            border: 1px solid var(--portal-border);
            border-radius: 10px;
            padding: 1rem 1.05rem;
            margin-bottom: 0.8rem;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
            overflow: visible;
        }

        .mscc-status-label {
            color: var(--portal-muted);
            font-size: 0.78rem;
            font-weight: 650;
        }

        .mscc-status-value {
            color: #172033;
            font-size: 1.05rem;
            font-weight: 740;
            line-height: 1.35;
            word-break: keep-all;
            overflow-wrap: normal;
            white-space: nowrap;
        }

        .mscc-status-description {
            color: var(--portal-muted);
            line-height: 1.5;
            word-break: keep-all;
            overflow-wrap: anywhere;
            white-space: normal;
        }

        .portal-card-title {
            color: #172033;
            font-size: 1.02rem;
            font-weight: 750;
            margin-bottom: 0.15rem;
        }

        .portal-card-subtitle {
            color: #64748b;
            font-size: 0.78rem;
            margin-bottom: 0.75rem;
            line-height: 1.5;
        }

        .answer-runtime-summary {
            display: flex;
            flex-wrap: wrap;
            gap: 0.45rem 0.9rem;
            align-items: center;
            width: fit-content;
            max-width: 100%;
            margin: 0.25rem 0 0.75rem 0;
            padding: 0.48rem 0.7rem;
            border: 1px solid #d9e2ec;
            border-radius: 8px;
            background: #f8fafc;
            color: #475569;
            font-size: 0.86rem;
            line-height: 1.45;
        }

        .answer-runtime-summary strong {
            color: #172033;
            font-weight: 720;
        }

        .info-card {
            min-height: 122px;
            height: auto;
            display: flex;
            align-items: flex-start;
            gap: 0.78rem;
            background: #ffffff;
            border: 1px solid #d9e2ec;
            border-radius: 12px;
            padding: 1rem;
            margin-bottom: 0.75rem;
            box-shadow: 0 3px 10px rgba(30, 52, 79, 0.06);
            overflow: visible;
        }

        .info-card-icon {
            width: 2.25rem;
            height: 2.25rem;
            flex: 0 0 2.25rem;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            border-radius: 9px;
            font-size: 1.05rem;
            font-weight: 800;
        }

        .info-card-body {
            min-width: 0;
            flex: 1;
        }

        .info-card-label {
            color: #64748b;
            font-size: 0.75rem;
            font-weight: 700;
            line-height: 1.35;
        }

        .info-card-value {
            color: #172033;
            font-size: 1.15rem;
            font-weight: 800;
            line-height: 1.35;
            margin-top: 0.28rem;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .info-card-subtitle {
            color: #718096;
            font-size: 0.72rem;
            line-height: 1.45;
            margin-top: 0.3rem;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .info-card-success {
            border-top: 3px solid #2d9b68;
        }

        .info-card-success .info-card-icon {
            color: #15734a;
            background: #e2f5eb;
        }

        .info-card-warning {
            border-top: 3px solid #d58a18;
        }

        .info-card-warning .info-card-icon {
            color: #9a5a07;
            background: #fff1d6;
        }

        .info-card-navy {
            border-top: 3px solid #315b88;
        }

        .info-card-navy .info-card-icon {
            color: #244b74;
            background: #e5edf6;
        }

        .info-card-teal {
            border-top: 3px solid #2b8790;
        }

        .info-card-teal .info-card-icon {
            color: #176a72;
            background: #e0f1f2;
        }

        .info-card-blue {
            border-top: 3px solid #3c71b5;
        }

        .info-card-blue .info-card-icon {
            color: #285f9e;
            background: #e4eefb;
        }

        .card-title-row {
            display: flex;
            align-items: center;
            gap: 0.48rem;
            margin-bottom: 0.15rem;
        }

        .card-title-icon {
            width: 1.65rem;
            height: 1.65rem;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            border-radius: 7px;
            font-size: 0.86rem;
            font-weight: 800;
        }

        .card-title-icon-success {
            color: #17744c;
            background: #dff4e8;
        }

        .card-title-icon-navy {
            color: #284f7b;
            background: #e5edf6;
        }

        .card-title-icon-warning {
            color: #995b08;
            background: #fff0d2;
        }

        .portal-summary-lead {
            color: #24364d;
            background: #f4f7fb;
            border-left: 4px solid #2f66a7;
            border-radius: 0 7px 7px 0;
            padding: 0.75rem 0.85rem;
            margin: 0.4rem 0 0.8rem 0;
            font-size: 0.9rem;
            line-height: 1.65;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .portal-summary-list {
            margin: 0.2rem 0 0 0;
            padding-left: 0;
            list-style: none;
        }

        .portal-summary-list li {
            color: #334155;
            margin: 0.35rem 0;
            line-height: 1.6;
            position: relative;
            padding-left: 1.55rem;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .portal-summary-list li::before {
            content: "✓";
            position: absolute;
            left: 0;
            top: 0.03rem;
            width: 1.05rem;
            height: 1.05rem;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            color: #ffffff;
            background: #2d9b68;
            border-radius: 50%;
            font-size: 0.67rem;
            font-weight: 800;
        }

        .portal-table-scroll {
            width: 100%;
            overflow-x: auto;
            overflow-y: visible;
            border: 1px solid #dbe3ec;
            border-radius: 8px;
            background: #ffffff;
        }

        .portal-data-table {
            width: 100%;
            min-width: 620px;
            table-layout: fixed;
            border-collapse: collapse;
            background: #ffffff;
        }

        .portal-data-table th {
            color: #29415f !important;
            background: #eef3f8 !important;
            border-bottom: 1px solid #d8e1eb !important;
            padding: 0.62rem 0.58rem !important;
            font-size: 0.76rem;
            font-weight: 720;
            line-height: 1.4;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .portal-data-table td {
            color: #334155 !important;
            background: #ffffff !important;
            border-bottom: 1px solid #e5eaf0 !important;
            padding: 0.62rem 0.58rem !important;
            font-size: 0.76rem;
            line-height: 1.55;
            vertical-align: top;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .portal-data-table tr:last-child td {
            border-bottom: 0 !important;
        }

        .portal-evidence-table {
            min-width: 610px;
        }

        .portal-checklist-table {
            width: 100%;
            min-width: 1040px;
        }

        .portal-badge-complete {
            color: #11633f;
            background: #def5e9;
            border: 1px solid #a8dec3;
        }

        .portal-badge-progress {
            color: #8a4a0b;
            background: #fff0d8;
            border: 1px solid #efc786;
        }

        .portal-badge-wait {
            color: #526171;
            background: #edf1f5;
            border: 1px solid #d2dae3;
        }

        .status-badge,
        .law-badge {
            display: inline-block;
            border-radius: 999px;
            padding: 0.18rem 0.48rem;
            font-size: 0.68rem;
            font-weight: 750;
            line-height: 1.3;
            white-space: nowrap;
        }

        .badge-success {
            color: #11633f;
            background: #def5e9;
            border: 1px solid #a8dec3;
        }

        .badge-warning {
            color: #8a4a0b;
            background: #fff0d8;
            border: 1px solid #efc786;
        }

        .badge-muted {
            color: #526171;
            background: #edf1f5;
            border: 1px solid #d2dae3;
        }

        .badge-danger {
            color: #a02d2d;
            background: #fde7e7;
            border: 1px solid #efb8b8;
        }

        .law-badge {
            color: #284f7b;
            background: #e6eef7;
            border: 1px solid #bfd0e3;
        }

        .law-badge-guideline {
            color: #3f5f7e;
            background: #edf2f7;
            border-color: #ccd8e5;
        }

        .law-badge-notice {
            color: #6d4b13;
            background: #fff3dc;
            border-color: #ecd29e;
        }

        .law-badge-reference {
            color: #536171;
            background: #f0f3f6;
            border-color: #d4dce5;
        }

        .law-badge-major {
            color: #234c7a;
            background: #dfeaf7;
            border-color: #adc6df;
        }

        .law-badge-faq {
            color: #6d4b13;
            background: #fff3dc;
            border-color: #ecd29e;
        }

        .law-badge-reply {
            color: #315d54;
            background: #e1f2ee;
            border-color: #b7dcd4;
        }

        .law-badge-commentary {
            color: #4e4074;
            background: #ece8f7;
            border-color: #cfc4e7;
        }

        .major-law-sidebar-card {
            background: #13294a;
            border: 1px solid #365072;
            border-radius: 10px;
            padding: 0.85rem;
            margin: 0.85rem 0;
            color: #d9e3f0;
        }

        .major-law-sidebar-title {
            color: #ffffff;
            font-weight: 780;
            font-size: 0.9rem;
            margin-bottom: 0.35rem;
        }

        .major-law-sidebar-line {
            color: #c8d5e7;
            font-size: 0.76rem;
            line-height: 1.5;
            margin: 0.16rem 0;
        }

        .db-path-detail {
            color: #d9e3f0;
            font-size: 0.74rem;
            line-height: 1.45;
            white-space: normal;
            overflow-wrap: anywhere;
            word-break: break-all;
        }

        .major-law-badge-row {
            display: flex;
            flex-wrap: wrap;
            gap: 0.32rem;
            margin: 0.5rem 0;
        }

        .major-law-mini-badge {
            display: inline-block;
            border-radius: 999px;
            padding: 0.18rem 0.48rem;
            font-size: 0.68rem;
            font-weight: 760;
            color: #dbeafe;
            background: #1f3f68;
            border: 1px solid #45668f;
        }

        .major-law-evidence-box {
            background: #ffffff;
            border: 1px solid #d9e2ec;
            border-radius: 12px;
            padding: 1rem 1.05rem;
            margin: 0.85rem 0 1rem 0;
            box-shadow: 0 8px 22px rgba(15, 23, 42, 0.06);
        }

        .major-law-evidence-title {
            color: #18324f;
            font-weight: 820;
            font-size: 1rem;
            margin-bottom: 0.35rem;
        }

        .major-law-evidence-text {
            color: #4a5f77;
            font-size: 0.86rem;
            line-height: 1.6;
            margin-bottom: 0.65rem;
        }

        .major-law-doc-list {
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 0.45rem;
        }

        .major-law-doc-item {
            background: #f8fafc;
            border: 1px solid #e1e8f0;
            border-radius: 8px;
            padding: 0.55rem 0.65rem;
            color: #29415f;
            font-size: 0.82rem;
            line-height: 1.45;
        }

        .portal-kras-note {
            color: #64748b;
            background: #f5f7fa;
            border: 1px solid #e0e6ed;
            border-radius: 7px;
            padding: 0.65rem 0.75rem;
            margin: 0.35rem 0 0.85rem 0;
            font-size: 0.78rem;
            line-height: 1.55;
            word-break: keep-all;
        }

        .portal-evidence-ribbon {
            display: flex;
            flex-wrap: wrap;
            gap: 0.45rem;
            align-items: center;
            background: #eaf1fb;
            border: 1px solid #cbd9eb;
            border-radius: 10px;
            padding: 0.8rem 0.9rem;
            margin: 0.65rem 0 0.8rem 0;
        }

        .portal-ribbon-label {
            color: #274060;
            font-size: 0.8rem;
            font-weight: 750;
            margin-right: 0.2rem;
        }

        .portal-source-chip {
            display: inline-block;
            color: #25476f;
            background: #ffffff;
            border: 1px solid #b9cbe2;
            border-radius: 999px;
            padding: 0.28rem 0.58rem;
            font-size: 0.74rem;
            line-height: 1.2;
        }

        .portal-badge {
            display: inline-block;
            border-radius: 999px;
            padding: 0.2rem 0.55rem;
            font-size: 0.72rem;
            font-weight: 700;
        }

        .portal-badge-wait {
            color: #7c4a03;
            background: #fff4d6;
            border: 1px solid #f0d28a;
        }

        .mscc-priority-strip {
            background: #fff8e8;
            border: 1px solid #ead7a2;
            border-left: 4px solid #c58a16;
            border-radius: 8px;
            color: #4b3a15;
        }

        .mscc-priority-strip strong {
            color: #795a10;
        }

        .mscc-section-label,
        .mscc-evidence-title {
            color: #1f3d64;
        }

        .mscc-evidence-meta {
            color: #64748b;
        }

        .mscc-evidence-preview {
            color: #334155;
        }

        [data-testid="stMetric"],
        [data-testid="stVerticalBlockBorderWrapper"],
        [data-testid="stExpander"] {
            background: var(--portal-card);
            border-color: #d9e2ec;
            border-radius: 12px;
            box-shadow: 0 3px 10px rgba(30, 52, 79, 0.05);
            height: auto;
            max-height: none;
            overflow: visible;
        }

        [data-testid="stVerticalBlockBorderWrapper"] > div,
        [data-testid="stExpander"] details,
        [data-testid="stExpander"] [data-testid="stExpanderDetails"] {
            height: auto;
            max-height: none;
            overflow: visible;
        }

        [data-testid="stMetricLabel"],
        [data-testid="stMetricLabel"] * {
            color: var(--portal-muted) !important;
        }

        [data-testid="stMetricValue"],
        [data-testid="stMetricValue"] * {
            color: #172033 !important;
        }

        [data-testid="stDataFrame"] {
            background: #ffffff;
            border-color: var(--portal-border);
            border-radius: 8px;
        }

        [data-testid="stTabs"] [role="tablist"] {
            border-bottom-color: var(--portal-border);
        }

        [data-testid="stTabs"] button[role="tab"] {
            color: #64748b;
            background: transparent;
        }

        [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
            color: #1e4f8f;
            background: #ffffff;
            border-bottom-color: #2f66a7;
        }

        [data-baseweb="tab-highlight"] {
            background-color: #2f66a7 !important;
        }

        .stButton > button,
        .stDownloadButton > button {
            border-radius: 7px;
            border: 1px solid #2f66a7;
            background: #2f66a7;
            color: #ffffff;
            font-weight: 650;
            box-shadow: none;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover {
            background: #234f84;
            border-color: #234f84;
            color: #ffffff;
        }

        [data-testid="stMain"] [data-testid="stTextArea"] textarea,
        [data-testid="stMain"] [data-testid="stTextInput"] input,
        [data-testid="stMain"] [data-baseweb="select"] > div,
        [data-testid="stMain"] [data-testid="stNumberInput"] input {
            background: #ffffff;
            color: var(--portal-text);
            border-color: #cbd5e1;
        }

        [data-testid="stMain"] [data-testid="stAlert"],
        [data-testid="stMain"] [data-testid="stAlert"] > div,
        [data-testid="stMain"] div[role="alert"] {
            color: #334155;
            background: #eef4fb !important;
            border-color: #cbd9eb;
        }

        [data-testid="stMain"] code {
            color: #1f3d64;
            background: #eef3f8;
        }

        [data-testid="stMain"] table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.82rem;
            table-layout: fixed;
        }

        [data-testid="stMain"] th {
            background: #edf3fa;
            color: #274060;
            border: 1px solid #d7e0eb;
            padding: 0.5rem 0.55rem;
            text-align: left;
        }

        [data-testid="stMain"] td {
            background: #ffffff;
            color: #334155;
            border: 1px solid #e1e7ef;
            padding: 0.5rem 0.55rem;
            vertical-align: top;
            line-height: 1.6;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        [data-testid="stMain"] th {
            line-height: 1.5;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        [data-testid="stMain"] [data-testid="stMarkdownContainer"] > table th:first-child,
        [data-testid="stMain"] [data-testid="stMarkdownContainer"] > table td:first-child {
            width: 180px;
        }

        [data-testid="stMain"] [data-testid="stMarkdownContainer"] {
            overflow: visible;
        }

        [data-testid="stMain"] [data-testid="stMarkdownContainer"] p,
        [data-testid="stMain"] [data-testid="stMarkdownContainer"] li {
            line-height: 1.62;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        [data-testid="stMain"] [data-testid="stExpander"] summary {
            color: #29415f;
            background: #f8fafc;
        }

        [data-testid="stMain"] [data-testid="stExpander"] details {
            background: #ffffff;
        }

        .mscc-sidebar-brand {
            border-bottom: 1px solid #304766;
            padding: 0.35rem 0 1rem 0;
            margin-bottom: 0.9rem;
        }

        .mscc-sidebar-brand strong {
            color: #ffffff;
            font-size: 1.05rem;
        }

        .portal-sidebar-en {
            color: #9fb2ca;
            font-size: 0.72rem;
            line-height: 1.35;
            margin-top: 0.2rem;
        }

        .portal-nav-group {
            color: #8fa4be;
            font-size: 0.7rem;
            font-weight: 700;
            margin: 1rem 0 0.35rem 0;
        }

        .portal-nav-item {
            color: #d9e3f0;
            border-radius: 6px;
            padding: 0.42rem 0.55rem;
            font-size: 0.82rem;
            margin: 0.12rem 0;
        }

        .portal-nav-item.active {
            color: #ffffff;
            background: #264b78;
            font-weight: 700;
        }

        .portal-system-state {
            background: var(--portal-sidebar-soft);
            border: 1px solid #365072;
            border-radius: 8px;
            padding: 0.75rem;
            margin-top: 1rem;
        }

        .portal-system-state-title {
            color: #a9bdd4;
            font-size: 0.72rem;
            margin-bottom: 0.25rem;
        }

        .portal-system-state-value {
            color: #dff7eb;
            font-size: 0.86rem;
            font-weight: 720;
        }

        .portal-system-state-time {
            color: #9fb2ca;
            font-size: 0.7rem;
            margin-top: 0.3rem;
        }

        .mscc-sidebar-note {
            background: var(--portal-sidebar-soft);
            border-color: #365072;
            color: #c8d5e7;
        }

        .mscc-footer {
            color: #718096;
        }

        @media (max-width: 900px) {
            .block-container {
                padding-left: 0.8rem;
                padding-right: 0.8rem;
            }

            .portal-page-title {
                font-size: 1.5rem;
            }

            .mscc-status-card {
                min-height: auto;
            }

            [data-testid="stMain"] [data-testid="stMarkdownContainer"] > table th:first-child,
            [data-testid="stMain"] [data-testid="stMarkdownContainer"] > table td:first-child {
                width: 120px;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_status_card(
    title: str,
    value: str,
    description: str | None = None,
    accent: str = "#64748b",
) -> None:
    description_html = (
        f'<div class="mscc-status-description">{escape(description)}</div>'
        if description
        else ""
    )
    st.markdown(
        (
            '<div class="mscc-status-card">'
            f'<div class="mscc-status-label">{escape(title)}</div>'
            f'<div class="mscc-status-value">{escape(value)}</div>'
            f"{description_html}</div>"
        ),
        unsafe_allow_html=True,
    )


def get_risk_level(situation_type: str) -> tuple[str, str]:
    risk_map = {
        ACCIDENT_RESPONSE_QUESTION_TYPE: ("높음", "#b45309"),
        COMPLEX_RISK_QUESTION_TYPE: ("높음", "#b45309"),
        BLASTING_QUESTION_TYPE: ("높음", "#b45309"),
        "환기/유해가스": ("높음", "#b45309"),
        "화재/폭발": ("매우 높음", "#b45309"),
        ELECTRICAL_QUESTION_TYPE: ("중간~높음", "#64748b"),
        PPE_DUST_QUESTION_TYPE: ("중간", "#64748b"),
        RISK_ASSESSMENT_QUESTION_TYPE: ("중간", "#64748b"),
        DAILY_INSPECTION_QUESTION_TYPE: ("낮음~중간", "#64748b"),
        PASSAGE_SAFETY_QUESTION_TYPE: ("중간", "#64748b"),
        TBM_QUESTION_TYPE: ("낮음~중간", "#64748b"),
        "장비/운반": ("중간~높음", "#64748b"),
        "낙반/붕락": ("높음", "#b45309"),
        "중대재해처벌법": ("높음", "#b45309"),
        PPE_GENERAL_INTENT: ("중간", "#64748b"),
        OUT_OF_SCOPE_INTENT: ("범위 밖", "#64748b"),
    }
    return risk_map.get(situation_type, ("보통", "#64748b"))


def get_priority_action(situation_type: str) -> str:
    priority_map = {
        ACCIDENT_RESPONSE_QUESTION_TYPE: "작업중지 · 구조 및 응급조치 · 119 연락 · 2차 사고 통제",
        COMPLEX_RISK_QUESTION_TYPE: "출입통제 후 불발·공기질·환기·갱도 상태를 순서대로 확인",
        BLASTING_QUESTION_TYPE: "임의 접근 금지 · 위험구역 통제 · 발파 책임자 확인",
        "환기/유해가스": "작업중지 · 대피 · 가스 측정 · 환기 확보",
        "화재/폭발": "즉시 대피 · 비상 신고 · 인원 확인 · 안전 범위 내 초기 대응",
        ELECTRICAL_QUESTION_TYPE: "전원 차단 · 잠금 및 표지 · 접지·절연·누전 보호 확인",
        PPE_DUST_QUESTION_TYPE: "분진 저감조치 확인 후 보호구 미착용자 작업 투입 금지",
        RISK_ASSESSMENT_QUESTION_TYPE: "변경 요인을 재평가하고 감소대책 이행 후 작업 시작",
        DAILY_INSPECTION_QUESTION_TYPE: "이상 발견 즉시 보고·조치하고 안전일지에 기록",
        PASSAGE_SAFETY_QUESTION_TYPE: "통로·조명·표지·대피로 장애요인을 제거한 뒤 통행 허용",
        TBM_QUESTION_TYPE: "당일 위험요인·역할·신호·대피 절차 공유 후 작업 시작",
        "장비/운반": "에너지 차단 · 접근통제 · 장비와 작업자 위치 확인",
        "낙반/붕락": "작업중지 · 출입통제 · 대피 · 천반·부석·지보공 확인",
        "중대재해처벌법": "작업중지 · 보고 · 원인조사 · 재발방지대책 검토",
        PPE_GENERAL_INTENT: "작업 위험에 맞는 보호구 선정 · 착용상태 확인 · 지급/교육 기록",
        OUT_OF_SCOPE_INTENT: "광산 안전관리와 연결된 질문으로 다시 입력",
    }
    return priority_map.get(
        situation_type,
        "검색 근거와 현장 조건을 대조하고 책임자 확인 후 작업 여부를 판단",
    )


def resolve_display_situation_type(question_type: str, answer: str) -> str:
    if question_type != "일반":
        return question_type
    for candidate in [
        "환기/유해가스",
        "화재/폭발",
        "장비/운반",
        "낙반/붕락",
        "중대재해처벌법",
    ]:
        if f"상황 유형: {candidate}" in answer:
            return candidate
    return "일반"


inject_dashboard_css()

if "new_question_token" not in st.session_state:
    st.session_state["new_question_token"] = 0

page_title_col, page_action_col = st.columns([5, 1])
with page_title_col:
    st.markdown(
        """
        <div class="portal-page-header">
            <div class="portal-breadcrumb">안전 질의 및 답변 &gt; 답변 상세</div>
            <div class="portal-page-title">MineSafe AI</div>
            <div class="portal-page-subtitle">
                광산 안전 지침 및 중대재해처벌법 대응 RAG 가상 안전관리자
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
with page_action_col:
    st.write("")
    if st.button("새 질문하기", use_container_width=True, key="new_question_top"):
        st.session_state["new_question_token"] += 1
        st.session_state.pop("scenario_test_result", None)
        st.rerun()


# ==============================
# 캐시 함수
# ==============================
@st.cache_resource(show_spinner=False)
def load_embedding_model() -> SentenceTransformer:
    return SentenceTransformer(EMBEDDING_MODEL_NAME)


@st.cache_resource(show_spinner=False)
def load_chroma_collection():
    if not VECTOR_DB_DIR.exists():
        return None, f"Vector DB 폴더를 찾을 수 없습니다: {VECTOR_DB_DIR}"

    try:
        client = chromadb.PersistentClient(path=str(VECTOR_DB_DIR))
        collection = client.get_collection(name=COLLECTION_NAME)
        return collection, None
    except Exception as e:
        return None, str(e)


@st.cache_resource(show_spinner=False)
def load_official_case_collection():
    if OFFICIAL_CASE_COLLECTION_NAME in {
        COLLECTION_NAME,
        UNVERIFIED_OFFICIAL_CASE_COLLECTION_NAME,
        AUTO_SCREENED_OFFICIAL_CASE_COLLECTION_NAME,
        TEXT_SAFE_OFFICIAL_CASE_COLLECTION_NAME,
    }:
        return None, "verified_collection_conflict"
    if not OFFICIAL_CASE_VECTOR_DB_DIR.exists():
        return None, "verified_not_created_zero_cases"
    try:
        client = chromadb.PersistentClient(path=str(OFFICIAL_CASE_VECTOR_DB_DIR))
        collection = client.get_collection(name=OFFICIAL_CASE_COLLECTION_NAME)
        return collection, None
    except Exception:
        return None, "verified_db_unavailable"


@st.cache_resource(show_spinner=False)
def load_auto_screened_case_collection():
    collection_names = {
        COLLECTION_NAME,
        UNVERIFIED_OFFICIAL_CASE_COLLECTION_NAME,
        OFFICIAL_CASE_COLLECTION_NAME,
        AUTO_SCREENED_OFFICIAL_CASE_COLLECTION_NAME,
        TEXT_SAFE_OFFICIAL_CASE_COLLECTION_NAME,
    }
    if len(collection_names) != 5:
        return None, "auto_screened_collection_conflict"
    if not AUTO_SCREENED_OFFICIAL_CASE_VECTOR_DB_DIR.exists():
        return None, "auto_screened_not_created_zero_cases"
    try:
        client = chromadb.PersistentClient(path=str(AUTO_SCREENED_OFFICIAL_CASE_VECTOR_DB_DIR))
        collection = client.get_collection(name=AUTO_SCREENED_OFFICIAL_CASE_COLLECTION_NAME)
        return collection, None
    except Exception:
        return None, "auto_screened_db_unavailable"


@st.cache_resource(show_spinner=False)
def load_text_safe_case_collection():
    collection_names = {
        COLLECTION_NAME,
        UNVERIFIED_OFFICIAL_CASE_COLLECTION_NAME,
        OFFICIAL_CASE_COLLECTION_NAME,
        AUTO_SCREENED_OFFICIAL_CASE_COLLECTION_NAME,
        TEXT_SAFE_OFFICIAL_CASE_COLLECTION_NAME,
    }
    if len(collection_names) != 5:
        return None, "text_safe_collection_conflict"
    if not TEXT_SAFE_OFFICIAL_CASE_VECTOR_DB_DIR.exists():
        return None, "text_safe_not_created_zero_cases"
    try:
        client = chromadb.PersistentClient(path=str(TEXT_SAFE_OFFICIAL_CASE_VECTOR_DB_DIR))
        collection = client.get_collection(name=TEXT_SAFE_OFFICIAL_CASE_COLLECTION_NAME)
        return collection, None
    except Exception:
        return None, "text_safe_db_unavailable"


def get_gemini_api_key() -> str | None:
    try:
        env_key = (os.getenv("GEMINI_API_KEY") or "").strip()
        if env_key:
            return env_key
    except Exception:
        pass

    try:
        secrets = st.secrets
        secret_value = secrets.get("GEMINI_API_KEY") if hasattr(secrets, "get") else None
        if isinstance(secret_value, str):
            secret_value = secret_value.strip()
            if secret_value:
                return secret_value
    except Exception:
        pass

    return None


@st.cache_resource(show_spinner=False)
def load_gemini_client():
    api_key = get_gemini_api_key()

    if not api_key:
        return None, "Gemini API 키 없음. 로컬 .env 또는 Streamlit Cloud Secrets의 GEMINI_API_KEY를 확인해 주세요."

    try:
        http_options = types.HttpOptions(
            timeout=GEMINI_HTTP_TIMEOUT_MS,
            retry_options=types.HttpRetryOptions(
                attempts=1,
                initial_delay=0.1,
                max_delay=0.2,
                http_status_codes=[429, 500, 502, 503, 504],
            ),
        )
        client = genai.Client(api_key=api_key, http_options=http_options)
        return client, None
    except Exception as e:
        return None, str(e)


@st.cache_data(show_spinner=False)

def normalize_question_id_for_display(value, fallback_index=None):
    """질문ID를 Q001 형태로 표시하기 위한 보정 함수."""
    text = str(value).strip()

    if text and text.lower() not in ["nan", "none"]:
        if text.upper().startswith("Q"):
            digits = "".join(ch for ch in text if ch.isdigit())
            if digits:
                return f"Q{int(digits):03d}"
            return text
        if text.isdigit():
            return f"Q{int(text):03d}"

    if fallback_index is not None:
        return f"Q{int(fallback_index) + 1:03d}"

    return "Q---"


def load_question_scenarios(path: str | Path | None = None) -> tuple[list[dict[str, str]], str | None]:
    scenario_path = Path(path) if path else SCENARIO_PATH
    if not scenario_path.exists():
        return [], f"질문 시나리오 파일을 찾을 수 없습니다: {scenario_path}"

    try:
        with open(scenario_path, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f, delimiter="\t"))
        return rows, None
    except Exception as e:
        return [], str(e)


@st.cache_data(show_spinner=False)
def load_evaluation_criteria() -> str | None:
    if not EVALUATION_CRITERIA_PATH.exists():
        return None

    try:
        return EVALUATION_CRITERIA_PATH.read_text(encoding="utf-8-sig")
    except Exception:
        return None


@st.cache_data(show_spinner=False)
def load_q01_evaluation_example() -> str | None:
    if not EVALUATION_EXAMPLE_Q01_PATH.exists():
        return None

    try:
        return EVALUATION_EXAMPLE_Q01_PATH.read_text(encoding="utf-8-sig")
    except Exception:
        return None


# ==============================
# 유틸 함수
# ==============================
def get_source(meta: dict[str, Any]) -> str:
    return (
        meta.get("source")
        or meta.get("file_name")
        or meta.get("doc_name")
        or meta.get("title")
        or meta.get("source_file")
        or "출처 정보 없음"
    )


def get_chunk_id(meta: dict[str, Any]) -> str:
    value = (
        meta.get("chunk_id")
        or meta.get("id")
        or meta.get("doc_id")
        or meta.get("chunk_index")
        or meta.get("index")
        or ""
    )
    return str(value) if value != "" else "정보 없음"


def clean_text(text: str) -> str:
    text = text.replace("\n", " ").replace("\t", " ")
    return " ".join(text.split())


def format_distance(distance: Any) -> str:
    if isinstance(distance, (float, int)):
        return f"{distance:.4f}"
    return "없음"


def is_temporary_gemini_error(error_message: str) -> bool:
    message = error_message.lower()
    keywords = [
        "503",
        "unavailable",
        "high demand",
        "overloaded",
        "temporarily",
        "timeout",
        "timed out",
        "deadline",
        "429",
        "resource exhausted",
        "rate limit",
    ]
    return any(keyword in message for keyword in keywords)


def is_timeout_error(error_message: str) -> bool:
    message = error_message.lower()
    return (
        "timeout" in message
        or "timed out" in message
        or "deadline" in message
        or "제한 시간" in message
    )


def classify_question_type(question: str) -> str:
    text = clean_text(question).lower()

    if any(keyword in text for keyword in COMPLEX_RISK_KEYWORDS):
        return COMPLEX_RISK_QUESTION_TYPE
    if any(keyword in text for keyword in TBM_KEYWORDS):
        return TBM_QUESTION_TYPE
    risk_keywords = [
        keyword for keyword in RISK_ASSESSMENT_KEYWORDS
        if keyword != "신규 작업"
    ]
    if any(keyword in text for keyword in risk_keywords) or (
        "신규 작업" in text and "신규 작업자" not in text
    ):
        return RISK_ASSESSMENT_QUESTION_TYPE
    if any(keyword in text for keyword in PPE_DUST_CORE_KEYWORDS):
        return PPE_DUST_QUESTION_TYPE
    strong_accident_keywords = [
        "부상",
        "다침",
        "접촉사고",
        "사망",
        "중대부상",
        "응급조치",
        "119",
        "구조",
        "구급",
        "현장 보존",
        "재해조사",
        "사고조사",
        "재발방지대책",
    ]
    has_accident_event = (
        "사고" in text
        and any(
            keyword in text
            for keyword in ["발생", "부상", "다침", "사망", "보고", "조사"]
        )
    )
    has_major_accident_event = (
        "중대재해" in text
        and any(keyword in text for keyword in ["발생", "사망", "부상", "사고"])
    )
    if (
        any(keyword in text for keyword in strong_accident_keywords)
        or has_accident_event
        or has_major_accident_event
    ):
        return ACCIDENT_RESPONSE_QUESTION_TYPE
    if "중대재해처벌법" in text:
        return "일반"
    if any(keyword in text for keyword in DAILY_INSPECTION_CORE_KEYWORDS):
        return DAILY_INSPECTION_QUESTION_TYPE
    if (
        "기록" in text
        and any(keyword in text for keyword in ["점검", "관리자", "안전관리"])
    ):
        return DAILY_INSPECTION_QUESTION_TYPE
    has_transport_context = any(
        keyword in text
        for keyword in [
            "운반차량",
            "덤프트럭",
            "광차",
            "운반장비",
            "차량",
            "보행 작업자",
            "충돌",
        ]
    )
    if (
        any(keyword in text for keyword in PASSAGE_SAFETY_KEYWORDS)
        and not has_transport_context
    ):
        return PASSAGE_SAFETY_QUESTION_TYPE
    if any(keyword in text for keyword in ELECTRICAL_KEYWORDS):
        return ELECTRICAL_QUESTION_TYPE
    if any(keyword in text for keyword in BLASTING_CORE_KEYWORDS):
        return BLASTING_QUESTION_TYPE
    return "일반"


PPE_GENERAL_INTENT = "보호구/PPE 일반 설명"
VENTILATION_GAS_INTENT = "환기/유해가스"
BLASTING_MISFIRE_INTENT = "발파/불발"
ROOF_FALL_INTENT = "낙반/붕락/지보"
DUST_RESPIRATORY_INTENT = "분진/호흡보호"
ELECTRICAL_SAFETY_INTENT = "전기안전"
EQUIPMENT_TRANSPORT_INTENT = "장비/운반/차량"
PREWORK_TBM_INTENT = "작업 전 점검/TBM/교육"
ACCIDENT_EMERGENCY_INTENT = "사고보고/응급조치"
KRAS_INTENT = "위험성평가/KRAS"
LAW_INTENT = "중대재해처벌법/법령 설명"
GENERAL_MINE_SAFETY_INTENT = "일반 광산 안전 질문"
FIRE_EMERGENCY_INTENT = "화재/소화/비상대응"
EVACUATION_ACCESS_SIGN_INTENT = "비상대피로/출입통제/표지"
LIGHTING_VISIBILITY_INTENT = "조명/시야/작업환경"
GAS_DETECTOR_INTENT = "가스측정기/계측기 관리"
VENTILATION_EQUIPMENT_INTENT = "환기설비/팬/통풍 이상"
FLOOD_DRAINAGE_INTENT = "침수/배수/수해"
WALKWAY_HOUSEKEEPING_INTENT = "이동통로/정리정돈/미끄럼"
BACKING_SIGNAL_INTENT = "장비 후진/신호수/충돌 방지"
CONVEYOR_ROTATING_INTENT = "컨베이어/회전체/끼임"
HOT_WORK_INTENT = "용접/절단/화기작업"
HEIGHT_WORK_INTENT = "고소작업/사다리/작업발판"
LIFTING_HEAVY_INTENT = "중량물/인양/크레인/호이스트"
FATIGUE_WORKER_INTENT = "피로/무리한 작업/작업자 상태"
CONTRACTOR_PERMIT_INTENT = "외주/협력업체/작업허가"
NEW_WORKER_TRAINING_INTENT = "교육/훈련/신규 작업자"
INSPECTION_RECORD_INTENT = "점검/순찰/기록관리"
ERGONOMICS_MANUAL_INTENT = "근골격계/수작업/작업자세"
NOISE_VIBRATION_INTENT = "소음/진동"
CHEMICAL_FUEL_INTENT = "화학물질/유류/가연물"
DOCUMENT_EVIDENCE_INTENT = "문서/증빙자료 추천"
OUT_OF_SCOPE_INTENT = "범위 밖 질문"

QUESTION_INTENT_KEYWORDS: dict[str, list[str]] = {
    PPE_GENERAL_INTENT: [
        "안전화", "안전모", "방진마스크", "마스크", "보호구", "ppe", "보안경",
        "귀마개", "귀덮개", "안전대", "장갑", "보호장갑", "작업복", "반사조끼",
    ],
    VENTILATION_GAS_INTENT: [
        "메탄", "메탄가스", "산소", "일산화탄소", "유해가스", "가스농도",
        "환기", "통기", "산소결핍", "질식", "폭발위험",
    ],
    BLASTING_MISFIRE_INTENT: [
        "발파", "폭약", "장약", "뇌관", "불발", "불발공", "대피",
        "발파 후 점검", "발파작업",
    ],
    ROOF_FALL_INTENT: [
        "낙반", "붕락", "지보", "천반", "측벽", "균열", "암반", "부석", "갱도 붕괴",
    ],
    DUST_RESPIRATORY_INTENT: [
        "분진", "먼지", "굴진", "천공", "파쇄", "살수", "집진", "호흡보호구",
    ],
    ELECTRICAL_SAFETY_INTENT: [
        "전기", "감전", "누전", "차단기", "접지", "케이블", "전원", "전기설비", "전기작업",
    ],
    EQUIPMENT_TRANSPORT_INTENT: [
        "장비", "굴삭기", "굴착기", "로더", "덤프", "덤프트럭", "운반", "차량",
        "후진", "협착", "끼임", "충돌", "적재", "운반로",
    ],
    PREWORK_TBM_INTENT: [
        "작업 전", "점검", "tbm", "툴박스", "안전회의", "교육", "작업허가",
        "체크리스트", "위험예지", "작업 시작 전",
    ],
    ACCIDENT_EMERGENCY_INTENT: [
        "사고", "다침", "부상", "응급", "구조", "신고", "보고", "119", "병원", "재해",
    ],
    KRAS_INTENT: [
        "위험성평가", "kras", "가능성", "중대성", "감소대책", "제거", "대체",
        "공학적 대책", "관리적 대책",
    ],
    LAW_INTENT: [
        "중대재해", "중대재해처벌법", "산업안전보건법", "광산안전법", "법령",
        "처벌", "의무", "안전보건관리체계", "경영책임자",
    ],
    FIRE_EMERGENCY_INTENT: [
        "화재", "불", "소화기", "소화전", "소방", "비상", "비상대응",
        "피난", "연기", "연소", "화재위험",
    ],
    EVACUATION_ACCESS_SIGN_INTENT: [
        "대피로", "비상구", "출입통제", "통로", "안전표지", "경고표지",
        "표지판", "접근금지", "위험구역", "안내표지",
    ],
    LIGHTING_VISIBILITY_INTENT: [
        "조명", "어두움", "어두워", "시야", "조도", "램프", "헤드랜턴",
        "작업환경", "가시거리",
    ],
    GAS_DETECTOR_INTENT: [
        "가스측정기", "측정기", "검지기", "센서", "교정", "보정",
        "알람", "경보", "측정값", "검교정",
    ],
    VENTILATION_EQUIPMENT_INTENT: [
        "환기팬", "팬", "통풍", "환기설비", "환기 불량", "풍량",
        "송풍", "배기", "정전", "팬 정지",
    ],
    FLOOD_DRAINAGE_INTENT: [
        "침수", "물", "배수", "양수기", "누수", "유입수", "물 고임",
        "갱내수", "수해", "펌프",
    ],
    WALKWAY_HOUSEKEEPING_INTENT: [
        "정리정돈", "미끄럼", "넘어짐", "걸림", "바닥", "적치물",
        "장애물", "통행로",
    ],
    BACKING_SIGNAL_INTENT: [
        "후진", "신호수", "경광등", "경보음", "사각지대", "유도자", "장비 접근",
    ],
    CONVEYOR_ROTATING_INTENT: [
        "컨베이어", "벨트", "회전체", "풀리", "롤러", "방호덮개", "가드",
    ],
    HOT_WORK_INTENT: [
        "용접", "절단", "그라인더", "불꽃", "화기작업", "스파크", "산소절단", "핫워크",
    ],
    HEIGHT_WORK_INTENT: [
        "고소작업", "사다리", "발판", "작업대", "추락", "난간", "개구부",
    ],
    LIFTING_HEAVY_INTENT: [
        "중량물", "인양", "크레인", "호이스트", "와이어로프", "슬링",
        "체인", "낙하", "매달림",
    ],
    FATIGUE_WORKER_INTENT: [
        "피로", "피곤", "졸림", "무리", "컨디션", "휴식", "장시간 작업", "야간작업", "집중력",
    ],
    CONTRACTOR_PERMIT_INTENT: [
        "협력업체", "외주", "작업허가", "허가서", "사전교육", "출입관리", "도급", "작업계획",
    ],
    NEW_WORKER_TRAINING_INTENT: [
        "신규", "초보", "훈련", "작업자 교육", "안전교육", "숙련도", "작업 절차",
    ],
    INSPECTION_RECORD_INTENT: [
        "순찰", "일지", "체크리스트", "개선조치", "이행확인",
    ],
    ERGONOMICS_MANUAL_INTENT: [
        "허리", "무릎", "반복작업", "수작업", "들기", "작업자세", "근골격계",
    ],
    NOISE_VIBRATION_INTENT: [
        "소음", "진동", "착암기", "파쇄기", "청력", "소음성 난청",
    ],
    CHEMICAL_FUEL_INTENT: [
        "화학물질", "유류", "기름", "연료", "가연물", "msds", "누출", "보관", "취급",
    ],
    DOCUMENT_EVIDENCE_INTENT: [
        "증빙자료", "어떤 기록", "무슨 서류", "사진", "보고서", "보관", "기록해야",
    ],
}

MINING_SAFETY_CONTEXT_KEYWORDS = [
    "광산", "갱내", "갱도", "채굴", "굴진", "작업장", "작업자", "근로자",
    "현장", "안전", "보건", "위험", "점검", "관리자", "작업", "설비",
    "기계", "보호", "교육", "기록", "대피", "출입통제", "통로", "조명",
    "화재", "소화기", "침수", "배수", "컨베이어", "용접", "사다리", "협력업체",
]

OUT_OF_SCOPE_HINTS = [
    "점심", "저녁", "아침", "메뉴", "맛집", "영화", "음악", "주식", "코인",
    "연애", "상담", "게임", "여행", "날씨", "축구", "야구", "로또", "복권",
]

INTENT_SEARCH_EXPANSIONS: dict[str, list[str]] = {
    PPE_GENERAL_INTENT: ["보호구", "착용 기준", "지급", "착용상태 확인", "작업 전 점검", "안전관리자 확인"],
    VENTILATION_GAS_INTENT: ["메탄가스", "산소농도", "유해가스 측정", "환기설비", "작업중지", "대피", "재측정"],
    BLASTING_MISFIRE_INTENT: ["불발공", "발파 후 점검", "출입통제", "대피", "폭약", "뇌관", "발파작업 안전"],
    ROOF_FALL_INTENT: ["천반 점검", "지보", "부석 제거", "균열 확인", "출입통제", "작업중지"],
    DUST_RESPIRATORY_INTENT: ["호흡보호구", "분진", "굴진", "천공", "파쇄", "살수", "집진", "작업환경측정"],
    ELECTRICAL_SAFETY_INTENT: ["감전 예방", "전원 차단", "잠금표지", "접지", "누전차단기", "케이블 점검"],
    EQUIPMENT_TRANSPORT_INTENT: ["협착", "충돌", "후진 경보", "신호수", "운반로", "장비 점검", "사각지대"],
    PREWORK_TBM_INTENT: ["작업 전 안전회의", "위험요인 공유", "작업허가", "점검표", "교육기록"],
    ACCIDENT_EMERGENCY_INTENT: ["응급조치", "사고보고", "작업중지", "현장보존", "재해자 구조", "기록"],
    KRAS_INTENT: ["가능성", "중대성", "위험도", "감소대책", "제거", "대체", "공학적 대책", "관리적 대책", "보호구"],
    LAW_INTENT: ["안전보건관리체계", "경영책임자 의무", "유해위험요인 확인", "개선조치", "이행점검"],
    FIRE_EMERGENCY_INTENT: ["소화기", "화재 예방", "가연물 제거", "비상대응", "대피", "소방설비", "점검기록"],
    EVACUATION_ACCESS_SIGN_INTENT: ["비상구", "대피로 확보", "출입통제", "안전표지", "경고표지", "통로 확보"],
    LIGHTING_VISIBILITY_INTENT: ["작업 조명", "조도", "시야 확보", "헤드랜턴", "갱내 조명", "어두운 작업장"],
    GAS_DETECTOR_INTENT: ["검지기", "측정기 교정", "보정", "알람 확인", "측정 기록", "가스 농도 측정"],
    VENTILATION_EQUIPMENT_INTENT: ["환기팬", "풍량", "송풍", "배기", "환기설비 점검", "팬 정지", "환기 불량"],
    FLOOD_DRAINAGE_INTENT: ["배수", "양수기", "누수", "물 유입", "갱내수", "침수 위험", "펌프 점검"],
    WALKWAY_HOUSEKEEPING_INTENT: ["통로 확보", "미끄럼 방지", "장애물 제거", "정리정돈", "넘어짐 예방"],
    BACKING_SIGNAL_INTENT: ["신호수", "후진 경보", "사각지대", "충돌 예방", "장비 접근통제"],
    CONVEYOR_ROTATING_INTENT: ["방호덮개", "가드", "회전체 접근금지", "끼임 예방", "정비 전 전원 차단"],
    HOT_WORK_INTENT: ["화기작업 허가", "불꽃 비산", "가연물 제거", "소화기 비치", "화재감시"],
    HEIGHT_WORK_INTENT: ["추락 방지", "안전대", "난간", "사다리 점검", "작업발판", "개구부 덮개"],
    LIFTING_HEAVY_INTENT: ["인양작업", "낙하물", "슬링 점검", "와이어로프", "작업반경 출입통제"],
    FATIGUE_WORKER_INTENT: ["작업자 건강상태", "휴식", "장시간 작업", "야간작업", "집중력 저하"],
    CONTRACTOR_PERMIT_INTENT: ["작업허가서", "사전교육", "출입관리", "위험요인 공유", "도급 작업 안전"],
    NEW_WORKER_TRAINING_INTENT: ["안전교육", "작업절차 교육", "신규 작업자", "보호구 착용 교육", "TBM"],
    INSPECTION_RECORD_INTENT: ["점검표", "작업일지", "개선조치", "이행확인", "사진 기록", "증빙자료"],
    ERGONOMICS_MANUAL_INTENT: ["중량물 취급", "작업자세", "반복작업", "허리 부상", "수작업 위험"],
    NOISE_VIBRATION_INTENT: ["청력보호구", "소음 측정", "귀마개", "귀덮개", "착암", "파쇄"],
    CHEMICAL_FUEL_INTENT: ["MSDS", "누출 대응", "보관", "취급", "가연물", "유류 화재"],
    DOCUMENT_EVIDENCE_INTENT: ["점검기록", "교육기록", "작업허가서", "작업중지 기록", "개선조치 사진", "TBM 기록"],
    GENERAL_MINE_SAFETY_INTENT: ["광산 안전", "작업 전 점검", "안전교육", "위험요인", "관리감독", "기록"],
}

PPE_ITEM_EXPANSIONS: dict[str, list[str]] = {
    "안전화": ["발 보호", "낙하물", "찔림", "미끄러짐", "끼임", "작업화", "장비 작업", "운반 작업"],
    "안전모": ["머리 보호", "낙하물", "낙반", "비래물", "충돌", "갱내 작업", "굴진 작업"],
    "방진마스크": ["호흡보호구", "분진", "굴진", "천공", "파쇄", "살수", "집진", "작업환경측정"],
    "마스크": ["호흡보호구", "분진", "굴진", "천공", "파쇄", "살수", "집진", "작업환경측정"],
    "보안경": ["눈 보호", "비래물", "파편", "분진", "절단", "연마", "천공"],
    "귀마개": ["청력 보호", "소음", "착암", "파쇄", "장비 작업"],
    "귀덮개": ["청력 보호", "소음", "착암", "파쇄", "장비 작업"],
    "안전대": ["추락 위험", "고소작업", "사다리", "작업발판", "개구부", "훅 체결"],
    "장갑": ["손 보호", "베임", "찔림", "마찰", "화상", "화학물질"],
    "보호장갑": ["손 보호", "베임", "찔림", "마찰", "화상", "화학물질"],
}

PPE_BASIC_KNOWLEDGE: dict[str, dict[str, list[str] | str]] = {
    "안전화": {
        "why": "발을 보호하기 위해 착용합니다. 낙하물, 찔림, 미끄러짐, 끼임, 장비 이동 중 발 부상 위험을 줄입니다.",
        "when": ["갱내 작업", "운반 작업", "장비 주변 작업", "굴진·파쇄 작업"],
        "checks": ["밑창 마모·미끄럼 방지 상태", "앞코와 발등 보호 상태", "찢김·구멍·젖음 여부"],
    },
    "안전모": {
        "why": "머리 부상을 막기 위해 착용합니다. 낙반, 낙하물, 비래물, 충돌, 낮은 천장이나 구조물 접촉 위험을 줄입니다.",
        "when": ["갱내 작업", "굴진 작업", "발파 후 점검", "운반·장비 작업 구역"],
        "checks": ["균열·파손 여부", "턱끈 체결 상태", "내피와 완충재 상태"],
    },
    "방진마스크": {
        "why": "분진 흡입을 줄이기 위해 착용합니다. 굴진·천공·파쇄·적재·운반 중 발생하는 먼지가 호흡기로 들어가는 것을 줄입니다.",
        "when": ["굴진", "천공", "파쇄", "적재", "운반 등 분진 발생 작업"],
        "checks": ["얼굴 밀착 상태", "필터 오염·교체 상태", "코편·끈 손상 여부"],
    },
    "보안경": {
        "why": "눈을 보호하기 위해 착용합니다. 파편, 비래물, 분진, 절단·연마·천공 중 튀는 물질로 인한 눈 손상을 줄입니다.",
        "when": ["절단", "연마", "천공", "파편·비래물·분진 발생 작업"],
        "checks": ["렌즈 손상·오염 여부", "김서림으로 시야가 가려지는지", "얼굴 밀착 상태"],
    },
    "귀마개/귀덮개": {
        "why": "소음으로부터 청력을 보호하기 위해 착용합니다. 착암, 파쇄, 장비 운전 등 고소음 작업의 청력 손상 위험을 줄입니다.",
        "when": ["착암", "파쇄", "장비 운전", "고소음 작업"],
        "checks": ["착용 밀착 상태", "오염·손상 여부", "작업 소음에 맞는 보호구인지"],
    },
    "안전대": {
        "why": "추락 위험이 있는 작업에서 몸을 지지해 추락 재해를 줄이기 위해 착용합니다.",
        "when": ["고소작업", "사다리 작업", "작업발판", "개구부 주변 작업"],
        "checks": ["걸이설비 상태", "훅 체결 상태", "벨트·죔줄 손상 여부"],
    },
    "장갑": {
        "why": "베임, 찔림, 마찰, 화상, 화학물질 접촉 등 손 위험을 줄이기 위해 착용합니다.",
        "when": ["자재 취급", "정비", "절단·연마", "거친 표면 취급", "화학물질 접촉 가능 작업"],
        "checks": ["작업 종류에 맞는 장갑인지", "찢김·마모·오염 여부", "회전체 말림 위험과의 충돌 여부"],
    },
}


RISK_SIGN_INTENTS = {
    VENTILATION_GAS_INTENT,
    BLASTING_MISFIRE_INTENT,
    ROOF_FALL_INTENT,
    ELECTRICAL_SAFETY_INTENT,
    EQUIPMENT_TRANSPORT_INTENT,
    ACCIDENT_EMERGENCY_INTENT,
    FIRE_EMERGENCY_INTENT,
    VENTILATION_EQUIPMENT_INTENT,
    FLOOD_DRAINAGE_INTENT,
    BACKING_SIGNAL_INTENT,
    CONVEYOR_ROTATING_INTENT,
    HOT_WORK_INTENT,
    HEIGHT_WORK_INTENT,
    LIFTING_HEAVY_INTENT,
    CHEMICAL_FUEL_INTENT,
}

MANAGEMENT_RECORD_INTENTS = {
    LAW_INTENT,
    DOCUMENT_EVIDENCE_INTENT,
    CONTRACTOR_PERMIT_INTENT,
    INSPECTION_RECORD_INTENT,
}

GENERAL_EXPLANATION_INTENTS = {
    PREWORK_TBM_INTENT,
    GENERAL_MINE_SAFETY_INTENT,
    DUST_RESPIRATORY_INTENT,
    EVACUATION_ACCESS_SIGN_INTENT,
    LIGHTING_VISIBILITY_INTENT,
    GAS_DETECTOR_INTENT,
    WALKWAY_HOUSEKEEPING_INTENT,
    FATIGUE_WORKER_INTENT,
    NEW_WORKER_TRAINING_INTENT,
    ERGONOMICS_MANUAL_INTENT,
    NOISE_VIBRATION_INTENT,
}

def detect_question_intent(question: str) -> str:
    text = clean_text(question).lower()
    if not text:
        return GENERAL_MINE_SAFETY_INTENT

    if any(keyword in text for keyword in OUT_OF_SCOPE_HINTS) and not any(
        keyword in text for keyword in MINING_SAFETY_CONTEXT_KEYWORDS
    ):
        return OUT_OF_SCOPE_INTENT

    ordered_intents = [
        PPE_GENERAL_INTENT,
        VENTILATION_EQUIPMENT_INTENT,
        GAS_DETECTOR_INTENT,
        VENTILATION_GAS_INTENT,
        BLASTING_MISFIRE_INTENT,
        ROOF_FALL_INTENT,
        DUST_RESPIRATORY_INTENT,
        ELECTRICAL_SAFETY_INTENT,
        CONVEYOR_ROTATING_INTENT,
        BACKING_SIGNAL_INTENT,
        EQUIPMENT_TRANSPORT_INTENT,
        HOT_WORK_INTENT,
        FIRE_EMERGENCY_INTENT,
        CHEMICAL_FUEL_INTENT,
        HEIGHT_WORK_INTENT,
        LIFTING_HEAVY_INTENT,
        FLOOD_DRAINAGE_INTENT,
        EVACUATION_ACCESS_SIGN_INTENT,
        WALKWAY_HOUSEKEEPING_INTENT,
        LIGHTING_VISIBILITY_INTENT,
        FATIGUE_WORKER_INTENT,
        CONTRACTOR_PERMIT_INTENT,
        NEW_WORKER_TRAINING_INTENT,
        INSPECTION_RECORD_INTENT,
        ERGONOMICS_MANUAL_INTENT,
        NOISE_VIBRATION_INTENT,
        DOCUMENT_EVIDENCE_INTENT,
        KRAS_INTENT,
        LAW_INTENT,
        PREWORK_TBM_INTENT,
        ACCIDENT_EMERGENCY_INTENT,
    ]
    for intent in ordered_intents:
        if any(keyword in text for keyword in QUESTION_INTENT_KEYWORDS[intent]):
            return intent

    if any(keyword in text for keyword in MINING_SAFETY_CONTEXT_KEYWORDS):
        return GENERAL_MINE_SAFETY_INTENT
    return OUT_OF_SCOPE_INTENT


def detect_ppe_item(question: str) -> str:
    text = clean_text(question).lower()
    if "안전화" in text or "작업화" in text:
        return "안전화"
    if "안전모" in text:
        return "안전모"
    if "방진마스크" in text or "호흡보호구" in text:
        return "방진마스크"
    if "마스크" in text:
        return "방진마스크"
    if "보안경" in text:
        return "보안경"
    if "귀마개" in text or "귀덮개" in text:
        return "귀마개/귀덮개"
    if "안전대" in text:
        return "안전대"
    if "장갑" in text:
        return "장갑"
    return "보호구"


def expand_search_query(question: str, intent: str | None = None) -> str:
    intent = intent or detect_question_intent(question)
    if intent == OUT_OF_SCOPE_INTENT:
        return question

    expansions = list(INTENT_SEARCH_EXPANSIONS.get(intent, []))
    ppe_item = detect_ppe_item(question)
    if ppe_item in PPE_ITEM_EXPANSIONS:
        expansions.extend(PPE_ITEM_EXPANSIONS[ppe_item])
    unique_terms = list(dict.fromkeys(term for term in expansions if term))
    if not unique_terms:
        return question
    return " ".join([question, *unique_terms])


def map_intent_to_situation_type(intent: str, question: str = "") -> str:
    mapping = {
        PPE_GENERAL_INTENT: PPE_GENERAL_INTENT,
        VENTILATION_GAS_INTENT: "환기/유해가스",
        BLASTING_MISFIRE_INTENT: BLASTING_QUESTION_TYPE,
        ROOF_FALL_INTENT: "낙반/붕락",
        DUST_RESPIRATORY_INTENT: PPE_DUST_QUESTION_TYPE,
        ELECTRICAL_SAFETY_INTENT: ELECTRICAL_QUESTION_TYPE,
        EQUIPMENT_TRANSPORT_INTENT: "장비/운반",
        PREWORK_TBM_INTENT: TBM_QUESTION_TYPE,
        ACCIDENT_EMERGENCY_INTENT: ACCIDENT_RESPONSE_QUESTION_TYPE,
        KRAS_INTENT: RISK_ASSESSMENT_QUESTION_TYPE,
        LAW_INTENT: "중대재해처벌법",
        GENERAL_MINE_SAFETY_INTENT: "일반 광산 안전",
        FIRE_EMERGENCY_INTENT: "화재/폭발",
        EVACUATION_ACCESS_SIGN_INTENT: PASSAGE_SAFETY_QUESTION_TYPE,
        LIGHTING_VISIBILITY_INTENT: PASSAGE_SAFETY_QUESTION_TYPE,
        GAS_DETECTOR_INTENT: "환기/유해가스",
        VENTILATION_EQUIPMENT_INTENT: "환기/유해가스",
        FLOOD_DRAINAGE_INTENT: "일반 광산 안전",
        WALKWAY_HOUSEKEEPING_INTENT: PASSAGE_SAFETY_QUESTION_TYPE,
        BACKING_SIGNAL_INTENT: "장비/운반",
        CONVEYOR_ROTATING_INTENT: "장비/운반",
        HOT_WORK_INTENT: "화재/폭발",
        HEIGHT_WORK_INTENT: "일반 광산 안전",
        LIFTING_HEAVY_INTENT: "장비/운반",
        FATIGUE_WORKER_INTENT: "일반 광산 안전",
        CONTRACTOR_PERMIT_INTENT: TBM_QUESTION_TYPE,
        NEW_WORKER_TRAINING_INTENT: TBM_QUESTION_TYPE,
        INSPECTION_RECORD_INTENT: DAILY_INSPECTION_QUESTION_TYPE,
        ERGONOMICS_MANUAL_INTENT: "일반 광산 안전",
        NOISE_VIBRATION_INTENT: PPE_DUST_QUESTION_TYPE,
        CHEMICAL_FUEL_INTENT: "화재/폭발",
        DOCUMENT_EVIDENCE_INTENT: "중대재해처벌법",
        OUT_OF_SCOPE_INTENT: OUT_OF_SCOPE_INTENT,
    }
    return mapping.get(intent, "일반 광산 안전")


def assess_rag_evidence_sufficiency(
    question: str,
    results: list[dict[str, Any]],
    intent: str | None = None,
) -> dict[str, Any]:
    """공식 RAG 검색 결과의 구조적 충분성을 보수적으로 분류합니다.

    이 결과는 정확도나 법적 신뢰도 점수가 아닙니다. 현재 collection의 distance
    의미를 단정하지 않고 문서·chunk 식별자, 비어 있지 않은 본문, 중복과 출처
    편중만 확인하는 RAG 근거 충분성 휴리스틱입니다.
    """
    official_chunk_count = len(results)
    valid_sources: list[str] = []
    seen_chunk_ids: set[str] = set()
    non_empty_chunk_count = 0
    duplicate_chunk_count = 0
    missing_source_count = 0
    missing_chunk_id_count = 0
    valid_identity_count = 0

    for result in results:
        source = str(result.get("source", "")).strip()
        chunk_id = str(result.get("chunk_id", "")).strip()
        text = str(result.get("text", "")).strip()
        source_is_valid = source not in {"", "출처 정보 없음", "None"}
        chunk_id_is_valid = chunk_id not in {"", "정보 없음", "None"}

        if text:
            non_empty_chunk_count += 1
        if source_is_valid:
            valid_sources.append(source)
        else:
            missing_source_count += 1
        if chunk_id_is_valid:
            if chunk_id in seen_chunk_ids:
                duplicate_chunk_count += 1
            else:
                seen_chunk_ids.add(chunk_id)
        else:
            missing_chunk_id_count += 1
        if source_is_valid and chunk_id_is_valid:
            valid_identity_count += 1

    unique_document_count = len(set(valid_sources))
    source_counts = Counter(valid_sources)
    largest_document_count = max(source_counts.values(), default=0)
    single_document_ratio = (
        largest_document_count / len(valid_sources)
        if valid_sources
        else 0.0
    )
    duplicate_ratio = (
        duplicate_chunk_count / official_chunk_count
        if official_chunk_count
        else 0.0
    )
    intent_match_count = sum(
        not intent
        or str(result.get("question_intent", "")).strip() in {"", intent}
        for result in results
    )

    severe_limitations = [
        official_chunk_count < 2,
        non_empty_chunk_count < 2,
        valid_identity_count < 2,
        unique_document_count < 1,
        duplicate_ratio > 0.5,
    ]
    sufficient_conditions = [
        official_chunk_count >= EVIDENCE_MIN_OFFICIAL_CHUNKS,
        non_empty_chunk_count >= EVIDENCE_MIN_NON_EMPTY_CHUNKS,
        unique_document_count >= EVIDENCE_MIN_UNIQUE_DOCUMENTS,
        missing_source_count == 0,
        missing_chunk_id_count == 0,
        duplicate_ratio <= EVIDENCE_MAX_DUPLICATE_RATIO,
        single_document_ratio <= EVIDENCE_MAX_SINGLE_DOCUMENT_RATIO,
    ]

    if any(severe_limitations):
        status = EVIDENCE_STATUS_INSUFFICIENT
    elif all(sufficient_conditions):
        status = EVIDENCE_STATUS_SUFFICIENT
    else:
        status = EVIDENCE_STATUS_NEEDS_REVIEW

    return {
        "status": status,
        "label": EVIDENCE_STATUS_LABELS[status],
        "reason": EVIDENCE_STATUS_REASONS[status],
        "official_chunk_count": official_chunk_count,
        "unique_document_count": unique_document_count,
        "non_empty_chunk_count": non_empty_chunk_count,
        "duplicate_chunk_count": duplicate_chunk_count,
        "diagnostic_details": {
            "valid_identity_count": valid_identity_count,
            "missing_source_count": missing_source_count,
            "missing_chunk_id_count": missing_chunk_id_count,
            "duplicate_ratio": round(duplicate_ratio, 3),
            "single_document_ratio": round(single_document_ratio, 3),
            "question_intent": intent or detect_question_intent(question),
            "intent_match_count": intent_match_count,
            "thresholds": {
                "min_official_chunks": EVIDENCE_MIN_OFFICIAL_CHUNKS,
                "min_unique_documents": EVIDENCE_MIN_UNIQUE_DOCUMENTS,
                "min_non_empty_chunks": EVIDENCE_MIN_NON_EMPTY_CHUNKS,
                "max_duplicate_ratio": EVIDENCE_MAX_DUPLICATE_RATIO,
                "max_single_document_ratio": EVIDENCE_MAX_SINGLE_DOCUMENT_RATIO,
            },
            "distance_interpretation": (
                "ChromaDB distance는 원시 진단값으로만 유지하며 충분성 판정에는 사용하지 않음"
            ),
        },
    }


def evidence_is_limited(results: list[dict[str, Any]]) -> bool:
    assessment = assess_rag_evidence_sufficiency("", results)
    return assessment["status"] != EVIDENCE_STATUS_SUFFICIENT


def evidence_limitation_notice(results: list[dict[str, Any]]) -> str:
    if evidence_is_limited(results):
        return "검색된 문서 근거가 제한적이므로, 구체적인 현장 기준은 사업장 안전보건관리규정과 담당 안전관리자 확인이 필요합니다."
    return ""


def build_evidence_guardrail_prompt_guidance(
    assessment: dict[str, Any] | None,
) -> str:
    assessment = assessment or {
        "status": EVIDENCE_STATUS_NEEDS_REVIEW,
        "label": EVIDENCE_STATUS_LABELS[EVIDENCE_STATUS_NEEDS_REVIEW],
        "reason": EVIDENCE_STATUS_REASONS[EVIDENCE_STATUS_NEEDS_REVIEW],
    }
    status = str(assessment.get("status", EVIDENCE_STATUS_NEEDS_REVIEW))
    label = str(assessment.get("label", EVIDENCE_STATUS_LABELS[EVIDENCE_STATUS_NEEDS_REVIEW]))
    reason = str(assessment.get("reason", EVIDENCE_STATUS_REASONS[EVIDENCE_STATUS_NEEDS_REVIEW]))
    lines = [
        "[공식 근거 검색 상태]",
        f"- 상태: {label}",
        f"- 판정 사유: {reason}",
        "- 이 상태는 정확도나 법적 신뢰도 점수가 아니라 RAG 검색 결과의 구조적 충분성을 보는 보조 휴리스틱입니다.",
        "- 제공된 RAG 근거에 없는 법령 조항 번호·수치·처벌·작업 재개 조건을 새로 만들지 마세요.",
        "- 공식 근거는 검색된 RAG 문서명과 chunk_id를 기준으로 하세요.",
        "- 뉴스는 공식 법령 판단 근거로 사용하지 마세요.",
        "- 사례 기반 주의 포인트는 공식 법령 판단 근거로 사용하지 마세요.",
        "- 중대재해처벌법 위반을 확정적으로 단정하지 마세요.",
    ]
    if status == EVIDENCE_STATUS_INSUFFICIENT:
        lines.extend(
            [
                f"- {EVIDENCE_INSUFFICIENT_GUIDANCE}",
                "- 즉시 작업중지, 위험구역 접근통제, 작업자 대피, 담당 안전관리자 보고, 재점검과 승인 전 작업 재개 금지처럼 보수적인 조치만 우선 안내하세요.",
            ]
        )
    return "\n".join(lines)


def enforce_rag_evidence_answer_guardrail(
    answer: str,
    assessment: dict[str, Any] | None,
    *,
    worker_easy: bool = False,
) -> str:
    """공식 근거 상태에 맞게 최종 사용자 답변의 단정 표현을 제한합니다."""
    guarded_answer = str(answer or "").strip()
    cautious_replacements = {
        "중대재해처벌법 위반 " + "확정": "중대재해처벌법상 쟁점이 될 수 있음",
        "처벌이 확정됨": "실제 처벌 여부는 사실관계와 법령 해석에 따라 달라질 수 있음",
        "사업주가 반드시 처벌됨": "사업주의 책임 여부는 사고 경위와 의무 이행 여부에 따라 달라질 수 있음",
        "해당 사고는 법 위반임": "해당 사고는 법적 검토가 필요한 사안일 수 있음",
    }
    for prohibited, cautious in cautious_replacements.items():
        guarded_answer = guarded_answer.replace(prohibited, cautious)

    status = (
        str(assessment.get("status", ""))
        if isinstance(assessment, dict)
        else ""
    )
    if status != EVIDENCE_STATUS_INSUFFICIENT:
        return guarded_answer

    reason = str(
        assessment.get(
            "reason",
            EVIDENCE_STATUS_REASONS[EVIDENCE_STATUS_INSUFFICIENT],
        )
    )
    if worker_easy:
        return "\n".join(
            [
                "### 지금 바로 해야 할 일",
                "- 작업을 멈추세요.",
                "- 위험한 곳에 다른 사람이 가까이 가지 못하게 막으세요.",
                "- 작업자는 안전한 곳으로 이동하세요.",
                "- 담당 안전관리자에게 바로 알리세요.",
                "",
                "### 왜 위험한가요?",
                f"{reason} 정확한 수치나 시간은 지금 확인된 근거만으로 말할 수 없습니다.",
                "",
                "### 작업을 다시 시작하기 전에",
                "- 가스, 환기와 설비 상태를 다시 확인하세요.",
                "- 위험요인이 없어졌는지 확인하세요.",
                "- 관리자가 안전을 확인하고 다시 작업해도 된다고 할 때까지 기다리세요.",
            ]
        )
    return "\n".join(
        [
            "### 공식 근거 검색 상태",
            f"{EVIDENCE_STATUS_LABELS[EVIDENCE_STATUS_INSUFFICIENT]}: {reason}",
            "",
            EVIDENCE_INSUFFICIENT_GUIDANCE,
            "",
            "### 우선 적용할 보수적 안전조치",
            "- 즉시 작업중지",
            "- 위험구역 접근통제와 작업자 대피",
            "- 담당 안전관리자에게 즉시 보고",
            "- 가스·환기·설비 상태 재점검",
            "- 관계기관 또는 전문가를 통한 현장 조건과 최신 법령 확인",
            "- 개선조치 완료 확인 및 책임자 승인 전 작업 재개 금지",
            "",
            "현재 검색 결과에 없는 법령 조항 번호, 수치 기준, 처벌 수준 또는 작업 재개 가능 여부는 단정하지 않습니다.",
        ]
    )


def gemini_intent_guidance(intent: str) -> str:
    if intent == PPE_GENERAL_INTENT:
        return "보호구 질문입니다. 작업중지부터 말하지 말고 착용 이유, 필요한 작업·상황, 착용 전 점검사항, 현장관리자 확인사항을 자연스럽게 설명하세요."
    if intent in RISK_SIGN_INTENTS:
        return "위험 징후 대응형 질문입니다. 즉시 판단, 작업중지 또는 출입통제, 대피, 점검·재측정, 재개 전 확인, 기록·보고를 빠뜨리지 마세요."
    if intent in MANAGEMENT_RECORD_INTENTS:
        return "관리·기록형 질문입니다. 필요한 기록, 현장 확인사항, 책임자 확인, 보관 증빙자료, 법령 단정 금지 안내를 중심으로 답하세요."
    if intent in GENERAL_EXPLANATION_INTENTS:
        return "일반 안전교육형 질문입니다. 왜 필요한지, 언제 적용하는지, 작업 전 확인사항과 교육·기록 사항을 설명하세요."
    if intent == KRAS_INTENT:
        return "위험성평가 질문입니다. 위험요인, 현재 위험성, 감소대책, 확인 기록과 KRAS 초안 연결을 중심으로 답하세요."
    if intent == LAW_INTENT:
        return "법령 설명 질문입니다. 법령 조항 번호·처벌 수위는 검색 근거 없이 만들지 말고, 현장관리자 조치와 증빙자료를 중심으로 설명하세요."
    return "광산 안전관리 범위의 질문으로 보고 검색 근거에 맞춰 답하세요."


def worker_easy_intent_guidance(intent: str) -> str:
    """질문 유형별 핵심 위험을 신규 근로자가 이해하기 쉬운 말로 안내합니다."""
    if intent == CONVEYOR_ROTATING_INTENT:
        return "손, 옷이나 장갑이 기계에 끌려 들어갈 수 있음을 설명하고, 정지 전 접근 금지와 전원 차단·재가동 방지를 먼저 안내하세요."
    if intent in {EQUIPMENT_TRANSPORT_INTENT, BACKING_SIGNAL_INTENT}:
        return "운전자의 사각지대 때문에 사람이 보이지 않을 수 있음을 설명하고, 후진 전 신호수와 작업자가 신호를 확인하도록 안내하세요."
    if intent in {VENTILATION_GAS_INTENT, VENTILATION_EQUIPMENT_INTENT}:
        return "가스 때문에 불이 붙거나 숨쉬기 어려울 수 있음을 설명하고, 불꽃을 만들지 말고 작업중지·대피·환기·재측정을 우선 안내하세요."
    if intent == BLASTING_MISFIRE_INTENT:
        return "불발공은 폭약이 터지지 않고 남아 있을 수 있는 구멍이라고 설명하고, 접근·접촉·재천공을 금지하며 발파책임자에게 알리도록 안내하세요."
    if intent in {DUST_RESPIRATORY_INTENT, PPE_GENERAL_INTENT}:
        return "눈에 잘 보이지 않는 먼지도 폐로 들어갈 수 있음을 설명하고, 살수·집진 상태와 방진마스크 밀착을 확인하도록 안내하세요."
    if intent == ROOF_FALL_INTENT:
        return "천장이나 벽에서 돌이 떨어질 수 있음을 설명하고, 위험구역 접근 금지와 균열·지보 상태 확인을 안내하세요."
    if intent == ELECTRICAL_SAFETY_INTENT:
        return "전원이 꺼져 보이더라도 전기가 남아 있을 수 있음을 설명하고, 차단·잠금·표지와 젖은 장소의 감전 위험을 안내하세요."
    if intent in MANAGEMENT_RECORD_INTENTS:
        return "처벌 설명은 길게 하지 말고, 작업자를 보호하려면 보고와 기록이 왜 필요한지 쉬운 말로 설명하세요. 실제 위반 여부는 단정하지 마세요."
    return "가장 중요한 행동을 먼저 말하고, 작업자가 다칠 수 있는 이유를 쉬운 결과 중심으로 설명하세요."


def build_worker_easy_evidence_guidance(
    assessment: dict[str, Any] | None,
) -> str:
    """이미 계산된 공식 근거 검색 상태를 근로자용 쉬운 문장으로 바꿉니다."""
    status = (
        str(assessment.get("status", EVIDENCE_STATUS_NEEDS_REVIEW))
        if isinstance(assessment, dict)
        else EVIDENCE_STATUS_NEEDS_REVIEW
    )
    if status == EVIDENCE_STATUS_SUFFICIENT:
        return "검색된 공식 문서가 주요 안전조치를 뒷받침합니다. 아래 행동을 먼저 지키세요."
    if status == EVIDENCE_STATUS_INSUFFICIENT:
        return "검색된 공식 문서만으로 정확한 수치나 시간을 확인하기 어렵습니다. 먼저 작업을 멈추고 안전관리자에게 확인하세요."
    return "관련 공식 문서는 찾았지만 세부 수치나 시간은 더 확인해야 합니다. 안전관리자에게 확인한 뒤 작업하세요."


def is_blasting_preferred_source(source: str) -> bool:
    normalized = clean_text(source).lower()
    return any(
        marker.lower() in normalized
        for marker in BLASTING_PREFERRED_SOURCE_MARKERS
    )


def is_blasting_general_source(source: str) -> bool:
    normalized = clean_text(source).lower()
    return any(
        marker.lower() in normalized
        for marker in BLASTING_GENERAL_SOURCE_MARKERS
    )


def is_electrical_preferred_source(source: str) -> bool:
    normalized = clean_text(source).lower()
    return any(
        marker.lower() in normalized
        for marker in ELECTRICAL_PREFERRED_SOURCE_MARKERS
    )


def is_electrical_general_source(source: str) -> bool:
    normalized = clean_text(source).lower()
    return any(
        marker.lower() in normalized
        for marker in ELECTRICAL_GENERAL_SOURCE_MARKERS
    )


def squared_l2_distance(
    query_embedding: list[float],
    document_embedding: Any,
) -> float | None:
    if document_embedding is None:
        return None
    try:
        return sum(
            (float(query_value) - float(document_value)) ** 2
            for query_value, document_value in zip(
                query_embedding,
                document_embedding,
            )
        )
    except (TypeError, ValueError):
        return None


def make_search_result(
    doc: str,
    meta: dict[str, Any],
    distance: Any,
    vector_rank: int | None = None,
) -> dict[str, Any]:
    return {
        "rank": vector_rank or 0,
        "vector_rank": vector_rank,
        "source": get_source(meta),
        "chunk_id": get_chunk_id(meta),
        "distance": distance,
        "text": clean_text(doc or ""),
        "metadata": meta,
    }


def add_blasting_source_candidates(
    collection,
    query_embedding: list[float],
    candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    try:
        preferred = collection.get(
            where={"source": {"$in": BLASTING_PREFERRED_SOURCE_FILES}},
            include=["documents", "metadatas", "embeddings"],
        )
    except Exception:
        return candidates

    existing_keys = {
        (result["source"], result["chunk_id"])
        for result in candidates
    }
    documents = preferred.get("documents") or []
    metadatas = preferred.get("metadatas") or []
    embeddings = preferred.get("embeddings")
    if embeddings is None:
        embeddings = [None] * len(documents)

    for doc, meta, embedding in zip(documents, metadatas, embeddings):
        safe_meta = meta if isinstance(meta, dict) else {}
        result = make_search_result(
            doc or "",
            safe_meta,
            squared_l2_distance(query_embedding, embedding),
        )
        key = (result["source"], result["chunk_id"])
        if key not in existing_keys:
            candidates.append(result)
            existing_keys.add(key)
    return candidates


def add_electrical_source_candidates(
    collection,
    query_embedding: list[float],
    candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    try:
        preferred = collection.get(
            where={"source": {"$in": ELECTRICAL_PREFERRED_SOURCE_FILES}},
            include=["documents", "metadatas", "embeddings"],
        )
    except Exception:
        return candidates

    existing_keys = {
        (result["source"], result["chunk_id"])
        for result in candidates
    }
    documents = preferred.get("documents") or []
    metadatas = preferred.get("metadatas") or []
    embeddings = preferred.get("embeddings")
    if embeddings is None:
        embeddings = [None] * len(documents)

    for doc, meta, embedding in zip(documents, metadatas, embeddings):
        safe_meta = meta if isinstance(meta, dict) else {}
        result = make_search_result(
            doc or "",
            safe_meta,
            squared_l2_distance(query_embedding, embedding),
        )
        key = (result["source"], result["chunk_id"])
        if key not in existing_keys:
            candidates.append(result)
            existing_keys.add(key)
    return candidates


def source_matches_markers(source: str, markers: list[str]) -> bool:
    normalized = clean_text(source).lower()
    return any(marker.lower() in normalized for marker in markers)


def add_type_source_candidates(
    collection,
    query_embedding: list[float],
    candidates: list[dict[str, Any]],
    preferred_source_files: list[str],
) -> list[dict[str, Any]]:
    if not preferred_source_files:
        return candidates

    try:
        preferred = collection.get(
            where={"source": {"$in": preferred_source_files}},
            include=["documents", "metadatas", "embeddings"],
        )
    except Exception:
        return candidates

    existing_keys = {
        (result["source"], result["chunk_id"])
        for result in candidates
    }
    documents = preferred.get("documents") or []
    metadatas = preferred.get("metadatas") or []
    embeddings = preferred.get("embeddings")
    if embeddings is None:
        embeddings = [None] * len(documents)

    for doc, meta, embedding in zip(documents, metadatas, embeddings):
        safe_meta = meta if isinstance(meta, dict) else {}
        result = make_search_result(
            doc or "",
            safe_meta,
            squared_l2_distance(query_embedding, embedding),
        )
        key = (result["source"], result["chunk_id"])
        if key not in existing_keys:
            candidates.append(result)
            existing_keys.add(key)
    return candidates


def rerank_search_results(
    question: str,
    candidates: list[dict[str, Any]],
    top_k: int,
) -> list[dict[str, Any]]:
    question_type = classify_question_type(question)
    question_intent = detect_question_intent(question)
    config = RERANK_TYPE_CONFIGS.get(question_type)
    reranking_applied = config is not None
    question_text = clean_text(question).lower()
    if config:
        active_keywords = [
            keyword for keyword in config["keywords"]
            if keyword in question_text
        ]
        reranking_label = config["label"]
    else:
        active_keywords = []
        reranking_label = "기본 벡터 유사도 정렬 적용"

    for result in candidates:
        source = result["source"]
        text = result["text"].lower()
        distance = result["distance"]
        score = float(distance) if isinstance(distance, (int, float)) else 99.0

        if config:
            if source in config["preferred_files"]:
                score -= 0.90
            elif source_matches_markers(
                source,
                config["preferred_markers"],
            ):
                score -= 0.55
            elif source_matches_markers(
                source,
                config["general_markers"],
            ):
                score += config["general_penalty"]

            keyword_hits = sum(keyword in text for keyword in active_keywords)
            score -= min(keyword_hits, 6) * config["keyword_weight"]

        if question_type == BLASTING_QUESTION_TYPE:
            if "불발" in question and "불발" in text:
                score -= 0.50
            if "발파 후" in question and "발파 후" in text:
                score -= 0.35
        elif question_type == ELECTRICAL_QUESTION_TYPE:
            if "누전" in question_text and "누전" in text:
                score -= 0.45
            if "감전" in question_text and "감전" in text:
                score -= 0.40
            if (
                ("습기" in question_text or "물기" in question_text)
                and ("습기" in text or "물기" in text)
            ):
                score -= 0.20
        elif question_type == RISK_ASSESSMENT_QUESTION_TYPE:
            if "위험성평가" in text:
                score -= 0.35
            if "감소대책" in text:
                score -= 0.25
        elif question_type == PPE_DUST_QUESTION_TYPE:
            if ("분진" in question_text or "먼지" in question_text) and (
                "분진" in text or "먼지" in text
            ):
                score -= 0.35
            if "보호구" in text or "방진마스크" in text:
                score -= 0.25
            direct_ppe_hits = sum(
                keyword in question_text and keyword in text
                for keyword in [
                    "안전모",
                    "방진마스크",
                    "호흡보호구",
                    "분진",
                    "먼지",
                ]
            )
            score -= min(direct_ppe_hits, 3) * 0.40
        elif question_type == ACCIDENT_RESPONSE_QUESTION_TYPE:
            if "재해조사" in text or "사고조사" in text:
                score -= 0.25
            if "재발방지" in text:
                score -= 0.25
        elif question_type == TBM_QUESTION_TYPE:
            if "tbm" in text or "작업 전 안전" in text:
                score -= 0.35
        elif question_type == COMPLEX_RISK_QUESTION_TYPE:
            complex_hits = sum(
                keyword in text
                for keyword in [
                    "불발",
                    "후가스",
                    "메탄",
                    "환기",
                    "낙반",
                    "부석",
                    "운반",
                ]
            )
            score -= min(complex_hits, 5) * 0.12

        result["rerank_score"] = score
        result["question_type"] = question_type
        result["question_intent"] = question_intent
        result["reranking_applied"] = reranking_applied
        result["reranking_label"] = reranking_label

    if not reranking_applied:
        selected = candidates[:top_k]
    else:
        ordered = sorted(
            candidates,
            key=lambda item: (
                item["rerank_score"],
                item["vector_rank"] or 9999,
            ),
        )
        selected = []
        selected_sources = set()

        # 같은 문서의 유사 chunk가 결과를 독점하지 않도록 먼저 출처를 다양화합니다.
        for result in ordered:
            if result["source"] in selected_sources:
                continue
            selected.append(result)
            selected_sources.add(result["source"])
            if len(selected) == top_k:
                break

        if len(selected) < top_k:
            selected_keys = {
                (result["source"], result["chunk_id"])
                for result in selected
            }
            for result in ordered:
                key = (result["source"], result["chunk_id"])
                if key in selected_keys:
                    continue
                selected.append(result)
                selected_keys.add(key)
                if len(selected) == top_k:
                    break

    for rank, result in enumerate(selected, start=1):
        result["rank"] = rank
    return selected


def search_vector_db(question: str, top_k: int = 5):
    collection, error = load_chroma_collection()
    if error:
        return [], error

    intent = detect_question_intent(question)
    expanded_question = expand_search_query(question, intent)
    if intent != OUT_OF_SCOPE_INTENT:
        top_k = max(top_k, 5)

    model = load_embedding_model()
    query_embedding = model.encode(
        [expanded_question],
        normalize_embeddings=True,
        show_progress_bar=False,
    ).tolist()

    question_type = classify_question_type(expanded_question)
    rerank_config = RERANK_TYPE_CONFIGS.get(question_type)
    internal_count = top_k
    if rerank_config:
        internal_count = max(
            top_k,
            int(rerank_config["internal_count"]),
        )
    try:
        internal_count = min(internal_count, collection.count())
    except Exception:
        pass

    try:
        results = collection.query(
            query_embeddings=query_embedding,
            n_results=internal_count,
            include=["documents", "metadatas", "distances"],
        )
    except Exception as e:
        return [], str(e)

    docs = results.get("documents", [[]])[0]
    metas = results.get("metadatas", [[]])[0]
    distances = results.get("distances", [[]])[0]

    search_candidates = []
    for i, doc in enumerate(docs):
        meta = metas[i] if i < len(metas) and isinstance(metas[i], dict) else {}
        distance = distances[i] if i < len(distances) else None
        search_candidates.append(
            make_search_result(
                doc or "",
                meta,
                distance,
                vector_rank=i + 1,
            )
        )

    if rerank_config:
        search_candidates = add_type_source_candidates(
            collection,
            query_embedding[0],
            search_candidates,
            rerank_config["preferred_files"],
        )

    selected = rerank_search_results(
        expanded_question,
        search_candidates,
        top_k,
    )
    for result in selected:
        result["question_intent"] = intent
        result["expanded_search_query"] = expanded_question
    return selected, None


def official_case_query_terms(question_type: str) -> list[str]:
    mapping = {
        CONVEYOR_ROTATING_INTENT: ["컨베이어", "끼임", "말림", "회전체", "재가동", "청소"],
        BACKING_SIGNAL_INTENT: ["후진", "신호수", "충돌", "깔림", "차량", "지게차", "덤프트럭"],
        EQUIPMENT_TRANSPORT_INTENT: ["후진", "충돌", "깔림", "차량", "지게차", "덤프트럭", "굴착기"],
        VENTILATION_GAS_INTENT: ["질식", "산소결핍", "유해가스", "밀폐공간", "중독", "환기"],
        VENTILATION_EQUIPMENT_INTENT: ["질식", "산소결핍", "유해가스", "밀폐공간", "환기"],
        GAS_DETECTOR_INTENT: ["질식", "산소결핍", "유해가스", "밀폐공간", "중독"],
        ROOF_FALL_INTENT: ["붕괴", "매몰", "낙하", "토사", "암석", "무너짐", "낙반"],
        ELECTRICAL_SAFETY_INTENT: ["감전", "누전", "전기설비", "전원", "정비", "에너지 차단"],
        BLASTING_MISFIRE_INTENT: ["발파", "화약", "폭약", "폭발", "불발"],
        DUST_RESPIRATORY_INTENT: ["분진", "집진", "호흡기", "방진마스크", "먼지"],
        HEIGHT_WORK_INTENT: ["추락", "떨어짐", "고소작업", "작업발판", "사다리"],
        FIRE_EMERGENCY_INTENT: ["화재", "폭발", "불꽃", "가연물"],
        HOT_WORK_INTENT: ["화재", "폭발", "용접", "불꽃", "가연물"],
        LIFTING_HEAVY_INTENT: ["인양", "크레인", "낙하", "맞음", "중량물"],
    }
    return mapping.get(question_type, [])


def official_case_allowed_accident_types(question_type: str) -> set[str]:
    mapping = {
        CONVEYOR_ROTATING_INTENT: {"끼임", "말림"},
        BACKING_SIGNAL_INTENT: {"충돌", "깔림", "끼임"},
        EQUIPMENT_TRANSPORT_INTENT: {"충돌", "깔림", "끼임"},
        VENTILATION_GAS_INTENT: {"질식", "중독", "폭발"},
        VENTILATION_EQUIPMENT_INTENT: {"질식", "중독", "폭발"},
        GAS_DETECTOR_INTENT: {"질식", "중독", "폭발"},
        ROOF_FALL_INTENT: {"붕괴", "매몰", "낙하", "맞음"},
        ELECTRICAL_SAFETY_INTENT: {"감전"},
        BLASTING_MISFIRE_INTENT: {"폭발"},
        HEIGHT_WORK_INTENT: {"추락", "떨어짐", "낙하"},
        FIRE_EMERGENCY_INTENT: {"화재", "폭발"},
        HOT_WORK_INTENT: {"화재", "폭발"},
    }
    return mapping.get(question_type, set())


def official_case_relation_terms(question_type: str) -> dict[str, tuple[str, ...]]:
    mapping = {
        CONVEYOR_ROTATING_INTENT: {
            "direct": ("컨베이어", "벨트", "끼임", "말림"),
            "analogous": (
                "회전체", "정비 중 재가동", "청소 중 기계 작동",
                "방호장치", "에너지 차단", "전원 미차단",
            ),
        },
        BACKING_SIGNAL_INTENT: {
            "direct": ("후진", "신호수", "덤프트럭", "굴착기", "지게차", "차량 충돌"),
            "analogous": ("이동식 장비", "사각지대", "깔림", "중장비", "작업자 충돌"),
        },
        EQUIPMENT_TRANSPORT_INTENT: {
            "direct": ("후진", "신호수", "덤프트럭", "굴착기", "지게차", "차량 충돌"),
            "analogous": ("이동식 장비", "사각지대", "깔림", "중장비", "작업자 충돌"),
        },
        ROOF_FALL_INTENT: {
            "direct": ("낙반", "암석 붕락", "갱도 붕괴"),
            "analogous": ("토사 붕괴", "굴착면 붕괴", "매몰", "구조물 무너짐", "물체 낙하"),
        },
        ELECTRICAL_SAFETY_INTENT: {
            "direct": ("감전", "전기설비", "누전"),
            "analogous": ("정비 중 전원 투입", "에너지 차단", "잠금", "표지", "설비 재가동"),
        },
        VENTILATION_GAS_INTENT: {
            "direct": ("산소결핍", "유해가스", "질식"),
            "analogous": ("밀폐공간", "환기 불량", "가스 중독", "탱크", "맨홀"),
        },
        VENTILATION_EQUIPMENT_INTENT: {
            "direct": ("산소결핍", "유해가스", "질식"),
            "analogous": ("밀폐공간", "환기 불량", "가스 중독", "탱크", "맨홀"),
        },
        GAS_DETECTOR_INTENT: {
            "direct": ("산소결핍", "유해가스", "질식"),
            "analogous": ("밀폐공간", "환기 불량", "가스 중독", "탱크", "맨홀"),
        },
        BLASTING_MISFIRE_INTENT: {
            "direct": ("발파", "불발공", "화약"),
            "analogous": ("폭발", "폭발물", "점화원", "화약류"),
        },
        HEIGHT_WORK_INTENT: {
            "direct": ("추락", "떨어짐", "고소작업"),
            "analogous": ("작업발판", "사다리", "개구부"),
        },
        FIRE_EMERGENCY_INTENT: {
            "direct": ("화재", "불꽃"),
            "analogous": ("폭발", "가연물", "점화원"),
        },
        HOT_WORK_INTENT: {
            "direct": ("용접", "화기작업", "화재"),
            "analogous": ("폭발", "가연물", "점화원"),
        },
    }
    return mapping.get(question_type, {"direct": (), "analogous": ()})


def official_case_risk_family(question_type: str) -> str:
    mapping = {
        CONVEYOR_ROTATING_INTENT: "mechanical_entanglement",
        BACKING_SIGNAL_INTENT: "vehicle_transport",
        EQUIPMENT_TRANSPORT_INTENT: "vehicle_transport",
        ROOF_FALL_INTENT: "collapse_falling",
        ELECTRICAL_SAFETY_INTENT: "electrical_energy",
        VENTILATION_GAS_INTENT: "asphyxiation_gas",
        VENTILATION_EQUIPMENT_INTENT: "asphyxiation_gas",
        GAS_DETECTOR_INTENT: "asphyxiation_gas",
        BLASTING_MISFIRE_INTENT: "fire_explosion",
        FIRE_EMERGENCY_INTENT: "fire_explosion",
        HOT_WORK_INTENT: "fire_explosion",
        HEIGHT_WORK_INTENT: "fall_from_height",
    }
    return mapping.get(question_type, "")


def initialize_official_case_search_diagnostic(
    question_type: str,
    expanded_query_terms: list[str],
) -> dict[str, Any]:
    return {
        "status": "not_started",
        "active_collection_name": "",
        "active_collection_names": [],
        "collection_total_count": 0,
        "collection_counts": {},
        "raw_query_result_count": 0,
        "after_duplicate_filter_count": 0,
        "after_text_safety_filter_count": 0,
        "after_verification_tier_filter_count": 0,
        "after_relation_filter_count": 0,
        "final_result_count": 0,
        "result_count": 0,
        "question_type": question_type,
        "expanded_query_terms": list(expanded_query_terms),
        "direct_candidate_count": 0,
        "analogous_candidate_count": 0,
        "broad_family_candidate_count": 0,
        "removal_reasons": {},
        "embedding_model": EMBEDDING_MODEL_NAME,
        "query_embedding_dimension": 0,
    }


def public_case_text_is_safe_for_search(case: dict[str, Any]) -> bool:
    sanitized = verified_review.sanitize_display_case(case)
    summary = str(sanitized.get("display_accident_summary", ""))
    compact_length = len("".join(summary.split()))
    if not 40 <= compact_length <= 450:
        return False
    if not str(case.get("source_document", "")).strip():
        return False
    if case.get("original_page_number") in (None, "") and case.get("page_start") in (None, ""):
        return False
    for field in (
        "display_accident_summary",
        "display_cause_summary",
        "display_prevention_summary",
    ):
        value = str(sanitized.get(field, ""))
        if value and verified_review.detect_corrupted_ocr_text(value):
            return False
    return True


def search_official_siren_cases(
    question: str,
    question_type: str,
    top_k: int = OFFICIAL_CASE_TOP_K,
    diagnostic: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    diagnostic = diagnostic if isinstance(diagnostic, dict) else {}
    relation_terms = official_case_relation_terms(question_type)
    direct_terms = relation_terms.get("direct", ())
    analogous_terms = relation_terms.get("analogous", ())
    query_terms = list(dict.fromkeys((*direct_terms, *analogous_terms)))
    diagnostic.clear()
    diagnostic.update(
        initialize_official_case_search_diagnostic(question_type, query_terms)
    )
    if question_type == OUT_OF_SCOPE_INTENT:
        diagnostic.update({"status": "skipped_out_of_scope"})
        return []
    if not direct_terms and not analogous_terms:
        diagnostic.update({"status": "no_supported_case_type"})
        return []
    risk_family = official_case_risk_family(question_type)

    verified_collection, verified_db_status = load_official_case_collection()
    auto_collection, auto_db_status = load_auto_screened_case_collection()
    text_safe_collection, text_safe_db_status = load_text_safe_case_collection()
    available_collections = [
        ("verified", OFFICIAL_CASE_COLLECTION_NAME, verified_collection),
        (
            verified_review.AUTO_SCREENED_PUBLIC_TIER,
            AUTO_SCREENED_OFFICIAL_CASE_COLLECTION_NAME,
            auto_collection,
        ),
        (
            verified_review.TEXT_SAFE_FALLBACK_TIER,
            TEXT_SAFE_OFFICIAL_CASE_COLLECTION_NAME,
            text_safe_collection,
        ),
    ]
    available_collections = [item for item in available_collections if item[2] is not None]
    diagnostic.update(
        {
            "verified_db_status": verified_db_status or "ready",
            "auto_screened_db_status": auto_db_status or "ready",
            "text_safe_db_status": text_safe_db_status or "ready",
            "active_collection_names": [item[1] for item in available_collections],
            "active_collection_name": " + ".join(item[1] for item in available_collections),
        }
    )
    if not available_collections:
        actual_errors = [
            status
            for status in (verified_db_status, auto_db_status, text_safe_db_status)
            if status and ("db_unavailable" in status or "conflict" in status)
        ]
        diagnostic.update(
            {
                "status": "db_unavailable" if actual_errors else "no_quality_screened_cases",
            }
        )
        return []

    query_payloads: list[tuple[str, str, dict[str, Any]]] = []
    search_errors: list[str] = []
    try:
        model = load_embedding_model()
        expanded_query = clean_text(f"{question} {' '.join(query_terms)}")
        query_embedding = model.encode(
            [expanded_query],
            normalize_embeddings=True,
            show_progress_bar=False,
        ).tolist()
        diagnostic["query_embedding_dimension"] = (
            len(query_embedding[0]) if query_embedding and query_embedding[0] else 0
        )
    except Exception:
        diagnostic.update({"status": "search_failed"})
        return []

    collection_counts: dict[str, int] = {}
    for expected_tier, collection_name, collection in available_collections:
        try:
            collection_count = int(collection.count())
            collection_counts[expected_tier] = collection_count
            if collection_count <= 0:
                continue
            payload = collection.query(
                query_embeddings=query_embedding,
                n_results=min(
                    max(int(top_k) * 6, OFFICIAL_CASE_INTERNAL_SEARCH_COUNT),
                    collection_count,
                ),
                include=["documents", "metadatas", "distances"],
            )
            query_payloads.append((expected_tier, collection_name, payload))
        except Exception:
            search_errors.append(expected_tier)
    diagnostic["collection_counts"] = collection_counts
    diagnostic["collection_total_count"] = sum(collection_counts.values())

    raw_rows: list[dict[str, Any]] = []
    for expected_tier, collection_name, payload in query_payloads:
        documents = payload.get("documents", [[]])[0]
        metadatas = payload.get("metadatas", [[]])[0]
        distances = payload.get("distances", [[]])[0]
        for index, document in enumerate(documents):
            metadata = (
                metadatas[index]
                if index < len(metadatas) and isinstance(metadatas[index], dict)
                else {}
            )
            raw_rows.append(
                {
                    "expected_tier": expected_tier,
                    "collection_name": collection_name,
                    "document": str(document or ""),
                    "metadata": metadata,
                    "distance": distances[index] if index < len(distances) else None,
                }
            )
    diagnostic["raw_query_result_count"] = len(raw_rows)
    removal_reasons: Counter[str] = Counter()

    duplicate_filtered: list[dict[str, Any]] = []
    seen_query_rows: set[tuple[str, str]] = set()
    for row in raw_rows:
        metadata = row["metadata"]
        case_id = str(metadata.get("case_id", "")).strip()
        if not case_id:
            removal_reasons["missing_case_id"] += 1
            continue
        duplicate_key = (str(row["expected_tier"]), case_id)
        if duplicate_key in seen_query_rows:
            removal_reasons["duplicate_case_id_in_collection"] += 1
            continue
        seen_query_rows.add(duplicate_key)
        duplicate_filtered.append(row)
    diagnostic["after_duplicate_filter_count"] = len(duplicate_filtered)

    text_safe_rows: list[dict[str, Any]] = []
    for row in duplicate_filtered:
        if not public_case_text_is_safe_for_search(row["metadata"]):
            removal_reasons["display_text_safety_failed"] += 1
            continue
        text_safe_rows.append(row)
    diagnostic["after_text_safety_filter_count"] = len(text_safe_rows)

    tier_filtered_rows: list[dict[str, Any]] = []
    for row in text_safe_rows:
        metadata = row["metadata"]
        expected_tier = str(row["expected_tier"])
        verification_status = str(metadata.get("verification_status", ""))
        if metadata.get("official_case") is not True:
            removal_reasons["not_official_case"] += 1
            continue
        if expected_tier == "verified" and verification_status != "verified":
            removal_reasons["verified_tier_mismatch"] += 1
            continue
        if (
            expected_tier == verified_review.AUTO_SCREENED_PUBLIC_TIER
            and verification_status != verified_review.AUTO_SCREENED_STATUS
        ):
            removal_reasons["auto_screened_tier_mismatch"] += 1
            continue
        if (
            expected_tier == verified_review.TEXT_SAFE_FALLBACK_TIER
            and verified_review.effective_public_case_tier(metadata)
            != verified_review.TEXT_SAFE_FALLBACK_TIER
        ):
            removal_reasons["text_safe_tier_mismatch"] += 1
            continue
        if (
            expected_tier in {"verified", verified_review.AUTO_SCREENED_PUBLIC_TIER}
            and str(metadata.get("mine_relevance", "")) not in {"high", "medium"}
        ):
            removal_reasons["mine_relevance_not_eligible"] += 1
            continue
        if not verified_review.is_public_display_safe_case(metadata):
            removal_reasons["public_tier_safety_failed"] += 1
            continue
        tier_filtered_rows.append(row)
    diagnostic["after_verification_tier_filter_count"] = len(tier_filtered_rows)

    candidates: list[dict[str, Any]] = []
    direct_candidate_count = 0
    analogous_candidate_count = 0
    broad_family_candidate_count = 0
    for row in tier_filtered_rows:
        metadata = row["metadata"]
        relation_type, matched_terms = verified_review.classify_public_case_relation(
            metadata,
            direct_terms,
            analogous_terms,
            risk_family,
        )
        if relation_type not in {"direct", "analogous", "broad_family"}:
            removal_reasons["relation_mismatch"] += 1
            continue
        if relation_type == "direct":
            direct_candidate_count += 1
        elif relation_type == "analogous":
            analogous_candidate_count += 1
        else:
            broad_family_candidate_count += 1
        item = verified_review.sanitize_display_case(metadata)
        public_tier = verified_review.effective_public_case_tier(item)
        item.update(
            {
                "distance": row["distance"],
                "matched_terms": matched_terms,
                "relation_type": relation_type,
                "public_case_tier": public_tier,
                "source_grade": {
                    "verified": "원문 대조 검증 완료 공식 사고사례",
                    verified_review.AUTO_SCREENED_PUBLIC_TIER: "엄격 자동 품질검사 통과 공식 사고사례",
                    verified_review.TEXT_SAFE_FALLBACK_TIER: "문자 품질검사 통과 공식 사고사례",
                }.get(public_tier, "공개 제외 공식 사고사례"),
            }
        )
        candidates.append(item)

    diagnostic.update(
        {
            "after_relation_filter_count": len(candidates),
            "direct_candidate_count": direct_candidate_count,
            "analogous_candidate_count": analogous_candidate_count,
            "broad_family_candidate_count": broad_family_candidate_count,
        }
    )

    selected = verified_review.rank_public_official_cases(
        candidates,
        max_results=max(1, min(int(top_k), OFFICIAL_CASE_TOP_K)),
    )
    final_status = "search_failed" if search_errors and not query_payloads else (
        "ready" if selected else "no_search_results"
    )
    diagnostic.update(
        {
            "status": final_status,
            "search_error_sources": search_errors,
            "safety_rejected_count": removal_reasons.get("display_text_safety_failed", 0)
            + removal_reasons.get("public_tier_safety_failed", 0),
            "removal_reasons": dict(sorted(removal_reasons.items())),
            "final_result_count": len(selected),
            "result_count": len(selected),
        }
    )
    return selected


def build_context(results: list[dict[str, Any]]) -> str:
    context_blocks = []
    for r in results[:GEMINI_CONTEXT_TOP_K]:
        evidence_text = r["text"]
        if len(evidence_text) > CONTEXT_CHUNK_CHAR_LIMIT:
            evidence_text = evidence_text[:CONTEXT_CHUNK_CHAR_LIMIT].rstrip() + "..."

        block = (
            f"[근거 {r['rank']}]\n"
            f"출처: {r['source']}\n"
            f"chunk_id: {r.get('chunk_id', '정보 없음')}\n"
            f"거리값: {format_distance(r['distance'])}\n"
            f"내용:\n{evidence_text}\n"
        )
        context_blocks.append(block)
    return "\n".join(context_blocks)



def build_prompt(
    question: str,
    results: list[dict[str, Any]],
    intent: str | None = None,
    reference_cases: list[dict[str, Any]] | None = None,
    live_news_cases: list[dict[str, Any]] | None = None,
    evidence_assessment: dict[str, Any] | None = None,
    official_cases: list[dict[str, Any]] | None = None,
) -> str:
    context = build_context(results)
    intent = intent or detect_question_intent(question)
    intent_guidance = worker_easy_intent_guidance(intent)
    reference_case_context = format_reference_cases_for_prompt(intent, reference_cases or [])
    live_news_context = format_live_news_cases_for_prompt(live_news_cases or [])
    official_case_context = format_official_cases_for_prompt(official_cases or [])
    evidence_guardrail = build_evidence_guardrail_prompt_guidance(evidence_assessment)
    easy_evidence_guidance = build_worker_easy_evidence_guidance(evidence_assessment)
    easy_terms = "\n".join(f"- {item}" for item in WORKER_EASY_TERM_EXPLANATIONS)
    return f"""
당신은 신규 광산 근로자와 비전문 작업자에게 설명하는 성인 안전교육 보조자입니다.
초등학생에게 말하듯 지나치게 유치하게 쓰지 말고, 성인 신규 근로자가 바로 이해할 수 있는 쉬운 한국어를 사용하세요.

아래 [검색된 근거 문서]는 사용자의 질문에 대해 Vector DB / ChromaDB에서 검색된 RAG 근거입니다.
제공된 RAG 근거 밖의 사실을 추가하지 마세요. 확인되지 않은 수치, 시간, 농도, 법령 조문을 만들지 마세요.
공식 근거가 부족하면 모른다고 분명하게 말하고 담당 안전관리자, 관계기관 또는 전문가의 확인을 안내하세요.

[질문 유형]
{intent}
{intent_guidance}

{evidence_guardrail}

[공식 근거 검색 상태의 쉬운 안내]
{easy_evidence_guidance}

[근로자 쉬운 설명 모드 지침]
- 가장 중요한 현장 행동부터 설명하세요. 작업중지, 접근통제, 대피, 보고의 우선순위를 유지하세요.
- "왜 위험한가요?"를 반드시 포함하고, 끼임·질식·폭발·낙하·감전·충돌처럼 작업자가 이해할 수 있는 결과로 설명하세요.
- 한 문장은 짧게 쓰고 한 문장에 여러 조치를 길게 연결하지 마세요.
- 전문용어는 가능한 한 쉬운 말로 바꾸고, 꼭 써야 하면 바로 뒤에 쉬운 설명을 붙이세요.
- 핵심 본문은 250~600자를 우선하고, 복잡한 질문도 불필요하게 1,000자를 넘기지 마세요. 필수 안전조치는 길이 때문에 빼지 마세요.
- bullet은 기본 3~7개로 제한하고, 한 문단은 2~4문장을 넘기지 마세요.
- 이해를 돕는 가상 예시는 1개만 우선하고 필요한 경우 최대 2개만 사용하세요. 실제 공식 사고사례처럼 표현하지 마세요.
- 법령명, 조문 번호와 행정 표현을 반복하지 마세요. 긴 법령·판례·처벌 설명은 본문에 쓰지 마세요.
- 근로자를 겁주거나 처벌을 강조하지 마세요. 중대재해처벌법 위반을 확정적으로 단정하지 마세요.
- 마크다운 표를 사용하지 말고 너무 긴 서론과 결론을 쓰지 마세요.

[전문용어를 쉬운 말로 설명하는 방향]
{easy_terms}

[공식 사고사례 사용 제한]
- 아래 자료는 사람 검증 완료 또는 자동 품질검사 통과 전용 collection에서 검색한 공식 사고사례이지만 법령·지침은 아닙니다.
- 각 사례의 검증 상태와 직접 관련·유사 위험 구분을 유지하고, 유사 위험 사례를 동일 작업 사고처럼 표현하지 마세요.
- 사고 상황과 예방사항을 이해하는 참고자료로만 사용하고, 법령 위반 여부나 처벌 여부를 확정하지 마세요.
- 공식 법령 근거는 mine_safety_docs의 문서명과 chunk_id입니다.
- 사고사례 전체를 옮기지 말고 필요한 사고 개요와 예방사항만 짧게 참고하세요.

[공식 사고사례]
{official_case_context}

[사례 기반 주의 포인트 사용 제한]
- 아래 [사례 기반 주의 포인트]는 공식 법령 근거가 아니라 유사 위험을 이해하기 위한 참고 예시입니다.
- 공식 판단 근거는 검색된 RAG 근거 문서와 chunk_id이며, 사례 내용을 공식 사고사례나 법령 근거처럼 출력하지 마세요.
- 출력 예시는 사례를 그대로 옮기지 말고 "이해를 돕는 가상 예시"로 일반화하세요.

[사례 기반 주의 포인트]
{reference_case_context}

[뉴스 검색 참고자료 사용 제한]
- 아래 [실시간 사례 검색 참고]는 네이버 뉴스 검색 기반 참고자료이며 공식 법령 판단 근거가 아닙니다.
- 뉴스는 공식 근거가 아니므로 사용자 답변의 법적 판단이나 가상 예시에 사용하지 마세요.

[실시간 사례 검색 참고]
{live_news_context}

[사용자 질문]
{question}

[검색된 근거 문서]
{context}

[출력 제목]
### 근로자 쉬운 설명

[가능한 한 아래 구조로 짧게 출력]
### 지금 바로 해야 할 일
- 가장 중요한 행동 2~4개

### 왜 위험한가요?
- 위험한 이유를 쉬운 문장 2~4개

### 쉬운 현장 예시
- "이해를 돕는 가상 예시"라고 표시한 예시 1개, 필요한 경우 최대 2개

### 작업을 다시 시작하기 전에
- 위험요인 제거, 설비·환경 재점검, 관리자 확인과 작업 재개 승인 확인
""".strip()


def build_hybrid_prompt(
    question: str,
    results: list[dict[str, Any]],
    stable_draft: str,
    intent: str | None = None,
    reference_cases: list[dict[str, Any]] | None = None,
    live_news_cases: list[dict[str, Any]] | None = None,
    evidence_assessment: dict[str, Any] | None = None,
    official_cases: list[dict[str, Any]] | None = None,
) -> str:
    context = build_context(results)
    intent = intent or detect_question_intent(question)
    intent_guidance = gemini_intent_guidance(intent)
    reference_case_context = format_reference_cases_for_prompt(intent, reference_cases or [])
    live_news_context = format_live_news_cases_for_prompt(live_news_cases or [])
    official_case_context = format_official_cases_for_prompt(official_cases or [])
    evidence_guardrail = build_evidence_guardrail_prompt_guidance(evidence_assessment)
    return f"""
당신은 광산 안전관리자를 돕는 MineSafe AI의 하이브리드 답변 보완 역할입니다.

아래 [안정형 조치 초안]은 외부 LLM 없이 RAG 검색 근거로 만든 공식 문서 기반 체크리스트형 답변입니다.
아래 [검색된 근거 문서]는 같은 질문에 대해 Vector DB / ChromaDB에서 검색된 근거입니다.

[질문 유형]
{intent}
{intent_guidance}

{evidence_guardrail}

[하이브리드 모드 답변 지침]
- 대상은 현장관리자와 안전관리자입니다. 근로자 쉬운 설명 모드보다 더 전문적이고 상세하게 작성하세요.
- 안정형 조치 초안의 핵심 안전조치 구조는 유지하세요.
- 각 조치마다 "이유"와 "현장 적용"을 자연어로 보완하세요.
- 점검·통제·보고·기록관리와 법적·관리적 맥락을 함께 설명하세요.
- 근로자 쉬운 설명 모드처럼 단순한 예시 중심으로 축약하지 말고, 항목 구조를 분명히 유지하세요.
- 근거 없는 법령 조항 번호·수치·출처를 만들지 마세요.
- 검색 근거만으로 단정하기 어려운 내용은 "검색된 근거만으로는 단정하기 어렵습니다"라고 표시하세요.
- 법령 해석이나 실제 작업재개 판단은 안전관리자, 관계기관, 전문가 확인이 필요하다고 안내하세요.

[공식 사고사례 사용 제한]
- 아래 자료는 사람 검증 완료 또는 자동 품질검사 통과 전용 collection에서 검색한 공식 사고사례이며 법령·지침 근거가 아닙니다.
- 각 사례의 검증 상태와 직접 관련·유사 위험 구분을 유지하고, 유사 위험 사례를 동일 작업 사고처럼 표현하지 마세요.
- 사고 상황과 예방사항 설명에만 짧게 사용하고, 법령 위반 여부·처벌·작업 재개 가능 여부를 확정하지 마세요.
- 공식 법령 근거는 mine_safety_docs의 문서명과 chunk_id입니다.

[공식 사고사례]
{official_case_context}

[사례 기반 주의 포인트 사용 제한]
- 아래 [사례 기반 주의 포인트]는 공식 법령 근거가 아니라 유사 위험을 이해하기 위한 참고 예시입니다.
- 안정형 핵심 조치 구조는 RAG 근거와 초안을 기준으로 유지하고, 주의 포인트는 이유 설명 보완에만 사용하세요.
- 중대재해처벌법 위반 여부, 처벌 여부, 법령 조항을 주의 포인트만으로 단정하지 마세요.

[사례 기반 주의 포인트]
{reference_case_context}

[뉴스 검색 참고자료 사용 제한]
- 아래 [실시간 사례 검색 참고]는 네이버 뉴스 검색 기반 참고자료이며 공식 법령 판단 근거가 아닙니다.
- 안정형 핵심 조치 구조는 RAG 근거와 초안을 기준으로 유지하고, 뉴스는 이유 설명 보완에만 사용하세요.
- 뉴스 제목과 요약에 없는 내용을 만들지 말고, 뉴스만으로 법령 위반 여부나 처벌 여부를 단정하지 마세요.

[실시간 사례 검색 참고]
{live_news_context}

[사용자 질문]
{question}

[안정형 조치 초안]
{stable_draft}

[검색된 근거 문서]
{context}

[출력 제목]
### 하이브리드 답변

[반드시 아래 구조로 출력]
## 핵심 판단

## 조치별 설명
1. 작업중지
   - 이유:
   - 현장 적용:
2. 대피 또는 출입통제
   - 이유:
   - 현장 적용:
3. 점검 및 재측정
   - 이유:
   - 현장 적용:
4. 기록 및 증빙자료
   - 이유:
   - 현장 적용:

## 관련 근거 요약
""".strip()

def make_preview(text: str, limit: int = 320) -> str:
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "..."


def make_answer_summary(answer: str, limit: int = 300) -> str:
    summary = clean_text(answer)
    if len(summary) <= limit:
        return summary
    return summary[:limit].rstrip() + "..."


def make_evidence_summary(results: list[dict[str, Any]]) -> str:
    evidence_items = []
    for r in results[:5]:
        evidence_items.append(
            f"근거 {r['rank']}: {r['source']} | chunk_id {r.get('chunk_id', '정보 없음')} | 거리값 {format_distance(r['distance'])}"
        )
    return " | ".join(evidence_items)


def calculate_judgment(total_score: int) -> str:
    if total_score >= 90:
        return "매우 우수"
    if total_score >= 80:
        return "우수"
    if total_score >= 70:
        return "보통"
    if total_score >= 60:
        return "보완 필요"
    return "미흡"


def classify_safety_situation(question: str) -> str:
    intent = detect_question_intent(question)
    if intent == OUT_OF_SCOPE_INTENT:
        return OUT_OF_SCOPE_INTENT
    mapped_intent = map_intent_to_situation_type(intent, question)
    if mapped_intent != "일반 광산 안전":
        return mapped_intent

    question_type = classify_question_type(question)
    if question_type != "일반":
        return question_type

    text = question.lower()
    keyword_groups = [
        (ELECTRICAL_QUESTION_TYPE, ELECTRICAL_KEYWORDS),
        ("낙반/붕락", ["낙반", "붕락", "천반", "균열", "암석", "부석", "지보"]),
        (
            BLASTING_QUESTION_TYPE,
            ["발파", "불발", "화약", "폭약", "장약", "점화", "굴진"],
        ),
        ("환기/유해가스", ["유해가스", "메탄", "산소", "환기", "질식"]),
        ("화재/폭발", ["화재", "폭발", "연기"]),
        ("중대재해처벌법", ["사망", "중대재해", "경영책임자", "사업주"]),
        ("장비/운반", ["굴착기", "덤프트럭", "장비", "정비", "끼임", "운반"]),
    ]
    for situation_type, keywords in keyword_groups:
        if any(keyword in text for keyword in keywords):
            return situation_type
    return "일반 광산 안전"


def build_immediate_judgment(situation_type: str) -> str:
    judgments = {
        PPE_DUST_QUESTION_TYPE: "분진 발생 작업에는 적절한 저감조치와 보호구 확인 없이 작업자를 투입해서는 안 됩니다. 살수·집진·환기 상태, 작업환경측정 결과, 방진마스크·호흡보호구와 안전모 착용 여부를 먼저 확인해야 합니다.",
        RISK_ASSESSMENT_QUESTION_TYPE: "작업 방식, 장비 또는 작업 구간이 달라졌다면 기존 위험성평가를 그대로 적용하지 말고 변경된 유해·위험요인을 다시 평가해야 합니다. 추가 감소대책을 이행하고 작업자에게 공유한 뒤 작업을 시작해야 합니다.",
        ACCIDENT_RESPONSE_QUESTION_TYPE: "사고가 발생하면 작업중지와 부상자 구조·응급조치를 먼저 실시하고 2차 사고를 방지해야 합니다. 현장을 통제·보존하면서 책임자와 필요한 관계기관에 보고하고 원인조사와 재발방지 절차로 이어가야 합니다.",
        DAILY_INSPECTION_QUESTION_TYPE: "일상 안전점검은 작업장 구조, 환기·가스·분진, 설비, 보호구, 통로와 대피로를 빠짐없이 확인하고 이상 사항을 즉시 조치한 뒤 안전일지에 기록하는 절차로 운영해야 합니다.",
        PASSAGE_SAFETY_QUESTION_TYPE: "통로·조명·표지·대피로에 이상이 있으면 정상 통행이나 작업을 허용하지 말고 장애물과 위험요인을 먼저 제거해야 합니다. 주요 통행장소와 운반갱도를 순회 점검하고 결과를 기록·보고해야 합니다.",
        TBM_QUESTION_TYPE: "작업 전 안전회의에서는 당일 작업과 핵심 위험요인, 역할·신호·대피 절차를 작업자 전원에게 공유하고 이해 여부를 확인해야 합니다. 회의와 개선조치 이력을 남긴 후 작업을 시작해야 합니다.",
        COMPLEX_RISK_QUESTION_TYPE: "발파 후 환기 미확보 상태에서 운반차량을 투입해서는 안 됩니다. 출입통제 후 불발·공기질·환기·갱도 상태를 순서대로 확인하고 차량과 보행자 위험을 분리한 뒤 책임자 승인에 따라 단계적으로 재개해야 합니다.",
        ELECTRICAL_QUESTION_TYPE: "습기나 누전 우려가 있는 갱내 전기설비는 바로 점검을 시작하지 말고 먼저 전원을 차단해야 합니다. 임의 투입 방지, 접근통제, 접지·절연·누전차단기·케이블 상태를 확인하고 책임자가 안전을 확인한 뒤 작업을 재개해야 합니다.",
        "낙반/붕락": "낙반/붕락 위험이 의심되므로 작업을 계속하지 말고 즉시 작업중지 여부를 검토해야 합니다. 위험 구역 출입통제, 근로자 대피, 천반·부석·지보공 상태 확인 후 안전관리자 또는 책임자가 작업 재개 여부를 판단해야 합니다.",
        BLASTING_QUESTION_TYPE: "불발이 의심되면 작업자의 임의 접근을 금지하고 위험구역 출입통제와 대피를 유지해야 합니다. 발파 책임자 또는 안전관리자가 불발 여부와 현장 안전을 확인하기 전에는 굴진·정리 작업과 재출입을 허용하지 않아야 합니다.",
        "환기/유해가스": "유해가스, 산소부족, 환기 이상이 의심되면 작업을 계속하지 말고 가스 측정, 환기, 대피를 우선해야 합니다.",
        "화재/폭발": "화재, 폭발, 연기가 의심되면 작업중지, 대피, 신고, 출입통제를 우선하고, 안전이 확보되는 범위에서만 소화기 등 초기 대응을 검토해야 합니다.",
        "장비/운반": "장비 운전·정비 또는 운반 위험이 의심되면 전원 차단, 에너지 격리, 임의가동 방지, 접근 통제를 우선해야 합니다.",
        "중대재해처벌법": "사망 또는 중대재해 가능성이 있는 상황에서는 작업중지, 인명 구조와 보고, 원인조사, 재발방지대책, 안전보건관리체계 이행 여부 점검이 필요합니다.",
    }
    return judgments.get(
        situation_type,
        "검색된 근거 문서를 기준으로 위험요인을 먼저 확인하고, 급박한 위험이 있으면 작업중지와 대피를 우선 검토해야 합니다.",
    )


def build_priority_actions(situation_type: str) -> list[str]:
    if situation_type == PPE_DUST_QUESTION_TYPE:
        return [
            "작업 전에 굴진·파쇄 등 분진이 발생하는 작업인지 확인합니다.",
            "살수, 집진, 환기 등 분진 저감설비와 조치가 정상인지 확인합니다.",
            "방진마스크 또는 호흡보호구가 적절히 지급되고 올바르게 착용되었는지 확인합니다.",
            "안전모 등 해당 작업에 필요한 필수 보호구 착용 상태를 확인합니다.",
            "작업환경측정 결과와 유해인자 노출기준 초과 여부를 확인합니다.",
            "필수 보호구를 착용하지 않은 작업자는 작업에 투입하지 않습니다.",
            "작업 후 쌓인 분진을 제거·청소하고 점검과 조치 결과를 기록·보고합니다.",
        ]
    if situation_type == RISK_ASSESSMENT_QUESTION_TYPE:
        return [
            "기존 위험성평가를 변경된 작업에 그대로 사용할 수 있는지 먼저 검토합니다.",
            "작업 방식 변경, 신규 장비 투입, 신규 굴진 구간 등 변경 요인을 확인합니다.",
            "변경 사항으로 새로 발생하거나 달라진 유해·위험요인을 파악합니다.",
            "파악한 위험요인의 위험성 수준을 결정합니다.",
            "기존 위험 감소대책이 실제로 효과가 있는지 검토합니다.",
            "필요한 추가 감소대책과 담당자·이행 시점을 정합니다.",
            "감소대책을 이행하고 효과를 확인한 뒤 작업을 시작합니다.",
            "평가 결과를 근로자에게 공유하고 작업 전 안전회의를 실시합니다.",
            "평가와 조치 결과를 기록하고 변경 또는 문제 발생 시 재평가합니다.",
        ]
    if situation_type == ACCIDENT_RESPONSE_QUESTION_TYPE:
        return [
            "즉시 작업을 중지하고 사고 구역의 추가 작업을 막습니다.",
            "부상자를 안전하게 구조하고 가능한 범위에서 응급조치를 실시합니다.",
            "119 또는 현장 구급 연락체계를 가동합니다.",
            "2차 사고를 막기 위해 위험원을 통제하고 관계자 외 접근을 제한합니다.",
            "구조와 추가 위험 제거에 필요한 경우를 제외하고 사고 현장을 보존합니다.",
            "안전관리자, 현장 책임자, 사업주 또는 경영책임자에게 보고합니다.",
            "사고의 성격에 따라 필요한 관계기관 보고 절차를 확인합니다.",
            "사고 경위와 직접·간접 원인을 조사합니다.",
            "조사 결과를 바탕으로 재발방지대책을 수립하고 이행합니다.",
            "안전보건관리체계와 기존 안전조치가 실제로 이행되었는지 점검합니다.",
            "보고, 구조, 통제, 조사와 개선조치 내용을 기록합니다.",
        ]
    if situation_type == DAILY_INSPECTION_QUESTION_TYPE:
        return [
            "작업장 천장, 측벽과 작업면의 이상 여부를 점검합니다.",
            "부석, 낙석 흔적과 낙반·붕괴 위험을 확인합니다.",
            "지보공의 변형, 손상과 보강 필요성을 확인합니다.",
            "환기 상태와 유해가스·분진 상태를 확인합니다.",
            "전기, 기계와 운반 설비의 이상 여부를 점검합니다.",
            "작업자에게 필요한 보호구가 지급·착용되었는지 확인합니다.",
            "통로, 조명, 안내·위험표지와 대피로 상태를 확인합니다.",
            "안전교육과 작업 전 안전회의 실시 여부를 확인합니다.",
            "이상을 발견하면 즉시 보고하고 작업중지·통제·보수 등 필요한 조치를 합니다.",
            "점검과 조치 결과를 안전일지 또는 기록문서에 남깁니다.",
        ]
    if situation_type == PASSAGE_SAFETY_QUESTION_TYPE:
        return [
            "작업자와 장비가 안전하게 통행할 수 있는 폭과 통로 확보 여부를 확인합니다.",
            "통로와 작업구간의 조명 상태와 사각지대를 확인합니다.",
            "안내표지와 위험표시가 필요한 위치에 잘 보이도록 설치되었는지 확인합니다.",
            "대피로가 확보되어 있고 장애물이나 폐쇄 구간이 없는지 확인합니다.",
            "주요 통행장소와 운반갱도를 순회 점검합니다.",
            "낙반, 붕괴, 화재, 출수 등 통행을 위협하는 위험성을 확인합니다.",
            "발견한 이상과 조치 결과를 안전일지 등에 기록하고 책임자에게 보고합니다.",
        ]
    if situation_type == TBM_QUESTION_TYPE:
        return [
            "당일 작업 내용, 작업 위치와 작업 순서를 참여자에게 공유합니다.",
            "당일 작업의 주요 유해·위험요인과 통제대책을 공유합니다.",
            "낙반, 환기, 유해가스, 발파, 운반, 전기와 분진 위험을 작업별로 확인합니다.",
            "작업자별 필수 보호구 착용 상태를 확인합니다.",
            "작업자 역할, 작업지휘 체계와 신호체계를 확인합니다.",
            "대피로, 집결지와 비상연락망을 확인합니다.",
            "전일 지적사항과 개선조치 완료 여부를 확인합니다.",
            "근로자의 질문과 현장 위험에 관한 의견을 듣고 필요한 조치를 반영합니다.",
            "참석자 서명, 사진, 회의 내용과 조치사항 등 실시 기록을 남깁니다.",
        ]
    if situation_type == COMPLEX_RISK_QUESTION_TYPE:
        return [
            "작업을 중지하고 발파·환기·운반 작업구역의 출입을 통제합니다.",
            "발파 후 불발, 잔류화약류와 발파모선 상태를 확인합니다.",
            "후가스, 메탄, 산소, 일산화탄소 등 필요한 공기질 항목을 측정합니다.",
            "환기설비의 정상 작동과 충분한 환기 여부를 확인합니다.",
            "낙반, 부석과 갱도·작업면 상태를 확인합니다.",
            "운반차량 투입 전에 통행로, 조명, 신호수와 작업자 위치를 확인합니다.",
            "차량 운행구역과 보행자 동선을 분리하고 접근을 통제합니다.",
            "각 위험요인 제거를 확인한 후 책임자 승인에 따라 작업을 단계적으로 재개합니다.",
            "측정값, 확인자, 승인과 조치 결과를 기록하고 필요한 보고를 실시합니다.",
        ]
    if situation_type == ELECTRICAL_QUESTION_TYPE:
        return [
            "작업 전에 전원을 차단하고 무전압 상태인지 확인합니다.",
            "잠금·표지 등으로 전원이 임의로 다시 투입되지 않도록 조치합니다.",
            "작업구역 접근을 통제하고 전기작업에 적합한 보호구를 착용합니다.",
            "습기 또는 물기를 제거하고 접지 상태와 절연상태를 확인합니다.",
            "누전차단기 등 보호장치의 설치 상태와 작동 여부를 확인합니다.",
            "케이블·배선의 피복 손상, 접속부 이상, 물기 접촉 여부를 점검합니다.",
            "필요한 경우 방폭형 또는 갱내용 전기설비의 현장 적합성을 확인합니다.",
            "전기안전 담당자 또는 책임자의 확인 후에만 작업과 전원 투입을 재개합니다.",
        ]
    if situation_type == "낙반/붕락":
        return [
            "작업을 계속하지 말고 즉시 작업중지 여부를 검토합니다.",
            "낙반 가능 구역의 출입을 통제하고 근로자를 안전한 장소로 대피시킵니다.",
            "천반 균열, 부석 또는 낙석 위험, 지보공 변형·손상 상태를 우선 확인합니다.",
            "갱내수, 풍화, 추가 균열 등 위험을 키울 수 있는 요인을 함께 확인합니다.",
            "안전관리자 또는 책임자의 확인 후 작업 재개 여부를 판단합니다.",
        ]
    if situation_type == BLASTING_QUESTION_TYPE:
        return [
            "불발 의심 장소에 작업자가 임의로 접근하지 못하게 합니다.",
            "위험구역을 설정해 출입을 통제하고 작업자 대피 상태를 유지합니다.",
            "발파 책임자 또는 안전관리자에게 즉시 알리고 현장 확인 절차를 따릅니다.",
            "불발 여부가 확인되기 전에는 굴진·정리 작업을 재개하지 않습니다.",
            "발파 후 재출입은 후가스, 낙반·부석, 불발화약류 등 안전 확인 이후에만 허용합니다.",
            "확인 결과와 조치 내용을 필요한 보고 체계에 따라 보고하고 기록합니다.",
        ]
    if situation_type == "환기/유해가스":
        return [
            "작업을 중지하고 작업자를 대피시킵니다.",
            "메탄, 산소, 일산화탄소 등 필요한 항목을 측정합니다.",
            "환기설비 상태와 재측정 결과를 확인한 뒤 재개 여부를 판단합니다.",
        ]
    if situation_type == "화재/폭발":
        return [
            "작업중지와 대피를 우선합니다.",
            "비상 연락, 신고, 인원 확인을 실시합니다.",
            "초기 대응은 안전이 확보되는 범위에서만 수행합니다.",
        ]
    if situation_type == "장비/운반":
        return [
            "전원 차단과 에너지 격리를 먼저 확인합니다.",
            "임의가동 방지 조치와 접근 통제를 실시합니다.",
            "정비 또는 운반 작업 재개 전 장비 상태와 작업자 위치를 확인합니다.",
        ]
    if situation_type == "중대재해처벌법":
        return [
            "작업중지, 인명 구조, 현장 보존, 보고 절차를 우선합니다.",
            "사고 원인 조사와 재발방지대책 수립 필요성을 확인합니다.",
            "경영책임자의 안전보건 확보의무와 안전보건관리체계 이행 여부를 점검합니다.",
        ]
    return [
        "질문과 관련된 작업 장소, 설비, 작업자 상태와 주변 위험요인을 확인합니다.",
        "급박한 위험이 있으면 작업중지, 출입통제, 대피를 먼저 검토합니다.",
        "작업 전 확인사항과 조치 결과를 기록하고 책임자 확인 후 작업 재개 여부를 판단합니다.",
    ]


def build_check_items(situation_type: str) -> list[str]:
    if situation_type == PPE_DUST_QUESTION_TYPE:
        return [
            "분진 발생원과 살수·집진·환기설비 작동 상태",
            "작업환경측정 결과와 노출기준 초과 여부",
            "방진마스크·호흡보호구의 종류, 지급, 밀착과 착용 상태",
            "안전모 등 필수 보호구 착용과 미착용자 작업 제한",
            "작업 후 청소, 보호구 관리와 조치 기록",
        ]
    if situation_type == RISK_ASSESSMENT_QUESTION_TYPE:
        return [
            "작업 방식, 장비, 장소와 작업자 구성의 변경 여부",
            "새로운 유해·위험요인과 위험성 수준",
            "기존 대책의 효과와 추가 감소대책 필요성",
            "감소대책 이행·확인과 작업 시작 승인",
            "근로자 공유, 작업 전 회의, 기록과 재평가",
        ]
    if situation_type == ACCIDENT_RESPONSE_QUESTION_TYPE:
        return [
            "작업중지, 구조·응급조치와 119 연락 상태",
            "2차 사고 방지, 접근통제와 현장 보존 상태",
            "내부 책임자와 필요한 관계기관 보고 여부",
            "사고 원인조사와 재발방지대책 수립·이행 여부",
            "안전보건관리체계 점검과 전체 조치 기록",
        ]
    if situation_type == DAILY_INSPECTION_QUESTION_TYPE:
        return [
            "천장·측벽·작업면, 부석·낙반과 지보공 상태",
            "환기·유해가스·분진과 전기·기계·운반 설비 상태",
            "보호구, 통로·조명·표지와 대피로 상태",
            "안전교육·작업 전 회의와 이상 사항 조치 여부",
            "안전일지 또는 기록문서 작성 여부",
        ]
    if situation_type == PASSAGE_SAFETY_QUESTION_TYPE:
        return [
            "통로 폭, 장애물과 안전한 통행 가능 여부",
            "조명, 안내표지와 위험표시 상태",
            "대피로 개방 상태와 비상시 이동 가능 여부",
            "주요 통행장소·운반갱도의 낙반·화재·출수 위험",
            "순회 점검 결과의 기록과 보고 여부",
        ]
    if situation_type == TBM_QUESTION_TYPE:
        return [
            "당일 작업과 작업별 유해·위험요인 공유 여부",
            "보호구, 역할·신호체계와 대피·비상연락 확인",
            "전일 지적사항과 개선조치 완료 여부",
            "근로자 질문·의견 청취와 반영 여부",
            "참석자와 회의·조치 내용 기록 여부",
        ]
    if situation_type == COMPLEX_RISK_QUESTION_TYPE:
        return [
            "불발·잔류화약류·발파모선 확인 결과",
            "후가스·메탄·산소·일산화탄소 측정과 환기 상태",
            "낙반·부석·갱도 상태와 통행로·조명 상태",
            "차량·보행자 동선 분리와 신호수·작업자 위치",
            "책임자 승인, 단계적 재개와 조치 기록·보고",
        ]
    if situation_type == ELECTRICAL_QUESTION_TYPE:
        return [
            "전원 차단, 무전압 확인, 잠금·표지 또는 임의 투입 방지 상태",
            "접지 연결 상태와 절연 손상·열화 여부",
            "누전차단기 등 보호장치의 설치 및 시험 결과",
            "습기·물기 제거와 케이블·배선·접속부 손상 여부",
            "방폭형 또는 갱내용 전기설비 적합성 확인 필요 여부",
            "접근통제, 보호구 착용, 책임자 작업 재개 승인 여부",
        ]
    if situation_type == "낙반/붕락":
        return [
            "천반 균열 상태와 균열 확대 여부",
            "부석, 낙석 흔적, 암석 박리 가능성",
            "지보공 설치 상태, 변형, 손상, 보강 필요성",
            "갱내수, 풍화, 진동 등 추가 위험요인",
            "출입통제, 대피, 작업 재개 승인 기록",
        ]
    if situation_type == BLASTING_QUESTION_TYPE:
        return [
            "불발화약류, 잔류약, 발파모선 등 불발 의심 상태",
            "위험구역 출입통제, 감시, 작업자 대피 유지 여부",
            "발파 책임자 또는 안전관리자의 현장 확인 여부",
            "후가스, 메탄가스, 낙반, 부석 등 발파 후 추가 위험",
            "굴진·정리 작업 및 재출입 승인, 보고·기록 여부",
        ]
    return [
        "검색된 근거 문서가 현재 작업 유형과 직접 관련되는지",
        "작업중지, 대피, 보고, 보호구, 장비 점검 기준이 명시되어 있는지",
        "통로, 조명, 환기, 설비, 작업자 상태 등 현장 조건상 추가 확인이 필요한 위험요인이 있는지",
        "점검표, 사진, 교육기록, 작업허가서 등 남겨야 할 증빙자료가 무엇인지",
        "최신 법령·지침 또는 내부 안전관리 기준 확인이 필요한지",
    ]


def infer_risk_level_for_kras(
    question: str,
    situation_type: str,
) -> tuple[str, str, str]:
    text = question.lower()
    if any(keyword in text for keyword in ["사망", "매몰", "폭발", "화재"]):
        return "높음", "매우 높음", "매우 높음"
    if situation_type in {
        "낙반/붕락",
        BLASTING_QUESTION_TYPE,
        "환기/유해가스",
        "화재/폭발",
        ACCIDENT_RESPONSE_QUESTION_TYPE,
        COMPLEX_RISK_QUESTION_TYPE,
        "중대재해처벌법",
    }:
        return "높음", "매우 높음", "매우 높음"
    if situation_type in {
        ELECTRICAL_QUESTION_TYPE,
        "장비/운반",
        PPE_DUST_QUESTION_TYPE,
        PASSAGE_SAFETY_QUESTION_TYPE,
    }:
        return "중간", "높음", "높음"
    return "중간", "중간", "중간"


def format_kras_table(headers: list[str], rows: list[list[Any]]) -> str:
    def format_cell(value: Any) -> str:
        return (
            str(value or "")
            .replace("|", "\\|")
            .replace("\r\n", " / ")
            .replace("\n", " / ")
        )

    header_line = "| " + " | ".join(format_cell(header) for header in headers) + " |"
    separator_line = "| " + " | ".join("---" for _ in headers) + " |"
    body_lines = [
        "| " + " | ".join(format_cell(value) for value in row) + " |"
        for row in rows
    ]
    return "\n".join([header_line, separator_line, *body_lines])


def build_kras_risk_assessment_section(
    question: str,
    results: list[dict[str, Any]],
    situation_type: str | None = None,
) -> str:
    situation_type = situation_type or classify_safety_situation(question)
    text = question.lower()

    if any(keyword in text for keyword in ["낙반", "붕락", "천반", "측벽", "부석", "지보"]):
        profile = "낙반/붕락"
    elif any(keyword in text for keyword in ["발파", "불발", "화약", "폭약", "장약"]):
        profile = BLASTING_QUESTION_TYPE
    elif any(keyword in text for keyword in ["메탄", "유해가스", "산소", "환기", "질식"]):
        profile = "환기/유해가스"
    elif any(keyword in text for keyword in ["전기", "누전", "감전", "접지", "절연"]):
        profile = ELECTRICAL_QUESTION_TYPE
    elif any(keyword in text for keyword in ["분진", "먼지", "방진마스크", "호흡보호구", "보호구"]):
        profile = PPE_DUST_QUESTION_TYPE
    elif any(keyword in text for keyword in ["장비", "정비", "운반", "덤프트럭", "끼임"]):
        profile = "장비/운반"
    else:
        profile = situation_type

    profiles: dict[str, dict[str, Any]] = {
        "낙반/붕락": {
            "work": "갱내 채굴·굴진 또는 점검 작업 중 천반·측벽과 지보 상태를 확인하는 작업",
            "hazards": "천반·측벽 균열, 부석·낙석, 지보재 변형, 위험구역 통제 미흡",
            "outcome": "낙반으로 작업자 매몰·중상·사망 및 구조자 2차 피해 가능",
            "measures": {
                "제거": "작업중지, 위험구역 출입금지, 안전 확보 후 부석·낙석 위험 제거",
                "대체": "인력 접근 대신 원격·무인 점검 또는 장비 사용 가능 여부 검토",
                "공학적 대책": "지보 보강, 낙석 방호, 균열·변형 감시와 갱도 안정성 확인",
                "관리적 대책": "대피·통제, 책임자 확인, 사고조사·재발방지대책, 작업재개 승인",
                "보호구/PPE": "안전모 등 작업별 필수 보호구 착용. 보호구만으로 낙반 위험을 대체하지 않음",
            },
            "residual": "지보 보강과 위험 제거 후에도 추가 낙반 가능성을 재평가하고 책임자 승인 전 작업재개 금지",
            "records": "천반·측벽·지보 점검, 작업중지·대피·통제, 사고 경위와 원인, 보강·재발방지 및 작업재개 승인 기록",
            "items": [
                ["1", "천반·측벽 균열 또는 지보재 변형", "낙반 발생, 작업자 매몰·중상·사망", "매우 높음", "작업중지, 출입통제, 지보 보강, 안정성 확인 후 작업재개"],
                ["2", "사고 후 접근통제 미흡", "추가 낙반으로 구조자 2차 피해", "높음", "위험구역 통제, 구조 인원 제한, 대피로 확보"],
                ["3", "원인조사·점검 기록 미흡", "재발방지대책 누락 및 이행 확인 곤란", "높음", "사고 경위, 직접·간접 원인, 조치와 승인 결과 기록"],
            ],
        },
        BLASTING_QUESTION_TYPE: {
            "work": "발파 후 작업면 확인, 불발 여부 점검 및 굴진·정리 작업 재개 준비",
            "hazards": "불발화약류, 잔류화약류, 발파모선, 후가스, 임의 접근",
            "outcome": "지연 폭발·폭발물 접촉, 유해가스 노출, 낙반으로 중대 부상 가능",
            "measures": {
                "제거": "불발 의심 구역 작업중지와 출입금지, 확인 전 굴진·정리 작업 금지",
                "대체": "인력 접근 전 원격 확인 또는 비접촉 점검 가능 여부 검토",
                "공학적 대책": "환기 유지, 위험구역 차단, 발파설비와 모선 상태 확인",
                "관리적 대책": "발파 책임자 확인, 대피 유지, 재출입 승인, 보고·기록",
                "보호구/PPE": "발파 후 점검 작업에 필요한 보호구 착용. 불발 통제를 보호구로 대체하지 않음",
            },
            "residual": "불발·후가스·낙반 위험 확인과 책임자 승인 후 잔여위험을 재평가",
            "records": "발파 시각, 대피·통제, 불발 확인, 환기·재출입 승인 및 후속조치 기록",
            "items": [
                ["1", "불발 또는 잔류화약류", "지연 폭발로 작업자 중상·사망", "매우 높음", "접근금지, 책임자 확인, 안전 확인 전 작업재개 금지"],
                ["2", "발파 후 후가스·환기 불충분", "유해가스 노출·질식", "높음", "환기 유지, 공기질 확인, 안전 확인 후 재출입"],
                ["3", "위험구역 통제·기록 미흡", "제3자 접근 및 동일 사고 반복", "높음", "출입통제, 승인자 지정, 확인·보고 내용 기록"],
            ],
        },
        "환기/유해가스": {
            "work": "갱내 작업 전·중 환기 상태와 유해가스·산소 상태를 확인하는 작업",
            "hazards": "메탄·일산화탄소 등 유해가스, 산소부족, 환기설비 이상",
            "outcome": "질식·중독·폭발 및 대피 지연으로 중대 피해 가능",
            "measures": {
                "제거": "작업중지와 오염 구역 출입금지, 작업자 대피",
                "대체": "인력 투입 전 원격 측정·감시 방식 적용 가능 여부 검토",
                "공학적 대책": "환기설비 가동·보수, 가스 측정·경보와 공기 흐름 확보",
                "관리적 대책": "측정 결과 확인, 재출입 승인, 비상연락·대피 절차와 기록",
                "보호구/PPE": "측정·구조 작업에 적합한 보호구를 현장 기준에 따라 선정·착용",
            },
            "residual": "환기와 재측정 후에도 측정 결과와 현장 조건을 재평가한 뒤 작업재개 결정",
            "records": "측정 항목·시각·위치·결과, 환기 조치, 대피·재출입 승인 기록",
            "items": [
                ["1", "유해가스 또는 산소부족", "질식·중독·의식상실", "매우 높음", "작업중지, 대피, 측정과 환기 후 재평가"],
                ["2", "환기설비 이상", "가스 체류·확산 및 폭발 위험 증가", "높음", "설비 점검·복구, 공기 흐름 확인"],
                ["3", "재출입 승인·기록 미흡", "위험 상태에서 조기 작업재개", "높음", "측정 결과 확인, 책임자 승인과 기록"],
            ],
        },
        ELECTRICAL_QUESTION_TYPE: {
            "work": "갱내 전기설비·케이블·배선의 점검·정비 작업",
            "hazards": "누전, 절연 손상, 접지 불량, 습기·물기, 임의 전원 투입",
            "outcome": "감전·화재·설비 손상 및 2차 사고 가능",
            "measures": {
                "제거": "작업 전 전원 차단과 무전압 확인, 습기·물기 제거",
                "대체": "무전압 점검 또는 원격 점검 방식 적용 가능 여부 검토",
                "공학적 대책": "접지·절연·누전차단기, 케이블·배선과 방폭 적합성 확인",
                "관리적 대책": "잠금·표지, 접근통제, 작업허가와 책임자 확인 후 재투입",
                "보호구/PPE": "전기작업에 적합한 절연 보호구 등 현장 지정 보호구 착용",
            },
            "residual": "보호장치 시험과 책임자 확인 후 잔여 감전·재투입 위험을 재평가",
            "records": "전원 차단·잠금, 무전압·접지·절연·차단기 확인과 재투입 승인 기록",
            "items": [
                ["1", "누전·절연 손상·습기", "감전·화재", "높음", "전원 차단, 물기 제거, 접지·절연 확인"],
                ["2", "전원 임의 재투입", "정비 작업자 감전·끼임", "높음", "잠금·표지, 접근통제, 재투입 승인"],
                ["3", "케이블·보호장치 이상", "누전 차단 실패와 설비 손상", "높음", "배선·차단기 점검과 이상 부품 조치"],
            ],
        },
        PPE_DUST_QUESTION_TYPE: {
            "work": "굴진·파쇄 등 분진 발생 작업과 보호구 지급·착용 점검",
            "hazards": "분진 비산, 환기·집진 미흡, 호흡보호구 부적합·미착용",
            "outcome": "분진 노출에 따른 건강장해 및 시야 저하로 인한 사고 가능",
            "measures": {
                "제거": "불필요한 분진 발생 작업·공정 제거와 퇴적 분진 청소",
                "대체": "저분진 공법·재료 또는 습식 작업 방식 적용 가능 여부 검토",
                "공학적 대책": "살수·집진·환기와 발생원 격리",
                "관리적 대책": "작업환경측정 확인, 노출시간 관리, 착용 교육·점검과 미착용자 작업 제한",
                "보호구/PPE": "작업과 분진 특성에 맞는 방진마스크·호흡보호구 및 안전모 지급·착용",
            },
            "residual": "저감조치 후 작업환경과 착용 상태를 다시 확인하고 노출 위험을 재평가",
            "records": "측정 결과, 저감설비 점검, 보호구 지급·착용·교체, 교육과 미착용 조치 기록",
            "items": [
                ["1", "분진 발생원 저감 미흡", "고농도 분진 노출", "높음", "살수·집진·환기와 발생원 격리"],
                ["2", "호흡보호구 부적합·미착용", "호흡기 노출 증가", "높음", "적정 보호구 지급, 밀착·착용 확인, 미착용자 투입 금지"],
                ["3", "측정·청소·기록 미흡", "노출 상태와 개선 효과 확인 곤란", "중간", "작업환경 확인, 퇴적 분진 청소와 조치 기록"],
            ],
        },
    }

    profile_data = profiles.get(
        profile,
        {
            "work": f"{situation_type} 관련 광산 현장 작업과 작업 전·중 안전 확인",
            "hazards": "작업구역의 유해·위험요인 미확인, 통제 미흡, 작업절차 미준수",
            "outcome": "작업자 부상, 설비 손상 또는 2차 사고 가능",
            "measures": {
                "제거": "위험 작업중지와 위험원 제거 가능 여부 확인",
                "대체": "위험이 낮은 작업방법·장비로 대체 가능 여부 검토",
                "공학적 대책": "방호·격리·환기·감시 등 작업 유형에 맞는 설비 대책",
                "관리적 대책": "작업절차, 출입통제, 점검, 교육, 책임자 승인과 기록",
                "보호구/PPE": "작업별 필수 보호구 선정·지급·착용 확인",
            },
            "residual": "감소대책 이행 후 현장 조건을 재확인하고 책임자가 잔여위험을 재평가",
            "records": "위험요인, 점검 결과, 감소대책, 담당자, 이행 시점과 작업재개 승인 기록",
            "items": [
                ["1", "유해·위험요인 확인 미흡", "작업자 부상 또는 설비 손상", "높음", "작업중지, 위험원 확인·제거 후 재개"],
                ["2", "출입·작업절차 통제 미흡", "제3자 노출 또는 2차 사고", "중간", "출입통제, 작업절차 공유와 책임자 확인"],
                ["3", "점검·조치 기록 미흡", "개선조치 이행 확인 곤란", "중간", "점검·조치·승인 결과 기록 및 보고"],
            ],
        },
    )

    likelihood, severity, risk_grade = infer_risk_level_for_kras(
        question,
        situation_type,
    )

    source_items: list[str] = []
    for result in results[:5]:
        source = str(result.get("source", "출처 정보 없음")).strip()
        chunk_id = str(result.get("chunk_id", "정보 없음")).strip()
        source_item = f"{source} (chunk_id: {chunk_id})"
        if source_item not in source_items:
            source_items.append(source_item)
    related_basis = " / ".join(source_items) if source_items else "검색 근거 문서 확인 필요"
    related_basis += (
        " / 조문 번호·법적 해석·처벌 수위는 검색 근거에서 명확히 확인되는 경우만 "
        "기입하며, 관계기관·전문가 확인이 필요합니다."
    )

    measures = profile_data["measures"]
    measure_text = " / ".join(
        f"**{level}:** {measures[level]}"
        for level in ["제거", "대체", "공학적 대책", "관리적 대책", "보호구/PPE"]
    )
    current_risk = (
        f"가능성 {likelihood} / 중대성 {severity} / 위험등급 {risk_grade}"
        " / 질문 상황을 기준으로 한 정성 초안이며 현장 기준에 따라 재평가해야 합니다."
    )

    detail_table = format_kras_table(
        ["항목", "기입 초안"],
        [
            ["세부 작업 내용", profile_data["work"]],
            ["잠재위험요인", profile_data["hazards"]],
            ["위험발생 상황 및 결과", profile_data["outcome"]],
            ["관련 근거 / 법적 기준", related_basis],
            ["현재 위험성", current_risk],
            ["위험성 감소대책", measure_text],
            ["조치 후 잔여위험성", profile_data["residual"]],
            ["기록·보고 사항", profile_data["records"]],
        ],
    )
    hierarchy_table = format_kras_table(
        ["우선순위", "감소대책 기입 초안"],
        [[level, measures[level]] for level in ["제거", "대체", "공학적 대책", "관리적 대책", "보호구/PPE"]],
    )
    summary_table = format_kras_table(
        ["번호", "잠재위험요인", "위험발생 상황 및 결과", "위험성", "감소대책"],
        profile_data["items"],
    )

    return "\n\n".join(
        [
            "### KRAS식 위험성평가 기록 초안",
            "> 검색 근거와 질문 상황을 바탕으로 작성한 기입 초안입니다. 실제 가능성·중대성·위험등급과 법적 판단은 현장 기준 및 관계기관·전문가 확인이 필요합니다.",
            detail_table,
            "### 위험성 감소대책 우선순위",
            hierarchy_table,
            "### KRAS 양식 기입용 요약표",
            summary_table,
        ]
    )


def normalize_answer_mode_label(mode_name: Any) -> str:
    """과거 모드 값을 삭제하지 않고 현재 사용자 표시 이름으로 변환합니다."""
    text = str(mode_name or "").strip()
    lowered = text.lower()
    if text in {GEMINI_MODE, "자연어 설명 모드", "Gemini 모드", WORKER_EASY_MODE_LABEL}:
        return WORKER_EASY_MODE_LABEL
    if lowered in {"natural", "gemini"}:
        return WORKER_EASY_MODE_LABEL
    return text


def answer_mode_option_label(answer_mode: str) -> str:
    """selectbox 내부 값은 유지하고 사용자에게 짧은 새 이름만 표시합니다."""
    return short_answer_mode(answer_mode)


def short_answer_mode(answer_mode: str) -> str:
    if answer_mode == STABLE_MODE:
        return "안정성 모드"
    if answer_mode == GEMINI_MODE:
        return WORKER_EASY_MODE_LABEL
    if answer_mode == HYBRID_MODE:
        return "하이브리드 모드"
    return normalize_answer_mode_label(answer_mode)


def answer_mode_description(mode_name: str) -> str:
    mode_name = normalize_answer_mode_label(mode_name)
    if mode_name == "안정성 모드":
        return STABLE_MODE_HELP
    if mode_name == WORKER_EASY_MODE_LABEL:
        return WORKER_EASY_MODE_HELP
    if mode_name == "하이브리드 모드":
        return HYBRID_MODE_HELP
    return "검색 근거를 바탕으로 안전 답변을 제공합니다."


def public_answer_mode_label(answer_status: dict[str, Any], mode_name: str) -> str:
    fallback_used = bool(
        answer_status.get(
            "used_fallback",
            answer_status.get("fallback_used", False),
        )
    )
    if fallback_used:
        return "안정성 모드"
    mode = str(answer_status.get("mode", "")).lower()
    prompt_kind = str(answer_status.get("prompt_kind", "")).lower()
    if mode == "stable":
        return "안정성 모드"
    if mode == "hybrid" or prompt_kind == "hybrid":
        return "하이브리드 모드"
    if prompt_kind == "gemini" or mode == "gemini":
        return WORKER_EASY_MODE_LABEL
    return normalize_answer_mode_label(mode_name)


def classify_gemini_status(answer_status: dict[str, Any]) -> str:
    if answer_status.get("gemini_status"):
        return str(answer_status["gemini_status"])

    structured_status = str(answer_status.get("status", "")).strip().lower()
    status_labels = {
        "success": "성공",
        "503": "503",
        "429": "429",
        "timeout": "timeout",
        "error": "기타 오류",
        "not_called": "호출 안 함",
    }
    if structured_status in status_labels:
        return status_labels[structured_status]

    if answer_status.get("mode") == "gemini":
        return "성공"

    reason = str(answer_status.get("reason", "")).lower()
    if not reason:
        return "호출 안 함"
    if "503" in reason or "unavailable" in reason or "high demand" in reason:
        return "503"
    if "429" in reason or "rate limit" in reason or "resource exhausted" in reason:
        return "429"
    if "504" in reason:
        return "504 timeout"
    if is_timeout_error(reason):
        return "timeout"
    if answer_status.get("mode") == "fallback":
        return "실패"
    return "기타 오류"


def format_gemini_status(status: str) -> str:
    if status == "503":
        return "503 UNAVAILABLE"
    if status == "429":
        return "429 RESOURCE_EXHAUSTED"
    return status


def classify_gemini_error(error_message: str) -> str:
    reason = str(error_message or "").lower()
    if "503" in reason or "unavailable" in reason or "high demand" in reason:
        return "503"
    if "429" in reason or "rate limit" in reason or "resource exhausted" in reason:
        return "429"
    if "504" in reason or is_timeout_error(reason):
        return "timeout"
    return "error"


def summarize_gemini_error(error_message: str, limit: int = 240) -> str:
    summary = clean_text(str(error_message or "알 수 없는 오류"))
    if len(summary) > limit:
        return summary[:limit].rstrip() + "..."
    return summary


def get_gemini_failure_message(status: str, error_message: str = "") -> str:
    if status == "503":
        return (
            "Gemini 서버가 일시적으로 혼잡하여 안정형 답변으로 전환했습니다. "
            "잠시 후 다시 시도하거나 다른 모델을 선택해 보세요."
        )
    if status == "429":
        return (
            "Gemini API 사용량 또는 호출 제한에 도달했습니다. "
            "잠시 후 다시 시도하거나 호출 횟수를 줄여 주세요."
        )
    if status in {"timeout", "504 timeout"}:
        return (
            "Gemini 응답 시간이 길어 제한 시간 내 완료되지 않았습니다. "
            "timeout을 늘리거나 출력 토큰 수를 줄여 주세요."
        )
    return (
        "Gemini 호출 중 오류가 발생하여 안정형 답변으로 전환했습니다. "
        f"오류 요약: {summarize_gemini_error(error_message)}"
    )


def read_evaluation_rows() -> list[dict[str, str]]:
    if not EVALUATION_PATH.exists():
        return []

    with open(EVALUATION_PATH, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f, delimiter="\t"))


def safe_score_value(value: Any) -> float:
    text = str(value or "").strip()
    if not text or text.startswith("="):
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


EVALUATION_NEW_SCORE_FIELDS = [
    "검색_적합성",
    "근거_기반성",
    "안전법령_판단정확성",
    "실무성",
]
EVALUATION_LEGACY_SCORE_FIELDS = [
    "검색정확도(0~5)",
    "답변정확도(0~5)",
    "근거성(0~5)",
    "환각억제(0~5)",
    "실무성(0~5)",
]


def normalize_evaluation_row(row: dict[str, str]) -> dict[str, str]:
    normalized = dict(row)
    for field in EVALUATION_NEW_SCORE_FIELDS:
        normalized.setdefault(field, "0")
    for field in ["총점", "판정", "메모", "검색된 주요 근거 문서", "Gemini 답변 요약", "답변 생성 방식", "Gemini 상태"]:
        normalized.setdefault(field, "")
    return normalized


def is_evaluation_completed(row: dict[str, str]) -> bool:
    normalized_row = normalize_evaluation_row(row)
    total_score = safe_score_value(normalized_row.get("총점"))
    has_memo = bool(str(normalized_row.get("메모", "")).strip())
    has_any_score = any(
        safe_score_value(normalized_row.get(field)) > 0 for field in EVALUATION_NEW_SCORE_FIELDS
    ) or any(
        safe_score_value(normalized_row.get(field)) > 0 for field in EVALUATION_LEGACY_SCORE_FIELDS
    )
    return total_score > 0 or has_memo or has_any_score


def has_saved_evaluation_result(row: dict[str, str] | None) -> bool:
    if not row:
        return False
    saved_text_fields = [
        "검색된 주요 근거 문서",
        "Gemini 답변 요약",
        "답변 생성 방식",
        "Gemini 상태",
        "메모",
    ]
    return is_evaluation_completed(row) or any(
        str(row.get(field, "")).strip()
        for field in saved_text_fields
    )


def get_evaluation_row(
    scenario_no: Any,
) -> tuple[dict[str, str] | None, str | None]:
    target_no = str(scenario_no or "").strip()
    if not target_no:
        return None, "질문 번호를 확인할 수 없습니다."
    if not EVALUATION_PATH.exists():
        return None, f"평가표 파일을 찾을 수 없습니다: {EVALUATION_PATH}"

    try:
        for row in read_evaluation_rows():
            if str(row.get("번호", "")).strip() == target_no:
                return row, None
        return None, f"평가표에서 Q{int(target_no):02d} 행을 찾을 수 없습니다."
    except Exception as e:
        return None, f"이전 평가 결과를 읽는 중 오류가 발생했습니다: {e}"


def classify_evidence_profile(evidence_summary: str) -> str:
    text = clean_text(evidence_summary).lower()
    preferred_count = sum(
        marker.lower() in text
        for marker in BLASTING_PREFERRED_SOURCE_MARKERS
    )
    general_count = sum(
        marker.lower() in text
        for marker in BLASTING_GENERAL_SOURCE_MARKERS
    )

    if preferred_count and preferred_count >= general_count:
        return "광산 특화 문서 우선 검색"
    if preferred_count:
        return "광산 특화 문서 포함"
    if general_count:
        return "일반 안전관리 문서 위주"
    if text:
        return "기타 근거 문서 중심"
    return "저장된 검색 결과 없음"


def format_comparison_score(value: Any) -> str:
    score = safe_score_value(value)
    if score.is_integer():
        return f"{int(score)}점"
    return f"{score:.1f}점"


def safe_auto_eval_value(row: dict[str, Any], *names: str, default: Any = "") -> Any:
    for name in names:
        if name in row:
            value = row.get(name)
            if value is None:
                continue
            if isinstance(value, str):
                value = value.strip()
                if value == "":
                    continue
            return value
    return default


def parse_auto_eval_score(row: dict[str, Any], *names: str, default: int = 0) -> int:
    raw_value = safe_auto_eval_value(row, *names, default=default)
    try:
        return int(float(str(raw_value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return int(default)


def normalize_auto_eval_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "question_id": normalize_question_id_for_display(
            safe_auto_eval_value(row, "question_id", "번호", default="")
        ),
        "category": str(safe_auto_eval_value(row, "category", "분류", default="")),
        "question": str(safe_auto_eval_value(row, "question", "질문", default="")),
        "검색_적합성": parse_auto_eval_score(row, "검색_적합성", default=0),
        "근거_기반성": parse_auto_eval_score(row, "근거_기반성", default=0),
        "안전법령_판단정확성": parse_auto_eval_score(row, "안전법령_판단정확성", default=0),
        "실무성": parse_auto_eval_score(row, "실무성", default=0),
        "총점": parse_auto_eval_score(row, "총점", default=0),
        "판정": str(safe_auto_eval_value(row, "판정", default="")).strip() or "미흡",
        "검토필요": str(safe_auto_eval_value(row, "검토필요", default="")).strip().upper(),
    }


def load_auto_eval_summary() -> tuple[dict[str, Any] | None, str | None]:
    if not AUTO_EVAL_SUMMARY_PATH.exists():
        return None, "아직 자동 평가 결과 파일이 없습니다. 먼저 자동 평가 스크립트를 실행해 주세요."

    try:
        with open(AUTO_EVAL_SUMMARY_PATH, "r", encoding="utf-8-sig", newline="") as f:
            summary_rows = list(csv.DictReader(f, delimiter="\t"))

        detail_rows: list[dict[str, Any]] = []
        if AUTO_EVAL_DETAIL_PATH.exists():
            with open(AUTO_EVAL_DETAIL_PATH, "r", encoding="utf-8-sig", newline="") as f:
                detail_rows = list(csv.DictReader(f, delimiter="\t"))
        else:
            for batch_path in AUTO_EVAL_BATCH_PATHS:
                if not batch_path.exists():
                    continue
                with open(batch_path, "r", encoding="utf-8-sig", newline="") as f:
                    detail_rows.extend(list(csv.DictReader(f, delimiter="\t")))

        if not detail_rows and summary_rows:
            detail_rows = summary_rows

        if not detail_rows:
            return None, "자동 평가 결과 파일이 비어 있습니다."

        normalized_rows_by_id: dict[str, dict[str, Any]] = {}
        for detail_row in detail_rows:
            normalized_row = normalize_auto_eval_row(detail_row)
            question_id = str(normalized_row.get("question_id", "")).strip()
            if question_id in {"", "Q---", "Q001-Q100", "Q001-Q110"}:
                continue
            if str(normalized_row.get("category", "")).strip() == "요약":
                continue
            normalized_rows_by_id[question_id] = normalized_row
        normalized_rows = list(normalized_rows_by_id.values())

        judgment_counts = {
            "매우 우수": 0,
            "우수": 0,
            "보통": 0,
            "보완 필요": 0,
            "미흡": 0,
        }
        for row in normalized_rows:
            judgment = str(row.get("판정", "미흡") or "미흡")
            judgment_counts[judgment] = judgment_counts.get(judgment, 0) + 1

        review_count = sum(1 for row in normalized_rows if str(row.get("검토필요", "")).upper() == "Y")
        total_score = sum(int(row.get("총점", 0)) for row in normalized_rows)
        question_count = len(normalized_rows)
        average_score = round(total_score / question_count, 2) if question_count else 0.0

        return {
            "summary_path": str(AUTO_EVAL_SUMMARY_PATH),
            "summary_rows": summary_rows,
            "rows": normalized_rows,
            "metrics": {
                "question_count": question_count,
                "average_score": average_score,
                "very_good_count": judgment_counts.get("매우 우수", 0),
                "good_count": judgment_counts.get("우수", 0),
                "average_count": judgment_counts.get("보통", 0),
                "needs_review_count": judgment_counts.get("보완 필요", 0),
                "poor_count": judgment_counts.get("미흡", 0),
                "review_needed_count": review_count,
            },
        }, None
    except Exception as e:
        return None, f"자동 평가 결과를 읽는 중 오류가 발생했습니다: {e}"


def render_auto_eval_summary() -> dict[str, Any] | None:
    st.subheader("통합 평가 결과 요약")
    st.caption(
        "전체 Q001\\~Q110, 총 110개 문항을 4개 기준 / 100점 만점으로 통합 관리합니다. "
        "Q001\\~Q030은 기존 평가를 100점 기준으로 표준화했고, "
        "Q031\\~Q110은 자동평가 결과를 통합했습니다."
    )

    auto_eval, error = load_auto_eval_summary()
    if error:
        st.info(error)
        return None

    if not auto_eval:
        st.info("자동 평가 결과를 불러올 수 없습니다.")
        return None

    metrics = auto_eval.get("metrics", {})
    summary_items = [
        ("평가 문항 수", f"{metrics.get('question_count', 0)}개"),
        ("평균 점수", f"{metrics.get('average_score', 0.0):.2f}"),
        ("매우 우수", f"{metrics.get('very_good_count', 0)}개"),
        ("우수", f"{metrics.get('good_count', 0)}개"),
        ("보통", f"{metrics.get('average_count', 0)}개"),
        ("보완 필요", f"{metrics.get('needs_review_count', 0)}개"),
        ("미흡", f"{metrics.get('poor_count', 0)}개"),
        ("검토필요", f"{metrics.get('review_needed_count', 0)}개"),
    ]

    for index, (label, value) in enumerate(summary_items):
        if index % 2 == 0:
            cols = st.columns(2)
            current_col = cols[0]
        else:
            current_col = cols[1]
        current_col.markdown(
            f"<div style='border:1px solid #e2e8f0; border-radius:8px; padding:10px 12px; margin-bottom:8px;'>"
            f"<div style='font-size:0.82rem; color:#64748b;'>{label}</div>"
            f"<div style='font-size:1.08rem; font-weight:600; color:#0f172a;'>{value}</div></div>",
            unsafe_allow_html=True,
        )

    rows = auto_eval.get("rows", [])
    low_score_rows = sorted(
        rows,
        key=lambda row: (int(row.get("총점", 0)), str(row.get("question_id", ""))),
    )[:10]

    display_rows = [
        {
            "question_id": row.get("question_id", ""),
            "category": row.get("category", ""),
            "question": row.get("question", ""),
            "검색_적합성": row.get("검색_적합성", ""),
            "근거_기반성": row.get("근거_기반성", ""),
            "안전법령_판단정확성": row.get("안전법령_판단정확성", ""),
            "실무성": row.get("실무성", ""),
            "총점": row.get("총점", ""),
            "판정": row.get("판정", ""),
            "검토필요": row.get("검토필요", ""),
        }
        for row in low_score_rows
    ]

    st.caption("총점 낮은 문항 Top 10")
    st.dataframe(
        [
            {
                "question_id": row.get("question_id", ""),
                "category": row.get("category", ""),
                "question": (row.get("question", "")[:100] + "...") if len(str(row.get("question", ""))) > 100 else row.get("question", ""),
                "총점": row.get("총점", ""),
                "판정": row.get("판정", ""),
                "검토필요": row.get("검토필요", ""),
            }
            for row in low_score_rows
        ],
        use_container_width=True,
        hide_index=True,
    )

    with st.expander("전체 자동 평가 결과 보기"):
        st.info(
            "이 표는 최종 평가가 아니라 rule-based 방식의 1차 자동 평가 결과입니다. "
            "검토필요 문항은 수동 검토 후 최종 점수로 확정하는 것을 권장합니다."
        )
        full_display_rows = [
            {
                "question_id": row.get("question_id", ""),
                "category": row.get("category", ""),
                "question": (row.get("question", "")[:100] + "...") if len(str(row.get("question", ""))) > 100 else row.get("question", ""),
                "검색_적합성": row.get("검색_적합성", ""),
                "근거_기반성": row.get("근거_기반성", ""),
                "안전법령_판단정확성": row.get("안전법령_판단정확성", ""),
                "실무성": row.get("실무성", ""),
                "총점": row.get("총점", ""),
                "판정": row.get("판정", ""),
                "검토필요": row.get("검토필요", ""),
            }
            for row in rows
        ]
        st.dataframe(full_display_rows, use_container_width=True, hide_index=True)

    return auto_eval


def load_evaluation_progress(
    scenario_rows: list[dict[str, str]] | None = None,
    include_auto_eval: bool = False,
) -> tuple[dict[str, Any] | None, str | None]:
    if not EVALUATION_PATH.exists():
        return None, f"평가표 파일을 찾을 수 없습니다: {EVALUATION_PATH}"

    try:
        manual_rows = read_evaluation_rows()
    except Exception as e:
        return None, f"평가표 파일을 읽는 중 오류가 발생했습니다: {e}"

    manual_rows_by_id = {
        normalize_question_id_for_display(row.get("번호", "")): row
        for row in manual_rows
    }

    auto_rows_by_id: dict[str, dict[str, Any]] = {}
    auto_eval_error = None
    if include_auto_eval:
        auto_eval, auto_eval_error = load_auto_eval_summary()
        if auto_eval:
            auto_rows_by_id = {
                str(row.get("question_id", "")).strip(): row
                for row in auto_eval.get("rows", [])
            }

    target_scenarios = scenario_rows or [
        {
            "번호": row.get("번호", ""),
            "분류": row.get("분류", ""),
            "난이도": row.get("난이도", ""),
            "질문 시나리오": row.get("질문", ""),
        }
        for row in manual_rows
    ]

    table_rows: list[dict[str, Any]] = []
    status_by_no: dict[str, str] = {}
    completed_count = 0

    def manual_score(
        normalized_row: dict[str, str],
        new_field: str,
        legacy_field: str,
    ) -> Any:
        new_value = normalized_row.get(new_field, "0")
        if safe_score_value(new_value) > 0:
            return new_value
        return normalized_row.get(legacy_field, "0")

    for scenario_row in target_scenarios:
        question_id = normalize_question_id_for_display(
            scenario_row.get("번호", "")
        )
        manual_row = manual_rows_by_id.get(question_id)
        auto_row = auto_rows_by_id.get(question_id)

        if auto_row:
            total_score = int(auto_row.get("총점", 0))
            judgment = str(auto_row.get("판정", "")).strip()
            review_needed = str(auto_row.get("검토필요", "")).upper() == "Y"
            evaluation_status = "완료 (검토필요)" if review_needed else "완료"
            dropdown_status = f"완료 | {total_score}점 | {judgment or '판정 없음'}"
            if review_needed:
                dropdown_status += " | 검토필요"
            display_row = {
                "번호": question_id,
                "분류": scenario_row.get("분류") or auto_row.get("category", ""),
                "난이도": scenario_row.get("난이도", ""),
                "질문": scenario_row.get("질문 시나리오") or auto_row.get("question", ""),
                "검색 적합성": auto_row.get("검색_적합성", 0),
                "근거 기반성": auto_row.get("근거_기반성", 0),
                "안전·법령 판단 정확성": auto_row.get("안전법령_판단정확성", 0),
                "실무성": auto_row.get("실무성", 0),
                "총점": total_score,
                "판정": judgment,
                "평가 여부": evaluation_status,
            }
            completed = True
        else:
            normalized_row = normalize_evaluation_row(manual_row or {})
            completed = bool(manual_row) and is_evaluation_completed(normalized_row)
            dropdown_status = "완료" if completed else "미평가"
            display_row = {
                "번호": question_id,
                "분류": scenario_row.get("분류") or normalized_row.get("분류", ""),
                "난이도": scenario_row.get("난이도") or normalized_row.get("난이도", ""),
                "질문": scenario_row.get("질문 시나리오") or normalized_row.get("질문", ""),
                "검색 적합성": manual_score(
                    normalized_row, "검색_적합성", "검색정확도(0~5)"
                ),
                "근거 기반성": manual_score(
                    normalized_row, "근거_기반성", "근거성(0~5)"
                ),
                "안전·법령 판단 정확성": manual_score(
                    normalized_row,
                    "안전법령_판단정확성",
                    "답변정확도(0~5)",
                ),
                "실무성": manual_score(
                    normalized_row, "실무성", "실무성(0~5)"
                ),
                "총점": normalized_row.get("총점", ""),
                "판정": normalized_row.get("판정", ""),
                "평가 여부": "완료" if completed else "미평가",
            }

        if completed:
            completed_count += 1
        table_rows.append(display_row)

        plain_number = str(scenario_row.get("번호", "")).strip()
        status_by_no[plain_number] = dropdown_status
        status_by_no[question_id] = dropdown_status

    total_count = len(table_rows)
    incomplete_count = total_count - completed_count
    completion_rate = (completed_count / total_count) if total_count else 0.0
    return {
        "rows": manual_rows,
        "table_rows": table_rows,
        "status_by_no": status_by_no,
        "total_count": total_count,
        "completed_count": completed_count,
        "incomplete_count": incomplete_count,
        "completion_rate": completion_rate,
        "auto_eval_error": auto_eval_error,
    }, None


def render_evaluation_progress(
    scenario_rows: list[dict[str, str]] | None = None,
    include_auto_eval: bool = False,
) -> dict[str, Any] | None:
    st.subheader("평가 진행 현황")

    progress, error = load_evaluation_progress(
        scenario_rows=scenario_rows,
        include_auto_eval=include_auto_eval,
    )
    if error:
        st.warning(error)
        return None

    if not progress:
        st.info("평가 진행 현황을 불러올 수 없습니다.")
        return None

    completed_count = sum(
        1
        for table_row in progress["table_rows"]
        if str(table_row.get("평가 여부", "")).strip().startswith("완료")
    )

    total_count = len(progress["table_rows"])
    incomplete_count = total_count - completed_count
    completion_rate = completed_count / total_count if total_count else 0.0

    progress["total_count"] = total_count
    progress["completed_count"] = completed_count
    progress["incomplete_count"] = incomplete_count
    progress["completion_rate"] = completion_rate

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("전체 질문 수", f"{total_count}개")
    col2.metric("평가 완료", f"{completed_count}개")
    col3.metric("미평가", f"{incomplete_count}개")
    col4.metric("완료율", f"{completion_rate * 100:.1f}%")

    st.progress(completion_rate)

    st.dataframe(
        progress["table_rows"],
        use_container_width=True,
        hide_index=True,
    )

    if progress.get("auto_eval_error") and include_auto_eval:
        st.warning(progress["auto_eval_error"])

    st.info(
        "전체 Q001\\~Q110은 auto_eval_Q001_Q110.tsv의 "
        "4개 기준 / 100점 만점 통합 평가 결과를 우선 표시합니다. "
        "검토필요 문항은 수동 검토 후 최종 점수로 확정하는 것을 권장합니다."
    )

    return progress

def write_evaluation_rows(rows: list[dict[str, str]]) -> None:
    headers = [
        "번호",
        "분류",
        "난이도",
        "질문",
        "기대 검색 문서",
        "검색된 주요 근거 문서",
        "Gemini 답변 요약",
        "답변 생성 방식",
        "Gemini 상태",
        "검색_적합성",
        "근거_기반성",
        "안전법령_판단정확성",
        "실무성",
        "총점",
        "판정",
        "메모",
        "검색정확도(0~5)",
        "답변정확도(0~5)",
        "근거성(0~5)",
        "환각억제(0~5)",
        "실무성(0~5)",
    ]

    normalized_rows = [normalize_evaluation_row(row) for row in rows]
    EVALUATION_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(EVALUATION_PATH, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(normalized_rows)


def update_evaluation_result(
    scenario: dict[str, str],
    results: list[dict[str, Any]],
    answer: str,
    answer_mode: str,
    gemini_status: str,
    scores: dict[str, int],
    memo: str,
) -> tuple[bool, str]:
    try:
        rows = read_evaluation_rows()
        if not rows:
            scenarios, scenario_error = load_question_scenarios(selected_scenario_path)
            if scenario_error:
                return False, scenario_error
            rows = [
                {
                    "번호": r.get("번호", ""),
                    "분류": r.get("분류", ""),
                    "난이도": r.get("난이도", ""),
                    "질문": r.get("질문 시나리오", ""),
                    "기대 검색 문서": r.get("기대 검색 문서", ""),
                    "검색된 주요 근거 문서": "",
                    "Gemini 답변 요약": "",
                    "답변 생성 방식": "",
                    "Gemini 상태": "",
                    "검색_적합성": "0",
                    "근거_기반성": "0",
                    "안전법령_판단정확성": "0",
                    "실무성": "0",
                    "총점": "",
                    "판정": "",
                    "메모": "",
                }
                for r in scenarios
            ]

        normalized_scores = {
            "검색_적합성": int(scores.get("검색_적합성", scores.get("검색정확도(0~5)", 0))),
            "근거_기반성": int(scores.get("근거_기반성", scores.get("근거성(0~5)", 0))),
            "안전법령_판단정확성": int(scores.get("안전법령_판단정확성", scores.get("답변정확도(0~5)", 0))),
            "실무성": int(scores.get("실무성", scores.get("실무성(0~5)", 0))),
        }
        total_score = sum(normalized_scores.values())
        judgment = calculate_judgment(total_score)
        target_no = str(scenario.get("번호", "")).strip()
        updated = False

        for row in rows:
            row.setdefault("답변 생성 방식", "")
            row.setdefault("Gemini 상태", "")

        for row in rows:
            if str(row.get("번호", "")).strip() == target_no:
                row["검색된 주요 근거 문서"] = make_evidence_summary(results)
                row["Gemini 답변 요약"] = make_answer_summary(answer)
                row["답변 생성 방식"] = answer_mode
                row["Gemini 상태"] = gemini_status
                row["검색_적합성"] = str(normalized_scores["검색_적합성"])
                row["근거_기반성"] = str(normalized_scores["근거_기반성"])
                row["안전법령_판단정확성"] = str(normalized_scores["안전법령_판단정확성"])
                row["실무성"] = str(normalized_scores["실무성"])
                row["총점"] = str(total_score)
                row["판정"] = judgment
                row["메모"] = memo
                updated = True
                break

        if not updated:
            return False, f"evaluation_template.tsv에서 번호 {target_no} 행을 찾을 수 없습니다."

        write_evaluation_rows(rows)
        return True, f"평가 결과 저장 완료: Q{int(target_no):02d}, 총점 {total_score}, 판정 {judgment}"
    except Exception as e:
        return False, str(e)



def format_evidence_lines(results: list[dict[str, Any]], limit: int = 5) -> list[str]:
    lines = []
    for r in results[:limit]:
        lines.append(
            f"- 근거 {r.get('rank', '-')} | {r.get('source', '출처 정보 없음')} | "
            f"chunk_id {r.get('chunk_id', '정보 없음')} | 거리값 {format_distance(r.get('distance'))}\n"
            f"  - {make_preview(str(r.get('text', '')), 220)}"
        )
    if not lines:
        lines.append("- 검색된 문서 근거가 없습니다. 일반 안전 원칙 수준으로만 안내합니다.")
    return lines


def format_source_lines(results: list[dict[str, Any]], limit: int = 8) -> list[str]:
    unique_sources = []
    for r in results:
        source_label = f"{r.get('source', '출처 정보 없음')} (chunk_id: {r.get('chunk_id', '정보 없음')})"
        if source_label not in unique_sources:
            unique_sources.append(source_label)
    return [f"- {source}" for source in unique_sources[:limit]] or ["- 검색 근거 문서 없음"]


def with_evidence_notice(
    lines: list[str],
    results: list[dict[str, Any]],
    evidence_assessment: dict[str, Any] | None = None,
) -> list[str]:
    if isinstance(evidence_assessment, dict):
        status = evidence_assessment.get("status")
        notice = (
            str(evidence_assessment.get("reason", ""))
            if status != EVIDENCE_STATUS_SUFFICIENT
            else ""
        )
    else:
        notice = evidence_limitation_notice(results)
    if notice:
        return [*lines, "", f"> {notice}"]
    return lines


def build_ppe_general_answer(
    question: str,
    results: list[dict[str, Any]],
    intent: str,
    evidence_assessment: dict[str, Any] | None = None,
) -> str:
    ppe_item = detect_ppe_item(question)
    knowledge = PPE_BASIC_KNOWLEDGE.get(ppe_item)
    if knowledge is None and ppe_item == "보호구":
        knowledge = {
            "why": "보호구는 작업 중 몸에 직접 닿는 위험을 줄이는 마지막 방어수단입니다. 위험 제거·격리·환기 같은 조치를 먼저 하고, 남는 위험에 맞는 보호구를 착용해야 합니다.",
            "when": ["갱내 작업", "장비 주변 작업", "굴진·천공·파쇄 작업", "분진·소음·비래물·추락 위험이 있는 작업"],
            "checks": ["작업 위험에 맞는 보호구인지", "손상·오염·유효 상태", "몸에 맞고 올바르게 착용되었는지"],
        }
    evidence_lines = format_evidence_lines(results)
    source_lines = format_source_lines(results)
    return "\n".join(
        with_evidence_notice(
            [
                "## 핵심 답변",
                f"- {ppe_item}는 작업 중 다치기 쉬운 신체 부위를 보호하기 위해 착용합니다.",
                "- 보호구는 위험을 없애는 조치를 대신하지는 못하지만, 남아 있는 위험으로부터 작업자를 보호하는 중요한 장치입니다.",
                "",
                "## 착용해야 하는 이유",
                f"- {knowledge['why']}",
                "- 보호구를 착용하지 않은 상태로 위험작업을 하려는 경우에는 작업 투입 전 착용 상태를 먼저 확인해야 합니다.",
                "",
                "## 착용이 필요한 작업/상황",
                *[f"- {item}" for item in knowledge["when"]],
                "",
                "## 착용 전 확인사항",
                *[f"- {item}" for item in knowledge["checks"]],
                "",
                "## 현장관리자 확인사항",
                "- 작업 위험에 맞는 보호구가 지급되었는지 확인합니다.",
                "- 착용 방법을 작업자가 이해했는지 확인합니다.",
                "- 손상되거나 오염된 보호구는 교체하도록 조치합니다.",
                "- 보호구 지급, 착용 점검, 교육 내용을 기록으로 남깁니다.",
                "",
                "## 관련 근거 요약",
                *evidence_lines,
                "",
                "## 관련 근거 문서",
                *source_lines,
            ],
            results,
            evidence_assessment,
        )
    )


def build_risk_response_answer(
    question: str,
    results: list[dict[str, Any]],
    situation_type: str,
    evidence_assessment: dict[str, Any] | None = None,
) -> str:
    immediate_judgment = build_immediate_judgment(situation_type)
    priority_actions = [f"{idx}. {item}" for idx, item in enumerate(build_priority_actions(situation_type), start=1)]
    check_items = [f"{idx}. {item}" for idx, item in enumerate(build_check_items(situation_type), start=1)]
    evidence_lines = format_evidence_lines(results)
    source_lines = format_source_lines(results)
    kras_section = build_kras_risk_assessment_section(question, results, situation_type)
    return "\n".join(
        with_evidence_notice(
            [
                "## 즉시 판단",
                f"- 상황 유형: {situation_type}",
                f"- {immediate_judgment}",
                "",
                "## 우선 조치",
                *priority_actions,
                "",
                "## 작업중지/대피 필요 여부",
                "- 급박한 위험이 있으면 작업을 먼저 중지합니다.",
                "- 유해가스, 불발, 낙반, 전기 이상, 장비 충돌 위험 등 즉시 위험이 있으면 대피 또는 출입통제를 우선합니다.",
                "- 작업재개는 점검, 재측정, 보강, 승인, 기록이 끝난 뒤 검토합니다.",
                "",
                "## 현장 점검사항",
                *check_items,
                "",
                "## 기록 및 증빙자료",
                "- 작업중지 지시와 전파 내용을 기록합니다.",
                "- 측정값, 점검표, 사진, TBM 회의록, 교육·보호구 지급 기록을 남깁니다.",
                "- 조치 완료 후 작업재개 승인 근거를 보관합니다.",
                "",
                "## 관련 근거 요약",
                *evidence_lines,
                "",
                "## 관련 근거 문서",
                *source_lines,
                "",
                kras_section,
            ],
            results,
            evidence_assessment,
        )
    )


def build_general_education_answer(
    question: str,
    results: list[dict[str, Any]],
    intent: str,
    situation_type: str,
    evidence_assessment: dict[str, Any] | None = None,
) -> str:
    evidence_lines = format_evidence_lines(results)
    source_lines = format_source_lines(results)
    if intent == PREWORK_TBM_INTENT:
        lead = "작업 전 TBM은 당일 작업의 위험요인과 조치방법을 작업자와 관리자가 함께 확인하는 절차입니다."
        why = [
            "작업자가 오늘 해야 할 일과 위험요인을 같은 기준으로 이해하게 합니다.",
            "보호구, 장비, 신호, 대피 절차를 작업 시작 전에 확인할 수 있습니다.",
            "전날 지적사항이나 변경된 작업 조건을 빠뜨리지 않게 합니다.",
        ]
        when = ["매일 작업 시작 전", "작업 위치·방법·장비가 바뀐 때", "신규 작업자나 협력업체가 투입될 때"]
    else:
        lead = "광산 안전관리는 작은 이상 징후를 작업 전에 확인하고 기록해 사고로 이어지지 않게 하는 활동입니다."
        why = [
            "갱내 환경은 환기, 조명, 통로, 장비 상태가 조금만 달라져도 위험이 커질 수 있습니다.",
            "작업 전 확인과 교육은 작업자별 역할과 대피 기준을 분명히 합니다.",
            "기록은 조치 이행 여부와 재발방지 근거가 됩니다.",
        ]
        when = ["작업 시작 전", "작업 조건이 바뀐 때", "이상 징후가 발견된 때"]
    return "\n".join(
        with_evidence_notice(
            [
                "## 핵심 설명",
                f"- {lead}",
                "",
                "## 왜 필요한지",
                *[f"- {item}" for item in why],
                "",
                "## 현장에서 언제 적용하는지",
                *[f"- {item}" for item in when],
                "",
                "## 작업 전 확인사항",
                *[f"- {item}" for item in build_check_items(situation_type)[:6]],
                "",
                "## 기록 또는 교육자료로 남길 사항",
                "- 참석자, 일시, 작업명, 공유한 위험요인과 조치사항",
                "- 보호구 착용 확인, 장비·환기·조명·통로 점검 결과",
                "- 작업자가 제기한 의견과 개선 조치",
                "",
                "## 관련 근거 요약",
                *evidence_lines,
                "",
                "## 관련 근거 문서",
                *source_lines,
            ],
            results,
            evidence_assessment,
        )
    )


def build_kras_intent_answer(
    question: str,
    results: list[dict[str, Any]],
    situation_type: str,
    evidence_assessment: dict[str, Any] | None = None,
) -> str:
    evidence_lines = format_evidence_lines(results)
    source_lines = format_source_lines(results)
    kras_section = build_kras_risk_assessment_section(question, results, situation_type)
    return "\n".join(
        with_evidence_notice(
            [
                "## 위험요인",
                "- 작업 장소, 장비, 작업방법, 작업자 상태에서 유해·위험요인을 먼저 찾습니다.",
                "- 갱내 작업에서는 낙반, 환기·유해가스, 분진, 전기, 장비 운반 위험을 함께 확인합니다.",
                "",
                "## 현재 위험성",
                "- 가능성과 중대성을 함께 보아 위험 수준을 정합니다.",
                "- 검색 근거만으로 수치화하기 어려운 경우 현장 기준에 따라 재평가해야 합니다.",
                "",
                "## 감소대책",
                "- 제거, 대체, 공학적 대책, 관리적 대책, 보호구 순서로 검토합니다.",
                "- 대책은 담당자, 완료 시점, 확인 방법까지 정해야 합니다.",
                "",
                "## 확인해야 할 기록",
                "- 위험성평가표, 감소대책 이행 기록, TBM 기록, 점검표, 사진, 교육자료",
                "",
                "## KRAS 초안 연결 안내",
                "- 아래 KRAS식 위험성평가 기록 초안을 현장 조건에 맞게 보완해 사용할 수 있습니다.",
                "",
                "## 관련 근거 요약",
                *evidence_lines,
                "",
                "## 관련 근거 문서",
                *source_lines,
                "",
                kras_section,
            ],
            results,
            evidence_assessment,
        )
    )


def build_law_explanation_answer(
    question: str,
    results: list[dict[str, Any]],
    situation_type: str,
    evidence_assessment: dict[str, Any] | None = None,
) -> str:
    evidence_lines = format_evidence_lines(results)
    source_lines = format_source_lines(results)
    return "\n".join(
        with_evidence_notice(
            [
                "## 쉬운 설명",
                "- 중대재해처벌법과 관련 법령 대응은 사고가 난 뒤 설명하는 것이 아니라, 평소 안전보건관리체계가 실제로 작동했다는 점을 자료로 남기는 일입니다.",
                "- 조항 번호, 처벌 여부, 법적 책임은 검색 근거가 명확하지 않으면 단정하지 않습니다.",
                "",
                "## 현장관리자가 해야 할 일",
                "- 작업 전 위험요인을 확인하고 필요한 개선조치를 요청·이행합니다.",
                "- TBM, 보호구 착용, 작업허가, 점검, 작업중지·재개 승인 절차를 운영합니다.",
                "- 위험 상황이나 사고 발생 시 보고, 현장 통제, 원인조사, 재발방지대책 수립을 지원합니다.",
                "",
                "## 증빙자료로 남겨야 할 것",
                "- 점검표, TBM 참석자 명부, 교육자료, 보호구 지급·착용 점검 기록",
                "- 위험성평가와 감소대책 이행 기록, 개선조치 완료 사진",
                "- 사고·아차사고 보고서, 원인조사, 재발방지대책과 이행 확인 자료",
                "",
                "## 단정 금지",
                "- 법령 해석, 위반 여부, 처벌 수위는 현장 사실관계와 최신 법령 확인이 필요합니다.",
                "- 필요한 경우 안전관리자, 법무 담당자, 관계기관 또는 전문가 확인을 받아야 합니다.",
                "",
                "## 관련 근거 요약",
                *evidence_lines,
                "",
                "## 관련 근거 문서",
                *source_lines,
            ],
            results,
            evidence_assessment,
        )
    )


def build_out_of_scope_answer(question: str) -> str:
    return "\n".join(
        [
            "## 답변 범위 안내",
            "- MineSafe AI는 광산 안전관리, 산업안전, 작업장 안전, 관련 법령·지침 기반 답변에 특화되어 있습니다.",
            "- 현재 질문은 광산 안전관리 범위와 직접 관련이 적어 안전 답변으로 억지로 연결하지 않겠습니다.",
            "",
            "## 다시 질문하는 방법",
            "- 광산 현장 작업, 보호구, 환기, 발파, 낙반, 전기, 장비, TBM, 위험성평가, 사고보고, 중대재해처벌법과 연결해 질문해 주세요.",
            "- 예: '갱내 작업 전 안전화 착용을 어떻게 확인해야 해?'",
        ]
    )


def generate_local_fallback_answer(
    question: str,
    results: list[dict[str, Any]],
    reason: str = "",
    intent: str | None = None,
    evidence_assessment: dict[str, Any] | None = None,
) -> str:
    intent = intent or detect_question_intent(question)
    if intent == OUT_OF_SCOPE_INTENT:
        return build_out_of_scope_answer(question)

    situation_type = map_intent_to_situation_type(intent, question)
    if intent == PPE_GENERAL_INTENT:
        return build_ppe_general_answer(question, results, intent, evidence_assessment)
    if intent in RISK_SIGN_INTENTS:
        return build_risk_response_answer(question, results, situation_type, evidence_assessment)
    if intent == KRAS_INTENT:
        return build_kras_intent_answer(question, results, situation_type, evidence_assessment)
    if intent in MANAGEMENT_RECORD_INTENTS:
        return build_law_explanation_answer(question, results, situation_type, evidence_assessment)
    return build_general_education_answer(
        question,
        results,
        intent,
        situation_type,
        evidence_assessment,
    )


def call_gemini_with_timeout(
    client,
    prompt: str,
    model_name: str,
    *,
    timeout_seconds: int = GEMINI_RESPONSE_TIMEOUT_SECONDS,
    max_output_tokens: int = GEMINI_MAX_OUTPUT_TOKENS,
) -> str:
    config = types.GenerateContentConfig(
        temperature=0.2,
        max_output_tokens=max_output_tokens,
    )

    executor = ThreadPoolExecutor(max_workers=1)
    future = executor.submit(
        client.models.generate_content,
        model=model_name,
        contents=prompt,
        config=config,
    )

    try:
        response = future.result(timeout=timeout_seconds)
    except FutureTimeoutError as e:
        future.cancel()
        raise TimeoutError(
            f"Gemini 응답 제한 시간 {timeout_seconds}초 초과"
        ) from e
    finally:
        executor.shutdown(wait=False, cancel_futures=True)

    text = getattr(response, "text", None)
    if not isinstance(text, str):
        raise RuntimeError("Gemini 응답에 text가 없습니다.")

    cleaned_text = text.strip()
    if not cleaned_text:
        raise RuntimeError("Gemini 응답이 비어 있습니다.")
    return cleaned_text


def execute_gemini_request(
    prompt: str,
    model_name: str,
    *,
    max_output_tokens: int = GEMINI_MAX_OUTPUT_TOKENS,
) -> dict[str, Any]:
    client, error = load_gemini_client()
    if error:
        error_summary = summarize_gemini_error(error)
        return {
            "called": False,
            "success": False,
            "status": "error",
            "message": (
                "Gemini API를 호출할 수 없습니다. "
                f"오류 요약: {error_summary}"
            ),
            "answer": "",
            "attempts": 0,
            "model": model_name,
            "used_fallback": False,
            "error": error_summary,
            "elapsed": 0.0,
        }

    start_time = time.monotonic()
    last_error = ""
    last_status = "error"
    attempts_used = 0

    for attempt in range(1, GEMINI_MAX_ATTEMPTS + 1):
        attempts_used = attempt
        try:
            answer = call_gemini_with_timeout(
                client,
                prompt,
                model_name,
                timeout_seconds=GEMINI_RESPONSE_TIMEOUT_SECONDS,
                max_output_tokens=max_output_tokens,
            )
            if not answer or not answer.strip():
                raise RuntimeError("Gemini 응답이 비어 있습니다.")
            return {
                "called": True,
                "success": True,
                "status": "success",
                "message": "Gemini 응답 생성에 성공했습니다.",
                "answer": answer,
                "attempts": attempt,
                "model": model_name,
                "used_fallback": False,
                "error": "",
                "elapsed": time.monotonic() - start_time,
            }
        except Exception as e:
            last_error = str(e)
            last_status = classify_gemini_error(last_error)
            should_retry = (
                last_status in {"503", "429", "timeout"}
                or is_temporary_gemini_error(last_error)
            )
            if attempt < GEMINI_MAX_ATTEMPTS and should_retry:
                time.sleep(GEMINI_RETRY_SLEEP_SECONDS)
                continue
            break

    return {
        "called": True,
        "success": False,
        "status": last_status,
        "message": get_gemini_failure_message(last_status, last_error),
        "answer": "",
        "attempts": attempts_used,
        "model": model_name,
        "used_fallback": False,
        "error": summarize_gemini_error(last_error),
        "elapsed": time.monotonic() - start_time,
    }


def test_gemini_connection(model_name: str) -> dict[str, Any]:
    return execute_gemini_request(
        "테스트입니다. 한 문장으로 응답해 주세요.",
        model_name,
        max_output_tokens=80,
    )


def build_legacy_gemini_status(
    result: dict[str, Any],
    *,
    fallback_used: bool,
) -> dict[str, Any]:
    success = bool(result.get("success", False))
    return {
        **result,
        "mode": "gemini" if success else "fallback",
        "reason": str(result.get("error", "")),
        "selected_model": str(result.get("model", GEMINI_MODEL_NAME)),
        "gemini_called": bool(result.get("called", False)),
        "fallback_used": fallback_used,
        "used_fallback": fallback_used,
        "gemini_status": classify_gemini_status(result),
    }



def generate_gemini_answer(
    question: str,
    results: list[dict[str, Any]],
    model_name: str = GEMINI_MODEL_NAME,
    *,
    prompt_kind: str = "gemini",
    stable_draft: str = "",
    reference_cases: list[dict[str, Any]] | None = None,
    live_news_cases: list[dict[str, Any]] | None = None,
    evidence_assessment: dict[str, Any] | None = None,
    official_cases: list[dict[str, Any]] | None = None,
):
    intent = detect_question_intent(question)
    if prompt_kind == "hybrid":
        prompt = build_hybrid_prompt(
            question,
            results,
            stable_draft,
            intent,
            reference_cases or [],
            live_news_cases or [],
            evidence_assessment,
            official_cases or [],
        )
    else:
        prompt = build_prompt(
            question,
            results,
            intent,
            reference_cases or [],
            live_news_cases or [],
            evidence_assessment,
            official_cases or [],
        )

    result = execute_gemini_request(prompt, model_name)

    if result.get("success"):
        answer = str(result.get("answer", "")).strip()
        if prompt_kind == "gemini" and "근로자 쉬운 설명" not in answer:
            answer = "\n\n".join(["### 근로자 쉬운 설명", answer])
        if prompt_kind == "hybrid" and "하이브리드 답변" not in answer:
            answer = "\n\n".join(["### 하이브리드 답변", answer])
        if "KRAS식 위험성평가 기록 초안" not in answer:
            answer = "\n\n".join(
                [
                    answer,
                    build_kras_risk_assessment_section(question, results),
                ]
            )
        status = build_legacy_gemini_status(result, fallback_used=False)
        status["prompt_kind"] = prompt_kind
        return answer, status

    fallback_answer = generate_local_fallback_answer(
        question,
        results,
        str(result.get("message") or result.get("error") or "Gemini 호출 실패"),
        detect_question_intent(question),
        evidence_assessment=evidence_assessment,
    )
    status = build_legacy_gemini_status(result, fallback_used=True)
    status["prompt_kind"] = prompt_kind
    return fallback_answer, status

# ==============================
# 법령 체크리스트·증빙자료·위험성평가·대화이력 기능
# ==============================
LEGAL_CHECKLIST_ITEMS = [
    ("안전보건 경영방침 수립 여부", "경영방침을 문서화하고 현장 구성원에게 공유했는지 확인"),
    ("안전보건 전담조직 지정 여부", "안전보건 업무 담당 조직과 역할이 지정되어 있는지 확인"),
    ("안전보건 예산 편성 및 집행 여부", "안전보건 예산 편성, 집행 내역, 개선 투자 기록 확인"),
    ("위험성평가 실시 여부", "정기·수시 위험성평가 실시 및 개선대책 수립 여부 확인"),
    ("개선조치 이행 여부", "위험성평가, 점검, 사고조사 후 개선조치 완료 여부 확인"),
    ("작업 전 TBM 기록 여부", "작업 전 위험요인 공유, 보호구, 작업중지 기준 안내 기록 확인"),
    ("보호구 지급대장 보유 여부", "보호구 지급, 교체, 착용 점검 기록 보유 여부 확인"),
    ("교육 기록 보유 여부", "정기교육, 특별교육, TBM 참석자 명부 및 교육자료 확인"),
    ("가스/분진/환기 측정 기록 보유 여부", "측정값, 측정자, 측정장비, 조치 내역 기록 확인"),
    ("사고 및 아차사고 보고서 작성 여부", "사고·아차사고 발생 보고, 원인, 조치 기록 확인"),
    ("재발방지대책 수립 여부", "원인조사 후 재발방지대책과 이행 결과 확인"),
    ("협력업체 안전관리 기록 여부", "협력업체 작업 전 교육, 위험성평가, 작업허가 기록 확인"),
]


def ensure_feature_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    FEATURE_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def default_legal_checklist_rows() -> list[dict[str, Any]]:
    return [
        {
            "항목명": item,
            "설명": description,
            "상태": "미작성",
            "담당자": "",
            "최근 작성일": "",
            "비고": "",
        }
        for item, description in LEGAL_CHECKLIST_ITEMS
    ]


def load_legal_checklist_status() -> list[dict[str, Any]]:
    ensure_feature_dirs()
    if not LEGAL_CHECKLIST_STATUS_PATH.exists():
        rows = default_legal_checklist_rows()
        save_legal_checklist_status(rows)
        return rows
    try:
        data = json.loads(LEGAL_CHECKLIST_STATUS_PATH.read_text(encoding="utf-8"))
        if isinstance(data, list) and data:
            return data
    except Exception:
        pass
    return default_legal_checklist_rows()


def save_legal_checklist_status(rows: list[dict[str, Any]]) -> None:
    ensure_feature_dirs()
    LEGAL_CHECKLIST_STATUS_PATH.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def normalize_excel_value(value: Any) -> Any:
    if isinstance(value, (list, tuple, set)):
        return " | ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    if value is None:
        return ""
    return value


def build_simple_xlsx_workbook(sheet_name: str, rows: list[dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    headers = list(rows[0].keys()) if rows else ["내용"]
    ws.append(headers)
    for row in rows:
        ws.append([normalize_excel_value(row.get(header, "")) for header in headers])
    style_excel_sheet(ws)
    return wb


def build_simple_xlsx_bytes(sheet_name: str, rows: list[dict[str, Any]]) -> bytes:
    wb = build_simple_xlsx_workbook(sheet_name, rows)
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer.getvalue()


def write_simple_xlsx(path: Path, sheet_name: str, rows: list[dict[str, Any]]) -> None:
    ensure_feature_dirs()
    wb = build_simple_xlsx_workbook(sheet_name, rows)
    wb.save(path)
    wb.close()


def style_excel_sheet(ws) -> None:
    thin = Side(style="thin", color="D9E2EC")
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx > 1:
            ws.row_dimensions[row_idx].height = 42
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=10)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx == 1:
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="17324D")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, idx).value or "")
        width = 16
        if "답변" in header or "answer" in header or "비고" in header or "메모" in header:
            width = 48
        elif "설명" in header or "증빙" in header or "근거" in header or "질문" in header:
            width = 36
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def export_legal_checklist_xlsx(rows: list[dict[str, Any]]) -> Path:
    write_simple_xlsx(LEGAL_CHECKLIST_EXPORT_PATH, "legal_checklist", rows)
    return LEGAL_CHECKLIST_EXPORT_PATH


def recommend_evidence_records(question: str, answer: str = "") -> list[str]:
    text = f"{question} {answer}".lower()
    rules = [
        (["메탄", "가스", "환기", "산소"], ["가스 측정 기록지", "환기설비 점검표", "작업중지 지시 기록", "작업재개 승인 기록", "TBM 회의록", "개선조치 완료 사진"]),
        (["발파", "불발", "장약", "화약"], ["발파작업 허가서", "발파 전 점검표", "화약류 사용 기록", "불발 장약 처리 기록", "대피 확인 기록", "작업재개 승인 기록"]),
        (["낙반", "붕락", "막장", "지보", "천반"], ["막장 점검표", "지보 상태 점검 기록", "작업중지 기록", "보강조치 완료 사진", "작업재개 승인 기록"]),
        (["분진", "방진", "마스크", "보호구", "ppe"], ["분진 측정 기록", "방진마스크 지급대장", "보호구 착용 점검표", "작업환경측정 결과", "교육 참석자 명부"]),
        (["중대재해", "사고", "응급", "경영책임자"], ["사고 발생 보고서", "응급조치 기록", "관계기관 보고 기록", "재발방지대책", "개선조치 이행 결과", "경영책임자 보고 기록"]),
        (["전기", "감전", "접지", "절연"], ["전기설비 점검표", "전원 차단 및 잠금표지 기록", "절연저항 측정 기록", "접지 상태 점검 기록", "작업허가서"]),
        (["운반", "장비", "협착", "통로", "차량"], ["장비 일상점검표", "작업동선 분리 계획", "신호수 배치 기록", "후진경보장치 점검표", "작업자 교육 기록"]),
    ]
    recommended: list[str] = []
    for keywords, records in rules:
        if any(keyword in text for keyword in keywords):
            for record in records:
                if record not in recommended:
                    recommended.append(record)
    if not recommended:
        recommended = ["작업 전 TBM 회의록", "위험성평가표", "작업중지·재개 판단 기록", "현장 점검 사진", "교육 참석자 명부"]
    return recommended


def render_recommended_evidence_records(records: list[str]) -> None:
    with st.container(border=True):
        st.markdown("#### 필요 증빙자료")
        st.caption("질문과 답변의 키워드를 기준으로 현장 이행자료 후보를 자동 추천합니다.")
        cols = st.columns(2)
        for idx, record in enumerate(records):
            with cols[idx % 2]:
                st.markdown(f"- {record}")
        st.info("AI 답변만으로 법적 책임이 면제되는 것은 아니며, 실제 점검표·사진·교육기록·작업중지기록 등 현장 이행자료와 함께 관리해야 합니다.")



def load_latest_reference_cases() -> list[dict[str, Any]]:
    try:
        if not LATEST_REFERENCE_CASES_PATH.exists():
            return []
        data = json.loads(LATEST_REFERENCE_CASES_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    valid_cases: list[dict[str, Any]] = []
    for item in data:
        if isinstance(item, dict):
            valid_cases.append(item)
    return valid_cases


def reference_case_categories_for_intent(intent: str) -> list[str]:
    mapping = {
        PPE_GENERAL_INTENT: ["보호구/PPE"],
        VENTILATION_GAS_INTENT: ["환기/유해가스"],
        VENTILATION_EQUIPMENT_INTENT: ["환기/유해가스"],
        GAS_DETECTOR_INTENT: ["환기/유해가스"],
        BLASTING_MISFIRE_INTENT: ["발파/불발"],
        ROOF_FALL_INTENT: ["낙반/붕락"],
        DUST_RESPIRATORY_INTENT: ["분진/호흡보호"],
        ELECTRICAL_SAFETY_INTENT: ["전기안전"],
        EQUIPMENT_TRANSPORT_INTENT: ["장비/운반/차량"],
        BACKING_SIGNAL_INTENT: ["장비/운반/차량"],
        FIRE_EMERGENCY_INTENT: ["화재/소화"],
        HOT_WORK_INTENT: ["화재/소화"],
        EVACUATION_ACCESS_SIGN_INTENT: ["대피로/표지"],
        FLOOD_DRAINAGE_INTENT: ["침수/배수"],
        WALKWAY_HOUSEKEEPING_INTENT: ["대피로/표지", "침수/배수"],
        CONVEYOR_ROTATING_INTENT: ["컨베이어/끼임"],
        LAW_INTENT: ["중대재해처벌법/증빙자료"],
        DOCUMENT_EVIDENCE_INTENT: ["중대재해처벌법/증빙자료"],
        KRAS_INTENT: ["중대재해처벌법/증빙자료"],
    }
    return mapping.get(intent, [intent])


def match_reference_cases(question: str, intent: str, max_cases: int = 2) -> list[dict[str, Any]]:
    if intent == OUT_OF_SCOPE_INTENT:
        return []
    question_text = clean_text(question).lower()
    target_categories = reference_case_categories_for_intent(intent)
    scored: list[tuple[int, dict[str, Any]]] = []
    for case in load_latest_reference_cases():
        score = 0
        category = str(case.get("category", ""))
        category_matched = False
        if category in target_categories:
            score += 8
            category_matched = True
        elif any(category in target or target in category for target in target_categories):
            score += 5
            category_matched = True
        keywords = case.get("keywords", [])
        if not isinstance(keywords, list):
            keywords = []
        keyword_hits = 0
        for keyword in keywords:
            key = str(keyword).strip().lower()
            if key and key in question_text:
                keyword_hits += 1
        score += keyword_hits * 3
        if category_matched or keyword_hits:
            source_type = str(case.get("source_type", "")).lower()
            if source_type == "official":
                score += 3
            elif source_type == "public_agency":
                score += 2
            elif source_type in {"manual", "sample"}:
                score += 1
        if score > 0:
            enriched = dict(case)
            enriched["match_score"] = score
            scored.append((score, enriched))
    scored.sort(key=lambda item: item[0], reverse=True)
    return [case for _, case in scored[:max_cases]]


def case_warning_points_for_intent(intent: str, reference_cases: list[dict[str, Any]] | None = None) -> list[str]:
    point_map = {
        PPE_GENERAL_INTENT: [
            "보호구 미착용 상태에서 위험작업을 진행하면 사고 발생 시 보호구 지급·착용 확인 여부가 쟁점이 될 수 있습니다.",
            "안전모, 안전화, 방진마스크 등 보호구의 상태와 착용 기록을 남기는 것이 중요합니다.",
            "반복 미착용을 방치하면 관리감독 미흡으로 지적될 수 있습니다.",
        ],
        VENTILATION_GAS_INTENT: [
            "유해가스 또는 산소결핍 위험이 있는데 작업을 계속하면 작업중지·대피 조치 여부가 쟁점이 될 수 있습니다.",
            "가스 측정 기록, 환기설비 점검 기록, 재측정 기록이 중요합니다.",
            "경보 발생 후 조치가 지연되면 관리상 미흡으로 볼 수 있습니다.",
        ],
        VENTILATION_EQUIPMENT_INTENT: [
            "환기팬 정지나 풍량 이상을 알고도 작업을 계속하면 작업중지·대피 판단이 쟁점이 될 수 있습니다.",
            "환기설비 점검 기록, 가스 재측정 기록, 작업재개 승인 기록이 중요합니다.",
            "경보나 설비 이상 후 조치가 지연되면 관리상 미흡으로 지적될 수 있습니다.",
        ],
        GAS_DETECTOR_INTENT: [
            "가스측정기 경보나 측정값 이상을 확인하고도 작업을 계속하면 현장 통제 여부가 쟁점이 될 수 있습니다.",
            "검교정, 알람 확인, 측정자와 측정값 기록이 중요합니다.",
            "측정장비 이상을 방치하면 관리상 미흡으로 볼 수 있습니다.",
        ],
        BLASTING_MISFIRE_INTENT: [
            "불발공 의심 상황에서 출입통제 없이 작업을 계속하면 중대한 위험으로 이어질 수 있습니다.",
            "발파 후 점검 기록, 통제 기록, 담당자 확인 기록이 중요합니다.",
            "미확인 상태에서 재작업하면 작업절차 미준수로 지적될 수 있습니다.",
        ],
        ROOF_FALL_INTENT: [
            "천반 균열, 부석, 지보 이상을 확인하고도 작업을 계속하면 작업중지 판단이 쟁점이 될 수 있습니다.",
            "지보 점검, 출입통제, 보강조치 기록이 중요합니다.",
            "반복되는 낙반 징후를 방치하면 위험성평가와 개선조치 미흡으로 지적될 수 있습니다.",
        ],
        ELECTRICAL_SAFETY_INTENT: [
            "전원 차단, 잠금표지, 접지 확인 없이 전기작업을 하면 감전 사고 시 관리상 미흡이 문제될 수 있습니다.",
            "전기설비 점검 기록과 작업허가 기록이 중요합니다.",
            "임시배선이나 손상 케이블을 방치하면 사고 예방조치 미흡으로 볼 수 있습니다.",
        ],
        EQUIPMENT_TRANSPORT_INTENT: [
            "후진, 협착, 충돌 위험 작업에서 신호수나 출입통제가 없으면 안전조치 미흡이 쟁점이 될 수 있습니다.",
            "장비 점검표, 작업반경 통제, 신호수 배치 기록이 중요합니다.",
            "사각지대 위험을 알면서도 개선하지 않으면 관리 미흡으로 지적될 수 있습니다.",
        ],
        BACKING_SIGNAL_INTENT: [
            "후진, 협착, 충돌 위험 작업에서 신호수나 출입통제가 없으면 안전조치 미흡이 쟁점이 될 수 있습니다.",
            "장비 점검표, 작업반경 통제, 신호수 배치 기록이 중요합니다.",
            "사각지대 위험을 알면서도 개선하지 않으면 관리 미흡으로 지적될 수 있습니다.",
        ],
        CONVEYOR_ROTATING_INTENT: [
            "회전체·컨베이어 주변에서 방호덮개나 접근통제가 없으면 끼임 사고 위험이 커질 수 있습니다.",
            "정비 전 전원 차단과 잠금표지 기록이 중요합니다.",
            "가동 중 청소·점검을 허용하면 절차 미흡으로 지적될 수 있습니다.",
        ],
        FIRE_EMERGENCY_INTENT: [
            "화기작업 전 가연물 제거, 소화기 비치, 화재감시가 없으면 화재 발생 시 예방조치가 쟁점이 될 수 있습니다.",
            "소화기 점검 기록과 화기작업 허가 기록이 중요합니다.",
            "불꽃 비산 위험을 방치하면 관리상 미흡으로 지적될 수 있습니다.",
        ],
        HOT_WORK_INTENT: [
            "화기작업 전 가연물 제거, 소화기 비치, 화재감시가 없으면 화재 발생 시 예방조치가 쟁점이 될 수 있습니다.",
            "소화기 점검 기록과 화기작업 허가 기록이 중요합니다.",
            "불꽃 비산 위험을 방치하면 관리상 미흡으로 지적될 수 있습니다.",
        ],
        PREWORK_TBM_INTENT: [
            "작업 전 위험요인을 공유하지 않으면 사고 발생 시 교육·위험성 전달 여부가 쟁점이 될 수 있습니다.",
            "TBM 기록, 참석자 서명, 위험요인 공유 내용이 중요합니다.",
            "신규 작업자나 협력업체 작업자는 별도 교육 기록을 남기는 것이 좋습니다.",
        ],
        NEW_WORKER_TRAINING_INTENT: [
            "작업 전 위험요인을 공유하지 않으면 사고 발생 시 교육·위험성 전달 여부가 쟁점이 될 수 있습니다.",
            "교육 기록, 참석자 서명, 작업절차 안내 내용이 중요합니다.",
            "신규 작업자나 협력업체 작업자는 별도 교육 기록을 남기는 것이 좋습니다.",
        ],
        CONTRACTOR_PERMIT_INTENT: [
            "협력업체 작업 전 위험요인을 공유하지 않으면 사고 발생 시 도급 작업 관리 여부가 쟁점이 될 수 있습니다.",
            "작업허가서, 사전교육, 출입관리, 위험요인 공유 기록이 중요합니다.",
            "작업 범위와 책임자를 불명확하게 두면 관리상 미흡으로 지적될 수 있습니다.",
        ],
        LAW_INTENT: [
            "실제 조치를 했더라도 기록이 없으면 이행 여부를 입증하기 어려울 수 있습니다.",
            "위험성평가, 작업중지 기록, 개선조치 사진, 교육기록, 점검표를 함께 관리해야 합니다.",
            "AI 답변이나 체크리스트만으로 법적 책임이 면제되는 것은 아닙니다.",
        ],
        DOCUMENT_EVIDENCE_INTENT: [
            "실제 조치를 했더라도 기록이 없으면 이행 여부를 입증하기 어려울 수 있습니다.",
            "위험성평가, 작업중지 기록, 개선조치 사진, 교육기록, 점검표를 함께 관리해야 합니다.",
            "AI 답변이나 체크리스트만으로 법적 책임이 면제되는 것은 아닙니다.",
        ],
        KRAS_INTENT: [
            "위험성평가를 형식적으로 작성하고 개선조치를 하지 않은 경우, 중대재해 발생 시 관리상 미흡으로 지적될 수 있습니다.",
            "감소대책 이행 여부와 개선조치 완료 사진, 담당자 확인 기록이 중요합니다.",
            "AI가 만든 초안만으로 위험성평가 이행이 완료되는 것은 아닙니다.",
        ],
    }
    if intent in point_map:
        return point_map[intent][:3]

    categories = {
        str(case.get("category", ""))
        for case in (reference_cases or [])
        if isinstance(case, dict)
    }
    if "침수/배수" in categories:
        return [
            "통로 물 고임이나 침수를 방치하면 미끄럼, 감전, 장비 이동 위험 관리 여부가 쟁점이 될 수 있습니다.",
            "배수 조치, 통행 제한, 전기설비 접촉 위험 확인 기록이 중요합니다.",
            "반복적인 누수나 배수 불량을 방치하면 개선조치 미흡으로 지적될 수 있습니다.",
        ]
    if "대피로/표지" in categories:
        return [
            "대피로, 비상구, 경고표지가 확보되지 않으면 비상 시 현장 통제 여부가 쟁점이 될 수 있습니다.",
            "통로 확보, 출입통제, 표지 점검 기록이 중요합니다.",
            "위험구역 표시가 불명확하면 관리상 미흡으로 지적될 수 있습니다.",
        ]
    return [
        "위험요인을 알고도 필요한 통제조치를 하지 않으면 사고 발생 시 관리상 미흡으로 지적될 수 있습니다.",
        "점검, 교육, 작업중지, 개선조치 기록을 함께 남기는 것이 중요합니다.",
        "중대재해처벌법 위반 여부는 사고 경위와 사업장 조치, 법령 해석에 따라 달라질 수 있습니다.",
    ]


def format_official_cases_for_prompt(official_cases: list[dict[str, Any]]) -> str:
    if not official_cases:
        return "현재 질문과 직접 관련된 공식 사고사례 검색 결과 없음."
    blocks: list[str] = []
    for case in official_cases[:2]:
        public_tier = verified_review.effective_public_case_tier(case)
        status_label = {
            verified_review.VERIFIED_PUBLIC_TIER: "원문 대조 검증 완료",
            verified_review.AUTO_SCREENED_PUBLIC_TIER: "엄격 자동 품질검사 통과",
            verified_review.TEXT_SAFE_FALLBACK_TIER: "문자 품질검사 통과",
        }.get(public_tier, "공개 제외")
        relation_label = {
            "direct": "직접 관련",
            "analogous": "유사 위험",
            "broad_family": "같은 위험군",
        }.get(str(case.get("relation_type", "")), "관련성 미확인")
        summary = make_preview(
            clean_text(str(case.get("display_accident_summary", ""))),
            350,
        )
        prevention = make_preview(
            clean_text(str(case.get("display_prevention_summary", ""))),
            220,
        )
        blocks.append(
            "\n".join(
                [
                    f"[공식 사고사례] {case.get('case_id', '사례 ID 없음')}",
                    f"- 검증 상태: {status_label}",
                    f"- 질문 관련성: {relation_label}",
                    f"- 사고 유형: {case.get('accident_type', '정보 없음')}",
                    f"- 사고 개요: {summary or '정보 없음'}",
                    f"- 예방사항: {prevention or '원문에 별도 예방사항 없음'}",
                    f"- 출처: {case.get('source_document', '출처 정보 없음')} / {case.get('page_start', '페이지 정보 없음')}쪽",
                ]
            )
        )
    return "\n\n".join(blocks)


def official_case_warning_points(
    official_cases: list[dict[str, Any]],
    intent: str,
    reference_cases: list[dict[str, Any]] | None = None,
) -> list[dict[str, str]]:
    points: list[dict[str, str]] = []
    safe_cases = [
        verified_review.sanitize_display_case(case)
        for case in official_cases
        if verified_review.is_public_display_safe_case(case)
    ]
    priority_steps = (
        (verified_review.VERIFIED_PUBLIC_TIER, "display_prevention_summary"),
        (verified_review.AUTO_SCREENED_PUBLIC_TIER, "display_prevention_summary"),
        (verified_review.TEXT_SAFE_FALLBACK_TIER, "display_prevention_summary"),
        (verified_review.VERIFIED_PUBLIC_TIER, "display_cause_summary"),
        (verified_review.AUTO_SCREENED_PUBLIC_TIER, "display_cause_summary"),
        (verified_review.TEXT_SAFE_FALLBACK_TIER, "display_cause_summary"),
    )
    for public_tier, field in priority_steps:
        for case in safe_cases:
            if verified_review.effective_public_case_tier(case) != public_tier:
                continue
            value = make_preview(
                clean_text(str(case.get(field, ""))),
                220,
            )
            if (
                value
                and not verified_review.detect_corrupted_ocr_text(value)
                and all(item["text"] != value for item in points)
            ):
                source_label = (
                    "원문 대조 검증 사례 기반"
                    if public_tier == verified_review.VERIFIED_PUBLIC_TIER
                    else (
                        "엄격 자동 품질검사 통과 사례 기반"
                        if public_tier == verified_review.AUTO_SCREENED_PUBLIC_TIER
                        else "문자 품질검사 통과 사례 기반"
                    )
                )
                points.append({"text": value, "source": source_label})
            if len(points) >= 3:
                return points
    for point in case_warning_points_for_intent(intent, reference_cases)[:3]:
        points.append({"text": point, "source": "기존 안전규칙 기반"})
    return points[:3]


def summarize_official_cases_for_history(official_cases: list[dict[str, Any]]) -> dict[str, list[str]]:
    return {
        "official_case_ids": [
            str(case.get("case_id", "")).strip()
            for case in official_cases
            if str(case.get("case_id", "")).strip()
        ],
        "official_case_titles": [
            str(case.get("source_document", "")).strip()
            for case in official_cases
            if str(case.get("source_document", "")).strip()
        ],
        "official_case_sources": [
            f"{case.get('source_document', '출처 정보 없음')} / {case.get('page_start', '페이지 정보 없음')}쪽 / {case.get('case_id', '사례 ID 없음')}"
            for case in official_cases
        ],
    }


def format_reference_cases_for_prompt(intent: str, reference_cases: list[dict[str, Any]]) -> str:
    if not reference_cases:
        return "관련 주의 포인트 없음."
    points = case_warning_points_for_intent(intent, reference_cases)
    source_labels = []
    for case in reference_cases[:2]:
        case_id = str(case.get("case_id", "CASE"))
        label = str(case.get("reliability_label", "참고자료"))
        source_labels.append(f"{case_id}({label})")
    lines = [
        "아래 내용은 공식 법령 판단 근거가 아니라 유사 위험을 이해하기 위한 참고 예시입니다.",
        f"참고 사례 데이터: {', '.join(source_labels)}",
    ]
    lines.extend(f"- {point}" for point in points[:3])
    lines.append("주의: 실제 위반 여부는 사고 경위, 사업장 조치, 법령 해석에 따라 달라질 수 있습니다.")
    return "\n".join(lines)

def summarize_reference_cases_for_history(reference_cases: list[dict[str, Any]]) -> list[str]:
    summaries = []
    for case in reference_cases:
        title = str(case.get("title", "제목 없음"))
        case_id = str(case.get("case_id", "CASE"))
        label = str(case.get("reliability_label", "참고자료"))
        summaries.append(f"참고 사례: {case_id} / {title} / {label}")
    return summaries


def render_official_siren_case_card(case: dict[str, Any]) -> None:
    verification_status = str(case.get("verification_status", ""))
    public_tier = verified_review.effective_public_case_tier(case)
    if (
        verification_status != "verified"
        and public_tier == verified_review.HIDDEN_PUBLIC_TIER
    ):
        return
    case = verified_review.sanitize_display_case(case)
    if not verified_review.is_public_display_safe_case(case):
        return
    accident_type = str(case.get("accident_type", "") or "정보 없음")
    accident_date = str(
        case.get("accident_date")
        or case.get("accident_year")
        or "정보 없음"
    )
    industry = str(case.get("industry", "") or "정보 없음")
    summary = make_preview(
        clean_text(str(case.get("display_accident_summary", ""))),
        450,
    )
    cause = make_preview(
        clean_text(
            str(
                case.get("display_cause_summary")
                or "공식 원문에서 별도 원인을 확인하지 못했습니다."
            )
        ),
        360,
    )
    prevention = make_preview(
        clean_text(
            str(
                case.get("display_prevention_summary")
                or "공식 원문에서 별도 예방사항을 확인하지 못했습니다."
            )
        ),
        360,
    )
    relation_type = str(case.get("relation_type", ""))
    relation_label = {
        "direct": "직접 관련 사례",
        "analogous": "유사 위험 사례",
        "broad_family": "같은 위험군의 참고 사례",
    }.get(relation_type, "관련성 확인 사례")
    source_document = str(case.get("source_document", "") or "출처 정보 없음")
    source_period = str(case.get("source_period", "") or "기간 정보 없음")
    page_start = str(
        case.get("original_page_number")
        or case.get("page_start")
        or "페이지 정보 없음"
    )
    case_id = str(case.get("case_id", "") or "사례 ID 없음")
    with st.container(border=True):
        if public_tier == verified_review.VERIFIED_PUBLIC_TIER:
            st.success("✓ 원본 PDF와 대조하여 내용 검증이 완료된 공식 사고사례입니다.")
        elif public_tier == verified_review.AUTO_SCREENED_PUBLIC_TIER:
            st.info(
                "엄격 자동 품질검사 통과 · 공식 원문에서 자동 추출한 뒤 문자 깨짐·혼합 문장·"
                "출처 정보에 대한 품질검사를 통과한 사례입니다. "
                "사람의 원문 대조 검수는 아직 완료되지 않았습니다."
            )
        else:
            st.info(
                "문자 품질검사 통과 · 공식 자료에서 추출한 사례로, 화면 문장과 출처 정보에 대한 "
                "문자 품질검사를 통과했습니다. 날짜·업종 등 일부 metadata는 아직 원문 대조 검수가 "
                "완료되지 않았을 수 있습니다."
            )
        st.caption(f"관련성: {relation_label}")
        if relation_type in {"analogous", "broad_family"}:
            st.warning(
                "동일한 작업이나 설비의 사고가 아니라, 사고 발생 원리와 위험요인이 "
                "유사한 공식 사례입니다. 같은 위험 유형의 참고 사례일 수 있습니다."
            )
        header_cols = st.columns(3)
        header_cols[0].markdown(f"**사고 유형**  \n{accident_type}")
        header_cols[1].markdown(f"**발생일·연도**  \n{accident_date}")
        header_cols[2].markdown(f"**업종**  \n{industry}")
        st.markdown(f"**사고 개요**  \n{summary or '원문에서 확인된 사고 개요가 없습니다.'}")
        st.markdown(f"**주요 원인**  \n{cause}")
        st.markdown(f"**예방 및 주의사항**  \n{prevention}")
        st.caption(
            f"출처: {source_document} · "
            f"기간: {source_period} · 페이지: {page_start} · case_id: {case_id}"
        )
        original_image = ROOT_DIR / str(case.get("original_page_image", ""))
        if public_tier == verified_review.VERIFIED_PUBLIC_TIER and original_image.is_file():
            with st.expander("원문 이미지 확인", expanded=False):
                st.image(str(original_image), caption=f"{source_document} · {page_start}쪽")


def render_official_siren_cases(
    official_cases: list[dict[str, Any]],
    diagnostic: dict[str, Any] | None = None,
) -> None:
    public_cases = [
        case
        for case in official_cases
        if verified_review.effective_public_case_tier(case)
        != verified_review.HIDDEN_PUBLIC_TIER
        and verified_review.is_public_display_safe_case(case)
    ]
    st.caption(
        "원본 PDF 대조 검증 사례와 엄격 자동검사 사례를 우선하고, 없으면 문자 품질검사를 통과한 "
        "같은 위험군의 공식 사고사례를 표시합니다. "
        "사고의 발생 상황과 예방사항을 이해하기 위한 참고자료이며, "
        "개별 광산의 법령 위반 여부를 확정하는 근거는 아닙니다."
    )
    if globals().get("is_admin_mode", False) and isinstance(diagnostic, dict):
        verified_db_label = {
            "verified_not_created_zero_cases": "검증 완료 사례 0건",
            "verified_db_unavailable": "실제 verified DB 연결 오류",
            "verified_collection_conflict": "verified collection 설정 충돌",
            "ready": "verified DB 준비됨",
        }.get(str(diagnostic.get("verified_db_status", "")), "verified DB 상태 확인 필요")
        auto_db_label = {
            "auto_screened_not_created_zero_cases": "자동 품질검사 통과 사례 0건",
            "auto_screened_db_unavailable": "실제 auto_screened DB 연결 오류",
            "auto_screened_collection_conflict": "auto_screened collection 설정 충돌",
            "ready": "auto_screened DB 준비됨",
        }.get(str(diagnostic.get("auto_screened_db_status", "")), "auto_screened DB 상태 확인 필요")
        text_safe_db_label = {
            "text_safe_not_created_zero_cases": "문자 안전 사례 0건으로 미생성",
            "text_safe_db_unavailable": "문자 안전 사례 DB 연결 오류",
            "text_safe_collection_conflict": "문자 안전 collection 설정 충돌",
            "ready": "문자 안전 사례 DB 준비됨",
        }.get(str(diagnostic.get("text_safe_db_status", "")), "문자 안전 사례 DB 상태 확인 필요")
        st.caption(
            "관리자 진단 · "
            f"{verified_db_label} · {auto_db_label} · {text_safe_db_label} · "
            f"검색 결과: {diagnostic.get('result_count', 0)}건"
        )
        with st.expander("검색 단계별 진단 보기", expanded=False):
            diagnostic_status_label = {
                "ready": "검색 정상",
                "no_search_results": "관련 결과 0건",
                "no_quality_screened_cases": "공개 가능 사례 DB 미생성",
                "search_failed": "사례 검색 연결 오류",
                "db_unavailable": "사례 DB 연결 오류",
            }.get(str(diagnostic.get("status", "")), "검색 상태 확인 필요")
            st.write(f"검색 상태: {diagnostic_status_label}")
            st.write(f"사용 collection: {diagnostic.get('active_collection_name') or '없음'}")
            st.write(f"collection 저장 수: {diagnostic.get('collection_total_count', 0)}건")
            st.write(f"원시 검색 결과: {diagnostic.get('raw_query_result_count', 0)}건")
            st.write(f"중복 제거 후: {diagnostic.get('after_duplicate_filter_count', 0)}건")
            st.write(f"문자 안전검사 후: {diagnostic.get('after_text_safety_filter_count', 0)}건")
            st.write(f"공개 등급검사 후: {diagnostic.get('after_verification_tier_filter_count', 0)}건")
            st.write(f"위험 관련성검사 후: {diagnostic.get('after_relation_filter_count', 0)}건")
            st.write(f"최종 표시 결과: {diagnostic.get('final_result_count', 0)}건")
            st.write(
                "관련성 후보: "
                f"직접 {diagnostic.get('direct_candidate_count', 0)}건 · "
                f"유사 {diagnostic.get('analogous_candidate_count', 0)}건 · "
                f"같은 위험군 {diagnostic.get('broad_family_candidate_count', 0)}건"
            )
            failed_sources = diagnostic.get("search_error_sources", [])
            if isinstance(failed_sources, list) and failed_sources:
                failed_labels = {
                    "verified": "검증 완료 사례",
                    verified_review.AUTO_SCREENED_PUBLIC_TIER: "엄격 자동검사 사례",
                    verified_review.TEXT_SAFE_FALLBACK_TIER: "문자 안전 사례",
                }
                st.write(
                    "검색 연결 실패 대상: "
                    + ", ".join(
                        failed_labels.get(str(source), "사례 DB")
                        for source in failed_sources
                    )
                )
            removal_reasons = diagnostic.get("removal_reasons", {})
            if isinstance(removal_reasons, dict) and removal_reasons:
                st.write("제거 사유별 수:")
                reason_labels = {
                    "missing_case_id": "사례 ID 누락",
                    "duplicate_case_id_in_collection": "동일 사례 중복",
                    "display_text_safety_failed": "문자 안전검사 미통과",
                    "not_official_case": "공식 사례 표시 누락",
                    "verified_tier_mismatch": "검증 완료 등급 불일치",
                    "auto_screened_tier_mismatch": "엄격 자동검사 등급 불일치",
                    "text_safe_tier_mismatch": "문자 안전 등급 불일치",
                    "mine_relevance_not_eligible": "광산 위험 관련성 기준 미충족",
                    "public_tier_safety_failed": "공개 등급 안전조건 미충족",
                    "relation_mismatch": "질문 위험 유형과 불일치",
                }
                for reason, count in sorted(removal_reasons.items()):
                    st.write(f"- {reason_labels.get(reason, '기타 안전조건 미충족')}: {count}건")
    if not public_cases:
        st.info("현재 질문과 관련해 품질검사를 통과한 공식 사고사례가 없습니다.")
        return
    render_official_siren_case_card(public_cases[0])
    for index, case in enumerate(public_cases[1:], start=2):
        label = str(case.get("accident_type", "") or "추가 공식 사고사례")
        with st.expander(f"추가 사례 {index - 1} · {label}", expanded=False):
            render_official_siren_case_card(case)


def admin_case_display_preview(case: dict[str, Any]) -> str:
    """관리자 검수 화면에서 정제본을 우선하고 OCR 원문 대조를 허용합니다."""
    return str(case.get("display_accident_summary") or case.get("accident_summary", ""))


def render_verified_official_case_review_page() -> None:
    st.header("공식 사고사례 검수")
    st.caption(
        "기존 OCR 사례는 모두 미검증 상태입니다. 원본 카드 이미지와 각 필드를 직접 비교한 뒤 "
        "관리자가 명시적으로 승인한 사례는 verified DB에 반영되고 자동검사 통과 사례보다 우선 표시됩니다. "
        "auto_screened 사례도 이 화면에서 원문과 대조해 verified로 승격할 수 있습니다."
    )
    try:
        records = verified_review.initialize_review_store()
    except verified_review.ReviewWorkflowBlocked as error:
        st.error(str(error))
        return
    counts = verified_review.review_status_counts(records)
    metric_columns = st.columns(6)
    metric_values = (
        ("전체 후보", counts["total"]),
        ("미검증", counts["unverified"]),
        ("검증 완료", counts["verified"]),
        ("자동 품질검사", counts["auto_screened"]),
        ("사용 제외", counts["rejected"]),
        ("수동검토", counts["manual_review"]),
    )
    for column, (label, value) in zip(metric_columns, metric_values):
        column.metric(label, f"{value}건")

    priority_records = verified_review.priority_review_candidates(records)
    priority_ids = {str(record.get("case_id", "")) for record in priority_records}
    st.info(
        f"발표 질문 우선 검수 대상: {len(priority_records)}건 · "
        "컨베이어, 차량 후진, 낙반, 전기 유형별 최대 3건"
    )
    if st.button("우선 검수 원본 이미지 준비", key="prepare_verified_case_images"):
        try:
            result = verified_review.generate_priority_card_images(records)
            st.success(
                f"원본 카드 이미지 {result.get('generated_image_count', 0)}건을 준비했습니다. "
                "자동 검증 완료 처리는 하지 않았습니다."
            )
            st.rerun()
        except verified_review.ReviewWorkflowBlocked as error:
            st.error(str(error))

    show_priority_only = st.checkbox(
        "발표용 우선 검수 대상만 보기",
        value=True,
        key="verified_case_priority_only",
    )
    visible_records = (
        [record for record in records if str(record.get("case_id", "")) in priority_ids]
        if show_priority_only
        else records
    )
    if not visible_records:
        st.warning("검수 가능한 공식 사고사례가 없습니다.")
        return
    record_map = {str(record.get("case_id", "")): record for record in visible_records}
    selected_case_id = st.selectbox(
        "검수할 case_id",
        list(record_map),
        format_func=lambda case_id: (
            f"{case_id} · {record_map[case_id].get('accident_type') or '유형 미확인'} · "
            f"{record_map[case_id].get('verification_status', 'unverified')}"
        ),
    )
    selected = record_map[selected_case_id]
    current_status = str(selected.get("verification_status", "unverified"))
    status_labels = {
        "unverified": "미검증",
        "verified": "검증 완료",
        "auto_screened": "자동 품질검사 통과",
        "rejected": "사용 제외",
        "manual_review": "수동 검토",
    }
    st.caption(f"현재 상태: {status_labels.get(current_status, current_status)}")

    with st.form("verified_official_case_review_form"):
        left_column, right_column = st.columns([1.05, 1])
        with left_column:
            image_path = ROOT_DIR / str(selected.get("original_page_image", ""))
            if image_path.is_file():
                st.image(
                    str(image_path),
                    caption=(
                        f"{selected.get('source_document', '문서명 없음')} · "
                        f"{selected.get('original_page_number', '페이지 없음')}쪽"
                    ),
                    use_container_width=True,
                )
            else:
                st.warning("원본 사고 카드 이미지가 아직 준비되지 않았습니다.")
            st.write(f"**문서명:** {selected.get('source_document') or '정보 없음'}")
            st.write(f"**페이지:** {selected.get('original_page_number') or '정보 없음'}")
            st.write(f"**case_id:** {selected_case_id}")
            admin_display_preview = admin_case_display_preview(selected)
            with st.expander("원문 OCR 내용 보기", expanded=False):
                st.text(
                    str(
                        selected.get("layout_ocr_text")
                        or selected.get("full_accident_summary")
                        or admin_display_preview
                        or "OCR 원문 없음"
                    )
                )
                st.caption("이 원문은 관리자 대조용이며 공개 화면에는 표시되지 않습니다.")

        with right_column:
            accident_date = st.text_input("발생일", value=str(selected.get("accident_date", "")))
            industry = st.text_input("업종", value=str(selected.get("industry", "")))
            accident_type = st.text_input("사고 유형", value=str(selected.get("accident_type", "")))
            accident_summary = st.text_area(
                "사고 개요",
                value=str(selected.get("accident_summary", "")),
                height=150,
            )
            cause_summary = st.text_area(
                "원인",
                value=str(selected.get("cause_summary", "")),
                height=90,
            )
            prevention_summary = st.text_area(
                "예방사항",
                value=str(selected.get("prevention_summary", "")),
                height=90,
            )
            verification_note = st.text_area(
                "검증 메모",
                value=str(selected.get("verification_note", "")),
                height=80,
            )
            rejection_reason = st.text_input(
                "사용 제외 사유",
                value=str(selected.get("rejection_reason", "")),
            )

        st.markdown("#### 검증 완료 전 필수 확인")
        existing_checks = set(selected.get("verified_fields", []))
        check_labels = (
            ("summary_matches_source", "원본 이미지와 사고 개요가 일치함"),
            ("date_matches_source", "날짜가 일치함"),
            ("industry_matches_source", "업종이 일치함"),
            ("accident_type_matches_source", "사고 유형이 일치함"),
            ("single_accident_only", "다른 사고 내용이 섞이지 않음"),
            ("no_ocr_noise", "의미 없는 OCR 문자열이 없음"),
        )
        checked: dict[str, bool] = {}
        check_columns = st.columns(2)
        for index, (field, label) in enumerate(check_labels):
            with check_columns[index % 2]:
                checked[field] = st.checkbox(
                    label,
                    value=field in existing_checks,
                    key=f"review_check_{selected_case_id}_{field}",
                )

        action_columns = st.columns(4)
        save_clicked = action_columns[0].form_submit_button("수정 저장", use_container_width=True)
        manual_clicked = action_columns[1].form_submit_button("수동 검토", use_container_width=True)
        reject_clicked = action_columns[2].form_submit_button("사용 제외", use_container_width=True)
        verify_clicked = action_columns[3].form_submit_button("검증 완료", use_container_width=True)

    if any((save_clicked, manual_clicked, reject_clicked, verify_clicked)):
        target_status = current_status
        if manual_clicked:
            target_status = "manual_review"
        elif reject_clicked:
            target_status = "rejected"
        elif verify_clicked:
            target_status = "verified"
        edited_fields = {
            "accident_date": accident_date,
            "industry": industry,
            "accident_type": accident_type,
            "accident_summary": accident_summary,
            "cause_summary": cause_summary,
            "prevention_summary": prevention_summary,
        }
        verified_fields = [field for field, is_checked in checked.items() if is_checked]
        try:
            verified_review.save_review_update(
                selected_case_id,
                edited_fields,
                target_status,
                verification_note,
                verified_fields,
                rejection_reason,
            )
            load_official_case_collection.clear()
            load_auto_screened_case_collection.clear()
            load_text_safe_case_collection.clear()
            gc.collect()
            updated_records = verified_review.load_review_records()
            db_result = verified_review.rebuild_verified_case_db(updated_records)
            auto_db_result = verified_review.rebuild_auto_screened_case_db(updated_records)
            text_safe_db_result = verified_review.rebuild_text_safe_case_db(updated_records)
            load_official_case_collection.clear()
            load_auto_screened_case_collection.clear()
            load_text_safe_case_collection.clear()
            st.success(
                f"검수 상태를 저장했습니다. verified DB 상태: "
                f"{db_result.get('status', '확인 필요')} · auto_screened DB 상태: "
                f"{auto_db_result.get('status', '확인 필요')} · text_safe DB 상태: "
                f"{text_safe_db_result.get('status', '확인 필요')}"
            )
            st.rerun()
        except verified_review.ReviewWorkflowBlocked as error:
            st.error(str(error))
        except OSError:
            st.error(
                "사례 DB 파일이 다른 프로세스에서 사용 중이어서 갱신하지 못했습니다. "
                "실행 중인 검색을 마친 뒤 다시 시도해 주세요."
            )


def render_latest_reference_cases(
    reference_cases: list[dict[str, Any]],
    intent: str | None = None,
    official_cases: list[dict[str, Any]] | None = None,
) -> None:
    points = official_case_warning_points(
        official_cases or [],
        intent or GENERAL_MINE_SAFETY_INTENT,
        reference_cases,
    )
    st.caption(
        "공식 사례의 예방사항·원인을 우선하며, 없으면 기존 안전규칙을 사용합니다. "
        "핵심 주의사항도 법적 판단 근거는 아닙니다."
    )
    if not points:
        st.info("현재 질문에 표시할 핵심 주의사항이 없습니다.")
        return
    for item in points[:3]:
        st.markdown(
            f"- {escape(item.get('text', ''))}  \n"
            f"  <span style='color:#64748b;font-size:0.82rem;'>출처 성격: {escape(item.get('source', '참고자료'))}</span>",
            unsafe_allow_html=True,
        )
    st.warning(
        "※ 공식 사고사례와 핵심 주의사항은 사고 상황과 예방사항을 이해하기 위한 참고자료입니다. "
        "법령 위반 여부와 처벌 여부는 공식 법령·지침 근거와 실제 현장 조치를 별도로 확인해야 합니다."
    )



def clean_html_text(text: str) -> str:
    try:
        import re
        from html import unescape

        cleaned = re.sub(r"<[^>]+>", " ", str(text or ""))
        return clean_text(unescape(cleaned))
    except Exception:
        return clean_text(str(text or ""))


def live_case_search_configured() -> bool:
    enabled = os.getenv("ENABLE_LIVE_CASE_SEARCH", "").strip().lower() == "true"
    provider = os.getenv("LIVE_CASE_SEARCH_PROVIDER", "naver").strip().lower()
    client_id = os.getenv("NAVER_CLIENT_ID", "").strip()
    client_secret = os.getenv("NAVER_CLIENT_SECRET", "").strip()
    return enabled and provider == "naver" and bool(client_id) and bool(client_secret)


def build_live_case_search_query(question: str, intent: str) -> str:
    query_map = {
        PPE_GENERAL_INTENT: "안전모 안전화 보호구 산업재해 사고 사례",
        VENTILATION_GAS_INTENT: "메탄가스 질식 폭발 환기 산업재해 사고 사례",
        VENTILATION_EQUIPMENT_INTENT: "환기팬 환기설비 질식 산업재해 사고 사례",
        GAS_DETECTOR_INTENT: "가스측정기 경보 질식 폭발 산업재해 사고 사례",
        BLASTING_MISFIRE_INTENT: "발파 불발공 폭약 사고 사례 광산",
        ROOF_FALL_INTENT: "광산 낙반 붕락 천반 사고 사례",
        DUST_RESPIRATORY_INTENT: "분진 호흡보호 방진마스크 산업재해 사례",
        ELECTRICAL_SAFETY_INTENT: "감전 누전 전기작업 산업재해 사고 사례",
        EQUIPMENT_TRANSPORT_INTENT: "장비 후진 협착 충돌 산업재해 사고 사례",
        BACKING_SIGNAL_INTENT: "장비 후진 신호수 협착 충돌 산업재해 사고 사례",
        FIRE_EMERGENCY_INTENT: "화재 소화기 화기작업 산업재해 사고 사례",
        HOT_WORK_INTENT: "용접 화기작업 화재 산업재해 사고 사례",
        CONVEYOR_ROTATING_INTENT: "컨베이어 끼임 협착 산업재해 사고 사례",
        HEIGHT_WORK_INTENT: "고소작업 추락 산업재해 사고 사례",
        FLOOD_DRAINAGE_INTENT: "침수 배수 갱내수 광산 사고 사례",
        LAW_INTENT: "중대재해처벌법 사고 사례 안전보건관리체계 증빙자료",
        DOCUMENT_EVIDENCE_INTENT: "중대재해처벌법 사고 사례 안전보건관리체계 증빙자료",
        KRAS_INTENT: "위험성평가 개선조치 산업재해 사고 사례",
    }
    if intent in query_map:
        return query_map[intent]
    question_terms = " ".join(clean_text(question).split()[:5])
    return clean_text(f"{question_terms} 산업재해 사고 사례")


def live_news_relevance_label(title: str, description: str) -> str:
    include_keywords = [
        "사고", "재해", "중대재해", "산업재해", "안전", "작업장", "광산",
        "협착", "추락", "감전", "질식", "화재", "폭발", "낙반", "컨베이어",
        "끼임", "장비", "환기",
    ]
    combined = f"{title} {description}"
    if any(keyword in combined for keyword in include_keywords):
        return "뉴스검색 참고"
    return "관련성이 낮을 수 있음"


def normalize_naver_news_item(item: dict[str, Any]) -> dict[str, Any] | None:
    title = clean_html_text(item.get("title", ""))
    description = clean_html_text(item.get("description", ""))
    combined = f"{title} {description}"
    exclude_keywords = ["광고", "채용", "주식", "연예", "맛집", "부동산", "쇼핑"]
    if not title or any(keyword in combined for keyword in exclude_keywords):
        return None
    return {
        "title": title,
        "summary": description,
        "pub_date": clean_html_text(item.get("pubDate", "")),
        "link": str(item.get("originallink") or item.get("link") or ""),
        "reliability_label": live_news_relevance_label(title, description),
        "source_name": "네이버 뉴스 검색",
    }


@st.cache_data(ttl=LIVE_CASE_SEARCH_TTL_SECONDS, show_spinner=False)
def search_naver_news_cases(query: str, display: int = 5, sort: str = "date") -> list[dict[str, Any]]:
    if not live_case_search_configured():
        return []

    client_id = os.getenv("NAVER_CLIENT_ID", "").strip()
    client_secret = os.getenv("NAVER_CLIENT_SECRET", "").strip()
    params = {"query": query, "display": int(display), "sort": sort}
    headers = {
        "X-Naver-Client-Id": client_id,
        "X-Naver-Client-Secret": client_secret,
    }
    try:
        try:
            import requests

            response = requests.get(
                NAVER_NEWS_API_URL,
                params=params,
                headers=headers,
                timeout=4,
            )
            response.raise_for_status()
            payload = response.json()
        except ImportError:
            from urllib import parse, request

            query_string = parse.urlencode(params)
            req = request.Request(
                f"{NAVER_NEWS_API_URL}?{query_string}",
                headers=headers,
                method="GET",
            )
            with request.urlopen(req, timeout=4) as response:
                payload = json.loads(response.read().decode("utf-8"))
    except Exception:
        return []

    items = payload.get("items", []) if isinstance(payload, dict) else []
    cases = []
    for item in items:
        if not isinstance(item, dict):
            continue
        normalized = normalize_naver_news_item(item)
        if normalized:
            cases.append(normalized)
    cases.sort(key=lambda case: case.get("reliability_label") != "뉴스검색 참고")
    return cases[: int(display)]


def get_live_reference_cases(question: str, intent: str, max_cases: int = 3) -> list[dict[str, Any]]:
    if intent == OUT_OF_SCOPE_INTENT or not live_case_search_configured():
        return []
    query = build_live_case_search_query(question, intent)
    return search_naver_news_cases(query, display=max(max_cases, 5), sort="date")[:max_cases]


def format_live_news_cases_for_prompt(live_news_cases: list[dict[str, Any]]) -> str:
    if not live_news_cases:
        return "관련 뉴스 검색 결과 없음."
    blocks = []
    for idx, case in enumerate(live_news_cases[:3], start=1):
        blocks.append(
            "\n".join(
                [
                    f"[뉴스 참고 {idx}] {case.get('title', '제목 없음')}",
                    f"- 게시일: {case.get('pub_date', '정보 없음')}",
                    f"- 요약: {case.get('summary', '')}",
                    f"- 링크: {case.get('link', '')}",
                    "- 주의: 네이버 뉴스 검색 기반 참고자료이며 공식 법령 판단 근거가 아닙니다.",
                ]
            )
        )
    return "\n\n".join(blocks)


def render_live_news_reference_cases(live_news_cases: list[dict[str, Any]]) -> None:
    if not live_news_cases:
        st.info("현재 표시할 최근 뉴스 검색 결과가 없습니다.")
        return
    st.caption("네이버 뉴스 검색 기반 참고자료이며, 공식 법령 판단 근거가 아닙니다.")
    for case in live_news_cases[:3]:
        title = escape(str(case.get("title", "제목 없음")))
        pub_date = escape(str(case.get("pub_date", "정보 없음")))
        summary = escape(str(case.get("summary", "")))
        label = escape(str(case.get("reliability_label", "뉴스검색 참고")))
        link = str(case.get("link", "") or "")
        link_html = (
            f'<a href="{escape(link)}" target="_blank" rel="noopener noreferrer">기사 보기</a>'
            if link
            else "링크 없음"
        )
        st.markdown(
            f"""
            <div style="border:1px solid #dbe4ef;border-radius:10px;padding:12px 14px;margin:8px 0;background:#ffffff;">
                <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:5px;">
                    <strong style="color:#17324d;">{title}</strong>
                    <span style="background:#eef2ff;color:#3730a3;border-radius:999px;padding:2px 8px;font-size:12px;font-weight:700;">{label}</span>
                </div>
                <div style="color:#64748b;font-size:13px;margin-bottom:7px;">게시일: {pub_date} · 출처: {link_html}</div>
                <div style="color:#334155;line-height:1.6;font-size:0.94rem;">{summary}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    st.warning(
        "※ 이 결과는 네이버 뉴스 검색 기반 참고자료이며, 공식 법령 판단 근거가 아닙니다. "
        "공식 답변 근거는 RAG 근거 문서와 chunk_id를 기준으로 확인해야 합니다."
    )

def source_chunk_labels(results: list[dict[str, Any]]) -> list[str]:
    labels: list[str] = []
    for item in results:
        source = str(item.get("source", "출처 정보 없음"))
        chunk_id = str(item.get("chunk_id", "정보 없음"))
        label = f"{source} / {chunk_id}"
        if label not in labels:
            labels.append(label)
    return labels


def append_conversation_history(row: dict[str, Any]) -> str:
    ensure_feature_dirs()
    history_id = row.get("history_id") or str(uuid.uuid4())
    row["history_id"] = history_id
    with CONVERSATION_HISTORY_PATH.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\n")
    return history_id


def load_conversation_history() -> list[dict[str, Any]]:
    ensure_feature_dirs()
    if not CONVERSATION_HISTORY_PATH.exists():
        CONVERSATION_HISTORY_PATH.write_text("", encoding="utf-8")
        return []
    rows = []
    for line in CONVERSATION_HISTORY_PATH.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            rows.append(json.loads(line))
        except Exception:
            continue
    return rows


def rewrite_conversation_history(rows: list[dict[str, Any]]) -> None:
    ensure_feature_dirs()
    with CONVERSATION_HISTORY_PATH.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def update_conversation_history(history_id: str, updates: dict[str, Any]) -> bool:
    rows = load_conversation_history()
    updated = False
    for row in rows:
        if str(row.get("history_id")) == str(history_id):
            row.update(updates)
            updated = True
            break
    if updated:
        rewrite_conversation_history(rows)
    return updated


def normalize_conversation_history_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        rows = [{"history_id": "", "created_at": "", "question": "", "answer": ""}]
    normalized_rows = []
    for row in rows:
        normalized_rows.append(
            {
                "history_id": row.get("history_id", ""),
                "created_at": row.get("created_at", ""),
                "question": row.get("question", ""),
                "answer": row.get("answer", ""),
                "situation_type": row.get("situation_type", ""),
                "risk_level": row.get("risk_level", ""),
                "답변 모드": normalize_answer_mode_label(row.get("answer_mode", "")),
                "공식 근거 검색 상태": row.get("evidence_label", ""),
                "공식 근거 판정 사유": row.get("evidence_reason", ""),
                "evidence_documents": " | ".join(row.get("evidence_documents", [])),
                "source_chunks": " | ".join(row.get("source_chunks", [])),
                "recommended_evidence_records": " | ".join(row.get("recommended_evidence_records", [])),
                "reference_cases": " | ".join(row.get("reference_cases", [])),
                "공식 사고사례 ID": " | ".join(normalize_history_display_list(row.get("official_case_ids", []))),
                "공식 사고사례 문서": " | ".join(normalize_history_display_list(row.get("official_case_titles", []))),
                "공식 사고사례 출처": " | ".join(normalize_history_display_list(row.get("official_case_sources", []))),
                "user_action_status": row.get("user_action_status", ""),
                "manager": row.get("manager", ""),
                "action_due_date": row.get("action_due_date", ""),
                "memo": row.get("memo", ""),
            }
        )
    return normalized_rows


def build_conversation_history_xlsx_bytes(rows: list[dict[str, Any]]) -> bytes:
    return build_simple_xlsx_bytes("conversation_history", normalize_conversation_history_rows(rows))


def normalize_history_display_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        items = [str(item).strip() for item in value if str(item).strip()]
    else:
        text = str(value).strip()
        if not text:
            return []
        separator = " | " if " | " in text else "," if "," in text else None
        items = [part.strip() for part in text.split(separator)] if separator else [text]
    return [item for item in items if item]


def short_history_label(row: dict[str, Any], limit: int = 52) -> str:
    question = " ".join(str(row.get("question", "")).split())
    if len(question) > limit:
        question = question[: limit - 3] + "..."
    created_at = str(row.get("created_at", ""))[:16]
    situation = str(row.get("situation_type", "상황 미분류") or "상황 미분류")
    return f"{created_at} | {situation} | {question}"


def render_history_list_card(title: str, items: list[str], empty_message: str = "기록 없음") -> None:
    with st.container(border=True):
        st.markdown(f"#### {title}")
        if not items:
            st.caption(empty_message)
            return
        cols = st.columns(2)
        for idx, item in enumerate(items):
            with cols[idx % 2]:
                st.markdown(
                    f"""
                    <div style="background:#f8fafc;border:1px solid #dbe4ef;border-radius:10px;
                                padding:10px 12px;margin:4px 0;color:#17324d;line-height:1.45;">
                        <strong style="color:#2563eb;">•</strong> {escape(item)}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )


def render_history_report_field(title: str, content: str) -> None:
    st.markdown(
        f"""
        <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;
                    padding:14px 16px;margin-bottom:10px;box-shadow:0 1px 2px rgba(15,23,42,0.04);">
            <div style="font-weight:700;color:#17324d;margin-bottom:6px;">{escape(title)}</div>
            <div style="color:#334155;line-height:1.6;white-space:pre-wrap;">{escape(content or "기록 없음")}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_history_report_card(
    selected: dict[str, Any],
    evidence_documents: list[str],
    recommended_records: list[str],
    reference_case_records: list[str] | None = None,
    official_case_records: list[str] | None = None,
) -> None:
    st.subheader("증빙자료 보고서용 보기")
    render_history_report_field("질문", str(selected.get("question", "")))
    c1, c2 = st.columns(2)
    with c1:
        render_history_report_field("상황 유형", str(selected.get("situation_type", "상황 미분류") or "상황 미분류"))
    with c2:
        render_history_report_field("위험도", str(selected.get("risk_level", "검토 필요") or "검토 필요"))
    render_history_report_field(
        "답변 모드",
        normalize_answer_mode_label(selected.get("answer_mode", "기록 없음")) or "기록 없음",
    )
    render_history_report_field(
        "공식 근거 검색 상태",
        str(selected.get("evidence_label", "기록 없음") or "기록 없음"),
    )
    render_history_report_field(
        "공식 근거 판정 사유",
        str(selected.get("evidence_reason", "기록 없음") or "기록 없음"),
    )
    render_history_report_field("근거 문서", ", ".join(evidence_documents) if evidence_documents else "기록 없음")
    render_history_report_field("필요 증빙자료", ", ".join(recommended_records) if recommended_records else "기록 없음")
    render_history_report_field("참고 사례", ", ".join(reference_case_records or []) if reference_case_records else "기록 없음")
    render_history_report_field("공식 사고사례", ", ".join(official_case_records or []) if official_case_records else "기록 없음")
    render_history_report_field("조치 상태", str(selected.get("user_action_status", "미조치") or "미조치"))


def strip_inline_markdown(text: str) -> str:
    return str(text or "").replace("**", "").replace("\\|", "|").strip()


def split_display_items(text: str) -> list[str]:
    cleaned = strip_inline_markdown(text)
    if not cleaned:
        return []
    parts = []
    for raw in cleaned.replace("\r\n", " / ").replace("\n", " / ").split(" / "):
        item = raw.strip(" -•")
        if item:
            parts.append(item)
    return parts or [cleaned]


def kras_risk_badge_html(level: str) -> str:
    styles = {
        "낮음": ("#dcfce7", "#166534", "#86efac"),
        "보통": ("#fef9c3", "#854d0e", "#fde047"),
        "높음": ("#ffedd5", "#9a3412", "#fdba74"),
        "매우 높음": ("#fee2e2", "#991b1b", "#fca5a5"),
    }
    bg, fg, border = styles.get(level, ("#e2e8f0", "#334155", "#cbd5e1"))
    return (
        f'<span style="display:inline-block;background:{bg};color:{fg};border:1px solid {border};'
        'border-radius:999px;padding:2px 9px;margin:0 3px;font-weight:700;white-space:nowrap;">'
        f'{escape(level)}</span>'
    )


def emphasize_risk_terms_html(text: str) -> str:
    cleaned = strip_inline_markdown(text)
    levels = ["매우 높음", "높음", "보통", "낮음"]
    parts: list[str] = []
    index = 0
    while index < len(cleaned):
        matched_level = next(
            (level for level in levels if cleaned.startswith(level, index)),
            None,
        )
        if matched_level:
            parts.append(kras_risk_badge_html(matched_level))
            index += len(matched_level)
        else:
            parts.append(escape(cleaned[index]))
            index += 1
    return "".join(parts)


def measure_label_html(label: str) -> str:
    styles = {
        "제거": ("#dbeafe", "#1d4ed8"),
        "대체": ("#e0e7ff", "#4338ca"),
        "공학적 대책": ("#ccfbf1", "#0f766e"),
        "관리적 대책": ("#fef3c7", "#92400e"),
        "보호구/PPE": ("#fce7f3", "#be185d"),
    }
    bg, fg = styles.get(label, ("#e2e8f0", "#334155"))
    return (
        f'<span style="display:inline-block;background:{bg};color:{fg};border-radius:8px;'
        'padding:2px 8px;margin-right:6px;font-weight:800;white-space:nowrap;">'
        f'{escape(label)}</span>'
    )


def emphasize_measure_item_html(item: str) -> str:
    cleaned = strip_inline_markdown(item)
    for label in ["공학적 대책", "관리적 대책", "보호구/PPE", "제거", "대체"]:
        for marker in [f"{label}:", f"{label}："]:
            if cleaned.startswith(marker):
                detail = cleaned[len(marker):].strip()
                return f"{measure_label_html(label)} {emphasize_risk_terms_html(detail)}"
    return emphasize_risk_terms_html(cleaned)


def render_kras_value_card(label: str, value: str) -> None:
    clean_label = strip_inline_markdown(label)
    items = split_display_items(value)
    if len(items) > 1:
        if clean_label == "위험성 감소대책":
            body = "".join(f"<li>{emphasize_measure_item_html(item)}</li>" for item in items)
        else:
            body = "".join(f"<li>{emphasize_risk_terms_html(item)}</li>" for item in items)
        value_html = f"<ul style='margin:0;padding-left:18px;line-height:1.8;'>{body}</ul>"
    else:
        item = items[0] if items else "기록 필요"
        if clean_label == "위험성 감소대책":
            value_html = f"<div style='line-height:1.8;white-space:pre-wrap;'>{emphasize_measure_item_html(item)}</div>"
        else:
            value_html = f"<div style='line-height:1.8;white-space:pre-wrap;'>{emphasize_risk_terms_html(item)}</div>"
    st.markdown(
        f"""
        <div style="background:#ffffff;border:1px solid #dbe4ef;border-radius:12px;
                    padding:13px 15px;margin-bottom:10px;">
            <div style="font-weight:800;color:#17324d;margin-bottom:8px;">• {escape(clean_label)}</div>
            <div style="color:#334155;">{value_html}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def parse_markdown_table_lines(lines: list[str]) -> tuple[list[str], list[list[str]]]:
    rows = []
    for line in lines:
        stripped = line.strip()
        if not stripped.startswith("|"):
            continue
        cells = [cell.strip() for cell in stripped.strip("|").split("|")]
        if cells and all(set(cell) <= {"-"} for cell in cells if cell):
            continue
        rows.append(cells)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def render_generic_kras_table(headers: list[str], rows: list[list[str]]) -> None:
    if not headers or not rows:
        return
    header_html = "".join(f"<th>{escape(strip_inline_markdown(header))}</th>" for header in headers)
    row_html = []
    for row in rows:
        cells = []
        for idx, cell in enumerate(row):
            value = strip_inline_markdown(cell)
            if " / " in value:
                items = split_display_items(value)
                cell_html = "<ul style='margin:0;padding-left:16px;'>" + "".join(f"<li>{escape(item)}</li>" for item in items) + "</ul>"
            else:
                cell_html = escape(value)
            cells.append(f"<td>{cell_html}</td>")
        row_html.append("<tr>" + "".join(cells) + "</tr>")
    st.markdown(
        (
            "<div class='portal-table-scroll table-wrap'>"
            "<table class='portal-data-table'>"
            f"<thead><tr>{header_html}</tr></thead>"
            f"<tbody>{''.join(row_html)}</tbody></table></div>"
        ),
        unsafe_allow_html=True,
    )


def render_kras_readable_markdown(kras_answer: str) -> None:
    content = strip_kras_title(kras_answer)
    lines = content.splitlines()
    idx = 0
    while idx < len(lines):
        line = lines[idx].strip()
        if not line:
            idx += 1
            continue
        if line.startswith(">"):
            st.info(line.lstrip("> ").strip())
            idx += 1
            continue
        if line.startswith("###"):
            st.markdown(f"#### {escape(line.lstrip('#').strip())}")
            idx += 1
            continue
        if line.startswith("|"):
            table_lines = []
            while idx < len(lines) and lines[idx].strip().startswith("|"):
                table_lines.append(lines[idx])
                idx += 1
            headers, rows = parse_markdown_table_lines(table_lines)
            if headers[:2] == ["항목", "기입 초안"]:
                for row in rows:
                    if len(row) >= 2:
                        render_kras_value_card(row[0], row[1])
            else:
                render_generic_kras_table(headers, rows)
            continue
        st.markdown(line)
        idx += 1


def render_export_report_page() -> None:
    st.header("Excel 내보내기 및 보고서")
    st.info("현장 점검 보조자료를 Excel로 내려받아 교수님 보고나 내부 기록 정리에 활용할 수 있습니다.")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("#### 중대재해처벌법 체크리스트")
        checklist_excel_bytes = build_simple_xlsx_bytes("legal_checklist", load_legal_checklist_status())
        st.download_button(
            label="체크리스트 Excel 다운로드",
            data=checklist_excel_bytes,
            file_name="legal_checklist_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="field_download_legal_checklist_excel",
            use_container_width=True,
        )
    with c2:
        st.markdown("#### 위험성평가 초안")
        rows = st.session_state.get(
            "risk_assessment_draft_rows",
            build_risk_assessment_draft("갱내 작업", "막장 또는 갱내 작업장", "현장 위험요인", "작업 전 위험성평가 필요"),
        )
        risk_excel_bytes = build_simple_xlsx_bytes("risk_assessment_draft", rows)
        st.download_button(
            label="위험성평가 초안 Excel 다운로드",
            data=risk_excel_bytes,
            file_name="risk_assessment_draft_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="field_download_risk_assessment_excel",
            use_container_width=True,
        )
    with c3:
        st.markdown("#### 대화 이력")
        history_excel_bytes = build_conversation_history_xlsx_bytes(load_conversation_history())
        st.download_button(
            label="대화 이력 Excel 다운로드",
            data=history_excel_bytes,
            file_name="conversation_history_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="field_download_conversation_history_excel",
            use_container_width=True,
        )
    st.caption("AI 답변과 내보낸 파일은 법적 책임 면제 자료가 아니며, 현장 점검표·사진·교육기록·작업중지 기록과 함께 관리해야 합니다.")


def export_conversation_history_xlsx(rows: list[dict[str, Any]]) -> Path:
    write_simple_xlsx(CONVERSATION_HISTORY_EXPORT_PATH, "conversation_history", normalize_conversation_history_rows(rows))
    return CONVERSATION_HISTORY_EXPORT_PATH


def auto_save_conversation_history(
    question: str,
    answer: str,
    situation_type: str,
    risk_level: str,
    results: list[dict[str, Any]],
    recommended_records: list[str],
    result_key: str,
    reference_cases: list[dict[str, Any]] | None = None,
    official_cases: list[dict[str, Any]] | None = None,
    evidence_assessment: dict[str, Any] | None = None,
    answer_mode: str = "",
) -> str | None:
    if not question.strip() or not answer.strip():
        return None
    history_marker = f"history_saved_{result_key}_{abs(hash(question + answer))}"
    if st.session_state.get(history_marker):
        return st.session_state[history_marker]
    evidence_assessment = (
        evidence_assessment
        if isinstance(evidence_assessment, dict)
        else {}
    )
    official_case_history = summarize_official_cases_for_history(official_cases or [])
    history_id = append_conversation_history(
        {
            "history_id": str(uuid.uuid4()),
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "question": question,
            "answer": answer,
            "situation_type": situation_type,
            "risk_level": risk_level,
            "answer_mode": normalize_answer_mode_label(answer_mode),
            "evidence_status": str(evidence_assessment.get("status", "")),
            "evidence_label": str(evidence_assessment.get("label", "")),
            "evidence_reason": str(evidence_assessment.get("reason", "")),
            "evidence_documents": list(dict.fromkeys(str(item.get("source", "출처 정보 없음")) for item in results)),
            "source_chunks": source_chunk_labels(results),
            "recommended_evidence_records": recommended_records,
            "reference_cases": summarize_reference_cases_for_history(reference_cases or []),
            **official_case_history,
            "user_action_status": "미조치",
            "manager": "",
            "action_due_date": "",
            "memo": "",
        }
    )
    st.session_state[history_marker] = history_id
    return history_id


def infer_risk_keywords(text: str) -> list[str]:
    lowered = text.lower()
    mapping = {
        "발파": ["대피", "출입통제", "발파 전 점검", "불발 확인"],
        "불발": ["대피", "출입통제", "불발 장약 처리", "작업재개 승인"],
        "메탄/가스": ["작업중지", "대피", "환기", "재측정", "작업재개 승인"],
        "낙반": ["지보 점검", "천반 확인", "작업중지", "보강 후 재개"],
        "분진": ["살수", "집진", "방진마스크", "작업환경측정"],
        "전기": ["전원 차단", "절연", "잠금표지", "점검자 지정"],
        "장비/운반": ["동선 분리", "신호수 배치", "속도 제한", "후진 경보"],
    }
    selected = []
    for key, actions in mapping.items():
        key_parts = key.lower().replace("/", " ").split()
        if any(part in lowered for part in key_parts):
            selected.extend(actions)
    return selected or ["작업중지 기준 확인", "위험성평가", "보호구 확인", "책임자 승인"]


def build_risk_assessment_draft(
    work_name: str,
    work_place: str,
    hazards: str,
    situation: str,
) -> list[dict[str, Any]]:
    combined = f"{work_name} {work_place} {hazards} {situation}"
    actions = infer_risk_keywords(combined)
    records = recommend_evidence_records(combined, situation)
    priority = "높음" if any(word in combined for word in ["메탄", "가스", "불발", "낙반", "중대재해", "감전"]) else "보통"
    return [
        {
            "작업명": work_name,
            "작업 장소": work_place,
            "위험요인": hazards or "현장 위험요인 확인 필요",
            "위험상황": situation or "작업 전 위험상황 설명 필요",
            "현재 안전조치": "작업 전 점검, TBM, 보호구 착용, 작업구역 확인",
            "추가 개선대책": ", ".join(actions),
            "필요 증빙자료": ", ".join(records),
            "조치 우선순위": priority,
            "담당자": "",
            "조치 기한": "",
        }
    ]


def export_risk_assessment_xlsx(rows: list[dict[str, Any]]) -> Path:
    write_simple_xlsx(RISK_ASSESSMENT_EXPORT_PATH, "risk_assessment_draft", rows)
    return RISK_ASSESSMENT_EXPORT_PATH


def render_legal_checklist_page() -> None:
    st.header("중대재해처벌법 대응 체크리스트")
    st.warning("본 기능은 법적 면책 자료가 아니라 안전보건 확보의무 이행 상태 점검 보조 기능입니다.")
    rows = load_legal_checklist_status()
    status_counts = Counter(row.get("상태", "미작성") for row in rows)
    completed = status_counts.get("완료", 0)
    insufficient = status_counts.get("미흡", 0)
    missing = status_counts.get("미작성", 0)
    total = len(rows) or 1
    response_rate = round(completed / total * 100, 1)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("전체 대응률", f"{response_rate}%")
    col2.metric("완료", completed)
    col3.metric("미흡", insufficient)
    col4.metric("미작성", missing)

    edited_rows = []
    with st.form("legal_checklist_form"):
        for idx, row in enumerate(rows):
            with st.expander(f"{idx + 1}. {row.get('항목명', '')}", expanded=False):
                st.caption(row.get("설명", ""))
                c1, c2, c3 = st.columns([1, 1, 2])
                status = c1.selectbox("상태", ["완료", "미흡", "미작성"], index=["완료", "미흡", "미작성"].index(row.get("상태", "미작성") if row.get("상태") in ["완료", "미흡", "미작성"] else "미작성"), key=f"legal_status_{idx}")
                manager = c2.text_input("담당자", value=row.get("담당자", ""), key=f"legal_manager_{idx}")
                recent_date = c3.text_input("최근 작성일", value=row.get("최근 작성일", ""), placeholder="YYYY-MM-DD", key=f"legal_date_{idx}")
                memo = st.text_area("비고", value=row.get("비고", ""), height=70, key=f"legal_memo_{idx}")
                edited_rows.append({**row, "상태": status, "담당자": manager, "최근 작성일": recent_date, "비고": memo})
        save_clicked = st.form_submit_button("체크리스트 저장", type="primary")
    if save_clicked:
        save_legal_checklist_status(edited_rows)
        st.success(f"체크리스트 저장 완료: {LEGAL_CHECKLIST_STATUS_PATH}")
        st.rerun()

    checklist_excel_bytes = build_simple_xlsx_bytes("legal_checklist", load_legal_checklist_status())
    st.download_button(
        label="체크리스트 Excel 다운로드",
        data=checklist_excel_bytes,
        file_name="legal_checklist_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_legal_checklist_excel",
        use_container_width=True,
    )


def render_risk_assessment_draft_page() -> None:
    st.header("위험성평가 초안")
    st.warning("본 초안은 안전관리자가 검토·수정해야 하며, 최종 법적 문서로 바로 사용할 수 없습니다.")
    with st.form("risk_assessment_form"):
        work_name = st.text_input("작업명", value="갱내 작업")
        work_place = st.text_input("작업 장소", value="막장 또는 갱내 작업장")
        hazards = st.text_area("주요 위험요인", value="메탄가스, 환기 불량, 장비 이동")
        situation = st.text_area("작업 상황 설명", value="작업 전 TBM에서 위험요인을 공유하고 작업 가능 여부를 판단해야 함")
        submitted = st.form_submit_button("위험성평가 초안 생성", type="primary")
    if submitted or "risk_assessment_draft_rows" not in st.session_state:
        st.session_state["risk_assessment_draft_rows"] = build_risk_assessment_draft(work_name, work_place, hazards, situation)
    rows = st.session_state.get("risk_assessment_draft_rows", [])
    st.dataframe(rows, use_container_width=True, hide_index=True)
    risk_excel_bytes = build_simple_xlsx_bytes("risk_assessment_draft", rows)
    st.download_button(
        label="위험성평가 초안 Excel 다운로드",
        data=risk_excel_bytes,
        file_name="risk_assessment_draft_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_risk_assessment_excel",
        use_container_width=True,
    )


def render_conversation_history_page() -> None:
    st.header("대화 이력")
    st.info("대화 이력은 AI 답변과 근거 문서, 추천 증빙자료를 함께 저장하여 안전관리 기록 보조자료로 활용할 수 있습니다. 단, 실제 법적 증빙은 현장 점검표, 사진, 교육기록, 작업중지 기록 등과 함께 관리해야 합니다.")
    rows = load_conversation_history()
    st.caption(f"저장 위치: {CONVERSATION_HISTORY_PATH}")
    if not rows:
        st.info("아직 저장된 대화 이력이 없습니다. 안전 질의 및 답변 메뉴에서 질문을 생성하면 자동 저장됩니다.")
        empty_history_excel_bytes = build_conversation_history_xlsx_bytes(rows)
        st.download_button(
            label="빈 대화 이력 Excel 다운로드",
            data=empty_history_excel_bytes,
            file_name="conversation_history_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_empty_conversation_history_excel",
            use_container_width=True,
        )
        return

    keyword = st.text_input("검색창: 질문 키워드")
    situation_options = ["전체"] + sorted(set(str(row.get("situation_type", "기타")) for row in rows if row.get("situation_type")))
    risk_options = ["전체"] + sorted(set(str(row.get("risk_level", "검토 필요")) for row in rows if row.get("risk_level")))
    f1, f2, f3 = st.columns(3)
    selected_situation = f1.selectbox("상황 유형 필터", situation_options)
    selected_risk = f2.selectbox("위험도 필터", risk_options)
    selected_date = f3.text_input("날짜 필터", placeholder="YYYY-MM-DD")
    filtered = []
    for row in rows:
        if keyword and keyword not in str(row.get("question", "")):
            continue
        if selected_situation != "전체" and row.get("situation_type") != selected_situation:
            continue
        if selected_risk != "전체" and row.get("risk_level") != selected_risk:
            continue
        if selected_date and not str(row.get("created_at", "")).startswith(selected_date):
            continue
        filtered.append(row)
    st.metric("조회 결과", f"{len(filtered)}건")
    if not filtered:
        st.info("조건에 맞는 대화 이력이 없습니다.")
        return
    labels = [short_history_label(row) for row in filtered]
    selected_label = st.selectbox("저장된 질문 목록", labels)
    selected = filtered[labels.index(selected_label)]

    evidence_documents = normalize_history_display_list(selected.get("evidence_documents", []))
    recommended_records = normalize_history_display_list(selected.get("recommended_evidence_records", []))
    reference_case_records = normalize_history_display_list(selected.get("reference_cases", []))
    official_case_records = normalize_history_display_list(selected.get("official_case_sources", []))

    with st.container(border=True):
        st.subheader("질문/답변")
        st.markdown("**질문**")
        st.markdown(
            f"""
            <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;
                        padding:14px 16px;color:#1e293b;line-height:1.6;white-space:pre-wrap;">
                {escape(str(selected.get("question", "")))}
            </div>
            """,
            unsafe_allow_html=True,
        )
        with st.expander("당시 답변 보기", expanded=True):
            st.markdown(selected.get("answer", ""))

    col_docs, col_records, col_cases = st.columns(3)
    with col_docs:
        render_history_list_card("근거 문서", evidence_documents, "저장된 근거 문서가 없습니다.")
    with col_records:
        render_history_list_card("추천 증빙자료", recommended_records, "저장된 추천 증빙자료가 없습니다.")
    with col_cases:
        render_history_list_card("참고 사례", reference_case_records, "저장된 참고 사례가 없습니다.")
    render_history_list_card("공식 사고사례", official_case_records, "저장된 공식 사고사례가 없습니다.")

    render_history_report_card(
        selected,
        evidence_documents,
        recommended_records,
        reference_case_records,
        official_case_records,
    )

    with st.form(f"history_update_{selected.get('history_id')}"):
        c1, c2, c3 = st.columns(3)
        action_status = c1.selectbox("조치 여부", ["미조치", "진행 중", "조치 완료", "보류"], index=["미조치", "진행 중", "조치 완료", "보류"].index(selected.get("user_action_status", "미조치") if selected.get("user_action_status") in ["미조치", "진행 중", "조치 완료", "보류"] else "미조치"))
        manager = c2.text_input("담당자", value=selected.get("manager", ""))
        due_date = c3.text_input("조치 완료일", value=selected.get("action_due_date", ""), placeholder="YYYY-MM-DD")
        memo = st.text_area("메모", value=selected.get("memo", ""), height=80)
        update_clicked = st.form_submit_button("이력 메타데이터 저장", type="primary")
    if update_clicked:
        ok = update_conversation_history(
            selected.get("history_id", ""),
            {
                "user_action_status": action_status,
                "manager": manager,
                "action_due_date": due_date,
                "memo": memo,
            },
        )
        if ok:
            st.success("대화 이력 메타데이터 저장 완료")
            st.rerun()
        else:
            st.error("대화 이력 저장 실패")
    history_excel_bytes = build_conversation_history_xlsx_bytes(rows)
    st.download_button(
        label="대화 이력 Excel 다운로드",
        data=history_excel_bytes,
        file_name="conversation_history_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_conversation_history_excel",
        use_container_width=True,
    )


def render_notice_guidance_page() -> None:
    st.header("공지 및 지침")
    st.info("MineSafe AI는 공식 광산 안전 문서와 중대재해처벌법 관련 자료를 함께 검색하여 답변합니다.")
    st.markdown(
        """
        - 중대재해처벌법 대응 체크리스트는 안전보건 확보의무 이행 상태를 정리하는 보조 기능입니다.
        - 질문 답변 후 추천되는 증빙자료는 현장 기록 관리 후보이며, 실제 법적 판단은 전문가 검토가 필요합니다.
        - 위험성평가 초안은 안전관리자가 검토·수정해야 하며 최종 법적 문서로 바로 사용할 수 없습니다.
        - 대화 이력은 AI 답변 기록이며, 실제 점검표·사진·교육기록 등 이행자료와 함께 관리해야 합니다.
        """
    )


def _case_database_state(
    collection: Any,
    internal_status: str | None,
    zero_status: str,
) -> dict[str, Any]:
    if collection is not None:
        try:
            count = int(collection.count())
            return {"state": "ready", "count": count}
        except Exception:
            return {"state": "error", "count": 0}
    if internal_status == zero_status:
        return {"state": "zero_not_created", "count": 0}
    return {"state": "error", "count": 0}


def build_database_status_view_model(
    law_collection: Any,
    law_error: str | None,
    verified_collection: Any,
    verified_status: str | None,
    auto_collection: Any,
    auto_status: str | None,
    text_safe_collection: Any,
    text_safe_status: str | None,
) -> dict[str, Any]:
    try:
        law_count = int(law_collection.count()) if law_collection is not None else 0
        law_state = "ready" if law_error is None and law_collection is not None else "error"
    except Exception:
        law_count = 0
        law_state = "error"
    verified_state = _case_database_state(
        verified_collection,
        verified_status,
        "verified_not_created_zero_cases",
    )
    auto_state = _case_database_state(
        auto_collection,
        auto_status,
        "auto_screened_not_created_zero_cases",
    )
    text_safe_state = _case_database_state(
        text_safe_collection,
        text_safe_status,
        "text_safe_not_created_zero_cases",
    )
    searchable_case_count = sum(
        int(item["count"])
        for item in (verified_state, auto_state, text_safe_state)
        if item["state"] == "ready"
    )
    return {
        "law": {"state": law_state, "count": law_count},
        "verified": verified_state,
        "auto_screened": auto_state,
        "text_safe": text_safe_state,
        "official_case_connected": any(
            item["state"] == "ready"
            for item in (verified_state, auto_state, text_safe_state)
        ),
        "searchable_case_count": searchable_case_count,
    }


def database_state_label(name: str, state: dict[str, Any]) -> str:
    state_name = str(state.get("state", "error"))
    count = int(state.get("count", 0) or 0)
    if state_name == "ready":
        return f"{name}: 정상 · {count}건"
    if state_name == "zero_not_created":
        return f"{name}: 대상 사례 0건으로 미생성"
    return f"{name}: 연결 오류"


def render_sidebar_database_status(
    is_admin: bool,
    law_collection: Any,
    law_error: str | None,
) -> dict[str, Any]:
    verified_collection, verified_status = load_official_case_collection()
    auto_collection, auto_status = load_auto_screened_case_collection()
    text_safe_collection, text_safe_status = load_text_safe_case_collection()
    model = build_database_status_view_model(
        law_collection,
        law_error,
        verified_collection,
        verified_status,
        auto_collection,
        auto_status,
        text_safe_collection,
        text_safe_status,
    )
    if is_admin:
        st.sidebar.markdown("#### DB 연결 상태")
        law_label = (
            f"법령 DB: 정상 · {model['law']['count']}건"
            if model["law"]["state"] == "ready"
            else "법령 DB: 연결 오류"
        )
        st.sidebar.write(law_label)
        st.sidebar.write(database_state_label("검증 완료 사례 DB", model["verified"]))
        st.sidebar.write(database_state_label("엄격 자동검사 사례 DB", model["auto_screened"]))
        st.sidebar.write(database_state_label("문자 안전 사례 DB", model["text_safe"]))
        with st.sidebar.expander("DB 경로 상세 보기", expanded=False):
            path_rows = (
                ("법령 DB", VECTOR_DB_DIR),
                ("검증 완료 사례 DB", OFFICIAL_CASE_VECTOR_DB_DIR),
                ("엄격 자동검사 사례 DB", AUTO_SCREENED_OFFICIAL_CASE_VECTOR_DB_DIR),
                ("문자 안전 사례 DB", TEXT_SAFE_OFFICIAL_CASE_VECTOR_DB_DIR),
            )
            for label, path in path_rows:
                st.markdown(
                    f'<div class="db-path-detail"><strong>{escape(label)}</strong><br>'
                    f'{escape(str(path))}</div>',
                    unsafe_allow_html=True,
                )
    else:
        law_connection = "연결됨" if model["law"]["state"] == "ready" else "연결 확인 필요"
        case_connection = "연결됨" if model["official_case_connected"] else "사용 가능 사례 없음"
        st.sidebar.markdown(
            (
                '<div class="major-law-sidebar-card">'
                '<div class="major-law-sidebar-title">공식 자료 연결 상태</div>'
                f'<div class="major-law-sidebar-line">공식 법령 DB: {law_connection}</div>'
                f'<div class="major-law-sidebar-line">공식 사례 DB: {case_connection}</div>'
                '<div class="major-law-sidebar-line">공식 사례 검색 가능 수: '
                f'{model["searchable_case_count"]}건</div>'
                '</div>'
            ),
            unsafe_allow_html=True,
        )
    return model


# ==============================
# 사이드바
# ==============================
audience_mode = st.sidebar.radio(
    "사용자 모드",
    ["현장관리자 모드", "관리자/개발자 모드"],
    index=0,
)
is_admin_mode = audience_mode == "관리자/개발자 모드"

if is_admin_mode:
    st.sidebar.markdown(
        """
        <div class="mscc-sidebar-brand">
            <strong>MineSafe AI</strong><br>
            <div class="portal-sidebar-en">공식 문서 기반 광산 안전관리 AI 시스템</div>
        </div>
        <div class="portal-nav-group">운영 및 답변</div>
        <div class="portal-nav-item active">안전 질의 및 답변</div>
        <div class="portal-nav-item">대화 이력</div>
        <div class="portal-nav-item">공지 및 지침</div>
        <div class="portal-nav-group">지식 관리</div>
        <div class="portal-nav-item">공식 사고사례 검수</div>
        <div class="portal-nav-item">Vector DB 관리</div>
        <div class="portal-nav-item">문서 관리</div>
        <div class="portal-nav-item">모델 관리</div>
        <div class="portal-nav-group">시스템 관리</div>
        <div class="portal-nav-item">사용자 관리</div>
        <div class="portal-nav-item">권한 관리</div>
        <div class="portal-nav-item">설정</div>
        """,
        unsafe_allow_html=True,
    )
else:
    st.sidebar.markdown(
        """
        <div class="mscc-sidebar-brand">
            <strong>MineSafe AI</strong><br>
            <div class="portal-sidebar-en">현장 안전관리자용 화면</div>
        </div>
        <div class="portal-nav-group">실무 메뉴</div>
        <div class="portal-nav-item active">안전 질문 답변</div>
        <div class="portal-nav-item">중대재해처벌법 대응 체크리스트</div>
        <div class="portal-nav-item">위험성평가 초안</div>
        <div class="portal-nav-item">대화 이력</div>
        <div class="portal-nav-item">Excel 내보내기 또는 보고서</div>
        """,
        unsafe_allow_html=True,
    )

chunk_count: int | str = 0
collection, db_error = load_chroma_collection()
database_status_model = render_sidebar_database_status(
    is_admin_mode,
    collection,
    db_error,
)
chunk_count = (
    int(database_status_model["law"]["count"])
    if database_status_model["law"]["state"] == "ready"
    else "확인 실패"
)

law_badges_sidebar = "".join(
    f'<span class="major-law-mini-badge">{escape(label)}</span>'
    for label, _ in MAJOR_ACCIDENT_LAW_DOCS
)
if is_admin_mode:
    st.sidebar.markdown(
        (
            '<div class="major-law-sidebar-card">'
            '<div class="major-law-sidebar-title">법령 자료 반영 현황</div>'
            '<div class="major-law-sidebar-line">중대재해처벌법 자료 반영 완료</div>'
            '<div class="major-law-sidebar-line">반영 자료: FAQ / 질의회시집 / 해설서 / 따라하기 안내서</div>'
            f'<div class="major-law-badge-row">{law_badges_sidebar}</div>'
            f'<div class="major-law-sidebar-line">통합 chunk 수: {MAJOR_ACCIDENT_DOC_TOTAL_CHUNKS}개</div>'
            f'<div class="major-law-sidebar-line">추가 자료 chunk 수: {MAJOR_ACCIDENT_DOC_ADDED_CHUNKS}개</div>'
            "</div>"
        ),
        unsafe_allow_html=True,
    )
else:
    st.sidebar.markdown(
        (
            '<div class="major-law-sidebar-card">'
            '<div class="major-law-sidebar-title">법령 자료 반영</div>'
            '<div class="major-law-sidebar-line">중대재해처벌법 공식 자료를 함께 활용합니다.</div>'
            f'<div class="major-law-badge-row">{law_badges_sidebar}</div>'
            "</div>"
        ),
        unsafe_allow_html=True,
    )

if is_admin_mode:
    selected_gemini_model = st.sidebar.selectbox(
        "Gemini 모델",
        GEMINI_MODEL_OPTIONS,
        index=0,
    )

    _, gemini_error = load_gemini_client()
    if gemini_error:
        st.sidebar.warning("Gemini API · 키 미감지")
        st.sidebar.caption("로컬 .env 또는 Streamlit Cloud Secrets의 GEMINI_API_KEY를 확인하세요.")
        st.sidebar.caption(gemini_error)
    else:
        st.sidebar.info("Gemini API · API 키 감지됨")
        st.sidebar.caption("실제 연결 상태는 연결 테스트 또는 답변 호출 결과로 확인합니다.")
        st.sidebar.caption(f"선택 모델: {selected_gemini_model}")
        st.sidebar.write(f"호출 제한: 시도당 약 {GEMINI_RESPONSE_TIMEOUT_SECONDS}초")
        st.sidebar.write(f"최대 시도: {GEMINI_MAX_ATTEMPTS}회")

    if st.sidebar.button("Gemini 연결 테스트", use_container_width=True):
        with st.sidebar:
            with st.spinner("선택한 모델로 연결을 확인하는 중입니다..."):
                test_result = test_gemini_connection(
                    selected_gemini_model
                )
        test_state = classify_gemini_status(test_result)
        if test_result.get("success"):
            st.sidebar.success(
                f"연결 테스트 성공 · {test_result.get('model')} · "
                f"{test_result.get('elapsed', 0.0):.1f}초"
            )
            st.sidebar.caption(f"Gemini 응답: {test_result.get('answer', '')}")
        else:
            st.sidebar.warning(
                f"연결 테스트 실패 · {test_result.get('model')} · "
                f"{format_gemini_status(test_state)}"
            )
            st.sidebar.caption(test_result.get("message", "Gemini 연결 테스트 실패"))
            with st.sidebar.expander("오류 상세"):
                st.write(test_result.get("error") or "오류 메시지 없음")
                st.write(f"시도 횟수: {test_result.get('attempts', 0)}")

    st.sidebar.divider()

    answer_mode = st.sidebar.selectbox(
        "답변 생성 방식",
        [STABLE_MODE, GEMINI_MODE, HYBRID_MODE],
        index=0,
        format_func=answer_mode_option_label,
    )

    if answer_mode == STABLE_MODE:
        st.sidebar.info(f"안정성 모드\n\n{STABLE_MODE_HELP}")
    elif answer_mode == GEMINI_MODE:
        st.sidebar.info(
            f"{WORKER_EASY_MODE_LABEL}\n\n{WORKER_EASY_MODE_HELP}\n\n"
            "쉬운 설명 생성에 실패하면 공식 문서 기반 안정형 답변으로 전환합니다."
        )
    else:
        st.sidebar.info(
            "하이브리드 모드\n\n"
            f"{HYBRID_MODE_HELP}\n\n"
            "- 1차: 검색 근거 기반 안정형 답변\n"
            "- 2차: Gemini 기반 보조 답변\n"
            "- Gemini 실패 시에도 1차 답변 유지"
        )

    with st.sidebar.expander("시연 권장 설정"):
        st.markdown(
            "- **답변 모드:** 하이브리드 모드 권장\n"
            "- **Gemini 모델:** `gemini-2.5-flash-lite` 권장\n"
            "- **이유:** 안정형 답변을 먼저 제공하고 Gemini 응답을 보조적으로 확인\n"
            "- Gemini가 실패해도 검색 근거 기반 답변은 계속 제공됩니다."
        )

    top_k = st.sidebar.slider(
        "검색할 근거 문서 수",
        min_value=3,
        max_value=5,
        value=3,
        step=1,
    )

    with st.sidebar.expander("시연 상태 체크"):
        st.markdown(
            "- Vector DB 로드 확인\n"
            "- Gemini 모델 선택 확인\n"
            "- 답변 모드는 하이브리드 모드 권장\n"
            "- 답변과 함께 근거 문서가 표시되는지 확인"
        )

    selected_scenario_label = st.sidebar.selectbox(
        "질문 시나리오 세트",
        list(SCENARIO_SET_OPTIONS.keys()),
        index=list(SCENARIO_SET_OPTIONS.keys()).index("110개 분진·보호구 보강 평가 세트"),
    )
    selected_scenario_path = SCENARIO_SET_OPTIONS[selected_scenario_label]

    st.sidebar.markdown(
        """
        <div class="mscc-sidebar-note">
            테스트 모드에서는 선택한 시나리오 세트의 질문 검색, 답변 검토, 평가 저장과 진행률 확인을 지원합니다.
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.sidebar.caption("시연 시에는 110개 분진·보호구 보강 평가 세트를 선택하는 것을 권장합니다.")
    feature_page = st.sidebar.radio(
        "기능 메뉴",
        [
            "안전 질의 및 답변",
            "중대재해처벌법 대응",
            "위험성평가 초안",
            "공식 사고사례 검수",
            "대화 이력",
            "공지 및 지침",
        ],
        index=0,
    )

    st.sidebar.markdown(
        f"""
        <div class="portal-system-state">
            <div class="portal-system-state-title">시스템 상태</div>
            <div class="portal-system-state-value">정상 운영 중</div>
            <div class="portal-system-state-time">
                마지막 동기화: {escape(time.strftime("%Y-%m-%d %H:%M"))}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
else:
    selected_gemini_model = GEMINI_MODEL_OPTIONS[0]
    gemini_error = "현장관리자 모드에서는 Gemini 상태를 숨깁니다."
    answer_mode = STABLE_MODE
    top_k = 3
    selected_scenario_label = "110개 분진·보호구 보강 평가 세트"
    selected_scenario_path = SCENARIO_SET_OPTIONS[selected_scenario_label]
    feature_page = st.sidebar.radio(
        "기능 메뉴",
        ["안전 질문 답변", "중대재해처벌법 대응 체크리스트", "위험성평가 초안", "대화 이력", "Excel 내보내기 또는 보고서"],
        index=0,
    )


if is_admin_mode and feature_page == "공식 사고사례 검수":
    render_verified_official_case_review_page()
    st.stop()


# ==============================
# 상단 운영 상태
# ==============================
scenario_rows, scenario_status_error = load_question_scenarios(selected_scenario_path)
scenario_count = len(scenario_rows) if not scenario_status_error else 0
db_status_text = "정상" if not db_error else "점검 필요"
db_status_accent = "#10b981" if not db_error else "#b45309"
gemini_status_text = (
    "관리자 모드에서 확인"
    if not is_admin_mode
    else ("API 키 감지됨" if not gemini_error else "키 미감지")
)

status_col1, status_col2, status_col3, status_col4 = st.columns(4)
with status_col1:
    render_status_card(
        "Vector DB",
        db_status_text,
        "공식 법령 DB 연결 상태",
        db_status_accent,
    )
with status_col2:
    render_status_card(
        "문서 Chunk",
        f"{chunk_count}개",
        "공식 문서 검색 대상",
        "#64748b",
    )
with status_col3:
    render_status_card(
        "답변 모드",
        short_answer_mode(answer_mode),
        answer_mode_description(short_answer_mode(answer_mode)),
        "#64748b",
    )
with status_col4:
    render_status_card(
        "평가 시나리오",
        f"{scenario_count}개",
        f"{selected_scenario_label}",
        "#64748b",
    )


# ==============================
# 메뉴별 화면 전환
# ==============================
if feature_page in {"중대재해처벌법 대응", "중대재해처벌법 대응 체크리스트"}:
    render_legal_checklist_page()
    st.stop()
if feature_page == "위험성평가 초안":
    render_risk_assessment_draft_page()
    st.stop()
if feature_page == "대화 이력":
    render_conversation_history_page()
    st.stop()
if feature_page == "공지 및 지침":
    render_notice_guidance_page()
    st.stop()
if feature_page == "Excel 내보내기 또는 보고서":
    render_export_report_page()
    st.stop()

# ==============================
# 질문 입력 + 테스트 모드
# ==============================

def run_rag_flow(
    question_text: str,
    top_k_value: int,
    answer_mode: str,
    selected_model: str = GEMINI_MODEL_NAME,
):
    intent = detect_question_intent(question_text)
    if intent == OUT_OF_SCOPE_INTENT:
        stable_answer = generate_local_fallback_answer(
            question_text,
            [],
            "범위 밖 질문",
            intent,
        )
        out_status = {
            "selected_answer_mode": short_answer_mode(answer_mode),
            "answer_mode": short_answer_mode(answer_mode),
            "rag_used": False,
            "retrieved_chunk_count": 0,
            "selected_model": selected_model,
            "model": selected_model,
            "called": False,
            "success": False,
            "status": "out_of_scope",
            "message": "광산 안전관리 범위 밖 질문입니다.",
            "answer": "",
            "used_fallback": False,
            "mode": "stable",
            "gemini_status": "호출 안 함",
            "reason": "범위 밖 질문",
            "attempts": 0,
            "elapsed": 0.0,
            "gemini_called": False,
            "gemini_call_success": False,
            "fallback_used": False,
            "actual_execution": "범위 밖 질문 안내",
            "mode_output_title": "답변 범위 안내",
            "question_intent": intent,
            "official_cases": [],
            "official_case_diagnostic": {"status": "skipped_out_of_scope", "result_count": 0},
            "reference_cases": [],
            "live_news_cases": [],
        }
        return [], stable_answer, out_status, None

    with st.spinner("1단계: Vector DB에서 관련 근거 문서를 검색 중입니다..."):
        rag_results, search_error = search_vector_db(question_text, top_k=top_k_value)

    if search_error:
        rag_results = []
    evidence_assessment = assess_rag_evidence_sufficiency(
        question_text,
        rag_results,
        intent,
    )
    stable_answer = generate_local_fallback_answer(
        question_text,
        rag_results,
        search_error or "안정 모드: Gemini API 호출 안 함",
        intent,
        evidence_assessment=evidence_assessment,
    )
    stable_answer = enforce_rag_evidence_answer_guardrail(
        stable_answer,
        evidence_assessment,
    )
    reference_cases = match_reference_cases(question_text, intent, max_cases=2)
    live_news_cases = get_live_reference_cases(question_text, intent, max_cases=3)
    official_case_diagnostic: dict[str, Any] = {}
    official_cases = search_official_siren_cases(
        question_text,
        intent,
        top_k=OFFICIAL_CASE_TOP_K,
        diagnostic=official_case_diagnostic,
    )

    base_status = {
        "selected_answer_mode": short_answer_mode(answer_mode),
        "answer_mode": short_answer_mode(answer_mode),
        "rag_used": True,
        "retrieved_chunk_count": len(rag_results),
        "selected_model": selected_model,
        "model": selected_model,
        "question_intent": intent,
        "expanded_search_query": expand_search_query(question_text, intent),
        "official_cases": official_cases,
        "official_case_diagnostic": official_case_diagnostic,
        "reference_cases": reference_cases,
        "live_news_cases": live_news_cases,
        "evidence_assessment": evidence_assessment,
        "evidence_status": evidence_assessment["status"],
        "evidence_label": evidence_assessment["label"],
        "evidence_reason": evidence_assessment["reason"],
    }

    if answer_mode == STABLE_MODE:
        rag_status = {
            **base_status,
            "called": False,
            "success": False,
            "status": "not_called",
            "message": "안정 모드에서는 Gemini API를 호출하지 않습니다.",
            "answer": "",
            "used_fallback": False,
            "mode": "stable",
            "gemini_status": "호출 안 함",
            "reason": "안정 모드에서는 Gemini API를 호출하지 않습니다.",
            "attempts": 0,
            "elapsed": 0.0,
            "gemini_called": False,
            "gemini_call_success": False,
            "fallback_used": False,
            "actual_execution": "안정성 모드 - 공식 문서 기반 체크리스트형",
            "mode_output_title": "안정성 모드",
        }
        return rag_results, stable_answer, rag_status, None

    if answer_mode == HYBRID_MODE:
        with st.spinner(
            f"하이브리드 모드: 안정형 조치 초안과 RAG 근거를 Gemini에 전달해 자연어 보완을 시도합니다. 최대 {GEMINI_RESPONSE_TIMEOUT_SECONDS}초..."
        ):
            hybrid_answer, gemini_status = generate_gemini_answer(
                question_text,
                rag_results,
                selected_model,
                prompt_kind="hybrid",
                stable_draft=stable_answer,
                reference_cases=reference_cases,
                live_news_cases=live_news_cases,
                evidence_assessment=evidence_assessment,
                official_cases=official_cases,
            )

        gemini_state = classify_gemini_status(gemini_status)
        success = gemini_state == "성공"
        if success:
            final_answer = enforce_rag_evidence_answer_guardrail(
                hybrid_answer,
                evidence_assessment,
            )
            actual_execution = "Gemini API 호출 성공 - 안정형 조치에 이유/현장 적용 보완"
        else:
            final_answer = stable_answer
            actual_execution = "Gemini API 호출 실패 → 안정형 fallback"

        gemini_status.update(
            {
                **base_status,
                "mode": "hybrid",
                "answer_mode": short_answer_mode(answer_mode),
                "gemini_status": gemini_state,
                "gemini_called": bool(gemini_status.get("called", True)),
                "gemini_call_success": success,
                "fallback_used": not success,
                "used_fallback": not success,
                "actual_execution": actual_execution,
                "mode_output_title": "하이브리드 답변" if success else "안정성 모드",
            }
        )
        return rag_results, final_answer, gemini_status, None

    with st.spinner(
        f"{WORKER_EASY_MODE_LABEL}: 검색 근거를 짧고 쉬운 안전교육 문장으로 바꿉니다. 최대 {GEMINI_RESPONSE_TIMEOUT_SECONDS}초..."
    ):
        rag_answer, rag_status = generate_gemini_answer(
            question_text,
            rag_results,
            selected_model,
            prompt_kind="gemini",
            reference_cases=reference_cases,
            live_news_cases=live_news_cases,
            evidence_assessment=evidence_assessment,
            official_cases=official_cases,
        )

    gemini_state = classify_gemini_status(rag_status)
    success = gemini_state == "성공"
    final_answer = (
        enforce_rag_evidence_answer_guardrail(
            rag_answer,
            evidence_assessment,
            worker_easy=True,
        )
        if success
        else stable_answer
    )
    rag_status.update(
        {
            **base_status,
            "answer_mode": short_answer_mode(answer_mode),
            "gemini_status": gemini_state,
            "gemini_called": bool(rag_status.get("called", True)),
            "gemini_call_success": success,
            "fallback_used": not success,
            "used_fallback": not success,
            "actual_execution": "Gemini API 호출 성공 - 신규 근로자용 쉬운 설명" if success else "Gemini API 호출 실패 → 안정형 fallback",
            "mode_output_title": WORKER_EASY_MODE_LABEL if success else "안정성 모드",
            "fallback_notice": (
                ""
                if success
                else "쉬운 설명을 생성하지 못해 공식 문서 기반 핵심조치로 대신 안내합니다."
            ),
        }
    )

    return rag_results, final_answer, rag_status, None

def render_evidence_card(result: dict[str, Any]) -> None:
    distance_text = format_distance(result["distance"])
    source = str(result.get("source", "출처 정보 없음"))
    chunk_id = str(result.get("chunk_id", "정보 없음"))
    rank = result.get("rank", "-")
    preview = make_preview(str(result.get("text", "")), limit=520)

    with st.container(border=True):
        st.markdown(
            (
                f'<div class="mscc-evidence-title">'
                f'근거 {escape(str(rank))} | {escape(source)}</div>'
                f'<div class="mscc-evidence-meta">'
                f'CHUNK {escape(chunk_id)} · DISTANCE {escape(distance_text)}</div>'
                f'<div class="mscc-evidence-preview">{escape(preview)}</div>'
            ),
            unsafe_allow_html=True,
        )
        with st.expander("전체 본문 및 메타데이터"):
            st.write(f"문서명: {source}")
            st.write(f"chunk_id: {chunk_id}")
            st.write(f"distance 또는 유사도: {distance_text}")
            st.markdown("**전체 본문**")
            st.write(result.get("text", ""))
            st.markdown("**메타데이터**")
            st.json(result.get("metadata", {}))


def split_answer_for_dashboard(answer: str) -> tuple[str, str, str]:
    kras_markers = [
        "### KRAS식 위험성평가 기록 초안",
        "## 2. KRAS식 위험성평가 기록 초안",
        "## KRAS식 위험성평가 기록 초안",
    ]
    kras_start = -1
    for marker in kras_markers:
        marker_index = answer.find(marker)
        if marker_index >= 0 and (kras_start < 0 or marker_index < kras_start):
            kras_start = marker_index

    supplement_markers = [
        "\n---\n## Gemini 추가 답변",
        "\n## Gemini 추가 답변",
        "\n---\n## Gemini 응답 실패",
        "\n## Gemini 응답 실패",
    ]
    supplement_start = -1
    for marker in supplement_markers:
        marker_index = answer.find(marker)
        if marker_index >= 0 and (
            supplement_start < 0 or marker_index < supplement_start
        ):
            supplement_start = marker_index

    if kras_start < 0:
        core_end = supplement_start if supplement_start >= 0 else len(answer)
        core_answer = answer[:core_end].strip()
        evidence_marker = core_answer.find("\n## 관련 근거 문서")
        if evidence_marker >= 0:
            core_answer = core_answer[:evidence_marker].strip()
        return core_answer, "", answer[core_end:].strip()

    kras_end = (
        supplement_start
        if supplement_start >= 0 and supplement_start > kras_start
        else len(answer)
    )
    core_answer = answer[:kras_start].strip()
    evidence_marker = core_answer.find("\n## 관련 근거 문서")
    if evidence_marker >= 0:
        core_answer = core_answer[:evidence_marker].strip()
    return (
        core_answer,
        answer[kras_start:kras_end].strip(),
        answer[kras_end:].strip(),
    )


def extract_markdown_section(markdown_text: str, heading: str) -> str:
    marker = f"## {heading}"
    start = markdown_text.find(marker)
    if start < 0:
        return ""
    content_start = start + len(marker)
    next_heading = markdown_text.find("\n## ", content_start)
    content_end = next_heading if next_heading >= 0 else len(markdown_text)
    return markdown_text[content_start:content_end].strip()


def strip_markdown_prefix(text: str) -> str:
    cleaned_lines = []
    for line in text.splitlines():
        cleaned = line.strip()
        if not cleaned:
            continue
        while cleaned.startswith(("- ", "* ", "• ")):
            cleaned = cleaned[2:].strip()
        cleaned_lines.append(cleaned)
    return " ".join(cleaned_lines)


def build_dashboard_summary(
    core_answer: str,
    situation_type: str,
) -> tuple[str, list[str]]:
    immediate_text = strip_markdown_prefix(
        extract_markdown_section(core_answer, "즉시 판단")
    )
    if not immediate_text:
        immediate_text = build_immediate_judgment(situation_type)

    action_items: list[str] = []
    for item in build_priority_actions(situation_type) + build_check_items(situation_type):
        normalized = clean_text(str(item))
        if normalized and normalized not in action_items:
            action_items.append(normalized)
        if len(action_items) >= 6:
            break
    return immediate_text, action_items


def strip_kras_title(kras_answer: str) -> str:
    lines = kras_answer.splitlines()
    while lines and (
        not lines[0].strip()
        or "KRAS식 위험성평가 기록 초안" in lines[0]
    ):
        lines.pop(0)
    return "\n".join(lines).strip()


def decorate_kras_markdown(kras_answer: str) -> str:
    decorated_lines = []
    for line in kras_answer.splitlines():
        if "현재 위험성" in line:
            line = line.replace(
                "매우 높음",
                '<span class="status-badge badge-danger">매우 높음</span>',
            ).replace(
                "Level 5",
                '<span class="status-badge badge-danger">Level 5</span>',
            )
        elif "조치 후 잔여위험성" in line:
            line = line.replace(
                "낮음",
                '<span class="status-badge badge-success">낮음</span>',
            ).replace(
                "Level 2",
                '<span class="status-badge badge-success">Level 2</span>',
            )
        decorated_lines.append(line)
    return "\n".join(decorated_lines)


def render_info_card(
    title: str,
    value: str,
    subtitle: str,
    icon: str,
    tone: str,
) -> None:
    st.markdown(
        (
            f'<div class="info-card info-card-{escape(tone)}">'
            f'<div class="info-card-icon">{escape(icon)}</div>'
            '<div class="info-card-body">'
            f'<div class="info-card-label">{escape(title)}</div>'
            f'<div class="info-card-value">{escape(value)}</div>'
            f'<div class="info-card-subtitle">{escape(subtitle)}</div>'
            "</div></div>"
        ),
        unsafe_allow_html=True,
    )


def classify_evidence_document(source: str) -> tuple[str, str]:
    normalized = source.replace(" ", "").lower()
    if "중대재해처벌법" in normalized or "중대산업재해" in normalized:
        if "faq" in normalized:
            return "FAQ", "law-badge law-badge-faq"
        if "질의회시" in normalized:
            return "질의회시", "law-badge law-badge-reply"
        if "해설서" in normalized:
            return "해설서", "law-badge law-badge-commentary"
        if "따라하기" in normalized or "안내서" in normalized:
            return "따라하기", "law-badge law-badge-major"
        return "중대재해법", "law-badge law-badge-major"
    if "법" in normalized and "지침" not in normalized:
        return "법령", "law-badge"
    if "지침" in normalized or "가이드" in normalized or "안내서" in normalized:
        return "지침", "law-badge law-badge-guideline"
    if "기준" in normalized or "고시" in normalized:
        return "고시", "law-badge law-badge-notice"
    return "참고", "law-badge law-badge-reference"


def render_evidence_table(results: list[dict[str, Any]]) -> None:
    row_parts = []
    for index, item in enumerate(results[:5], start=1):
        source = str(item.get("source", "출처 정보 없음"))
        document_type, badge_class = classify_evidence_document(source)
        status_text = "적용" if index <= 2 else "참고"
        status_class = "badge-success" if index <= 2 else "badge-muted"
        row_parts.append(
            (
                "<tr>"
                f"<td>{escape(str(item.get('rank', index)))}</td>"
                f'<td><span class="{badge_class}">{escape(document_type)}</span></td>'
                f"<td>{escape(source)}</td>"
                f"<td>{escape(str(item.get('chunk_id', '정보 없음')))}</td>"
                f"<td>{escape(format_distance(item.get('distance')))}</td>"
                f'<td><span class="status-badge {status_class}">{status_text}</span></td>'
                "</tr>"
            )
        )
    rows_html = "".join(row_parts)
    st.markdown(
        (
            '<div class="portal-table-scroll table-wrap">'
            '<table class="portal-data-table portal-evidence-table">'
            "<colgroup>"
            '<col style="width:7%"><col style="width:11%">'
            '<col style="width:39%"><col style="width:23%">'
            '<col style="width:11%"><col style="width:9%">'
            "</colgroup>"
            "<thead><tr><th>근거</th><th>구분</th><th>문서명</th>"
            "<th>chunk_id</th><th>거리값</th><th>상태</th></tr></thead>"
            f"<tbody>{rows_html}</tbody></table></div>"
        ),
        unsafe_allow_html=True,
    )


def render_major_accident_law_evidence_panel() -> None:
    badge_html = "".join(
        f'<span class="law-badge law-badge-major">{escape(label)}</span>'
        for label, _ in MAJOR_ACCIDENT_LAW_DOCS
    )
    docs_html = "".join(
        (
            '<div class="major-law-doc-item">'
            f'<span class="law-badge law-badge-major">{escape(label)}</span> '
            f"{escape(doc_name)}</div>"
        )
        for label, doc_name in MAJOR_ACCIDENT_LAW_DOCS
    )
    warning_items = [
        "유해가스 농도 초과 상태인데 작업을 계속한 경우",
        "위험성평가를 하지 않거나 형식적으로만 작성한 경우",
        "작업중지 필요 상황인데 즉시 중지하지 않은 경우",
        "보호구 지급·착용 확인 없이 작업자를 투입한 경우",
        "안전교육, 점검, 개선조치 기록이 남아 있지 않은 경우",
    ]
    warning_html = "".join(f"<li>{escape(item)}</li>" for item in warning_items)
    st.markdown(
        (
            '<div class="major-law-evidence-box">'
            '<div class="major-law-evidence-title">중대재해처벌법 활용 근거</div>'
            '<div class="major-law-evidence-text">'
            "본 시스템은 광산 안전 지침뿐 아니라 중대재해처벌법 관련 공식 자료를 "
            "함께 검색하여 답변 생성에 활용합니다."
            "</div>"
            f'<div class="major-law-badge-row">{badge_html}'
            '<span class="law-badge law-badge-major">중대재해처벌법</span></div>'
            f'<div class="major-law-doc-list">{docs_html}</div>'
            "</div>"
        ),
        unsafe_allow_html=True,
    )
    st.markdown(
        (
            '<div class="major-law-evidence-box" style="margin-top:12px;">'
            '<div class="major-law-evidence-title">중대재해처벌법 유의사항</div>'
            '<div class="major-law-evidence-text">'
            "아래 사례는 법률 위반 여부를 단정하는 내용이 아니라, 현장에서 "
            "중대재해처벌법 대응 측면에서 주의가 필요한 대표 상황을 안내하기 위한 것입니다."
            "</div>"
            f'<ul style="margin:10px 0 10px 18px;line-height:1.7;">{warning_html}</ul>'
            '<div class="major-law-evidence-text" style="margin-top:8px;">'
            "AI 답변이나 체크리스트만으로 법적 책임이 면제되는 것은 아니며, "
            "점검표, 교육기록, 작업중지 기록, 개선조치 사진, 확인 서명 등 "
            "실제 이행자료를 함께 관리하는 것이 중요합니다."
            "</div>"
            "</div>"
        ),
        unsafe_allow_html=True,
    )


def render_checklist_table(situation_type: str) -> None:
    rows_html = "".join(
        (
            "<tr>"
            f"<td>{index}</td>"
            f"<td>{escape(clean_text(str(item)))}</td>"
            '<td><span class="portal-badge portal-badge-wait">대기</span></td>'
            "<td>-</td>"
            "<td>현장 책임자</td>"
            "</tr>"
        )
        for index, item in enumerate(
            build_check_items(situation_type)[:7],
            start=1,
        )
    )
    st.markdown(
        (
            '<div class="portal-table-scroll table-wrap">'
            '<table class="portal-data-table portal-checklist-table">'
            "<colgroup>"
            '<col style="width:6%"><col style="width:55%">'
            '<col style="width:12%"><col style="width:12%"><col style="width:15%">'
            "</colgroup>"
            "<thead><tr><th>번호</th><th>조치 항목</th><th>이행 상태</th>"
            "<th>이행일</th><th>담당자</th></tr></thead>"
            f"<tbody>{rows_html}</tbody></table></div>"
        ),
        unsafe_allow_html=True,
    )



def render_gemini_runtime_status(
    answer_status: dict[str, Any],
    selected_model: str,
    mode_name: str,
) -> None:
    gemini_state = classify_gemini_status(answer_status)
    gemini_called = bool(
        answer_status.get(
            "called",
            answer_status.get("gemini_called", False),
        )
    )
    fallback_used = bool(
        answer_status.get(
            "used_fallback",
            answer_status.get("fallback_used", False),
        )
    )
    gemini_success = bool(answer_status.get("gemini_call_success", gemini_state == "성공"))
    actual_execution = str(answer_status.get("actual_execution", "실행 정보 없음"))
    chunk_count = int(answer_status.get("retrieved_chunk_count", 0) or 0)
    public_mode = public_answer_mode_label(answer_status, mode_name)

    st.markdown(
        (
            '<div class="answer-runtime-summary">'
            f'<span>답변 모드: <strong>{escape(public_mode)}</strong></span>'
            f'<span>검색 근거: <strong>{chunk_count}건</strong></span>'
            + (
                f'<span>질문 유형: <strong>{escape(str(answer_status.get("question_intent")))}</strong></span>'
                if answer_status.get("question_intent")
                else ""
            )
            + "</div>"
        ),
        unsafe_allow_html=True,
    )

    if fallback_used:
        fallback_notice = str(
            answer_status.get(
                "fallback_notice",
                "외부 LLM 호출이 원활하지 않아 안정형 답변으로 전환되었습니다.",
            )
        ).strip()
        st.warning(fallback_notice)

    if not is_admin_mode:
        return

    with st.expander("개발자용 실행 정보", expanded=False):
        detail_rows = [
            ("선택한 답변 모드", mode_name),
            ("실제 실행된 답변 생성 방식", actual_execution),
            ("RAG 검색 사용 여부", "사용"),
            ("검색된 근거 chunk 수", str(chunk_count)),
            ("Gemini API 호출 여부", "예" if gemini_called else "아니오"),
            ("Gemini API 호출 성공 여부", "예" if gemini_success else "아니오"),
            ("fallback 사용 여부", "예" if fallback_used else "아니오"),
            ("사용 모델명", selected_model),
            ("질문 유형", str(answer_status.get("question_intent", ""))),
            ("확장 검색어", str(answer_status.get("expanded_search_query", ""))),
        ]
        detail_html = "".join(
            "<tr><th>{}</th><td>{}</td></tr>".format(escape(label), escape(value))
            for label, value in detail_rows
        )
        st.markdown(
            (
                '<div class="portal-table-wrap"><table class="portal-table">'
                "<tbody>"
                f"{detail_html}"
                "</tbody></table></div>"
            ),
            unsafe_allow_html=True,
        )
        if answer_status.get("reason"):
            st.caption(f"오류 정보: {answer_status.get('reason')}")


def render_rag_evidence_guardrail(
    assessment: dict[str, Any] | None,
) -> None:
    if not isinstance(assessment, dict) or not assessment:
        return

    status = str(assessment.get("status", EVIDENCE_STATUS_NEEDS_REVIEW))
    label = str(assessment.get("label", EVIDENCE_STATUS_LABELS[EVIDENCE_STATUS_NEEDS_REVIEW]))
    reason = str(assessment.get("reason", EVIDENCE_STATUS_REASONS[EVIDENCE_STATUS_NEEDS_REVIEW]))
    st.markdown(
        (
            '<div class="answer-runtime-summary">'
            '<span>공식 근거 검색 상태: '
            f'<strong>{escape(label)}</strong></span>'
            '<span>RAG 근거 충분성 휴리스틱</span>'
            '</div>'
        ),
        unsafe_allow_html=True,
    )
    st.caption(reason)
    if status == EVIDENCE_STATUS_INSUFFICIENT:
        st.warning(EVIDENCE_INSUFFICIENT_GUIDANCE)

    if not is_admin_mode:
        return

    details = assessment.get("diagnostic_details", {})
    if not isinstance(details, dict):
        details = {}
    thresholds = details.get("thresholds", {})
    if not isinstance(thresholds, dict):
        thresholds = {}
    with st.expander("RAG 근거 충분성 휴리스틱 진단", expanded=False):
        detail_rows = [
            ("공식 chunk 수", str(assessment.get("official_chunk_count", 0))),
            ("고유 문서 수", str(assessment.get("unique_document_count", 0))),
            ("비어 있지 않은 chunk 수", str(assessment.get("non_empty_chunk_count", 0))),
            ("중복 결과 수", str(assessment.get("duplicate_chunk_count", 0))),
            ("문서명 누락 수", str(details.get("missing_source_count", 0))),
            ("chunk_id 누락 수", str(details.get("missing_chunk_id_count", 0))),
            ("판정 사유", reason),
            (
                "검색 점수 해석 방식",
                str(
                    details.get(
                        "distance_interpretation",
                        "검색 점수는 충분성 판정에 사용하지 않음",
                    )
                ),
            ),
        ]
        detail_html = "".join(
            "<tr><th>{}</th><td>{}</td></tr>".format(escape(name), escape(value))
            for name, value in detail_rows
        )
        st.markdown(
            (
                '<div class="portal-table-wrap"><table class="portal-table">'
                f'<tbody>{detail_html}</tbody></table></div>'
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            "**사용된 판정 기준**\n\n"
            f"- 최소 공식 chunk 수: {thresholds.get('min_official_chunks', EVIDENCE_MIN_OFFICIAL_CHUNKS)}\n"
            f"- 최소 고유 문서 수: {thresholds.get('min_unique_documents', EVIDENCE_MIN_UNIQUE_DOCUMENTS)}\n"
            f"- 최소 비어 있지 않은 chunk 수: {thresholds.get('min_non_empty_chunks', EVIDENCE_MIN_NON_EMPTY_CHUNKS)}\n"
            f"- 허용 중복 비율: {thresholds.get('max_duplicate_ratio', EVIDENCE_MAX_DUPLICATE_RATIO)} 이하\n"
            f"- 단일 문서 최대 편중 비율: {thresholds.get('max_single_document_ratio', EVIDENCE_MAX_SINGLE_DOCUMENT_RATIO)} 이하"
        )
        st.caption("이 진단은 정확도 점수나 법적 신뢰도 점수가 아닙니다.")


def render_rag_result(
    answer: str,
    answer_status: dict[str, Any],
    results: list[dict[str, Any]],
    result_key: str = "rag",
    question_text: str = "",
    auto_save_history: bool = True,
) -> None:
    st.subheader("답변 상세")
    mode_name = answer_status.get("answer_mode", "알 수 없음")
    question_type = (
        results[0].get("question_type", "일반")
        if results
        else "일반"
    )
    reranking_applied = bool(
        results and results[0].get("reranking_applied")
    )
    situation_type = resolve_display_situation_type(question_type, answer)
    risk_level, risk_accent = get_risk_level(situation_type)
    priority_action = get_priority_action(situation_type)
    selected_model = str(
        answer_status.get(
            "model",
            answer_status.get("selected_model", GEMINI_MODEL_NAME),
        )
    )
    elapsed = float(answer_status.get("elapsed", 0.0) or 0.0)
    unique_sources = list(
        dict.fromkeys(str(item.get("source", "출처 정보 없음")) for item in results)
    )
    evidence_assessment = answer_status.get("evidence_assessment", {})
    if not isinstance(evidence_assessment, dict):
        evidence_assessment = {}
    evidence_status = str(
        evidence_assessment.get("status", EVIDENCE_STATUS_NEEDS_REVIEW)
    )
    evidence_label = str(
        evidence_assessment.get(
            "label",
            EVIDENCE_STATUS_LABELS[EVIDENCE_STATUS_NEEDS_REVIEW],
        )
    )
    evidence_tone = "success" if evidence_status == EVIDENCE_STATUS_SUFFICIENT else "warning"
    generation_time = f"{elapsed:.1f}초" if elapsed > 0 else "즉시 생성"
    core_answer, kras_answer, supplement_answer = split_answer_for_dashboard(answer)
    effective_question = question_text.strip() or extract_markdown_section(answer, "질문") or "질문 정보 없음"
    recommended_records = recommend_evidence_records(effective_question, answer)
    reference_cases = answer_status.get("reference_cases", [])
    if not isinstance(reference_cases, list):
        reference_cases = []
    live_news_cases = answer_status.get("live_news_cases", [])
    if not isinstance(live_news_cases, list):
        live_news_cases = []
    official_cases = answer_status.get("official_cases", [])
    if not isinstance(official_cases, list):
        official_cases = []
    official_case_diagnostic = answer_status.get("official_case_diagnostic", {})
    if not isinstance(official_case_diagnostic, dict):
        official_case_diagnostic = {}

    if answer_status.get("question_intent") == OUT_OF_SCOPE_INTENT:
        render_gemini_runtime_status(answer_status, selected_model, mode_name)
        with st.container(border=True):
            st.markdown(
                '<div class="card-title-row">'
                '<span class="card-title-icon card-title-icon-navy">i</span>'
                '<div class="portal-card-title">답변 범위 안내</div></div>'
                '<div class="portal-card-subtitle">광산 안전관리와 관련된 질문으로 다시 입력해 주세요.</div>',
                unsafe_allow_html=True,
            )
            st.markdown(answer)
        return

    public_mode = public_answer_mode_label(answer_status, mode_name)
    info_cols = st.columns(5)
    info_cards = [
        ("공식 근거 상태", evidence_label, "RAG 근거 충분성 휴리스틱", "✓", evidence_tone),
        ("위험 등급", risk_level, situation_type, "!", "warning"),
        ("관련 법령", f"{len(unique_sources)}개 문서", "Vector DB 검색 기준", "§", "navy"),
        ("답변 생성 시간", generation_time, "답변 준비 완료", "⏱", "teal"),
        ("답변 모드", public_mode, answer_mode_description(public_mode), "M", "blue"),
    ]
    for info_col, (title, value, description, icon, tone) in zip(info_cols, info_cards):
        with info_col:
            render_info_card(title, value, description, icon, tone)

    st.markdown(
        (
            '<div class="mscc-priority-strip">'
            f'<strong>우선 조치</strong> · {escape(priority_action)}'
            "</div>"
        ),
        unsafe_allow_html=True,
    )
    if reranking_applied:
        st.caption(
            results[0].get(
                "reranking_label",
                "광산 특화 문서 우선 정렬 적용",
            )
        )
    else:
        st.caption("기본 벡터 유사도 정렬 적용")

    render_gemini_runtime_status(answer_status, selected_model, mode_name)
    render_rag_evidence_guardrail(evidence_assessment)

    summary_lead, summary_actions = build_dashboard_summary(
        core_answer,
        situation_type,
    )
    summary_items_html = "".join(
        f"<li>{escape(item)}</li>" for item in summary_actions
    )

    answer_col, evidence_col = st.columns([3, 2], gap="large")
    with answer_col:
        with st.container(border=True):
            st.markdown(
                '<div class="card-title-row">'
                '<span class="card-title-icon card-title-icon-success">✓</span>'
                '<div class="portal-card-title">답변 요약</div></div>'
                '<div class="portal-card-subtitle">Evidence-based safety response</div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                (
                    f'<div class="portal-summary-lead">{escape(summary_lead)}</div>'
                    f'<ul class="portal-summary-list">{summary_items_html}</ul>'
                ),
                unsafe_allow_html=True,
            )

    with evidence_col:
        with st.container(border=True):
            st.markdown(
                '<div class="card-title-row">'
                '<span class="card-title-icon card-title-icon-navy">§</span>'
                '<div class="portal-card-title">관련 법령 및 문서 근거</div></div>'
                '<div class="portal-card-subtitle">검색 상위 3~5개 공식 문서</div>',
                unsafe_allow_html=True,
            )
            render_evidence_table(results)

    with st.container(border=True):
        mode_output_title = answer_status.get("mode_output_title", mode_name)
        st.markdown(
            '<div class="card-title-row">'
            '<span class="card-title-icon card-title-icon-teal">M</span>'
            f'<div class="portal-card-title">{escape(str(mode_output_title))}</div></div>'
            '<div class="portal-card-subtitle">검색 근거를 바탕으로 정리한 답변입니다.</div>',
            unsafe_allow_html=True,
        )
        st.markdown(core_answer or answer)

    with st.container(border=True):
        st.markdown(
            '<div class="card-title-row">'
            '<span class="card-title-icon card-title-icon-warning">✓</span>'
            '<div class="portal-card-title">조치 체크리스트</div></div>'
            '<div class="portal-card-subtitle">'
            "사업주·경영책임자 및 현장 관리자 확인용"
            "</div>",
            unsafe_allow_html=True,
        )
        render_checklist_table(situation_type)
        st.caption("이행 상태·이행일·담당자는 현장 확인 후 기록합니다.")

    with st.container(border=True):
        st.markdown(
            '<div class="card-title-row">'
            '<span class="card-title-icon card-title-icon-navy">▣</span>'
            '<div class="portal-card-title">KRAS식 위험성평가 기록 초안</div></div>'
            '<div class="portal-kras-note">'
            "검색 근거와 질문 상황을 바탕으로 작성한 기입 초안이며, "
            "최종 가능성·중대성·위험등급은 현장 기준에 따라 재평가해야 합니다."
            "</div>",
            unsafe_allow_html=True,
        )
        if kras_answer:
            render_kras_readable_markdown(kras_answer)
        else:
            st.info("KRAS 초안이 포함되지 않은 답변입니다. 검색 근거를 확인해 주세요.")

    if supplement_answer:
        with st.expander("Gemini 보조 답변 보기", expanded=False):
            st.markdown(supplement_answer)

    with st.expander("전체 답변 원문 보기", expanded=False):
        st.markdown(answer)

    render_major_accident_law_evidence_panel()

    render_recommended_evidence_records(recommended_records)
    st.markdown("### 사례 및 참고자료")
    news_reference_tab, official_case_tab, warning_points_tab = st.tabs(
        ["최근 뉴스 참고", "공식 재해사례", "핵심 주의사항"]
    )
    with news_reference_tab:
        render_live_news_reference_cases(live_news_cases)
    with official_case_tab:
        render_official_siren_cases(official_cases, official_case_diagnostic)
    with warning_points_tab:
        render_latest_reference_cases(
            reference_cases,
            answer_status.get("question_intent"),
            official_cases,
        )
    if auto_save_history:
        saved_history_id = auto_save_conversation_history(
            effective_question,
            answer,
            situation_type,
            risk_level,
            results,
            recommended_records,
            result_key,
            reference_cases=reference_cases,
            official_cases=official_cases,
            evidence_assessment=evidence_assessment,
            answer_mode=str(answer_status.get("selected_answer_mode", mode_name)),
        )
        if saved_history_id:
            st.caption(f"대화 이력 자동 저장 완료: {saved_history_id}")

    ribbon_items = "".join(
        (
            '<span class="portal-source-chip">'
            f'{escape(str(item.get("source", "출처 정보 없음")))} · '
            f'{escape(str(item.get("chunk_id", "정보 없음")))}'
            "</span>"
        )
        for item in results[:5]
    )
    st.markdown(
        (
            '<div class="portal-evidence-ribbon">'
            '<span class="portal-ribbon-label">문서 근거 · Evidence Ribbon</span>'
            f"{ribbon_items}</div>"
        ),
        unsafe_allow_html=True,
    )

    action_col1, action_col2, action_col3, action_spacer = st.columns([1, 1, 1.2, 3])
    if action_col1.button("인쇄", key=f"print_{result_key}", use_container_width=True):
        st.info("브라우저 인쇄 기능에서 현재 답변 화면을 인쇄할 수 있습니다.")
    if action_col2.button(
        "PDF 내보내기",
        key=f"pdf_{result_key}",
        use_container_width=True,
    ):
        st.info("브라우저 인쇄 메뉴에서 'PDF로 저장'을 선택해 주세요.")
    if action_col3.button(
        "즐겨찾기 추가",
        key=f"favorite_{result_key}",
        use_container_width=True,
    ):
        st.session_state["favorite_answer"] = {
            "answer": answer,
            "sources": unique_sources,
            "saved_at": time.strftime("%Y-%m-%d %H:%M"),
        }
        st.success("현재 답변을 세션 즐겨찾기에 추가했습니다.")

    with st.expander(f"근거 문서 상세 보기 ({len(results)}개)"):
        for result in results:
            render_evidence_card(result)


direct_tab, scenario_tab = st.tabs(["직접 질문", "질문 시나리오 테스트 모드"])

with direct_tab:
    st.subheader("질문 입력")

    demo_questions = [
        "갱내에서 메탄가스 농도가 높게 감지되면 어떻게 조치해야 해?",
        "발파 후 불발이 의심될 때 작업자는 어떻게 대응해야 해?",
        "분진이 많은 굴진 작업에서 방진마스크를 어떤 기준으로 지급하고 착용 상태를 확인해야 해?",
        "광산 현장에서 위험성평가를 실시할 때 어떤 절차로 진행해야 해?",
        "전기설비 점검 중 감전 위험이 있을 때 작업 전 확인해야 할 사항은 뭐야?",
    ]

    with st.expander("시연용 질문 예시"):
        for index, demo_question in enumerate(demo_questions, start=1):
            st.markdown(f"{index}. {demo_question}")

    example_questions = [
        "갱내 메탄가스 기준은 어떻게 확인해야 해?",
        "발파 작업 전에 현장 관리자가 확인해야 할 안전 체크리스트 알려줘",
        "갱내 작업 전에 환기 상태는 어떻게 점검해야 해?",
        "낙반 위험이 있을 때 작업을 중지해야 하는 기준은 뭐야?",
        "중대재해가 발생하면 현장 관리자는 먼저 무엇을 해야 해?",
        "중대재해처벌법상 경영책임자의 안전보건 확보의무를 요약해줘",
    ]

    selected_example = st.selectbox(
        "예시 질문 선택",
        ["직접 입력"] + example_questions,
    )

    default_question = (
        "갱내 메탄가스 기준은 어떻게 확인해야 해?"
        if selected_example == "직접 입력"
        else selected_example
    )

    question = st.text_area(
        "현장 관리자 질문",
        value=default_question,
        height=100,
        key=f"direct_question_{st.session_state['new_question_token']}",
    )

    col1, col2 = st.columns([1, 3])
    with col1:
        search_button = st.button("RAG 답변 생성", type="primary")
    with col2:
        st.caption(
            "Vector DB에서 근거 문서를 찾고 답변을 생성합니다. "
            "외부 호출이 원활하지 않을 때도 안정형 답변을 제공합니다."
        )

    if search_button:
        if not question.strip():
            st.warning("질문을 입력해 주세요.")
        elif db_error:
            st.error("Vector DB를 로드할 수 없어 검색을 진행할 수 없습니다.")
            st.caption("DB 연결 상태를 확인한 뒤 다시 시도해 주세요.")
        else:
            results, answer, answer_status, error = run_rag_flow(
                question,
                top_k,
                answer_mode,
                selected_gemini_model,
            )
            if error:
                st.error(error)
            else:
                if answer_status.get("question_intent") == OUT_OF_SCOPE_INTENT:
                    st.info("답변 범위 안내를 생성했습니다.")
                else:
                    st.success(f"Vector DB 검색 완료: 관련 근거 {len(results)}개")
                render_rag_result(
                    answer,
                    answer_status,
                    results,
                    result_key="direct",
                    question_text=question,
                )

with scenario_tab:
    st.subheader("질문 시나리오 테스트 모드")
    st.caption(
        f"현재 선택된 질문 세트: {selected_scenario_label}"
    )
    st.caption(SCENARIO_SET_DESCRIPTIONS.get(selected_scenario_label, ""))
    scenarios, scenario_error = load_question_scenarios(selected_scenario_path)
    progress_info = render_evaluation_progress(
        scenario_rows=scenarios if not scenario_error else None,
        include_auto_eval=bool(not scenario_error and len(scenarios) > 30),
    )
    render_auto_eval_summary()

    if scenario_error:
        st.error("질문 시나리오를 불러오지 못했습니다.")
        st.write(scenario_error)
    elif not scenarios:
        st.warning("질문 시나리오가 없습니다.")
    else:
        st.success(f"질문 시나리오 {len(scenarios)}개 로드 완료")

        status_by_no = progress_info.get("status_by_no", {}) if progress_info else {}
        scenario_options = {}
        for row in scenarios:
            no_text = str(row.get("번호", "")).strip()
            status_text = status_by_no.get(no_text, "미평가")
            label = f"Q{int(row.get('번호', 0)):02d} | {row.get('분류', '')} | {status_text}"
            scenario_options[label] = row

        selected_label = st.selectbox("테스트 질문 선택", list(scenario_options.keys()))
        selected_scenario = scenario_options[selected_label]

        info_col1, info_col2, info_col3 = st.columns(3)
        info_col1.metric("번호", f"Q{int(selected_scenario.get('번호', 0)):02d}")
        info_col2.metric("분류", selected_scenario.get("분류", ""))
        info_col3.metric("난이도", selected_scenario.get("난이도", ""))

        st.markdown("**질문 시나리오**")
        st.write(selected_scenario.get("질문 시나리오", ""))
        st.markdown("**기대 검색 문서**")
        st.write(selected_scenario.get("기대 검색 문서", ""))
        st.markdown("**정답에 포함되어야 할 핵심 요소**")
        st.write(selected_scenario.get("정답에 포함되어야 할 핵심 요소", ""))

        if st.button("선택한 질문으로 테스트 실행", type="primary"):
            if db_error:
                st.error("Vector DB를 로드할 수 없어 테스트를 진행할 수 없습니다.")
                st.caption("DB 연결 상태를 확인한 뒤 다시 시도해 주세요.")
            else:
                scenario_question = selected_scenario.get("질문 시나리오", "")
                results, answer, answer_status, error = run_rag_flow(
                    scenario_question,
                    top_k,
                    answer_mode,
                    selected_gemini_model,
                )
                if error:
                    st.error(error)
                else:
                    st.session_state["scenario_test_result"] = {
                        "scenario_no": selected_scenario.get("번호", ""),
                        "scenario": selected_scenario,
                        "results": results,
                        "answer": answer,
                        "answer_status": answer_status,
                        "answer_mode": short_answer_mode(answer_mode),
                        "gemini_status": classify_gemini_status(answer_status),
                    }
                    st.success(f"선택 질문 RAG 테스트 완료: 관련 근거 {len(results)}개")

        stored_result = st.session_state.get("scenario_test_result")
        if stored_result:
            stored_scenario = stored_result["scenario"]
            stored_no = int(stored_scenario.get("번호", 0))
            previous_evaluation, previous_evaluation_error = get_evaluation_row(
                stored_no
            )
            has_previous_evaluation = has_saved_evaluation_result(
                previous_evaluation
            )
            st.divider()
            st.caption(f"현재 평가 대상: Q{stored_no:02d} | {stored_scenario.get('분류', '')}")

            render_rag_result(
                stored_result["answer"],
                stored_result["answer_status"],
                stored_result["results"],
                result_key=f"scenario_{stored_no}",
                question_text=stored_scenario.get("질문 시나리오", ""),
            )

            st.subheader("평가 입력")
            with st.expander("평가 기준 보기"):
                st.info(
                    "평가 기준은 4개 항목, 총 100점 만점입니다. "
                    "환각억제는 별도 항목으로 분리하지 않고, 근거 기반성과 안전·법령 판단 정확성에서 함께 감점합니다."
                )
                st.markdown("**검색 적합성**")
                st.caption("질문과 관련된 공식 문서 및 chunk를 제대로 검색했는지 평가한다.")
                st.markdown("**근거 기반성**")
                st.caption("답변이 검색된 근거 문서와 명확히 연결되는지 평가한다.")
                st.markdown("**안전·법령 판단 정확성**")
                st.caption("위험 상황 판단, 작업 중지, 대피, 보고, 보호구, 법령·지침 적용이 정확한지 평가한다.")
                st.markdown("**실무성**")
                st.caption("현장 관리자가 답변을 보고 바로 조치할 수 있을 만큼 구체적인지 평가한다.")
                criteria_text = load_evaluation_criteria()
                if criteria_text:
                    st.markdown(criteria_text)
                else:
                    st.warning("평가 기준 파일을 찾을 수 없습니다.")

            with st.expander("Q01 평가 예시 보기"):
                q01_example_text = load_q01_evaluation_example()
                if q01_example_text:
                    st.markdown(q01_example_text)
                else:
                    st.warning("Q01 평가 예시 파일을 찾을 수 없습니다.")

            score_col1, score_col2, score_col3, score_col4 = st.columns(4)
            scores = {
                "검색_적합성": score_col1.number_input(
                    "검색 적합성 (0~25)", min_value=0, max_value=25, value=0, step=1
                ),
                "근거_기반성": score_col2.number_input(
                    "근거 기반성 (0~25)", min_value=0, max_value=25, value=0, step=1
                ),
                "안전법령_판단정확성": score_col3.number_input(
                    "안전·법령 판단 정확성 (0~25)", min_value=0, max_value=25, value=0, step=1
                ),
                "실무성": score_col4.number_input(
                    "실무성 (0~25)", min_value=0, max_value=25, value=0, step=1
                ),
            }
            memo = st.text_area("메모", height=100)
            total_score = sum(scores.values())
            judgment = calculate_judgment(total_score)

            eval_col1, eval_col2 = st.columns(2)
            eval_col1.metric("총점", total_score)
            eval_col2.metric("판정", judgment)

            st.subheader("재테스트 비교")
            current_evidence = make_evidence_summary(stored_result["results"])
            current_profile = classify_evidence_profile(current_evidence)

            if previous_evaluation_error:
                st.warning(previous_evaluation_error)

            if has_previous_evaluation and previous_evaluation:
                previous_evidence = previous_evaluation.get(
                    "검색된 주요 근거 문서",
                    "",
                )
                previous_profile = classify_evidence_profile(previous_evidence)
                previous_col, current_col = st.columns(2)

                with previous_col:
                    st.markdown("**이전 저장 결과**")
                    st.write(f"이전 상태: {previous_profile}")
                    st.metric(
                        "이전 총점",
                        format_comparison_score(
                            previous_evaluation.get("총점")
                        ),
                    )
                    st.markdown("**이전 검색된 주요 근거 문서**")
                    st.write(previous_evidence or "저장된 검색 결과가 없습니다.")

                with current_col:
                    st.markdown("**현재 재테스트 결과**")
                    st.write(f"현재 상태: {current_profile}")
                    st.metric("현재 입력 예정 총점", f"{total_score}점")
                    st.markdown("**현재 검색된 주요 근거 문서**")
                    st.write(current_evidence or "현재 검색 결과가 없습니다.")

                save_policy = st.radio(
                    "기존 평가 처리 방식",
                    ["기존 평가 유지", "기존 평가 덮어쓰기"],
                    index=0,
                    horizontal=True,
                    key=f"evaluation_save_policy_{stored_no}",
                )
                if save_policy == "기존 평가 유지":
                    st.info(
                        "기존 평가 유지가 선택되어 있습니다. "
                        "저장 버튼을 눌러도 TSV의 기존 평가 결과는 변경되지 않습니다."
                    )
                else:
                    st.warning(
                        "기존 평가 덮어쓰기가 선택되어 있습니다. "
                        "저장하면 현재 근거 문서, 점수와 메모로 기존 행이 갱신됩니다."
                    )
            else:
                st.info("이 질문에는 비교할 기존 저장 평가 결과가 없습니다.")
                st.write(f"현재 상태: {current_profile}")
                st.metric("현재 입력 예정 총점", f"{total_score}점")
                st.markdown("**현재 검색된 주요 근거 문서**")
                st.write(current_evidence or "현재 검색 결과가 없습니다.")
                save_policy = "새 평가 저장"

            if st.button("평가 결과 저장"):
                if (
                    has_previous_evaluation
                    and save_policy == "기존 평가 유지"
                ):
                    st.success(
                        f"Q{stored_no:02d} 기존 평가를 유지했습니다. "
                        "evaluation_template.tsv는 변경하지 않았습니다."
                    )
                else:
                    ok, message = update_evaluation_result(
                        stored_scenario,
                        stored_result["results"],
                        stored_result["answer"],
                        stored_result.get("answer_mode", stored_result.get("answer_status", {}).get("answer_mode", "")),
                        stored_result.get("gemini_status", classify_gemini_status(stored_result.get("answer_status", {}))),
                        scores,
                        memo,
                    )
                    if ok:
                        st.success(message)
                        st.caption(f"저장 파일: {EVALUATION_PATH}")
                        st.rerun()
                    else:
                        st.error("평가 결과 저장 실패")
                        st.write(message)


# ==============================
# 하단 설명
# ==============================
st.divider()

st.markdown(
    """
<div class="mscc-footer">
    MineSafe AI · 공식 안전 문서 검색 기반 현장 의사결정 지원
</div>
"""
    ,
    unsafe_allow_html=True,
)
