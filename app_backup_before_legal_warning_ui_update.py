from __future__ import annotations

from collections import Counter
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
import csv
from io import BytesIO
import json
import uuid
from datetime import date, datetime
from html import escape
from pathlib import Path
import os
import sys
import time
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    __import__("pysqlite3")
    sys.modules["sqlite3"] = sys.modules.pop("pysqlite3")
except Exception:
    pass

import chromadb
import streamlit as st
from dotenv import load_dotenv
from google import genai
from google.genai import types
from sentence_transformers import SentenceTransformer


# ==============================
# 기본 설정
# ==============================
ROOT_DIR = Path(__file__).resolve().parent
load_dotenv(ROOT_DIR / ".env", override=True)

VECTOR_DB_DIR = ROOT_DIR / "10_vector_db_with_major_accident_docs"
SCENARIO_PATH = ROOT_DIR / "02_질문시나리오" / "question_scenarios_30.tsv"
SCENARIO_PATH_65 = ROOT_DIR / "02_질문시나리오" / "question_scenarios_65.tsv"
SCENARIO_PATH_100 = ROOT_DIR / "02_질문시나리오" / "question_scenarios_100.tsv"
SCENARIO_PATH_110 = ROOT_DIR / "02_질문시나리오" / "question_scenarios_110.tsv"
SCENARIO_SET_OPTIONS = {
    "30개 기본 평가 세트": SCENARIO_PATH,
    "65개 1차 확장 평가 세트": SCENARIO_PATH_65,
    "100개 전체 확장 평가 세트": SCENARIO_PATH_100,
    "110개 분진·보호구 보강 평가 세트": SCENARIO_PATH_110,
}
SCENARIO_SET_DESCRIPTIONS = {
    "30개 기본 평가 세트": "기존 30개 질문 시나리오를 기준으로 한 기본 평가 세트입니다.",
    "65개 1차 확장 평가 세트": "기존 30개 질문에 추가 35개 질문을 더한 확장 세트입니다.",
    "100개 전체 확장 평가 세트": "기존 30개 질문에 작업 전 상황 중심 질문 70개를 추가하여 총 100개 질문 시나리오로 구성했습니다.",
    "110개 분진·보호구 보강 평가 세트": "기존 100개 질문에 분진 관리 5개와 보호구/PPE 5개를 추가하여 총 110개 질문 시나리오로 구성했습니다.",
}
EVALUATION_PATH = ROOT_DIR / "09_answer_tests" / "evaluation_template.tsv"
EVALUATION_CRITERIA_PATH = ROOT_DIR / "09_answer_tests" / "evaluation_criteria.md"
EVALUATION_EXAMPLE_Q01_PATH = ROOT_DIR / "09_answer_tests" / "evaluation_example_Q01.md"
AUTO_EVAL_SUMMARY_PATH = ROOT_DIR / "09_answer_tests" / "auto_eval_110_full_summary.tsv"
AUTO_EVAL_DETAIL_PATH = ROOT_DIR / "09_answer_tests" / "auto_eval_Q001_Q110.tsv"
AUTO_EVAL_BATCH_PATHS = [
    ROOT_DIR / "09_answer_tests" / "auto_eval_Q001_Q030.tsv",
    ROOT_DIR / "09_answer_tests" / "auto_eval_Q031_Q110.tsv",
]
COLLECTION_NAME = "mine_safety_docs"
EMBEDDING_MODEL_NAME = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"
GEMINI_MODEL_NAME = "gemini-2.5-flash-lite"
GEMINI_MODEL_OPTIONS = [
    "gemini-2.5-flash-lite",
    "gemini-2.5-flash",
    "gemini-2.5-pro",
]

MAJOR_ACCIDENT_DOC_TOTAL_CHUNKS = 820
MAJOR_ACCIDENT_DOC_ADDED_CHUNKS = 333
MAJOR_ACCIDENT_LAW_DOCS = [
    ("FAQ", "중대재해처벌법령 FAQ 중대산업재해"),
    ("질의회시", "중대재해처벌법 질의회시집"),
    ("해설서", "중대재해처벌법 해설서"),
    ("따라하기", "중대재해처벌법 따라하기 안내서"),
]

# 발표/평가 중에는 외부 API 때문에 화면이 오래 멈추지 않도록 짧게 제한합니다.
GEMINI_HTTP_TIMEOUT_MS = 60_000
GEMINI_RESPONSE_TIMEOUT_SECONDS = 60
GEMINI_MAX_ATTEMPTS = 3
GEMINI_RETRY_SLEEP_SECONDS = 2.0
GEMINI_MAX_OUTPUT_TOKENS = 800
GEMINI_CONTEXT_TOP_K = 3
CONTEXT_CHUNK_CHAR_LIMIT = 800

STABLE_MODE = "안정 모드: 검색 근거 기반 답변만 생성"
GEMINI_MODE = "Gemini 모드: Gemini 답변 생성 시도"
HYBRID_MODE = "하이브리드 모드: 검색 근거 답변을 먼저 표시하고 Gemini 답변도 추가 시도"

DATA_DIR = ROOT_DIR / "data"
FEATURE_OUTPUT_DIR = ROOT_DIR / "18_legal_evidence_features"
LEGAL_CHECKLIST_STATUS_PATH = DATA_DIR / "legal_checklist_status.json"
CONVERSATION_HISTORY_PATH = DATA_DIR / "conversation_history.jsonl"
LEGAL_CHECKLIST_EXPORT_PATH = FEATURE_OUTPUT_DIR / "legal_checklist_export.xlsx"
RISK_ASSESSMENT_EXPORT_PATH = FEATURE_OUTPUT_DIR / "risk_assessment_draft_export.xlsx"
CONVERSATION_HISTORY_EXPORT_PATH = FEATURE_OUTPUT_DIR / "conversation_history_export.xlsx"
FEATURE_REPORT_PATH = FEATURE_OUTPUT_DIR / "legal_evidence_history_feature_report.txt"

BLASTING_QUESTION_TYPE = "발파/불발"
BLASTING_KEYWORDS = [
    "발파",
    "불발",
    "화약",
    "폭약",
    "장약",
    "점화",
    "발파 후",
    "굴진",
    "대피",
    "위험구역",
    "출입통제",
]
BLASTING_CORE_KEYWORDS = [
    "발파",
    "불발",
    "화약",
    "폭약",
    "장약",
    "점화",
    "발파 후",
]
BLASTING_PREFERRED_SOURCE_FILES = [
    "06_광산안전법.txt",
    "08_광산안전법_시행규칙.txt",
    "09_광산안전기술기준_제10차개정_2024-12-19.txt",
    "10_광산안전업무_처리지침_2025-35_latest.txt",
]
BLASTING_PREFERRED_SOURCE_MARKERS = [
    "광산안전기술기준",
    "광산안전업무",
    "광산안전업무_처리지침",
    "광산안전법",
    "광산안전법_시행규칙",
]
BLASTING_GENERAL_SOURCE_MARKERS = [
    "안전보건관리체계",
    "산업안전보건법",
    "산업안전보건법_시행령",
    "산업안전보건법_시행규칙",
]
BLASTING_INTERNAL_SEARCH_COUNT = 20

ELECTRICAL_QUESTION_TYPE = "전기안전"
ELECTRICAL_KEYWORDS = [
    "전기설비",
    "전기",
    "누전",
    "감전",
    "접지",
    "절연",
    "누전차단기",
    "전원 차단",
    "방폭",
    "케이블",
    "배선",
    "습기",
    "물기",
    "전기안전",
]
ELECTRICAL_PREFERRED_SOURCE_FILES = [
    "05_산업안전보건법_시행규칙.txt",
    "06_광산안전법.txt",
    "07_광산안전법_시행령.txt",
    "08_광산안전법_시행규칙.txt",
    "09_광산안전기술기준_제10차개정_2024-12-19.txt",
    "10_광산안전업무_처리지침_2025-35_latest.txt",
]
ELECTRICAL_PREFERRED_SOURCE_MARKERS = [
    "광산안전기술기준",
    "광산안전법",
    "광산안전법_시행규칙",
    "광산안전업무",
    "광산안전업무_처리지침",
    "산업안전보건법_시행규칙",
]
ELECTRICAL_GENERAL_SOURCE_MARKERS = [
    "중대재해처벌법",
    "중대재해처벌법_시행령",
    "중대재해_처벌",
    "안전보건관리체계",
    "위험성평가",
]
ELECTRICAL_INTERNAL_SEARCH_COUNT = 20

PPE_DUST_QUESTION_TYPE = "보호구/PPE/분진"
PPE_DUST_KEYWORDS = [
    "안전모",
    "방진마스크",
    "보호구",
    "호흡보호구",
    "분진",
    "먼지",
    "굴진",
    "파쇄",
    "작업환경측정",
    "노출농도",
    "유해인자",
    "착용",
    "지급",
]
PPE_DUST_CORE_KEYWORDS = [
    "안전모",
    "방진마스크",
    "보호구",
    "호흡보호구",
    "분진",
    "먼지",
    "작업환경측정",
    "노출농도",
    "유해인자",
]

RISK_ASSESSMENT_QUESTION_TYPE = "위험성평가"
RISK_ASSESSMENT_KEYWORDS = [
    "위험성평가",
    "유해위험요인",
    "유해·위험요인",
    "감소대책",
    "추가 평가",
    "재검토",
    "새 장비",
    "작업 방식 변경",
    "신규 작업",
    "새로운 굴진 구간",
    "작업 전 평가",
]

ACCIDENT_RESPONSE_QUESTION_TYPE = "사고보고/응급조치/중대재해 대응"
ACCIDENT_RESPONSE_KEYWORDS = [
    "사고",
    "부상",
    "다침",
    "접촉사고",
    "사망",
    "중대부상",
    "중대재해",
    "응급조치",
    "119",
    "구조",
    "구급",
    "보고",
    "현장 보존",
    "재해조사",
    "사고조사",
    "재발방지대책",
]
ACCIDENT_RESPONSE_CORE_KEYWORDS = [
    keyword for keyword in ACCIDENT_RESPONSE_KEYWORDS
    if keyword != "보고"
]

DAILY_INSPECTION_QUESTION_TYPE = "일상 안전점검/기록"
DAILY_INSPECTION_KEYWORDS = [
    "일상 점검",
    "매일 점검",
    "안전일지",
    "기록",
    "안전관리 사항",
    "광산 관리자",
    "점검하고 기록",
]
DAILY_INSPECTION_CORE_KEYWORDS = [
    "일상 점검",
    "매일 점검",
    "안전일지",
    "안전관리 사항",
    "광산 관리자",
    "점검하고 기록",
]

PASSAGE_SAFETY_QUESTION_TYPE = "통로/조명/표지/대피로 점검"
PASSAGE_SAFETY_KEYWORDS = [
    "통로",
    "조명",
    "표지판",
    "안내표지",
    "위험표시",
    "대피로",
    "운반갱도",
    "주요 통행장소",
]

TBM_QUESTION_TYPE = "작업 전 안전회의/TBM"
TBM_KEYWORDS = [
    "작업 전 안전회의",
    "안전점검회의",
    "tbm",
    "매일 작업 시작 전",
    "관리감독자",
    "회의 내용",
    "작업 전 회의",
]

COMPLEX_RISK_QUESTION_TYPE = "복합위험 통제"
COMPLEX_RISK_KEYWORDS = [
    "복합 작업",
    "복합위험",
    "발파 후 환기",
    "환기가 충분하지",
    "운반차량 투입",
    "순서로 위험 통제",
]

RERANK_TYPE_CONFIGS = {
    BLASTING_QUESTION_TYPE: {
        "keywords": BLASTING_KEYWORDS,
        "preferred_files": BLASTING_PREFERRED_SOURCE_FILES,
        "preferred_markers": BLASTING_PREFERRED_SOURCE_MARKERS,
        "general_markers": BLASTING_GENERAL_SOURCE_MARKERS,
        "internal_count": BLASTING_INTERNAL_SEARCH_COUNT,
        "label": "광산 특화 문서 우선 정렬 적용",
        "keyword_weight": 0.35,
        "general_penalty": 0.35,
    },
    ELECTRICAL_QUESTION_TYPE: {
        "keywords": ELECTRICAL_KEYWORDS,
        "preferred_files": ELECTRICAL_PREFERRED_SOURCE_FILES,
        "preferred_markers": ELECTRICAL_PREFERRED_SOURCE_MARKERS,
        "general_markers": ELECTRICAL_GENERAL_SOURCE_MARKERS,
        "internal_count": ELECTRICAL_INTERNAL_SEARCH_COUNT,
        "label": "전기안전 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.30,
        "general_penalty": 0.45,
    },
    PPE_DUST_QUESTION_TYPE: {
        "keywords": PPE_DUST_KEYWORDS,
        "preferred_files": [
            "05_산업안전보건법_시행규칙.txt",
            "08_광산안전법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "10_광산안전업무_처리지침_2025-35_latest.txt",
            "13_2023_새로운_위험성평가_안내서.txt",
        ],
        "preferred_markers": [
            "광산안전법_시행규칙",
            "광산안전업무",
            "광산안전기술기준",
            "산업안전보건법_시행규칙",
            "위험성평가_안내서",
            "새로운_위험성평가",
        ],
        "general_markers": ["중대재해처벌법", "중대재해_처벌"],
        "internal_count": 20,
        "label": "보호구·분진 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.32,
        "general_penalty": 0.40,
    },
    RISK_ASSESSMENT_QUESTION_TYPE: {
        "keywords": RISK_ASSESSMENT_KEYWORDS,
        "preferred_files": [
            "03_산업안전보건법.txt",
            "05_산업안전보건법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "11_안전보건관리체계_구축_가이드북.txt",
            "13_2023_새로운_위험성평가_안내서.txt",
        ],
        "preferred_markers": [
            "새로운_위험성평가",
            "위험성평가_안내서",
            "안전보건관리체계_구축",
            "산업안전보건법",
            "산업안전보건법_시행규칙",
            "광산안전기술기준",
        ],
        "general_markers": ["중대재해처벌법", "중대재해_처벌"],
        "internal_count": 20,
        "label": "위험성평가 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.35,
        "general_penalty": 0.35,
    },
    ACCIDENT_RESPONSE_QUESTION_TYPE: {
        "keywords": ACCIDENT_RESPONSE_KEYWORDS,
        "preferred_files": [
            "01_중대재해_처벌_등에_관한_법률.txt",
            "02_중대재해처벌법_시행령.txt",
            "03_산업안전보건법.txt",
            "04_산업안전보건법_시행령.txt",
            "10_광산안전업무_처리지침_2025-35_latest.txt",
            "11_안전보건관리체계_구축_가이드북.txt",
        ],
        "preferred_markers": [
            "중대재해_처벌",
            "중대재해처벌법",
            "산업안전보건법",
            "산업안전보건법_시행령",
            "광산안전업무",
            "안전보건관리체계_구축",
        ],
        "general_markers": ["위험성평가_안내서", "새로운_위험성평가"],
        "internal_count": 20,
        "label": "사고보고·응급조치 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.32,
        "general_penalty": 0.30,
    },
    DAILY_INSPECTION_QUESTION_TYPE: {
        "keywords": DAILY_INSPECTION_KEYWORDS,
        "preferred_files": [
            "08_광산안전법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "10_광산안전업무_처리지침_2025-35_latest.txt",
        ],
        "preferred_markers": [
            "광산안전기술기준",
            "광산안전법_시행규칙",
            "광산안전업무",
        ],
        "general_markers": [
            "중대재해처벌법",
            "중대재해_처벌",
            "안전보건관리체계",
            "위험성평가",
        ],
        "internal_count": 20,
        "label": "광산 일상점검·기록 문서 우선 정렬 적용",
        "keyword_weight": 0.30,
        "general_penalty": 0.40,
    },
    PASSAGE_SAFETY_QUESTION_TYPE: {
        "keywords": PASSAGE_SAFETY_KEYWORDS,
        "preferred_files": [
            "08_광산안전법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "10_광산안전업무_처리지침_2025-35_latest.txt",
        ],
        "preferred_markers": [
            "광산안전기술기준",
            "광산안전법_시행규칙",
            "광산안전업무",
        ],
        "general_markers": [
            "중대재해처벌법",
            "중대재해_처벌",
            "안전보건관리체계",
            "위험성평가",
        ],
        "internal_count": 20,
        "label": "통로·조명·대피로 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.34,
        "general_penalty": 0.40,
    },
    TBM_QUESTION_TYPE: {
        "keywords": TBM_KEYWORDS,
        "preferred_files": [
            "08_광산안전법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "11_안전보건관리체계_구축_가이드북.txt",
            "13_2023_새로운_위험성평가_안내서.txt",
        ],
        "preferred_markers": [
            "새로운_위험성평가",
            "위험성평가_안내서",
            "안전보건관리체계_구축",
            "광산안전법_시행규칙",
            "광산안전기술기준",
        ],
        "general_markers": ["중대재해처벌법", "중대재해_처벌"],
        "internal_count": 20,
        "label": "작업 전 안전회의·TBM 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.38,
        "general_penalty": 0.35,
    },
    COMPLEX_RISK_QUESTION_TYPE: {
        "keywords": COMPLEX_RISK_KEYWORDS + [
            "불발",
            "잔류화약류",
            "후가스",
            "메탄",
            "산소",
            "일산화탄소",
            "환기",
            "낙반",
            "부석",
            "운반차량",
        ],
        "preferred_files": [
            "08_광산안전법_시행규칙.txt",
            "09_광산안전기술기준_제10차개정_2024-12-19.txt",
            "10_광산안전업무_처리지침_2025-35_latest.txt",
        ],
        "preferred_markers": [
            "광산안전기술기준",
            "광산안전법_시행규칙",
            "광산안전업무",
        ],
        "general_markers": [
            "중대재해처벌법",
            "중대재해_처벌",
            "안전보건관리체계",
            "위험성평가",
        ],
        "internal_count": 24,
        "label": "복합위험 통제 관련 문서 우선 정렬 적용",
        "keyword_weight": 0.30,
        "general_penalty": 0.40,
    },
}


# ==============================
# 페이지 설정
# ==============================
st.set_page_config(
    page_title="MineSafe AI",
    page_icon="⛑️",
    layout="wide",
)


def inject_dashboard_css() -> None:
    st.markdown(
        """
        <style>
        :root {
            --ms-bg: #0f172a;
            --ms-panel: #182033;
            --ms-panel-2: #1f2937;
            --ms-border: #334155;
            --ms-text: #f8fafc;
            --ms-muted: #cbd5e1;
            --ms-accent: #b45309;
            --ms-accent-hover: #92400e;
            --ms-success: #10b981;
        }

        .stApp {
            background: var(--ms-bg);
            color: var(--ms-text);
        }

        [data-testid="stHeader"] {
            background: rgba(15, 23, 42, 0.96);
            border-bottom: 1px solid var(--ms-border);
        }

        [data-testid="stSidebar"] {
            background: #111827;
            border-right: 1px solid var(--ms-border);
        }

        [data-testid="stSidebar"] * {
            color: var(--ms-text);
        }

        .block-container {
            max-width: 1500px;
            padding-top: 2rem;
            padding-bottom: 2rem;
        }

        h1, h2, h3, h4, p, li, label, [data-testid="stMarkdownContainer"] {
            color: var(--ms-text);
        }

        h1, h2, h3 {
            letter-spacing: 0;
        }

        h2 {
            font-size: 1.5rem;
            line-height: 1.3;
        }

        h3 {
            font-size: 1.3rem;
            line-height: 1.35;
        }

        h2, h3 {
            border-left: 0;
            border-bottom: 1px solid var(--ms-border);
            padding-left: 0;
            padding-bottom: 0.45rem;
            margin-top: 1.4rem;
            margin-bottom: 0.75rem;
        }

        .mscc-header {
            border-bottom: 1px solid var(--ms-border);
            padding: 0.8rem 0 1.1rem 0;
            margin-bottom: 1.1rem;
        }

        .mscc-kicker {
            color: #d97706;
            font-size: 0.76rem;
            font-weight: 650;
            margin-bottom: 0.35rem;
        }

        .mscc-title {
            color: #ffffff;
            font-size: 2.15rem;
            line-height: 1.18;
            font-weight: 760;
            margin: 0;
        }

        .mscc-title-ko {
            color: #d7dde4;
            font-size: 1.05rem;
            font-weight: 600;
            margin-top: 0.35rem;
        }

        .mscc-subtitle {
            color: var(--ms-muted);
            font-size: 0.92rem;
            margin-top: 0.55rem;
        }

        .mscc-status-card {
            min-height: 112px;
            background: var(--ms-panel);
            border: 1px solid var(--ms-border);
            border-radius: 6px;
            padding: 0.85rem 0.95rem;
            margin-bottom: 0.4rem;
            box-shadow: none;
        }

        .mscc-status-label {
            color: var(--ms-muted);
            font-size: 0.8rem;
            font-weight: 600;
            text-transform: none;
        }

        .mscc-status-value {
            color: var(--ms-text);
            font-size: 1.28rem;
            font-weight: 750;
            line-height: 1.25;
            margin-top: 0.35rem;
            overflow-wrap: anywhere;
        }

        .mscc-status-description {
            color: var(--ms-muted);
            font-size: 0.78rem;
            margin-top: 0.35rem;
        }

        .mscc-priority-strip {
            background: var(--ms-panel-2);
            border: 1px solid var(--ms-border);
            border-left: 3px solid var(--ms-accent);
            border-radius: 6px;
            padding: 0.8rem 0.95rem;
            margin: 0.35rem 0 1rem 0;
        }

        .mscc-priority-strip strong {
            color: #f1c38f;
        }

        .mscc-section-label {
            color: #d9a66f;
            font-size: 0.78rem;
            font-weight: 650;
            margin-bottom: 0.45rem;
        }

        .mscc-evidence-title {
            color: var(--ms-text);
            font-size: 1rem;
            font-weight: 720;
            margin-bottom: 0.25rem;
        }

        .mscc-evidence-meta {
            color: #a7b4c5;
            font-size: 0.78rem;
            margin-bottom: 0.65rem;
        }

        .mscc-evidence-preview {
            color: #d7dde4;
            font-size: 0.88rem;
            line-height: 1.58;
        }

        [data-testid="stMetric"] {
            background: var(--ms-panel);
            border: 1px solid var(--ms-border);
            border-radius: 6px;
            padding: 0.75rem 0.85rem;
            box-shadow: none;
        }

        [data-testid="stMetricLabel"] {
            color: var(--ms-muted);
        }

        [data-testid="stMetricValue"] {
            color: var(--ms-text);
        }

        [data-testid="stVerticalBlockBorderWrapper"] {
            background: var(--ms-panel);
            border-color: var(--ms-border);
            border-radius: 6px;
        }

        [data-testid="stExpander"] {
            background: var(--ms-panel-2);
            border: 1px solid var(--ms-border);
            border-radius: 6px;
        }

        [data-testid="stDataFrame"] {
            border: 1px solid var(--ms-border);
            border-radius: 6px;
            overflow: hidden;
        }

        [data-testid="stTabs"] [role="tablist"] {
            border-bottom: 1px solid var(--ms-border);
            gap: 0.25rem;
        }

        [data-testid="stTabs"] button[role="tab"] {
            color: var(--ms-muted);
            border-radius: 4px 4px 0 0;
            padding: 0.65rem 1rem;
        }

        [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
            color: #e5e7eb;
            background: var(--ms-panel);
            border-bottom-color: var(--ms-accent);
        }

        [data-baseweb="tab-highlight"] {
            background-color: var(--ms-accent) !important;
        }

        .stButton > button {
            border-radius: 5px;
            border: 1px solid var(--ms-accent);
            background: var(--ms-accent);
            color: #ffffff;
            font-weight: 650;
            box-shadow: none;
        }

        .stButton > button:hover {
            background: var(--ms-accent-hover);
            border-color: var(--ms-accent-hover);
            color: #ffffff;
        }

        [data-testid="stTextArea"] textarea,
        [data-testid="stTextInput"] input,
        [data-baseweb="select"] > div {
            background: var(--ms-panel);
            color: var(--ms-text);
            border-color: var(--ms-border);
            box-shadow: none;
        }

        [data-testid="stTextArea"] textarea:focus,
        [data-testid="stTextInput"] input:focus,
        [data-baseweb="select"] > div:focus-within {
            border-color: #64748b !important;
            box-shadow: 0 0 0 1px #64748b !important;
        }

        [data-testid="stAlert"] {
            background-color: var(--ms-panel-2) !important;
            border-color: var(--ms-border);
            color: var(--ms-text);
            box-shadow: none;
        }

        [data-testid="stAlert"] > div,
        div[role="alert"] {
            background-color: var(--ms-panel-2) !important;
        }

        code {
            color: #e2e8f0;
            background: #111827;
        }

        hr {
            border-color: var(--ms-border);
        }

        .mscc-sidebar-brand {
            border-bottom: 1px solid var(--ms-border);
            padding-bottom: 0.8rem;
            margin-bottom: 0.8rem;
        }

        .mscc-sidebar-brand strong {
            color: var(--ms-text);
            font-size: 1rem;
        }

        .mscc-sidebar-note {
            background: var(--ms-panel);
            border: 1px solid var(--ms-border);
            border-left: 2px solid #64748b;
            border-radius: 5px;
            color: var(--ms-muted);
            font-size: 0.78rem;
            padding: 0.65rem 0.75rem;
            margin-top: 0.75rem;
        }

        .mscc-footer {
            color: #7f8b97;
            font-size: 0.76rem;
            text-align: center;
            padding-top: 0.7rem;
        }

        @media (max-width: 800px) {
            .block-container {
                padding-left: 1rem;
                padding-right: 1rem;
            }

            .mscc-title {
                font-size: 1.65rem;
            }

            .mscc-status-card {
                min-height: auto;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <style>
        :root {
            --portal-bg: #f4f7fb;
            --portal-card: #ffffff;
            --portal-border: #dce3ec;
            --portal-text: #1f2937;
            --portal-muted: #64748b;
            --portal-blue: #1d4ed8;
            --portal-blue-dark: #173a6b;
            --portal-sidebar: #10213d;
            --portal-sidebar-soft: #172c4d;
            --portal-success: #16845b;
            --portal-warning: #b45309;
        }

        .stApp {
            background: var(--portal-bg);
            color: var(--portal-text);
        }

        [data-testid="stHeader"] {
            background: rgba(244, 247, 251, 0.96);
            border-bottom: 1px solid var(--portal-border);
        }

        [data-testid="stSidebar"] {
            background: var(--portal-sidebar);
            border-right: 1px solid #253b5d;
        }

        [data-testid="stSidebar"] * {
            color: #e8eef7;
        }

        [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] .stCaption {
            color: #c8d5e7;
        }

        [data-testid="stSidebar"] [data-baseweb="select"] > div,
        [data-testid="stSidebar"] [data-testid="stTextInput"] input {
            background: var(--portal-sidebar-soft);
            border-color: #365072;
            color: #f8fafc;
        }

        [data-testid="stSidebar"] [data-testid="stAlert"],
        [data-testid="stSidebar"] [data-testid="stExpander"] {
            background: var(--portal-sidebar-soft) !important;
            border-color: #365072;
        }

        .block-container {
            box-sizing: border-box;
            width: 100%;
            max-width: 1650px;
            padding-top: 1.75rem;
            padding-bottom: 2.5rem;
            padding-left: 2rem;
            padding-right: 2rem;
            overflow: visible;
        }

        [data-testid="stMain"] h1,
        [data-testid="stMain"] h2,
        [data-testid="stMain"] h3,
        [data-testid="stMain"] h4,
        [data-testid="stMain"] p,
        [data-testid="stMain"] li,
        [data-testid="stMain"] label,
        [data-testid="stMain"] [data-testid="stMarkdownContainer"] {
            color: var(--portal-text);
        }

        [data-testid="stMain"] h2,
        [data-testid="stMain"] h3 {
            border-bottom: 0;
            padding-bottom: 0;
        }

        .portal-page-header {
            margin-bottom: 0.5rem;
        }

        .portal-breadcrumb {
            color: #64748b;
            font-size: 0.82rem;
            margin-bottom: 0.45rem;
        }

        .portal-page-title {
            color: #172033;
            font-size: 1.8rem;
            font-weight: 760;
            line-height: 1.25;
            margin: 0;
        }

        .portal-page-subtitle {
            color: #64748b;
            font-size: 0.9rem;
            margin-top: 0.35rem;
        }

        .mscc-header {
            display: none;
        }

        .mscc-status-card {
            min-height: 112px;
            height: auto;
            background: var(--portal-card);
            border: 1px solid var(--portal-border);
            border-radius: 10px;
            padding: 1rem 1.05rem;
            margin-bottom: 0.8rem;
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
            overflow: visible;
        }

        .mscc-status-label {
            color: var(--portal-muted);
            font-size: 0.78rem;
            font-weight: 650;
        }

        .mscc-status-value {
            color: #172033;
            font-size: 1.22rem;
            font-weight: 760;
            line-height: 1.35;
            word-break: keep-all;
            overflow-wrap: anywhere;
            white-space: normal;
        }

        .mscc-status-description {
            color: var(--portal-muted);
            line-height: 1.5;
            word-break: keep-all;
            overflow-wrap: anywhere;
            white-space: normal;
        }

        .portal-card-title {
            color: #172033;
            font-size: 1.02rem;
            font-weight: 750;
            margin-bottom: 0.15rem;
        }

        .portal-card-subtitle {
            color: #64748b;
            font-size: 0.78rem;
            margin-bottom: 0.75rem;
            line-height: 1.5;
        }

        .info-card {
            min-height: 122px;
            height: auto;
            display: flex;
            align-items: flex-start;
            gap: 0.78rem;
            background: #ffffff;
            border: 1px solid #d9e2ec;
            border-radius: 12px;
            padding: 1rem;
            margin-bottom: 0.75rem;
            box-shadow: 0 3px 10px rgba(30, 52, 79, 0.06);
            overflow: visible;
        }

        .info-card-icon {
            width: 2.25rem;
            height: 2.25rem;
            flex: 0 0 2.25rem;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            border-radius: 9px;
            font-size: 1.05rem;
            font-weight: 800;
        }

        .info-card-body {
            min-width: 0;
            flex: 1;
        }

        .info-card-label {
            color: #64748b;
            font-size: 0.75rem;
            font-weight: 700;
            line-height: 1.35;
        }

        .info-card-value {
            color: #172033;
            font-size: 1.15rem;
            font-weight: 800;
            line-height: 1.35;
            margin-top: 0.28rem;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .info-card-subtitle {
            color: #718096;
            font-size: 0.72rem;
            line-height: 1.45;
            margin-top: 0.3rem;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .info-card-success {
            border-top: 3px solid #2d9b68;
        }

        .info-card-success .info-card-icon {
            color: #15734a;
            background: #e2f5eb;
        }

        .info-card-warning {
            border-top: 3px solid #d58a18;
        }

        .info-card-warning .info-card-icon {
            color: #9a5a07;
            background: #fff1d6;
        }

        .info-card-navy {
            border-top: 3px solid #315b88;
        }

        .info-card-navy .info-card-icon {
            color: #244b74;
            background: #e5edf6;
        }

        .info-card-teal {
            border-top: 3px solid #2b8790;
        }

        .info-card-teal .info-card-icon {
            color: #176a72;
            background: #e0f1f2;
        }

        .info-card-blue {
            border-top: 3px solid #3c71b5;
        }

        .info-card-blue .info-card-icon {
            color: #285f9e;
            background: #e4eefb;
        }

        .card-title-row {
            display: flex;
            align-items: center;
            gap: 0.48rem;
            margin-bottom: 0.15rem;
        }

        .card-title-icon {
            width: 1.65rem;
            height: 1.65rem;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            border-radius: 7px;
            font-size: 0.86rem;
            font-weight: 800;
        }

        .card-title-icon-success {
            color: #17744c;
            background: #dff4e8;
        }

        .card-title-icon-navy {
            color: #284f7b;
            background: #e5edf6;
        }

        .card-title-icon-warning {
            color: #995b08;
            background: #fff0d2;
        }

        .portal-summary-lead {
            color: #24364d;
            background: #f4f7fb;
            border-left: 4px solid #2f66a7;
            border-radius: 0 7px 7px 0;
            padding: 0.75rem 0.85rem;
            margin: 0.4rem 0 0.8rem 0;
            font-size: 0.9rem;
            line-height: 1.65;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .portal-summary-list {
            margin: 0.2rem 0 0 0;
            padding-left: 0;
            list-style: none;
        }

        .portal-summary-list li {
            color: #334155;
            margin: 0.35rem 0;
            line-height: 1.6;
            position: relative;
            padding-left: 1.55rem;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .portal-summary-list li::before {
            content: "✓";
            position: absolute;
            left: 0;
            top: 0.03rem;
            width: 1.05rem;
            height: 1.05rem;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            color: #ffffff;
            background: #2d9b68;
            border-radius: 50%;
            font-size: 0.67rem;
            font-weight: 800;
        }

        .portal-table-scroll {
            width: 100%;
            overflow-x: auto;
            overflow-y: visible;
            border: 1px solid #dbe3ec;
            border-radius: 8px;
            background: #ffffff;
        }

        .portal-data-table {
            width: 100%;
            min-width: 620px;
            table-layout: fixed;
            border-collapse: collapse;
            background: #ffffff;
        }

        .portal-data-table th {
            color: #29415f !important;
            background: #eef3f8 !important;
            border-bottom: 1px solid #d8e1eb !important;
            padding: 0.62rem 0.58rem !important;
            font-size: 0.76rem;
            font-weight: 720;
            line-height: 1.4;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .portal-data-table td {
            color: #334155 !important;
            background: #ffffff !important;
            border-bottom: 1px solid #e5eaf0 !important;
            padding: 0.62rem 0.58rem !important;
            font-size: 0.76rem;
            line-height: 1.55;
            vertical-align: top;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        .portal-data-table tr:last-child td {
            border-bottom: 0 !important;
        }

        .portal-evidence-table {
            min-width: 610px;
        }

        .portal-checklist-table {
            width: 100%;
            min-width: 1040px;
        }

        .portal-badge-complete {
            color: #11633f;
            background: #def5e9;
            border: 1px solid #a8dec3;
        }

        .portal-badge-progress {
            color: #8a4a0b;
            background: #fff0d8;
            border: 1px solid #efc786;
        }

        .portal-badge-wait {
            color: #526171;
            background: #edf1f5;
            border: 1px solid #d2dae3;
        }

        .status-badge,
        .law-badge {
            display: inline-block;
            border-radius: 999px;
            padding: 0.18rem 0.48rem;
            font-size: 0.68rem;
            font-weight: 750;
            line-height: 1.3;
            white-space: nowrap;
        }

        .badge-success {
            color: #11633f;
            background: #def5e9;
            border: 1px solid #a8dec3;
        }

        .badge-warning {
            color: #8a4a0b;
            background: #fff0d8;
            border: 1px solid #efc786;
        }

        .badge-muted {
            color: #526171;
            background: #edf1f5;
            border: 1px solid #d2dae3;
        }

        .badge-danger {
            color: #a02d2d;
            background: #fde7e7;
            border: 1px solid #efb8b8;
        }

        .law-badge {
            color: #284f7b;
            background: #e6eef7;
            border: 1px solid #bfd0e3;
        }

        .law-badge-guideline {
            color: #3f5f7e;
            background: #edf2f7;
            border-color: #ccd8e5;
        }

        .law-badge-notice {
            color: #6d4b13;
            background: #fff3dc;
            border-color: #ecd29e;
        }

        .law-badge-reference {
            color: #536171;
            background: #f0f3f6;
            border-color: #d4dce5;
        }

        .law-badge-major {
            color: #234c7a;
            background: #dfeaf7;
            border-color: #adc6df;
        }

        .law-badge-faq {
            color: #6d4b13;
            background: #fff3dc;
            border-color: #ecd29e;
        }

        .law-badge-reply {
            color: #315d54;
            background: #e1f2ee;
            border-color: #b7dcd4;
        }

        .law-badge-commentary {
            color: #4e4074;
            background: #ece8f7;
            border-color: #cfc4e7;
        }

        .major-law-sidebar-card {
            background: #13294a;
            border: 1px solid #365072;
            border-radius: 10px;
            padding: 0.85rem;
            margin: 0.85rem 0;
            color: #d9e3f0;
        }

        .major-law-sidebar-title {
            color: #ffffff;
            font-weight: 780;
            font-size: 0.9rem;
            margin-bottom: 0.35rem;
        }

        .major-law-sidebar-line {
            color: #c8d5e7;
            font-size: 0.76rem;
            line-height: 1.5;
            margin: 0.16rem 0;
        }

        .major-law-badge-row {
            display: flex;
            flex-wrap: wrap;
            gap: 0.32rem;
            margin: 0.5rem 0;
        }

        .major-law-mini-badge {
            display: inline-block;
            border-radius: 999px;
            padding: 0.18rem 0.48rem;
            font-size: 0.68rem;
            font-weight: 760;
            color: #dbeafe;
            background: #1f3f68;
            border: 1px solid #45668f;
        }

        .major-law-evidence-box {
            background: #ffffff;
            border: 1px solid #d9e2ec;
            border-radius: 12px;
            padding: 1rem 1.05rem;
            margin: 0.85rem 0 1rem 0;
            box-shadow: 0 8px 22px rgba(15, 23, 42, 0.06);
        }

        .major-law-evidence-title {
            color: #18324f;
            font-weight: 820;
            font-size: 1rem;
            margin-bottom: 0.35rem;
        }

        .major-law-evidence-text {
            color: #4a5f77;
            font-size: 0.86rem;
            line-height: 1.6;
            margin-bottom: 0.65rem;
        }

        .major-law-doc-list {
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 0.45rem;
        }

        .major-law-doc-item {
            background: #f8fafc;
            border: 1px solid #e1e8f0;
            border-radius: 8px;
            padding: 0.55rem 0.65rem;
            color: #29415f;
            font-size: 0.82rem;
            line-height: 1.45;
        }

        .portal-kras-note {
            color: #64748b;
            background: #f5f7fa;
            border: 1px solid #e0e6ed;
            border-radius: 7px;
            padding: 0.65rem 0.75rem;
            margin: 0.35rem 0 0.85rem 0;
            font-size: 0.78rem;
            line-height: 1.55;
            word-break: keep-all;
        }

        .portal-evidence-ribbon {
            display: flex;
            flex-wrap: wrap;
            gap: 0.45rem;
            align-items: center;
            background: #eaf1fb;
            border: 1px solid #cbd9eb;
            border-radius: 10px;
            padding: 0.8rem 0.9rem;
            margin: 0.65rem 0 0.8rem 0;
        }

        .portal-ribbon-label {
            color: #274060;
            font-size: 0.8rem;
            font-weight: 750;
            margin-right: 0.2rem;
        }

        .portal-source-chip {
            display: inline-block;
            color: #25476f;
            background: #ffffff;
            border: 1px solid #b9cbe2;
            border-radius: 999px;
            padding: 0.28rem 0.58rem;
            font-size: 0.74rem;
            line-height: 1.2;
        }

        .portal-badge {
            display: inline-block;
            border-radius: 999px;
            padding: 0.2rem 0.55rem;
            font-size: 0.72rem;
            font-weight: 700;
        }

        .portal-badge-wait {
            color: #7c4a03;
            background: #fff4d6;
            border: 1px solid #f0d28a;
        }

        .mscc-priority-strip {
            background: #fff8e8;
            border: 1px solid #ead7a2;
            border-left: 4px solid #c58a16;
            border-radius: 8px;
            color: #4b3a15;
        }

        .mscc-priority-strip strong {
            color: #795a10;
        }

        .mscc-section-label,
        .mscc-evidence-title {
            color: #1f3d64;
        }

        .mscc-evidence-meta {
            color: #64748b;
        }

        .mscc-evidence-preview {
            color: #334155;
        }

        [data-testid="stMetric"],
        [data-testid="stVerticalBlockBorderWrapper"],
        [data-testid="stExpander"] {
            background: var(--portal-card);
            border-color: #d9e2ec;
            border-radius: 12px;
            box-shadow: 0 3px 10px rgba(30, 52, 79, 0.05);
            height: auto;
            max-height: none;
            overflow: visible;
        }

        [data-testid="stVerticalBlockBorderWrapper"] > div,
        [data-testid="stExpander"] details,
        [data-testid="stExpander"] [data-testid="stExpanderDetails"] {
            height: auto;
            max-height: none;
            overflow: visible;
        }

        [data-testid="stMetricLabel"],
        [data-testid="stMetricLabel"] * {
            color: var(--portal-muted) !important;
        }

        [data-testid="stMetricValue"],
        [data-testid="stMetricValue"] * {
            color: #172033 !important;
        }

        [data-testid="stDataFrame"] {
            background: #ffffff;
            border-color: var(--portal-border);
            border-radius: 8px;
        }

        [data-testid="stTabs"] [role="tablist"] {
            border-bottom-color: var(--portal-border);
        }

        [data-testid="stTabs"] button[role="tab"] {
            color: #64748b;
            background: transparent;
        }

        [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
            color: #1e4f8f;
            background: #ffffff;
            border-bottom-color: #2f66a7;
        }

        [data-baseweb="tab-highlight"] {
            background-color: #2f66a7 !important;
        }

        .stButton > button,
        .stDownloadButton > button {
            border-radius: 7px;
            border: 1px solid #2f66a7;
            background: #2f66a7;
            color: #ffffff;
            font-weight: 650;
            box-shadow: none;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover {
            background: #234f84;
            border-color: #234f84;
            color: #ffffff;
        }

        [data-testid="stMain"] [data-testid="stTextArea"] textarea,
        [data-testid="stMain"] [data-testid="stTextInput"] input,
        [data-testid="stMain"] [data-baseweb="select"] > div,
        [data-testid="stMain"] [data-testid="stNumberInput"] input {
            background: #ffffff;
            color: var(--portal-text);
            border-color: #cbd5e1;
        }

        [data-testid="stMain"] [data-testid="stAlert"],
        [data-testid="stMain"] [data-testid="stAlert"] > div,
        [data-testid="stMain"] div[role="alert"] {
            color: #334155;
            background: #eef4fb !important;
            border-color: #cbd9eb;
        }

        [data-testid="stMain"] code {
            color: #1f3d64;
            background: #eef3f8;
        }

        [data-testid="stMain"] table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.82rem;
            table-layout: fixed;
        }

        [data-testid="stMain"] th {
            background: #edf3fa;
            color: #274060;
            border: 1px solid #d7e0eb;
            padding: 0.5rem 0.55rem;
            text-align: left;
        }

        [data-testid="stMain"] td {
            background: #ffffff;
            color: #334155;
            border: 1px solid #e1e7ef;
            padding: 0.5rem 0.55rem;
            vertical-align: top;
            line-height: 1.6;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        [data-testid="stMain"] th {
            line-height: 1.5;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        [data-testid="stMain"] [data-testid="stMarkdownContainer"] > table th:first-child,
        [data-testid="stMain"] [data-testid="stMarkdownContainer"] > table td:first-child {
            width: 180px;
        }

        [data-testid="stMain"] [data-testid="stMarkdownContainer"] {
            overflow: visible;
        }

        [data-testid="stMain"] [data-testid="stMarkdownContainer"] p,
        [data-testid="stMain"] [data-testid="stMarkdownContainer"] li {
            line-height: 1.62;
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: anywhere;
        }

        [data-testid="stMain"] [data-testid="stExpander"] summary {
            color: #29415f;
            background: #f8fafc;
        }

        [data-testid="stMain"] [data-testid="stExpander"] details {
            background: #ffffff;
        }

        .mscc-sidebar-brand {
            border-bottom: 1px solid #304766;
            padding: 0.35rem 0 1rem 0;
            margin-bottom: 0.9rem;
        }

        .mscc-sidebar-brand strong {
            color: #ffffff;
            font-size: 1.05rem;
        }

        .portal-sidebar-en {
            color: #9fb2ca;
            font-size: 0.72rem;
            line-height: 1.35;
            margin-top: 0.2rem;
        }

        .portal-nav-group {
            color: #8fa4be;
            font-size: 0.7rem;
            font-weight: 700;
            margin: 1rem 0 0.35rem 0;
        }

        .portal-nav-item {
            color: #d9e3f0;
            border-radius: 6px;
            padding: 0.42rem 0.55rem;
            font-size: 0.82rem;
            margin: 0.12rem 0;
        }

        .portal-nav-item.active {
            color: #ffffff;
            background: #264b78;
            font-weight: 700;
        }

        .portal-system-state {
            background: var(--portal-sidebar-soft);
            border: 1px solid #365072;
            border-radius: 8px;
            padding: 0.75rem;
            margin-top: 1rem;
        }

        .portal-system-state-title {
            color: #a9bdd4;
            font-size: 0.72rem;
            margin-bottom: 0.25rem;
        }

        .portal-system-state-value {
            color: #dff7eb;
            font-size: 0.86rem;
            font-weight: 720;
        }

        .portal-system-state-time {
            color: #9fb2ca;
            font-size: 0.7rem;
            margin-top: 0.3rem;
        }

        .mscc-sidebar-note {
            background: var(--portal-sidebar-soft);
            border-color: #365072;
            color: #c8d5e7;
        }

        .mscc-footer {
            color: #718096;
        }

        @media (max-width: 900px) {
            .block-container {
                padding-left: 0.8rem;
                padding-right: 0.8rem;
            }

            .portal-page-title {
                font-size: 1.5rem;
            }

            .mscc-status-card {
                min-height: auto;
            }

            [data-testid="stMain"] [data-testid="stMarkdownContainer"] > table th:first-child,
            [data-testid="stMain"] [data-testid="stMarkdownContainer"] > table td:first-child {
                width: 120px;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_status_card(
    title: str,
    value: str,
    description: str | None = None,
    accent: str = "#64748b",
) -> None:
    description_html = (
        f'<div class="mscc-status-description">{escape(description)}</div>'
        if description
        else ""
    )
    st.markdown(
        (
            '<div class="mscc-status-card">'
            f'<div class="mscc-status-label">{escape(title)}</div>'
            f'<div class="mscc-status-value">{escape(value)}</div>'
            f"{description_html}</div>"
        ),
        unsafe_allow_html=True,
    )


def get_risk_level(situation_type: str) -> tuple[str, str]:
    risk_map = {
        ACCIDENT_RESPONSE_QUESTION_TYPE: ("높음", "#b45309"),
        COMPLEX_RISK_QUESTION_TYPE: ("높음", "#b45309"),
        BLASTING_QUESTION_TYPE: ("높음", "#b45309"),
        "환기/유해가스": ("높음", "#b45309"),
        "화재/폭발": ("매우 높음", "#b45309"),
        ELECTRICAL_QUESTION_TYPE: ("중간~높음", "#64748b"),
        PPE_DUST_QUESTION_TYPE: ("중간", "#64748b"),
        RISK_ASSESSMENT_QUESTION_TYPE: ("중간", "#64748b"),
        DAILY_INSPECTION_QUESTION_TYPE: ("낮음~중간", "#64748b"),
        PASSAGE_SAFETY_QUESTION_TYPE: ("중간", "#64748b"),
        TBM_QUESTION_TYPE: ("낮음~중간", "#64748b"),
        "장비/운반": ("중간~높음", "#64748b"),
        "낙반/붕락": ("높음", "#b45309"),
        "중대재해처벌법": ("높음", "#b45309"),
    }
    return risk_map.get(situation_type, ("보통", "#64748b"))


def get_priority_action(situation_type: str) -> str:
    priority_map = {
        ACCIDENT_RESPONSE_QUESTION_TYPE: "작업중지 · 구조 및 응급조치 · 119 연락 · 2차 사고 통제",
        COMPLEX_RISK_QUESTION_TYPE: "출입통제 후 불발·공기질·환기·갱도 상태를 순서대로 확인",
        BLASTING_QUESTION_TYPE: "임의 접근 금지 · 위험구역 통제 · 발파 책임자 확인",
        "환기/유해가스": "작업중지 · 대피 · 가스 측정 · 환기 확보",
        "화재/폭발": "즉시 대피 · 비상 신고 · 인원 확인 · 안전 범위 내 초기 대응",
        ELECTRICAL_QUESTION_TYPE: "전원 차단 · 잠금 및 표지 · 접지·절연·누전 보호 확인",
        PPE_DUST_QUESTION_TYPE: "분진 저감조치 확인 후 보호구 미착용자 작업 투입 금지",
        RISK_ASSESSMENT_QUESTION_TYPE: "변경 요인을 재평가하고 감소대책 이행 후 작업 시작",
        DAILY_INSPECTION_QUESTION_TYPE: "이상 발견 즉시 보고·조치하고 안전일지에 기록",
        PASSAGE_SAFETY_QUESTION_TYPE: "통로·조명·표지·대피로 장애요인을 제거한 뒤 통행 허용",
        TBM_QUESTION_TYPE: "당일 위험요인·역할·신호·대피 절차 공유 후 작업 시작",
        "장비/운반": "에너지 차단 · 접근통제 · 장비와 작업자 위치 확인",
        "낙반/붕락": "작업중지 · 출입통제 · 대피 · 천반·부석·지보공 확인",
        "중대재해처벌법": "작업중지 · 보고 · 원인조사 · 재발방지대책 검토",
    }
    return priority_map.get(
        situation_type,
        "검색 근거와 현장 조건을 대조하고 책임자 확인 후 작업 여부를 판단",
    )


def resolve_display_situation_type(question_type: str, answer: str) -> str:
    if question_type != "일반":
        return question_type
    for candidate in [
        "환기/유해가스",
        "화재/폭발",
        "장비/운반",
        "낙반/붕락",
        "중대재해처벌법",
    ]:
        if f"상황 유형: {candidate}" in answer:
            return candidate
    return "일반"


inject_dashboard_css()

if "new_question_token" not in st.session_state:
    st.session_state["new_question_token"] = 0

page_title_col, page_action_col = st.columns([5, 1])
with page_title_col:
    st.markdown(
        """
        <div class="portal-page-header">
            <div class="portal-breadcrumb">안전 질의 및 답변 &gt; 답변 상세</div>
            <div class="portal-page-title">MineSafe AI</div>
            <div class="portal-page-subtitle">
                광산 안전 지침 및 중대재해처벌법 대응 RAG 가상 안전관리자
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
with page_action_col:
    st.write("")
    if st.button("새 질문하기", use_container_width=True, key="new_question_top"):
        st.session_state["new_question_token"] += 1
        st.session_state.pop("scenario_test_result", None)
        st.rerun()


# ==============================
# 캐시 함수
# ==============================
@st.cache_resource(show_spinner=False)
def load_embedding_model() -> SentenceTransformer:
    return SentenceTransformer(EMBEDDING_MODEL_NAME)


@st.cache_resource(show_spinner=False)
def load_chroma_collection():
    if not VECTOR_DB_DIR.exists():
        return None, f"Vector DB 폴더를 찾을 수 없습니다: {VECTOR_DB_DIR}"

    try:
        client = chromadb.PersistentClient(path=str(VECTOR_DB_DIR))
        collection = client.get_collection(name=COLLECTION_NAME)
        return collection, None
    except Exception as e:
        return None, str(e)


def get_gemini_api_key() -> str | None:
    try:
        env_key = (os.getenv("GEMINI_API_KEY") or "").strip()
        if env_key:
            return env_key
    except Exception:
        pass

    try:
        secrets = st.secrets
        secret_value = secrets.get("GEMINI_API_KEY") if hasattr(secrets, "get") else None
        if isinstance(secret_value, str):
            secret_value = secret_value.strip()
            if secret_value:
                return secret_value
    except Exception:
        pass

    return None


@st.cache_resource(show_spinner=False)
def load_gemini_client():
    api_key = get_gemini_api_key()

    if not api_key:
        return None, "Gemini API 키 없음. 로컬 .env 또는 Streamlit Cloud Secrets의 GEMINI_API_KEY를 확인해 주세요."

    try:
        http_options = types.HttpOptions(
            timeout=GEMINI_HTTP_TIMEOUT_MS,
            retry_options=types.HttpRetryOptions(
                attempts=1,
                initial_delay=0.1,
                max_delay=0.2,
                http_status_codes=[429, 500, 502, 503, 504],
            ),
        )
        client = genai.Client(api_key=api_key, http_options=http_options)
        return client, None
    except Exception as e:
        return None, str(e)


@st.cache_data(show_spinner=False)

def normalize_question_id_for_display(value, fallback_index=None):
    """질문ID를 Q001 형태로 표시하기 위한 보정 함수."""
    text = str(value).strip()

    if text and text.lower() not in ["nan", "none"]:
        if text.upper().startswith("Q"):
            digits = "".join(ch for ch in text if ch.isdigit())
            if digits:
                return f"Q{int(digits):03d}"
            return text
        if text.isdigit():
            return f"Q{int(text):03d}"

    if fallback_index is not None:
        return f"Q{int(fallback_index) + 1:03d}"

    return "Q---"


def load_question_scenarios(path: str | Path | None = None) -> tuple[list[dict[str, str]], str | None]:
    scenario_path = Path(path) if path else SCENARIO_PATH
    if not scenario_path.exists():
        return [], f"질문 시나리오 파일을 찾을 수 없습니다: {scenario_path}"

    try:
        with open(scenario_path, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f, delimiter="\t"))
        return rows, None
    except Exception as e:
        return [], str(e)


@st.cache_data(show_spinner=False)
def load_evaluation_criteria() -> str | None:
    if not EVALUATION_CRITERIA_PATH.exists():
        return None

    try:
        return EVALUATION_CRITERIA_PATH.read_text(encoding="utf-8-sig")
    except Exception:
        return None


@st.cache_data(show_spinner=False)
def load_q01_evaluation_example() -> str | None:
    if not EVALUATION_EXAMPLE_Q01_PATH.exists():
        return None

    try:
        return EVALUATION_EXAMPLE_Q01_PATH.read_text(encoding="utf-8-sig")
    except Exception:
        return None


# ==============================
# 유틸 함수
# ==============================
def get_source(meta: dict[str, Any]) -> str:
    return (
        meta.get("source")
        or meta.get("file_name")
        or meta.get("doc_name")
        or meta.get("title")
        or meta.get("source_file")
        or "출처 정보 없음"
    )


def get_chunk_id(meta: dict[str, Any]) -> str:
    value = (
        meta.get("chunk_id")
        or meta.get("id")
        or meta.get("doc_id")
        or meta.get("chunk_index")
        or meta.get("index")
        or ""
    )
    return str(value) if value != "" else "정보 없음"


def clean_text(text: str) -> str:
    text = text.replace("\n", " ").replace("\t", " ")
    return " ".join(text.split())


def format_distance(distance: Any) -> str:
    if isinstance(distance, (float, int)):
        return f"{distance:.4f}"
    return "없음"


def is_temporary_gemini_error(error_message: str) -> bool:
    message = error_message.lower()
    keywords = [
        "503",
        "unavailable",
        "high demand",
        "overloaded",
        "temporarily",
        "timeout",
        "timed out",
        "deadline",
        "429",
        "resource exhausted",
        "rate limit",
    ]
    return any(keyword in message for keyword in keywords)


def is_timeout_error(error_message: str) -> bool:
    message = error_message.lower()
    return (
        "timeout" in message
        or "timed out" in message
        or "deadline" in message
        or "제한 시간" in message
    )


def classify_question_type(question: str) -> str:
    text = clean_text(question).lower()

    if any(keyword in text for keyword in COMPLEX_RISK_KEYWORDS):
        return COMPLEX_RISK_QUESTION_TYPE
    if any(keyword in text for keyword in TBM_KEYWORDS):
        return TBM_QUESTION_TYPE
    risk_keywords = [
        keyword for keyword in RISK_ASSESSMENT_KEYWORDS
        if keyword != "신규 작업"
    ]
    if any(keyword in text for keyword in risk_keywords) or (
        "신규 작업" in text and "신규 작업자" not in text
    ):
        return RISK_ASSESSMENT_QUESTION_TYPE
    if any(keyword in text for keyword in PPE_DUST_CORE_KEYWORDS):
        return PPE_DUST_QUESTION_TYPE
    strong_accident_keywords = [
        "부상",
        "다침",
        "접촉사고",
        "사망",
        "중대부상",
        "응급조치",
        "119",
        "구조",
        "구급",
        "현장 보존",
        "재해조사",
        "사고조사",
        "재발방지대책",
    ]
    has_accident_event = (
        "사고" in text
        and any(
            keyword in text
            for keyword in ["발생", "부상", "다침", "사망", "보고", "조사"]
        )
    )
    has_major_accident_event = (
        "중대재해" in text
        and any(keyword in text for keyword in ["발생", "사망", "부상", "사고"])
    )
    if (
        any(keyword in text for keyword in strong_accident_keywords)
        or has_accident_event
        or has_major_accident_event
    ):
        return ACCIDENT_RESPONSE_QUESTION_TYPE
    if "중대재해처벌법" in text:
        return "일반"
    if any(keyword in text for keyword in DAILY_INSPECTION_CORE_KEYWORDS):
        return DAILY_INSPECTION_QUESTION_TYPE
    if (
        "기록" in text
        and any(keyword in text for keyword in ["점검", "관리자", "안전관리"])
    ):
        return DAILY_INSPECTION_QUESTION_TYPE
    has_transport_context = any(
        keyword in text
        for keyword in [
            "운반차량",
            "덤프트럭",
            "광차",
            "운반장비",
            "차량",
            "보행 작업자",
            "충돌",
        ]
    )
    if (
        any(keyword in text for keyword in PASSAGE_SAFETY_KEYWORDS)
        and not has_transport_context
    ):
        return PASSAGE_SAFETY_QUESTION_TYPE
    if any(keyword in text for keyword in ELECTRICAL_KEYWORDS):
        return ELECTRICAL_QUESTION_TYPE
    if any(keyword in text for keyword in BLASTING_CORE_KEYWORDS):
        return BLASTING_QUESTION_TYPE
    return "일반"


def is_blasting_preferred_source(source: str) -> bool:
    normalized = clean_text(source).lower()
    return any(
        marker.lower() in normalized
        for marker in BLASTING_PREFERRED_SOURCE_MARKERS
    )


def is_blasting_general_source(source: str) -> bool:
    normalized = clean_text(source).lower()
    return any(
        marker.lower() in normalized
        for marker in BLASTING_GENERAL_SOURCE_MARKERS
    )


def is_electrical_preferred_source(source: str) -> bool:
    normalized = clean_text(source).lower()
    return any(
        marker.lower() in normalized
        for marker in ELECTRICAL_PREFERRED_SOURCE_MARKERS
    )


def is_electrical_general_source(source: str) -> bool:
    normalized = clean_text(source).lower()
    return any(
        marker.lower() in normalized
        for marker in ELECTRICAL_GENERAL_SOURCE_MARKERS
    )


def squared_l2_distance(
    query_embedding: list[float],
    document_embedding: Any,
) -> float | None:
    if document_embedding is None:
        return None
    try:
        return sum(
            (float(query_value) - float(document_value)) ** 2
            for query_value, document_value in zip(
                query_embedding,
                document_embedding,
            )
        )
    except (TypeError, ValueError):
        return None


def make_search_result(
    doc: str,
    meta: dict[str, Any],
    distance: Any,
    vector_rank: int | None = None,
) -> dict[str, Any]:
    return {
        "rank": vector_rank or 0,
        "vector_rank": vector_rank,
        "source": get_source(meta),
        "chunk_id": get_chunk_id(meta),
        "distance": distance,
        "text": clean_text(doc or ""),
        "metadata": meta,
    }


def add_blasting_source_candidates(
    collection,
    query_embedding: list[float],
    candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    try:
        preferred = collection.get(
            where={"source": {"$in": BLASTING_PREFERRED_SOURCE_FILES}},
            include=["documents", "metadatas", "embeddings"],
        )
    except Exception:
        return candidates

    existing_keys = {
        (result["source"], result["chunk_id"])
        for result in candidates
    }
    documents = preferred.get("documents") or []
    metadatas = preferred.get("metadatas") or []
    embeddings = preferred.get("embeddings")
    if embeddings is None:
        embeddings = [None] * len(documents)

    for doc, meta, embedding in zip(documents, metadatas, embeddings):
        safe_meta = meta if isinstance(meta, dict) else {}
        result = make_search_result(
            doc or "",
            safe_meta,
            squared_l2_distance(query_embedding, embedding),
        )
        key = (result["source"], result["chunk_id"])
        if key not in existing_keys:
            candidates.append(result)
            existing_keys.add(key)
    return candidates


def add_electrical_source_candidates(
    collection,
    query_embedding: list[float],
    candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    try:
        preferred = collection.get(
            where={"source": {"$in": ELECTRICAL_PREFERRED_SOURCE_FILES}},
            include=["documents", "metadatas", "embeddings"],
        )
    except Exception:
        return candidates

    existing_keys = {
        (result["source"], result["chunk_id"])
        for result in candidates
    }
    documents = preferred.get("documents") or []
    metadatas = preferred.get("metadatas") or []
    embeddings = preferred.get("embeddings")
    if embeddings is None:
        embeddings = [None] * len(documents)

    for doc, meta, embedding in zip(documents, metadatas, embeddings):
        safe_meta = meta if isinstance(meta, dict) else {}
        result = make_search_result(
            doc or "",
            safe_meta,
            squared_l2_distance(query_embedding, embedding),
        )
        key = (result["source"], result["chunk_id"])
        if key not in existing_keys:
            candidates.append(result)
            existing_keys.add(key)
    return candidates


def source_matches_markers(source: str, markers: list[str]) -> bool:
    normalized = clean_text(source).lower()
    return any(marker.lower() in normalized for marker in markers)


def add_type_source_candidates(
    collection,
    query_embedding: list[float],
    candidates: list[dict[str, Any]],
    preferred_source_files: list[str],
) -> list[dict[str, Any]]:
    if not preferred_source_files:
        return candidates

    try:
        preferred = collection.get(
            where={"source": {"$in": preferred_source_files}},
            include=["documents", "metadatas", "embeddings"],
        )
    except Exception:
        return candidates

    existing_keys = {
        (result["source"], result["chunk_id"])
        for result in candidates
    }
    documents = preferred.get("documents") or []
    metadatas = preferred.get("metadatas") or []
    embeddings = preferred.get("embeddings")
    if embeddings is None:
        embeddings = [None] * len(documents)

    for doc, meta, embedding in zip(documents, metadatas, embeddings):
        safe_meta = meta if isinstance(meta, dict) else {}
        result = make_search_result(
            doc or "",
            safe_meta,
            squared_l2_distance(query_embedding, embedding),
        )
        key = (result["source"], result["chunk_id"])
        if key not in existing_keys:
            candidates.append(result)
            existing_keys.add(key)
    return candidates


def rerank_search_results(
    question: str,
    candidates: list[dict[str, Any]],
    top_k: int,
) -> list[dict[str, Any]]:
    question_type = classify_question_type(question)
    config = RERANK_TYPE_CONFIGS.get(question_type)
    reranking_applied = config is not None
    question_text = clean_text(question).lower()
    if config:
        active_keywords = [
            keyword for keyword in config["keywords"]
            if keyword in question_text
        ]
        reranking_label = config["label"]
    else:
        active_keywords = []
        reranking_label = "기본 벡터 유사도 정렬 적용"

    for result in candidates:
        source = result["source"]
        text = result["text"].lower()
        distance = result["distance"]
        score = float(distance) if isinstance(distance, (int, float)) else 99.0

        if config:
            if source in config["preferred_files"]:
                score -= 0.90
            elif source_matches_markers(
                source,
                config["preferred_markers"],
            ):
                score -= 0.55
            elif source_matches_markers(
                source,
                config["general_markers"],
            ):
                score += config["general_penalty"]

            keyword_hits = sum(keyword in text for keyword in active_keywords)
            score -= min(keyword_hits, 6) * config["keyword_weight"]

        if question_type == BLASTING_QUESTION_TYPE:
            if "불발" in question and "불발" in text:
                score -= 0.50
            if "발파 후" in question and "발파 후" in text:
                score -= 0.35
        elif question_type == ELECTRICAL_QUESTION_TYPE:
            if "누전" in question_text and "누전" in text:
                score -= 0.45
            if "감전" in question_text and "감전" in text:
                score -= 0.40
            if (
                ("습기" in question_text or "물기" in question_text)
                and ("습기" in text or "물기" in text)
            ):
                score -= 0.20
        elif question_type == RISK_ASSESSMENT_QUESTION_TYPE:
            if "위험성평가" in text:
                score -= 0.35
            if "감소대책" in text:
                score -= 0.25
        elif question_type == PPE_DUST_QUESTION_TYPE:
            if ("분진" in question_text or "먼지" in question_text) and (
                "분진" in text or "먼지" in text
            ):
                score -= 0.35
            if "보호구" in text or "방진마스크" in text:
                score -= 0.25
            direct_ppe_hits = sum(
                keyword in question_text and keyword in text
                for keyword in [
                    "안전모",
                    "방진마스크",
                    "호흡보호구",
                    "분진",
                    "먼지",
                ]
            )
            score -= min(direct_ppe_hits, 3) * 0.40
        elif question_type == ACCIDENT_RESPONSE_QUESTION_TYPE:
            if "재해조사" in text or "사고조사" in text:
                score -= 0.25
            if "재발방지" in text:
                score -= 0.25
        elif question_type == TBM_QUESTION_TYPE:
            if "tbm" in text or "작업 전 안전" in text:
                score -= 0.35
        elif question_type == COMPLEX_RISK_QUESTION_TYPE:
            complex_hits = sum(
                keyword in text
                for keyword in [
                    "불발",
                    "후가스",
                    "메탄",
                    "환기",
                    "낙반",
                    "부석",
                    "운반",
                ]
            )
            score -= min(complex_hits, 5) * 0.12

        result["rerank_score"] = score
        result["question_type"] = question_type
        result["reranking_applied"] = reranking_applied
        result["reranking_label"] = reranking_label

    if not reranking_applied:
        selected = candidates[:top_k]
    else:
        ordered = sorted(
            candidates,
            key=lambda item: (
                item["rerank_score"],
                item["vector_rank"] or 9999,
            ),
        )
        selected = []
        selected_sources = set()

        # 같은 문서의 유사 chunk가 결과를 독점하지 않도록 먼저 출처를 다양화합니다.
        for result in ordered:
            if result["source"] in selected_sources:
                continue
            selected.append(result)
            selected_sources.add(result["source"])
            if len(selected) == top_k:
                break

        if len(selected) < top_k:
            selected_keys = {
                (result["source"], result["chunk_id"])
                for result in selected
            }
            for result in ordered:
                key = (result["source"], result["chunk_id"])
                if key in selected_keys:
                    continue
                selected.append(result)
                selected_keys.add(key)
                if len(selected) == top_k:
                    break

    for rank, result in enumerate(selected, start=1):
        result["rank"] = rank
    return selected


def search_vector_db(question: str, top_k: int = 5):
    collection, error = load_chroma_collection()
    if error:
        return [], error

    model = load_embedding_model()
    query_embedding = model.encode(
        [question],
        normalize_embeddings=True,
        show_progress_bar=False,
    ).tolist()

    question_type = classify_question_type(question)
    rerank_config = RERANK_TYPE_CONFIGS.get(question_type)
    internal_count = top_k
    if rerank_config:
        internal_count = max(
            top_k,
            int(rerank_config["internal_count"]),
        )
    try:
        internal_count = min(internal_count, collection.count())
    except Exception:
        pass

    try:
        results = collection.query(
            query_embeddings=query_embedding,
            n_results=internal_count,
            include=["documents", "metadatas", "distances"],
        )
    except Exception as e:
        return [], str(e)

    docs = results.get("documents", [[]])[0]
    metas = results.get("metadatas", [[]])[0]
    distances = results.get("distances", [[]])[0]

    search_candidates = []
    for i, doc in enumerate(docs):
        meta = metas[i] if i < len(metas) and isinstance(metas[i], dict) else {}
        distance = distances[i] if i < len(distances) else None
        search_candidates.append(
            make_search_result(
                doc or "",
                meta,
                distance,
                vector_rank=i + 1,
            )
        )

    if rerank_config:
        search_candidates = add_type_source_candidates(
            collection,
            query_embedding[0],
            search_candidates,
            rerank_config["preferred_files"],
        )

    return rerank_search_results(
        question,
        search_candidates,
        top_k,
    ), None


def build_context(results: list[dict[str, Any]]) -> str:
    context_blocks = []
    for r in results[:GEMINI_CONTEXT_TOP_K]:
        evidence_text = r["text"]
        if len(evidence_text) > CONTEXT_CHUNK_CHAR_LIMIT:
            evidence_text = evidence_text[:CONTEXT_CHUNK_CHAR_LIMIT].rstrip() + "..."

        block = (
            f"[근거 {r['rank']}]\n"
            f"출처: {r['source']}\n"
            f"거리값: {format_distance(r['distance'])}\n"
            f"내용:\n{evidence_text}\n"
        )
        context_blocks.append(block)
    return "\n".join(context_blocks)


def build_prompt(question: str, results: list[dict[str, Any]]) -> str:
    context = build_context(results)
    return f"""
당신은 광산 안전 지침과 중대재해처벌법 대응을 돕는 안전관리자용 AI입니다.

반드시 아래 [검색된 근거 문서]의 내용만 근거로 답변하세요.
근거 문서에 없는 내용은 추측하지 말고 "제공된 근거 문서만으로는 확인하기 어렵습니다"라고 말하세요.
검색된 근거 외의 법령명, 조문 번호, 수치, 처벌 수위 또는 의무를 생성하지 마세요.
조문 번호가 검색 근거에 명확하지 않으면 "검색 근거 기준" 또는 "관련 문서 확인 필요"라고 표시하세요.
답변은 현장 관리자가 바로 확인할 수 있도록 체크리스트와 실행 순서 중심으로 작성하세요.
법령 해석이나 실제 현장 조치가 필요한 경우에는 최종 판단을 해당 안전관리자, 관계 기관, 전문가에게 확인해야 한다고 안내하세요.
답변 마지막에는 사용한 근거 문서명을 정리하세요.

[사용자 질문]
{question}

[검색된 근거 문서]
{context}

[답변 형식]
## 1. 핵심 답변
질문에 대한 핵심 답변을 3~5문장으로 작성합니다.

## 2. KRAS식 위험성평가 기록 초안
반드시 markdown 표로 아래 항목을 작성합니다.
- 세부 작업 내용
- 잠재위험요인
- 위험발생 상황 및 결과
- 관련 근거 / 법적 기준
- 현재 위험성: 가능성, 중대성, 위험등급
- 위험성 감소대책: 제거 → 대체 → 공학적 대책 → 관리적 대책 → 보호구/PPE 순서
- 조치 후 잔여위험성
- 기록·보고 사항
- KRAS 양식 기입용 요약표
관련 근거 / 법적 기준에는 검색된 문서명과 chunk_id만 사용하고, 확인되지 않은 조문은 만들지 않습니다.
위험성 등급은 현장 확인 전 정성 초안임을 명시합니다.

## 3. 현장 안전 체크리스트
- 작업 전 확인사항
- 작업 중 확인사항
- 이상 상황 발생 시 조치사항

## 4. 관련 법령 및 지침
검색된 근거 문서에서 확인 가능한 법령, 지침, 기준을 정리합니다.

## 5. 현장 관리자 조치사항
현장 관리자가 해야 할 일을 실행 순서로 정리합니다.

## 6. 추가 확인 필요사항
근거가 부족하거나 현장 조건에 따라 추가 확인이 필요한 사항을 적습니다.

## 7. 근거 문서
답변에 사용한 근거 문서명을 bullet 형태로 정리합니다.
""".strip()


def make_preview(text: str, limit: int = 320) -> str:
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "..."


def make_answer_summary(answer: str, limit: int = 300) -> str:
    summary = clean_text(answer)
    if len(summary) <= limit:
        return summary
    return summary[:limit].rstrip() + "..."


def make_evidence_summary(results: list[dict[str, Any]]) -> str:
    evidence_items = []
    for r in results[:5]:
        evidence_items.append(
            f"근거 {r['rank']}: {r['source']} | chunk_id {r.get('chunk_id', '정보 없음')} | 거리값 {format_distance(r['distance'])}"
        )
    return " | ".join(evidence_items)


def calculate_judgment(total_score: int) -> str:
    if total_score >= 90:
        return "매우 우수"
    if total_score >= 80:
        return "우수"
    if total_score >= 70:
        return "보통"
    if total_score >= 60:
        return "보완 필요"
    return "미흡"


def classify_safety_situation(question: str) -> str:
    question_type = classify_question_type(question)
    if question_type != "일반":
        return question_type

    text = question.lower()
    keyword_groups = [
        (ELECTRICAL_QUESTION_TYPE, ELECTRICAL_KEYWORDS),
        ("낙반/붕락", ["낙반", "붕락", "천반", "균열", "암석", "부석", "지보"]),
        (
            BLASTING_QUESTION_TYPE,
            ["발파", "불발", "화약", "폭약", "장약", "점화", "굴진"],
        ),
        ("환기/유해가스", ["유해가스", "메탄", "산소", "환기", "질식"]),
        ("화재/폭발", ["화재", "폭발", "연기"]),
        ("중대재해처벌법", ["사망", "중대재해", "경영책임자", "사업주"]),
        ("장비/운반", ["굴착기", "덤프트럭", "장비", "정비", "끼임", "운반"]),
    ]
    for situation_type, keywords in keyword_groups:
        if any(keyword in text for keyword in keywords):
            return situation_type
    return "일반 광산 안전"


def build_immediate_judgment(situation_type: str) -> str:
    judgments = {
        PPE_DUST_QUESTION_TYPE: "분진 발생 작업에는 적절한 저감조치와 보호구 확인 없이 작업자를 투입해서는 안 됩니다. 살수·집진·환기 상태, 작업환경측정 결과, 방진마스크·호흡보호구와 안전모 착용 여부를 먼저 확인해야 합니다.",
        RISK_ASSESSMENT_QUESTION_TYPE: "작업 방식, 장비 또는 작업 구간이 달라졌다면 기존 위험성평가를 그대로 적용하지 말고 변경된 유해·위험요인을 다시 평가해야 합니다. 추가 감소대책을 이행하고 작업자에게 공유한 뒤 작업을 시작해야 합니다.",
        ACCIDENT_RESPONSE_QUESTION_TYPE: "사고가 발생하면 작업중지와 부상자 구조·응급조치를 먼저 실시하고 2차 사고를 방지해야 합니다. 현장을 통제·보존하면서 책임자와 필요한 관계기관에 보고하고 원인조사와 재발방지 절차로 이어가야 합니다.",
        DAILY_INSPECTION_QUESTION_TYPE: "일상 안전점검은 작업장 구조, 환기·가스·분진, 설비, 보호구, 통로와 대피로를 빠짐없이 확인하고 이상 사항을 즉시 조치한 뒤 안전일지에 기록하는 절차로 운영해야 합니다.",
        PASSAGE_SAFETY_QUESTION_TYPE: "통로·조명·표지·대피로에 이상이 있으면 정상 통행이나 작업을 허용하지 말고 장애물과 위험요인을 먼저 제거해야 합니다. 주요 통행장소와 운반갱도를 순회 점검하고 결과를 기록·보고해야 합니다.",
        TBM_QUESTION_TYPE: "작업 전 안전회의에서는 당일 작업과 핵심 위험요인, 역할·신호·대피 절차를 작업자 전원에게 공유하고 이해 여부를 확인해야 합니다. 회의와 개선조치 이력을 남긴 후 작업을 시작해야 합니다.",
        COMPLEX_RISK_QUESTION_TYPE: "발파 후 환기 미확보 상태에서 운반차량을 투입해서는 안 됩니다. 출입통제 후 불발·공기질·환기·갱도 상태를 순서대로 확인하고 차량과 보행자 위험을 분리한 뒤 책임자 승인에 따라 단계적으로 재개해야 합니다.",
        ELECTRICAL_QUESTION_TYPE: "습기나 누전 우려가 있는 갱내 전기설비는 바로 점검을 시작하지 말고 먼저 전원을 차단해야 합니다. 임의 투입 방지, 접근통제, 접지·절연·누전차단기·케이블 상태를 확인하고 책임자가 안전을 확인한 뒤 작업을 재개해야 합니다.",
        "낙반/붕락": "낙반/붕락 위험이 의심되므로 작업을 계속하지 말고 즉시 작업중지 여부를 검토해야 합니다. 위험 구역 출입통제, 근로자 대피, 천반·부석·지보공 상태 확인 후 안전관리자 또는 책임자가 작업 재개 여부를 판단해야 합니다.",
        BLASTING_QUESTION_TYPE: "불발이 의심되면 작업자의 임의 접근을 금지하고 위험구역 출입통제와 대피를 유지해야 합니다. 발파 책임자 또는 안전관리자가 불발 여부와 현장 안전을 확인하기 전에는 굴진·정리 작업과 재출입을 허용하지 않아야 합니다.",
        "환기/유해가스": "유해가스, 산소부족, 환기 이상이 의심되면 작업을 계속하지 말고 가스 측정, 환기, 대피를 우선해야 합니다.",
        "화재/폭발": "화재, 폭발, 연기가 의심되면 작업중지, 대피, 신고, 가능한 범위의 초기 대응을 우선해야 합니다.",
        "장비/운반": "장비 운전·정비 또는 운반 위험이 의심되면 전원 차단, 에너지 격리, 임의가동 방지, 접근 통제를 우선해야 합니다.",
        "중대재해처벌법": "사망 또는 중대재해 가능성이 있는 상황에서는 작업중지, 인명 구조와 보고, 원인조사, 재발방지대책, 안전보건관리체계 이행 여부 점검이 필요합니다.",
    }
    return judgments.get(
        situation_type,
        "검색된 근거 문서를 기준으로 위험요인을 먼저 확인하고, 급박한 위험이 있으면 작업중지와 대피를 우선 검토해야 합니다.",
    )


def build_priority_actions(situation_type: str) -> list[str]:
    if situation_type == PPE_DUST_QUESTION_TYPE:
        return [
            "작업 전에 굴진·파쇄 등 분진이 발생하는 작업인지 확인합니다.",
            "살수, 집진, 환기 등 분진 저감설비와 조치가 정상인지 확인합니다.",
            "방진마스크 또는 호흡보호구가 적절히 지급되고 올바르게 착용되었는지 확인합니다.",
            "안전모 등 해당 작업에 필요한 필수 보호구 착용 상태를 확인합니다.",
            "작업환경측정 결과와 유해인자 노출기준 초과 여부를 확인합니다.",
            "필수 보호구를 착용하지 않은 작업자는 작업에 투입하지 않습니다.",
            "작업 후 쌓인 분진을 제거·청소하고 점검과 조치 결과를 기록·보고합니다.",
        ]
    if situation_type == RISK_ASSESSMENT_QUESTION_TYPE:
        return [
            "기존 위험성평가를 변경된 작업에 그대로 사용할 수 있는지 먼저 검토합니다.",
            "작업 방식 변경, 신규 장비 투입, 신규 굴진 구간 등 변경 요인을 확인합니다.",
            "변경 사항으로 새로 발생하거나 달라진 유해·위험요인을 파악합니다.",
            "파악한 위험요인의 위험성 수준을 결정합니다.",
            "기존 위험 감소대책이 실제로 효과가 있는지 검토합니다.",
            "필요한 추가 감소대책과 담당자·이행 시점을 정합니다.",
            "감소대책을 이행하고 효과를 확인한 뒤 작업을 시작합니다.",
            "평가 결과를 근로자에게 공유하고 작업 전 안전회의를 실시합니다.",
            "평가와 조치 결과를 기록하고 변경 또는 문제 발생 시 재평가합니다.",
        ]
    if situation_type == ACCIDENT_RESPONSE_QUESTION_TYPE:
        return [
            "즉시 작업을 중지하고 사고 구역의 추가 작업을 막습니다.",
            "부상자를 안전하게 구조하고 가능한 범위에서 응급조치를 실시합니다.",
            "119 또는 현장 구급 연락체계를 가동합니다.",
            "2차 사고를 막기 위해 위험원을 통제하고 관계자 외 접근을 제한합니다.",
            "구조와 추가 위험 제거에 필요한 경우를 제외하고 사고 현장을 보존합니다.",
            "안전관리자, 현장 책임자, 사업주 또는 경영책임자에게 보고합니다.",
            "사고의 성격에 따라 필요한 관계기관 보고 절차를 확인합니다.",
            "사고 경위와 직접·간접 원인을 조사합니다.",
            "조사 결과를 바탕으로 재발방지대책을 수립하고 이행합니다.",
            "안전보건관리체계와 기존 안전조치가 실제로 이행되었는지 점검합니다.",
            "보고, 구조, 통제, 조사와 개선조치 내용을 기록합니다.",
        ]
    if situation_type == DAILY_INSPECTION_QUESTION_TYPE:
        return [
            "작업장 천장, 측벽과 작업면의 이상 여부를 점검합니다.",
            "부석, 낙석 흔적과 낙반·붕괴 위험을 확인합니다.",
            "지보공의 변형, 손상과 보강 필요성을 확인합니다.",
            "환기 상태와 유해가스·분진 상태를 확인합니다.",
            "전기, 기계와 운반 설비의 이상 여부를 점검합니다.",
            "작업자에게 필요한 보호구가 지급·착용되었는지 확인합니다.",
            "통로, 조명, 안내·위험표지와 대피로 상태를 확인합니다.",
            "안전교육과 작업 전 안전회의 실시 여부를 확인합니다.",
            "이상을 발견하면 즉시 보고하고 작업중지·통제·보수 등 필요한 조치를 합니다.",
            "점검과 조치 결과를 안전일지 또는 기록문서에 남깁니다.",
        ]
    if situation_type == PASSAGE_SAFETY_QUESTION_TYPE:
        return [
            "작업자와 장비가 안전하게 통행할 수 있는 폭과 통로 확보 여부를 확인합니다.",
            "통로와 작업구간의 조명 상태와 사각지대를 확인합니다.",
            "안내표지와 위험표시가 필요한 위치에 잘 보이도록 설치되었는지 확인합니다.",
            "대피로가 확보되어 있고 장애물이나 폐쇄 구간이 없는지 확인합니다.",
            "주요 통행장소와 운반갱도를 순회 점검합니다.",
            "낙반, 붕괴, 화재, 출수 등 통행을 위협하는 위험성을 확인합니다.",
            "발견한 이상과 조치 결과를 안전일지 등에 기록하고 책임자에게 보고합니다.",
        ]
    if situation_type == TBM_QUESTION_TYPE:
        return [
            "당일 작업 내용, 작업 위치와 작업 순서를 참여자에게 공유합니다.",
            "당일 작업의 주요 유해·위험요인과 통제대책을 공유합니다.",
            "낙반, 환기, 유해가스, 발파, 운반, 전기와 분진 위험을 작업별로 확인합니다.",
            "작업자별 필수 보호구 착용 상태를 확인합니다.",
            "작업자 역할, 작업지휘 체계와 신호체계를 확인합니다.",
            "대피로, 집결지와 비상연락망을 확인합니다.",
            "전일 지적사항과 개선조치 완료 여부를 확인합니다.",
            "근로자의 질문과 현장 위험에 관한 의견을 듣고 필요한 조치를 반영합니다.",
            "참석자 서명, 사진, 회의 내용과 조치사항 등 실시 기록을 남깁니다.",
        ]
    if situation_type == COMPLEX_RISK_QUESTION_TYPE:
        return [
            "작업을 중지하고 발파·환기·운반 작업구역의 출입을 통제합니다.",
            "발파 후 불발, 잔류화약류와 발파모선 상태를 확인합니다.",
            "후가스, 메탄, 산소, 일산화탄소 등 필요한 공기질 항목을 측정합니다.",
            "환기설비의 정상 작동과 충분한 환기 여부를 확인합니다.",
            "낙반, 부석과 갱도·작업면 상태를 확인합니다.",
            "운반차량 투입 전에 통행로, 조명, 신호수와 작업자 위치를 확인합니다.",
            "차량 운행구역과 보행자 동선을 분리하고 접근을 통제합니다.",
            "각 위험요인 제거를 확인한 후 책임자 승인에 따라 작업을 단계적으로 재개합니다.",
            "측정값, 확인자, 승인과 조치 결과를 기록하고 필요한 보고를 실시합니다.",
        ]
    if situation_type == ELECTRICAL_QUESTION_TYPE:
        return [
            "작업 전에 전원을 차단하고 무전압 상태인지 확인합니다.",
            "잠금·표지 등으로 전원이 임의로 다시 투입되지 않도록 조치합니다.",
            "작업구역 접근을 통제하고 전기작업에 적합한 보호구를 착용합니다.",
            "습기 또는 물기를 제거하고 접지 상태와 절연상태를 확인합니다.",
            "누전차단기 등 보호장치의 설치 상태와 작동 여부를 확인합니다.",
            "케이블·배선의 피복 손상, 접속부 이상, 물기 접촉 여부를 점검합니다.",
            "필요한 경우 방폭형 또는 갱내용 전기설비의 현장 적합성을 확인합니다.",
            "전기안전 담당자 또는 책임자의 확인 후에만 작업과 전원 투입을 재개합니다.",
        ]
    if situation_type == "낙반/붕락":
        return [
            "작업을 계속하지 말고 즉시 작업중지 여부를 검토합니다.",
            "낙반 가능 구역의 출입을 통제하고 근로자를 안전한 장소로 대피시킵니다.",
            "천반 균열, 부석 또는 낙석 위험, 지보공 변형·손상 상태를 우선 확인합니다.",
            "갱내수, 풍화, 추가 균열 등 위험을 키울 수 있는 요인을 함께 확인합니다.",
            "안전관리자 또는 책임자의 확인 후 작업 재개 여부를 판단합니다.",
        ]
    if situation_type == BLASTING_QUESTION_TYPE:
        return [
            "불발 의심 장소에 작업자가 임의로 접근하지 못하게 합니다.",
            "위험구역을 설정해 출입을 통제하고 작업자 대피 상태를 유지합니다.",
            "발파 책임자 또는 안전관리자에게 즉시 알리고 현장 확인 절차를 따릅니다.",
            "불발 여부가 확인되기 전에는 굴진·정리 작업을 재개하지 않습니다.",
            "발파 후 재출입은 후가스, 낙반·부석, 불발화약류 등 안전 확인 이후에만 허용합니다.",
            "확인 결과와 조치 내용을 필요한 보고 체계에 따라 보고하고 기록합니다.",
        ]
    if situation_type == "환기/유해가스":
        return [
            "작업을 중지하고 작업자를 대피시킵니다.",
            "메탄, 산소, 일산화탄소 등 필요한 항목을 측정합니다.",
            "환기설비 상태와 재측정 결과를 확인한 뒤 재개 여부를 판단합니다.",
        ]
    if situation_type == "화재/폭발":
        return [
            "작업중지와 대피를 우선합니다.",
            "비상 연락, 신고, 인원 확인을 실시합니다.",
            "초기 대응은 안전이 확보되는 범위에서만 수행합니다.",
        ]
    if situation_type == "장비/운반":
        return [
            "전원 차단과 에너지 격리를 먼저 확인합니다.",
            "임의가동 방지 조치와 접근 통제를 실시합니다.",
            "정비 또는 운반 작업 재개 전 장비 상태와 작업자 위치를 확인합니다.",
        ]
    if situation_type == "중대재해처벌법":
        return [
            "작업중지, 인명 구조, 현장 보존, 보고 절차를 우선합니다.",
            "사고 원인 조사와 재발방지대책 수립 필요성을 확인합니다.",
            "경영책임자의 안전보건 확보의무와 안전보건관리체계 이행 여부를 점검합니다.",
        ]
    return [
        "검색 근거와 현장 상황을 대조합니다.",
        "급박한 위험이 있으면 작업중지, 출입통제, 대피를 먼저 검토합니다.",
        "책임자 확인 후 작업 재개 여부를 판단합니다.",
    ]


def build_check_items(situation_type: str) -> list[str]:
    if situation_type == PPE_DUST_QUESTION_TYPE:
        return [
            "분진 발생원과 살수·집진·환기설비 작동 상태",
            "작업환경측정 결과와 노출기준 초과 여부",
            "방진마스크·호흡보호구의 종류, 지급, 밀착과 착용 상태",
            "안전모 등 필수 보호구 착용과 미착용자 작업 제한",
            "작업 후 청소, 보호구 관리와 조치 기록",
        ]
    if situation_type == RISK_ASSESSMENT_QUESTION_TYPE:
        return [
            "작업 방식, 장비, 장소와 작업자 구성의 변경 여부",
            "새로운 유해·위험요인과 위험성 수준",
            "기존 대책의 효과와 추가 감소대책 필요성",
            "감소대책 이행·확인과 작업 시작 승인",
            "근로자 공유, 작업 전 회의, 기록과 재평가",
        ]
    if situation_type == ACCIDENT_RESPONSE_QUESTION_TYPE:
        return [
            "작업중지, 구조·응급조치와 119 연락 상태",
            "2차 사고 방지, 접근통제와 현장 보존 상태",
            "내부 책임자와 필요한 관계기관 보고 여부",
            "사고 원인조사와 재발방지대책 수립·이행 여부",
            "안전보건관리체계 점검과 전체 조치 기록",
        ]
    if situation_type == DAILY_INSPECTION_QUESTION_TYPE:
        return [
            "천장·측벽·작업면, 부석·낙반과 지보공 상태",
            "환기·유해가스·분진과 전기·기계·운반 설비 상태",
            "보호구, 통로·조명·표지와 대피로 상태",
            "안전교육·작업 전 회의와 이상 사항 조치 여부",
            "안전일지 또는 기록문서 작성 여부",
        ]
    if situation_type == PASSAGE_SAFETY_QUESTION_TYPE:
        return [
            "통로 폭, 장애물과 안전한 통행 가능 여부",
            "조명, 안내표지와 위험표시 상태",
            "대피로 개방 상태와 비상시 이동 가능 여부",
            "주요 통행장소·운반갱도의 낙반·화재·출수 위험",
            "순회 점검 결과의 기록과 보고 여부",
        ]
    if situation_type == TBM_QUESTION_TYPE:
        return [
            "당일 작업과 작업별 유해·위험요인 공유 여부",
            "보호구, 역할·신호체계와 대피·비상연락 확인",
            "전일 지적사항과 개선조치 완료 여부",
            "근로자 질문·의견 청취와 반영 여부",
            "참석자와 회의·조치 내용 기록 여부",
        ]
    if situation_type == COMPLEX_RISK_QUESTION_TYPE:
        return [
            "불발·잔류화약류·발파모선 확인 결과",
            "후가스·메탄·산소·일산화탄소 측정과 환기 상태",
            "낙반·부석·갱도 상태와 통행로·조명 상태",
            "차량·보행자 동선 분리와 신호수·작업자 위치",
            "책임자 승인, 단계적 재개와 조치 기록·보고",
        ]
    if situation_type == ELECTRICAL_QUESTION_TYPE:
        return [
            "전원 차단, 무전압 확인, 잠금·표지 또는 임의 투입 방지 상태",
            "접지 연결 상태와 절연 손상·열화 여부",
            "누전차단기 등 보호장치의 설치 및 시험 결과",
            "습기·물기 제거와 케이블·배선·접속부 손상 여부",
            "방폭형 또는 갱내용 전기설비 적합성 확인 필요 여부",
            "접근통제, 보호구 착용, 책임자 작업 재개 승인 여부",
        ]
    if situation_type == "낙반/붕락":
        return [
            "천반 균열 상태와 균열 확대 여부",
            "부석, 낙석 흔적, 암석 박리 가능성",
            "지보공 설치 상태, 변형, 손상, 보강 필요성",
            "갱내수, 풍화, 진동 등 추가 위험요인",
            "출입통제, 대피, 작업 재개 승인 기록",
        ]
    if situation_type == BLASTING_QUESTION_TYPE:
        return [
            "불발화약류, 잔류약, 발파모선 등 불발 의심 상태",
            "위험구역 출입통제, 감시, 작업자 대피 유지 여부",
            "발파 책임자 또는 안전관리자의 현장 확인 여부",
            "후가스, 메탄가스, 낙반, 부석 등 발파 후 추가 위험",
            "굴진·정리 작업 및 재출입 승인, 보고·기록 여부",
        ]
    return [
        "검색된 근거 문서가 현재 작업 유형과 직접 관련되는지",
        "작업중지, 대피, 보고, 보호구, 장비 점검 기준이 명시되어 있는지",
        "현장 조건상 추가 확인이 필요한 위험요인이 있는지",
        "최신 법령·지침 또는 내부 안전관리 기준 확인이 필요한지",
    ]


def infer_risk_level_for_kras(
    question: str,
    situation_type: str,
) -> tuple[str, str, str]:
    text = question.lower()
    if any(keyword in text for keyword in ["사망", "매몰", "폭발", "화재"]):
        return "높음", "매우 높음", "매우 높음"
    if situation_type in {
        "낙반/붕락",
        BLASTING_QUESTION_TYPE,
        "환기/유해가스",
        "화재/폭발",
        ACCIDENT_RESPONSE_QUESTION_TYPE,
        COMPLEX_RISK_QUESTION_TYPE,
        "중대재해처벌법",
    }:
        return "높음", "매우 높음", "매우 높음"
    if situation_type in {
        ELECTRICAL_QUESTION_TYPE,
        "장비/운반",
        PPE_DUST_QUESTION_TYPE,
        PASSAGE_SAFETY_QUESTION_TYPE,
    }:
        return "중간", "높음", "높음"
    return "중간", "중간", "중간"


def format_kras_table(headers: list[str], rows: list[list[Any]]) -> str:
    def format_cell(value: Any) -> str:
        return (
            str(value or "")
            .replace("|", "\\|")
            .replace("\r\n", " / ")
            .replace("\n", " / ")
        )

    header_line = "| " + " | ".join(format_cell(header) for header in headers) + " |"
    separator_line = "| " + " | ".join("---" for _ in headers) + " |"
    body_lines = [
        "| " + " | ".join(format_cell(value) for value in row) + " |"
        for row in rows
    ]
    return "\n".join([header_line, separator_line, *body_lines])


def build_kras_risk_assessment_section(
    question: str,
    results: list[dict[str, Any]],
    situation_type: str | None = None,
) -> str:
    situation_type = situation_type or classify_safety_situation(question)
    text = question.lower()

    if any(keyword in text for keyword in ["낙반", "붕락", "천반", "측벽", "부석", "지보"]):
        profile = "낙반/붕락"
    elif any(keyword in text for keyword in ["발파", "불발", "화약", "폭약", "장약"]):
        profile = BLASTING_QUESTION_TYPE
    elif any(keyword in text for keyword in ["메탄", "유해가스", "산소", "환기", "질식"]):
        profile = "환기/유해가스"
    elif any(keyword in text for keyword in ["전기", "누전", "감전", "접지", "절연"]):
        profile = ELECTRICAL_QUESTION_TYPE
    elif any(keyword in text for keyword in ["분진", "먼지", "방진마스크", "호흡보호구", "보호구"]):
        profile = PPE_DUST_QUESTION_TYPE
    elif any(keyword in text for keyword in ["장비", "정비", "운반", "덤프트럭", "끼임"]):
        profile = "장비/운반"
    else:
        profile = situation_type

    profiles: dict[str, dict[str, Any]] = {
        "낙반/붕락": {
            "work": "갱내 채굴·굴진 또는 점검 작업 중 천반·측벽과 지보 상태를 확인하는 작업",
            "hazards": "천반·측벽 균열, 부석·낙석, 지보재 변형, 위험구역 통제 미흡",
            "outcome": "낙반으로 작업자 매몰·중상·사망 및 구조자 2차 피해 가능",
            "measures": {
                "제거": "작업중지, 위험구역 출입금지, 안전 확보 후 부석·낙석 위험 제거",
                "대체": "인력 접근 대신 원격·무인 점검 또는 장비 사용 가능 여부 검토",
                "공학적 대책": "지보 보강, 낙석 방호, 균열·변형 감시와 갱도 안정성 확인",
                "관리적 대책": "대피·통제, 책임자 확인, 사고조사·재발방지대책, 작업재개 승인",
                "보호구/PPE": "안전모 등 작업별 필수 보호구 착용. 보호구만으로 낙반 위험을 대체하지 않음",
            },
            "residual": "지보 보강과 위험 제거 후에도 추가 낙반 가능성을 재평가하고 책임자 승인 전 작업재개 금지",
            "records": "천반·측벽·지보 점검, 작업중지·대피·통제, 사고 경위와 원인, 보강·재발방지 및 작업재개 승인 기록",
            "items": [
                ["1", "천반·측벽 균열 또는 지보재 변형", "낙반 발생, 작업자 매몰·중상·사망", "매우 높음", "작업중지, 출입통제, 지보 보강, 안정성 확인 후 작업재개"],
                ["2", "사고 후 접근통제 미흡", "추가 낙반으로 구조자 2차 피해", "높음", "위험구역 통제, 구조 인원 제한, 대피로 확보"],
                ["3", "원인조사·점검 기록 미흡", "재발방지대책 누락 및 이행 확인 곤란", "높음", "사고 경위, 직접·간접 원인, 조치와 승인 결과 기록"],
            ],
        },
        BLASTING_QUESTION_TYPE: {
            "work": "발파 후 작업면 확인, 불발 여부 점검 및 굴진·정리 작업 재개 준비",
            "hazards": "불발화약류, 잔류화약류, 발파모선, 후가스, 임의 접근",
            "outcome": "지연 폭발·폭발물 접촉, 유해가스 노출, 낙반으로 중대 부상 가능",
            "measures": {
                "제거": "불발 의심 구역 작업중지와 출입금지, 확인 전 굴진·정리 작업 금지",
                "대체": "인력 접근 전 원격 확인 또는 비접촉 점검 가능 여부 검토",
                "공학적 대책": "환기 유지, 위험구역 차단, 발파설비와 모선 상태 확인",
                "관리적 대책": "발파 책임자 확인, 대피 유지, 재출입 승인, 보고·기록",
                "보호구/PPE": "발파 후 점검 작업에 필요한 보호구 착용. 불발 통제를 보호구로 대체하지 않음",
            },
            "residual": "불발·후가스·낙반 위험 확인과 책임자 승인 후 잔여위험을 재평가",
            "records": "발파 시각, 대피·통제, 불발 확인, 환기·재출입 승인 및 후속조치 기록",
            "items": [
                ["1", "불발 또는 잔류화약류", "지연 폭발로 작업자 중상·사망", "매우 높음", "접근금지, 책임자 확인, 안전 확인 전 작업재개 금지"],
                ["2", "발파 후 후가스·환기 불충분", "유해가스 노출·질식", "높음", "환기 유지, 공기질 확인, 안전 확인 후 재출입"],
                ["3", "위험구역 통제·기록 미흡", "제3자 접근 및 동일 사고 반복", "높음", "출입통제, 승인자 지정, 확인·보고 내용 기록"],
            ],
        },
        "환기/유해가스": {
            "work": "갱내 작업 전·중 환기 상태와 유해가스·산소 상태를 확인하는 작업",
            "hazards": "메탄·일산화탄소 등 유해가스, 산소부족, 환기설비 이상",
            "outcome": "질식·중독·폭발 및 대피 지연으로 중대 피해 가능",
            "measures": {
                "제거": "작업중지와 오염 구역 출입금지, 작업자 대피",
                "대체": "인력 투입 전 원격 측정·감시 방식 적용 가능 여부 검토",
                "공학적 대책": "환기설비 가동·보수, 가스 측정·경보와 공기 흐름 확보",
                "관리적 대책": "측정 결과 확인, 재출입 승인, 비상연락·대피 절차와 기록",
                "보호구/PPE": "측정·구조 작업에 적합한 보호구를 현장 기준에 따라 선정·착용",
            },
            "residual": "환기와 재측정 후에도 측정 결과와 현장 조건을 재평가한 뒤 작업재개 결정",
            "records": "측정 항목·시각·위치·결과, 환기 조치, 대피·재출입 승인 기록",
            "items": [
                ["1", "유해가스 또는 산소부족", "질식·중독·의식상실", "매우 높음", "작업중지, 대피, 측정과 환기 후 재평가"],
                ["2", "환기설비 이상", "가스 체류·확산 및 폭발 위험 증가", "높음", "설비 점검·복구, 공기 흐름 확인"],
                ["3", "재출입 승인·기록 미흡", "위험 상태에서 조기 작업재개", "높음", "측정 결과 확인, 책임자 승인과 기록"],
            ],
        },
        ELECTRICAL_QUESTION_TYPE: {
            "work": "갱내 전기설비·케이블·배선의 점검·정비 작업",
            "hazards": "누전, 절연 손상, 접지 불량, 습기·물기, 임의 전원 투입",
            "outcome": "감전·화재·설비 손상 및 2차 사고 가능",
            "measures": {
                "제거": "작업 전 전원 차단과 무전압 확인, 습기·물기 제거",
                "대체": "무전압 점검 또는 원격 점검 방식 적용 가능 여부 검토",
                "공학적 대책": "접지·절연·누전차단기, 케이블·배선과 방폭 적합성 확인",
                "관리적 대책": "잠금·표지, 접근통제, 작업허가와 책임자 확인 후 재투입",
                "보호구/PPE": "전기작업에 적합한 절연 보호구 등 현장 지정 보호구 착용",
            },
            "residual": "보호장치 시험과 책임자 확인 후 잔여 감전·재투입 위험을 재평가",
            "records": "전원 차단·잠금, 무전압·접지·절연·차단기 확인과 재투입 승인 기록",
            "items": [
                ["1", "누전·절연 손상·습기", "감전·화재", "높음", "전원 차단, 물기 제거, 접지·절연 확인"],
                ["2", "전원 임의 재투입", "정비 작업자 감전·끼임", "높음", "잠금·표지, 접근통제, 재투입 승인"],
                ["3", "케이블·보호장치 이상", "누전 차단 실패와 설비 손상", "높음", "배선·차단기 점검과 이상 부품 조치"],
            ],
        },
        PPE_DUST_QUESTION_TYPE: {
            "work": "굴진·파쇄 등 분진 발생 작업과 보호구 지급·착용 점검",
            "hazards": "분진 비산, 환기·집진 미흡, 호흡보호구 부적합·미착용",
            "outcome": "분진 노출에 따른 건강장해 및 시야 저하로 인한 사고 가능",
            "measures": {
                "제거": "불필요한 분진 발생 작업·공정 제거와 퇴적 분진 청소",
                "대체": "저분진 공법·재료 또는 습식 작업 방식 적용 가능 여부 검토",
                "공학적 대책": "살수·집진·환기와 발생원 격리",
                "관리적 대책": "작업환경측정 확인, 노출시간 관리, 착용 교육·점검과 미착용자 작업 제한",
                "보호구/PPE": "작업과 분진 특성에 맞는 방진마스크·호흡보호구 및 안전모 지급·착용",
            },
            "residual": "저감조치 후 작업환경과 착용 상태를 다시 확인하고 노출 위험을 재평가",
            "records": "측정 결과, 저감설비 점검, 보호구 지급·착용·교체, 교육과 미착용 조치 기록",
            "items": [
                ["1", "분진 발생원 저감 미흡", "고농도 분진 노출", "높음", "살수·집진·환기와 발생원 격리"],
                ["2", "호흡보호구 부적합·미착용", "호흡기 노출 증가", "높음", "적정 보호구 지급, 밀착·착용 확인, 미착용자 투입 금지"],
                ["3", "측정·청소·기록 미흡", "노출 상태와 개선 효과 확인 곤란", "중간", "작업환경 확인, 퇴적 분진 청소와 조치 기록"],
            ],
        },
    }

    profile_data = profiles.get(
        profile,
        {
            "work": f"{situation_type} 관련 광산 현장 작업과 작업 전·중 안전 확인",
            "hazards": "작업구역의 유해·위험요인 미확인, 통제 미흡, 작업절차 미준수",
            "outcome": "작업자 부상, 설비 손상 또는 2차 사고 가능",
            "measures": {
                "제거": "위험 작업중지와 위험원 제거 가능 여부 확인",
                "대체": "위험이 낮은 작업방법·장비로 대체 가능 여부 검토",
                "공학적 대책": "방호·격리·환기·감시 등 작업 유형에 맞는 설비 대책",
                "관리적 대책": "작업절차, 출입통제, 점검, 교육, 책임자 승인과 기록",
                "보호구/PPE": "작업별 필수 보호구 선정·지급·착용 확인",
            },
            "residual": "감소대책 이행 후 현장 조건을 재확인하고 책임자가 잔여위험을 재평가",
            "records": "위험요인, 점검 결과, 감소대책, 담당자, 이행 시점과 작업재개 승인 기록",
            "items": [
                ["1", "유해·위험요인 확인 미흡", "작업자 부상 또는 설비 손상", "높음", "작업중지, 위험원 확인·제거 후 재개"],
                ["2", "출입·작업절차 통제 미흡", "제3자 노출 또는 2차 사고", "중간", "출입통제, 작업절차 공유와 책임자 확인"],
                ["3", "점검·조치 기록 미흡", "개선조치 이행 확인 곤란", "중간", "점검·조치·승인 결과 기록 및 보고"],
            ],
        },
    )

    likelihood, severity, risk_grade = infer_risk_level_for_kras(
        question,
        situation_type,
    )

    source_items: list[str] = []
    for result in results[:5]:
        source = str(result.get("source", "출처 정보 없음")).strip()
        chunk_id = str(result.get("chunk_id", "정보 없음")).strip()
        source_item = f"{source} (chunk_id: {chunk_id})"
        if source_item not in source_items:
            source_items.append(source_item)
    related_basis = " / ".join(source_items) if source_items else "검색 근거 문서 확인 필요"
    related_basis += (
        " / 조문 번호·법적 해석·처벌 수위는 검색 근거에서 명확히 확인되는 경우만 "
        "기입하며, 관계기관·전문가 확인이 필요합니다."
    )

    measures = profile_data["measures"]
    measure_text = " / ".join(
        f"**{level}:** {measures[level]}"
        for level in ["제거", "대체", "공학적 대책", "관리적 대책", "보호구/PPE"]
    )
    current_risk = (
        f"가능성 {likelihood} / 중대성 {severity} / 위험등급 {risk_grade}"
        " / 질문 상황을 기준으로 한 정성 초안이며 현장 기준에 따라 재평가해야 합니다."
    )

    detail_table = format_kras_table(
        ["항목", "기입 초안"],
        [
            ["세부 작업 내용", profile_data["work"]],
            ["잠재위험요인", profile_data["hazards"]],
            ["위험발생 상황 및 결과", profile_data["outcome"]],
            ["관련 근거 / 법적 기준", related_basis],
            ["현재 위험성", current_risk],
            ["위험성 감소대책", measure_text],
            ["조치 후 잔여위험성", profile_data["residual"]],
            ["기록·보고 사항", profile_data["records"]],
        ],
    )
    hierarchy_table = format_kras_table(
        ["우선순위", "감소대책 기입 초안"],
        [[level, measures[level]] for level in ["제거", "대체", "공학적 대책", "관리적 대책", "보호구/PPE"]],
    )
    summary_table = format_kras_table(
        ["번호", "잠재위험요인", "위험발생 상황 및 결과", "위험성", "감소대책"],
        profile_data["items"],
    )

    return "\n\n".join(
        [
            "### KRAS식 위험성평가 기록 초안",
            "> 검색 근거와 질문 상황을 바탕으로 작성한 기입 초안입니다. 실제 가능성·중대성·위험등급과 법적 판단은 현장 기준 및 관계기관·전문가 확인이 필요합니다.",
            detail_table,
            "### 위험성 감소대책 우선순위",
            hierarchy_table,
            "### KRAS 양식 기입용 요약표",
            summary_table,
        ]
    )


def short_answer_mode(answer_mode: str) -> str:
    if answer_mode == STABLE_MODE:
        return "안정 모드"
    if answer_mode == GEMINI_MODE:
        return "Gemini 모드"
    if answer_mode == HYBRID_MODE:
        return "하이브리드 모드"
    return answer_mode


def classify_gemini_status(answer_status: dict[str, Any]) -> str:
    if answer_status.get("gemini_status"):
        return str(answer_status["gemini_status"])

    structured_status = str(answer_status.get("status", "")).strip().lower()
    status_labels = {
        "success": "성공",
        "503": "503",
        "429": "429",
        "timeout": "timeout",
        "error": "기타 오류",
        "not_called": "호출 안 함",
    }
    if structured_status in status_labels:
        return status_labels[structured_status]

    if answer_status.get("mode") == "gemini":
        return "성공"

    reason = str(answer_status.get("reason", "")).lower()
    if not reason:
        return "호출 안 함"
    if "503" in reason or "unavailable" in reason or "high demand" in reason:
        return "503"
    if "429" in reason or "rate limit" in reason or "resource exhausted" in reason:
        return "429"
    if "504" in reason:
        return "504 timeout"
    if is_timeout_error(reason):
        return "timeout"
    if answer_status.get("mode") == "fallback":
        return "실패"
    return "기타 오류"


def format_gemini_status(status: str) -> str:
    if status == "503":
        return "503 UNAVAILABLE"
    if status == "429":
        return "429 RESOURCE_EXHAUSTED"
    return status


def classify_gemini_error(error_message: str) -> str:
    reason = str(error_message or "").lower()
    if "503" in reason or "unavailable" in reason or "high demand" in reason:
        return "503"
    if "429" in reason or "rate limit" in reason or "resource exhausted" in reason:
        return "429"
    if "504" in reason or is_timeout_error(reason):
        return "timeout"
    return "error"


def summarize_gemini_error(error_message: str, limit: int = 240) -> str:
    summary = clean_text(str(error_message or "알 수 없는 오류"))
    if len(summary) > limit:
        return summary[:limit].rstrip() + "..."
    return summary


def get_gemini_failure_message(status: str, error_message: str = "") -> str:
    if status == "503":
        return (
            "Gemini 서버가 일시적으로 혼잡하여 안정형 답변으로 전환했습니다. "
            "잠시 후 다시 시도하거나 다른 모델을 선택해 보세요."
        )
    if status == "429":
        return (
            "Gemini API 사용량 또는 호출 제한에 도달했습니다. "
            "잠시 후 다시 시도하거나 호출 횟수를 줄여 주세요."
        )
    if status in {"timeout", "504 timeout"}:
        return (
            "Gemini 응답 시간이 길어 제한 시간 내 완료되지 않았습니다. "
            "timeout을 늘리거나 출력 토큰 수를 줄여 주세요."
        )
    return (
        "Gemini 호출 중 오류가 발생하여 안정형 답변으로 전환했습니다. "
        f"오류 요약: {summarize_gemini_error(error_message)}"
    )


def read_evaluation_rows() -> list[dict[str, str]]:
    if not EVALUATION_PATH.exists():
        return []

    with open(EVALUATION_PATH, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f, delimiter="\t"))


def safe_score_value(value: Any) -> float:
    text = str(value or "").strip()
    if not text or text.startswith("="):
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


EVALUATION_NEW_SCORE_FIELDS = [
    "검색_적합성",
    "근거_기반성",
    "안전법령_판단정확성",
    "실무성",
]
EVALUATION_LEGACY_SCORE_FIELDS = [
    "검색정확도(0~5)",
    "답변정확도(0~5)",
    "근거성(0~5)",
    "환각억제(0~5)",
    "실무성(0~5)",
]


def normalize_evaluation_row(row: dict[str, str]) -> dict[str, str]:
    normalized = dict(row)
    for field in EVALUATION_NEW_SCORE_FIELDS:
        normalized.setdefault(field, "0")
    for field in ["총점", "판정", "메모", "검색된 주요 근거 문서", "Gemini 답변 요약", "답변 생성 방식", "Gemini 상태"]:
        normalized.setdefault(field, "")
    return normalized


def is_evaluation_completed(row: dict[str, str]) -> bool:
    normalized_row = normalize_evaluation_row(row)
    total_score = safe_score_value(normalized_row.get("총점"))
    has_memo = bool(str(normalized_row.get("메모", "")).strip())
    has_any_score = any(
        safe_score_value(normalized_row.get(field)) > 0 for field in EVALUATION_NEW_SCORE_FIELDS
    ) or any(
        safe_score_value(normalized_row.get(field)) > 0 for field in EVALUATION_LEGACY_SCORE_FIELDS
    )
    return total_score > 0 or has_memo or has_any_score


def has_saved_evaluation_result(row: dict[str, str] | None) -> bool:
    if not row:
        return False
    saved_text_fields = [
        "검색된 주요 근거 문서",
        "Gemini 답변 요약",
        "답변 생성 방식",
        "Gemini 상태",
        "메모",
    ]
    return is_evaluation_completed(row) or any(
        str(row.get(field, "")).strip()
        for field in saved_text_fields
    )


def get_evaluation_row(
    scenario_no: Any,
) -> tuple[dict[str, str] | None, str | None]:
    target_no = str(scenario_no or "").strip()
    if not target_no:
        return None, "질문 번호를 확인할 수 없습니다."
    if not EVALUATION_PATH.exists():
        return None, f"평가표 파일을 찾을 수 없습니다: {EVALUATION_PATH}"

    try:
        for row in read_evaluation_rows():
            if str(row.get("번호", "")).strip() == target_no:
                return row, None
        return None, f"평가표에서 Q{int(target_no):02d} 행을 찾을 수 없습니다."
    except Exception as e:
        return None, f"이전 평가 결과를 읽는 중 오류가 발생했습니다: {e}"


def classify_evidence_profile(evidence_summary: str) -> str:
    text = clean_text(evidence_summary).lower()
    preferred_count = sum(
        marker.lower() in text
        for marker in BLASTING_PREFERRED_SOURCE_MARKERS
    )
    general_count = sum(
        marker.lower() in text
        for marker in BLASTING_GENERAL_SOURCE_MARKERS
    )

    if preferred_count and preferred_count >= general_count:
        return "광산 특화 문서 우선 검색"
    if preferred_count:
        return "광산 특화 문서 포함"
    if general_count:
        return "일반 안전관리 문서 위주"
    if text:
        return "기타 근거 문서 중심"
    return "저장된 검색 결과 없음"


def format_comparison_score(value: Any) -> str:
    score = safe_score_value(value)
    if score.is_integer():
        return f"{int(score)}점"
    return f"{score:.1f}점"


def safe_auto_eval_value(row: dict[str, Any], *names: str, default: Any = "") -> Any:
    for name in names:
        if name in row:
            value = row.get(name)
            if value is None:
                continue
            if isinstance(value, str):
                value = value.strip()
                if value == "":
                    continue
            return value
    return default


def parse_auto_eval_score(row: dict[str, Any], *names: str, default: int = 0) -> int:
    raw_value = safe_auto_eval_value(row, *names, default=default)
    try:
        return int(float(str(raw_value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return int(default)


def normalize_auto_eval_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "question_id": normalize_question_id_for_display(
            safe_auto_eval_value(row, "question_id", "번호", default="")
        ),
        "category": str(safe_auto_eval_value(row, "category", "분류", default="")),
        "question": str(safe_auto_eval_value(row, "question", "질문", default="")),
        "검색_적합성": parse_auto_eval_score(row, "검색_적합성", default=0),
        "근거_기반성": parse_auto_eval_score(row, "근거_기반성", default=0),
        "안전법령_판단정확성": parse_auto_eval_score(row, "안전법령_판단정확성", default=0),
        "실무성": parse_auto_eval_score(row, "실무성", default=0),
        "총점": parse_auto_eval_score(row, "총점", default=0),
        "판정": str(safe_auto_eval_value(row, "판정", default="")).strip() or "미흡",
        "검토필요": str(safe_auto_eval_value(row, "검토필요", default="")).strip().upper(),
    }


def load_auto_eval_summary() -> tuple[dict[str, Any] | None, str | None]:
    if not AUTO_EVAL_SUMMARY_PATH.exists():
        return None, "아직 자동 평가 결과 파일이 없습니다. 먼저 자동 평가 스크립트를 실행해 주세요."

    try:
        with open(AUTO_EVAL_SUMMARY_PATH, "r", encoding="utf-8-sig", newline="") as f:
            summary_rows = list(csv.DictReader(f, delimiter="\t"))

        detail_rows: list[dict[str, Any]] = []
        if AUTO_EVAL_DETAIL_PATH.exists():
            with open(AUTO_EVAL_DETAIL_PATH, "r", encoding="utf-8-sig", newline="") as f:
                detail_rows = list(csv.DictReader(f, delimiter="\t"))
        else:
            for batch_path in AUTO_EVAL_BATCH_PATHS:
                if not batch_path.exists():
                    continue
                with open(batch_path, "r", encoding="utf-8-sig", newline="") as f:
                    detail_rows.extend(list(csv.DictReader(f, delimiter="\t")))

        if not detail_rows and summary_rows:
            detail_rows = summary_rows

        if not detail_rows:
            return None, "자동 평가 결과 파일이 비어 있습니다."

        normalized_rows_by_id: dict[str, dict[str, Any]] = {}
        for detail_row in detail_rows:
            normalized_row = normalize_auto_eval_row(detail_row)
            question_id = str(normalized_row.get("question_id", "")).strip()
            if question_id in {"", "Q---", "Q001-Q100", "Q001-Q110"}:
                continue
            if str(normalized_row.get("category", "")).strip() == "요약":
                continue
            normalized_rows_by_id[question_id] = normalized_row
        normalized_rows = list(normalized_rows_by_id.values())

        judgment_counts = {
            "매우 우수": 0,
            "우수": 0,
            "보통": 0,
            "보완 필요": 0,
            "미흡": 0,
        }
        for row in normalized_rows:
            judgment = str(row.get("판정", "미흡") or "미흡")
            judgment_counts[judgment] = judgment_counts.get(judgment, 0) + 1

        review_count = sum(1 for row in normalized_rows if str(row.get("검토필요", "")).upper() == "Y")
        total_score = sum(int(row.get("총점", 0)) for row in normalized_rows)
        question_count = len(normalized_rows)
        average_score = round(total_score / question_count, 2) if question_count else 0.0

        return {
            "summary_path": str(AUTO_EVAL_SUMMARY_PATH),
            "summary_rows": summary_rows,
            "rows": normalized_rows,
            "metrics": {
                "question_count": question_count,
                "average_score": average_score,
                "very_good_count": judgment_counts.get("매우 우수", 0),
                "good_count": judgment_counts.get("우수", 0),
                "average_count": judgment_counts.get("보통", 0),
                "needs_review_count": judgment_counts.get("보완 필요", 0),
                "poor_count": judgment_counts.get("미흡", 0),
                "review_needed_count": review_count,
            },
        }, None
    except Exception as e:
        return None, f"자동 평가 결과를 읽는 중 오류가 발생했습니다: {e}"


def render_auto_eval_summary() -> dict[str, Any] | None:
    st.subheader("통합 평가 결과 요약")
    st.caption(
        "전체 Q001\\~Q110, 총 110개 문항을 4개 기준 / 100점 만점으로 통합 관리합니다. "
        "Q001\\~Q030은 기존 평가를 100점 기준으로 표준화했고, "
        "Q031\\~Q110은 자동평가 결과를 통합했습니다."
    )

    auto_eval, error = load_auto_eval_summary()
    if error:
        st.info(error)
        return None

    if not auto_eval:
        st.info("자동 평가 결과를 불러올 수 없습니다.")
        return None

    metrics = auto_eval.get("metrics", {})
    summary_items = [
        ("평가 문항 수", f"{metrics.get('question_count', 0)}개"),
        ("평균 점수", f"{metrics.get('average_score', 0.0):.2f}"),
        ("매우 우수", f"{metrics.get('very_good_count', 0)}개"),
        ("우수", f"{metrics.get('good_count', 0)}개"),
        ("보통", f"{metrics.get('average_count', 0)}개"),
        ("보완 필요", f"{metrics.get('needs_review_count', 0)}개"),
        ("미흡", f"{metrics.get('poor_count', 0)}개"),
        ("검토필요", f"{metrics.get('review_needed_count', 0)}개"),
    ]

    for index, (label, value) in enumerate(summary_items):
        if index % 2 == 0:
            cols = st.columns(2)
            current_col = cols[0]
        else:
            current_col = cols[1]
        current_col.markdown(
            f"<div style='border:1px solid #e2e8f0; border-radius:8px; padding:10px 12px; margin-bottom:8px;'>"
            f"<div style='font-size:0.82rem; color:#64748b;'>{label}</div>"
            f"<div style='font-size:1.08rem; font-weight:600; color:#0f172a;'>{value}</div></div>",
            unsafe_allow_html=True,
        )

    rows = auto_eval.get("rows", [])
    low_score_rows = sorted(
        rows,
        key=lambda row: (int(row.get("총점", 0)), str(row.get("question_id", ""))),
    )[:10]

    display_rows = [
        {
            "question_id": row.get("question_id", ""),
            "category": row.get("category", ""),
            "question": row.get("question", ""),
            "검색_적합성": row.get("검색_적합성", ""),
            "근거_기반성": row.get("근거_기반성", ""),
            "안전법령_판단정확성": row.get("안전법령_판단정확성", ""),
            "실무성": row.get("실무성", ""),
            "총점": row.get("총점", ""),
            "판정": row.get("판정", ""),
            "검토필요": row.get("검토필요", ""),
        }
        for row in low_score_rows
    ]

    st.caption("총점 낮은 문항 Top 10")
    st.dataframe(
        [
            {
                "question_id": row.get("question_id", ""),
                "category": row.get("category", ""),
                "question": (row.get("question", "")[:100] + "...") if len(str(row.get("question", ""))) > 100 else row.get("question", ""),
                "총점": row.get("총점", ""),
                "판정": row.get("판정", ""),
                "검토필요": row.get("검토필요", ""),
            }
            for row in low_score_rows
        ],
        use_container_width=True,
        hide_index=True,
    )

    with st.expander("전체 자동 평가 결과 보기"):
        st.info(
            "이 표는 최종 평가가 아니라 rule-based 방식의 1차 자동 평가 결과입니다. "
            "검토필요 문항은 수동 검토 후 최종 점수로 확정하는 것을 권장합니다."
        )
        full_display_rows = [
            {
                "question_id": row.get("question_id", ""),
                "category": row.get("category", ""),
                "question": (row.get("question", "")[:100] + "...") if len(str(row.get("question", ""))) > 100 else row.get("question", ""),
                "검색_적합성": row.get("검색_적합성", ""),
                "근거_기반성": row.get("근거_기반성", ""),
                "안전법령_판단정확성": row.get("안전법령_판단정확성", ""),
                "실무성": row.get("실무성", ""),
                "총점": row.get("총점", ""),
                "판정": row.get("판정", ""),
                "검토필요": row.get("검토필요", ""),
            }
            for row in rows
        ]
        st.dataframe(full_display_rows, use_container_width=True, hide_index=True)

    return auto_eval


def load_evaluation_progress(
    scenario_rows: list[dict[str, str]] | None = None,
    include_auto_eval: bool = False,
) -> tuple[dict[str, Any] | None, str | None]:
    if not EVALUATION_PATH.exists():
        return None, f"평가표 파일을 찾을 수 없습니다: {EVALUATION_PATH}"

    try:
        manual_rows = read_evaluation_rows()
    except Exception as e:
        return None, f"평가표 파일을 읽는 중 오류가 발생했습니다: {e}"

    manual_rows_by_id = {
        normalize_question_id_for_display(row.get("번호", "")): row
        for row in manual_rows
    }

    auto_rows_by_id: dict[str, dict[str, Any]] = {}
    auto_eval_error = None
    if include_auto_eval:
        auto_eval, auto_eval_error = load_auto_eval_summary()
        if auto_eval:
            auto_rows_by_id = {
                str(row.get("question_id", "")).strip(): row
                for row in auto_eval.get("rows", [])
            }

    target_scenarios = scenario_rows or [
        {
            "번호": row.get("번호", ""),
            "분류": row.get("분류", ""),
            "난이도": row.get("난이도", ""),
            "질문 시나리오": row.get("질문", ""),
        }
        for row in manual_rows
    ]

    table_rows: list[dict[str, Any]] = []
    status_by_no: dict[str, str] = {}
    completed_count = 0

    def manual_score(
        normalized_row: dict[str, str],
        new_field: str,
        legacy_field: str,
    ) -> Any:
        new_value = normalized_row.get(new_field, "0")
        if safe_score_value(new_value) > 0:
            return new_value
        return normalized_row.get(legacy_field, "0")

    for scenario_row in target_scenarios:
        question_id = normalize_question_id_for_display(
            scenario_row.get("번호", "")
        )
        manual_row = manual_rows_by_id.get(question_id)
        auto_row = auto_rows_by_id.get(question_id)

        if auto_row:
            total_score = int(auto_row.get("총점", 0))
            judgment = str(auto_row.get("판정", "")).strip()
            review_needed = str(auto_row.get("검토필요", "")).upper() == "Y"
            evaluation_status = "완료 (검토필요)" if review_needed else "완료"
            dropdown_status = f"완료 | {total_score}점 | {judgment or '판정 없음'}"
            if review_needed:
                dropdown_status += " | 검토필요"
            display_row = {
                "번호": question_id,
                "분류": scenario_row.get("분류") or auto_row.get("category", ""),
                "난이도": scenario_row.get("난이도", ""),
                "질문": scenario_row.get("질문 시나리오") or auto_row.get("question", ""),
                "검색 적합성": auto_row.get("검색_적합성", 0),
                "근거 기반성": auto_row.get("근거_기반성", 0),
                "안전·법령 판단 정확성": auto_row.get("안전법령_판단정확성", 0),
                "실무성": auto_row.get("실무성", 0),
                "총점": total_score,
                "판정": judgment,
                "평가 여부": evaluation_status,
            }
            completed = True
        else:
            normalized_row = normalize_evaluation_row(manual_row or {})
            completed = bool(manual_row) and is_evaluation_completed(normalized_row)
            dropdown_status = "완료" if completed else "미평가"
            display_row = {
                "번호": question_id,
                "분류": scenario_row.get("분류") or normalized_row.get("분류", ""),
                "난이도": scenario_row.get("난이도") or normalized_row.get("난이도", ""),
                "질문": scenario_row.get("질문 시나리오") or normalized_row.get("질문", ""),
                "검색 적합성": manual_score(
                    normalized_row, "검색_적합성", "검색정확도(0~5)"
                ),
                "근거 기반성": manual_score(
                    normalized_row, "근거_기반성", "근거성(0~5)"
                ),
                "안전·법령 판단 정확성": manual_score(
                    normalized_row,
                    "안전법령_판단정확성",
                    "답변정확도(0~5)",
                ),
                "실무성": manual_score(
                    normalized_row, "실무성", "실무성(0~5)"
                ),
                "총점": normalized_row.get("총점", ""),
                "판정": normalized_row.get("판정", ""),
                "평가 여부": "완료" if completed else "미평가",
            }

        if completed:
            completed_count += 1
        table_rows.append(display_row)

        plain_number = str(scenario_row.get("번호", "")).strip()
        status_by_no[plain_number] = dropdown_status
        status_by_no[question_id] = dropdown_status

    total_count = len(table_rows)
    incomplete_count = total_count - completed_count
    completion_rate = (completed_count / total_count) if total_count else 0.0
    return {
        "rows": manual_rows,
        "table_rows": table_rows,
        "status_by_no": status_by_no,
        "total_count": total_count,
        "completed_count": completed_count,
        "incomplete_count": incomplete_count,
        "completion_rate": completion_rate,
        "auto_eval_error": auto_eval_error,
    }, None


def render_evaluation_progress(
    scenario_rows: list[dict[str, str]] | None = None,
    include_auto_eval: bool = False,
) -> dict[str, Any] | None:
    st.subheader("평가 진행 현황")

    progress, error = load_evaluation_progress(
        scenario_rows=scenario_rows,
        include_auto_eval=include_auto_eval,
    )
    if error:
        st.warning(error)
        return None

    if not progress:
        st.info("평가 진행 현황을 불러올 수 없습니다.")
        return None

    completed_count = sum(
        1
        for table_row in progress["table_rows"]
        if str(table_row.get("평가 여부", "")).strip().startswith("완료")
    )

    total_count = len(progress["table_rows"])
    incomplete_count = total_count - completed_count
    completion_rate = completed_count / total_count if total_count else 0.0

    progress["total_count"] = total_count
    progress["completed_count"] = completed_count
    progress["incomplete_count"] = incomplete_count
    progress["completion_rate"] = completion_rate

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("전체 질문 수", f"{total_count}개")
    col2.metric("평가 완료", f"{completed_count}개")
    col3.metric("미평가", f"{incomplete_count}개")
    col4.metric("완료율", f"{completion_rate * 100:.1f}%")

    st.progress(completion_rate)

    st.dataframe(
        progress["table_rows"],
        use_container_width=True,
        hide_index=True,
    )

    if progress.get("auto_eval_error") and include_auto_eval:
        st.warning(progress["auto_eval_error"])

    st.info(
        "전체 Q001\\~Q110은 auto_eval_Q001_Q110.tsv의 "
        "4개 기준 / 100점 만점 통합 평가 결과를 우선 표시합니다. "
        "검토필요 문항은 수동 검토 후 최종 점수로 확정하는 것을 권장합니다."
    )

    return progress

def write_evaluation_rows(rows: list[dict[str, str]]) -> None:
    headers = [
        "번호",
        "분류",
        "난이도",
        "질문",
        "기대 검색 문서",
        "검색된 주요 근거 문서",
        "Gemini 답변 요약",
        "답변 생성 방식",
        "Gemini 상태",
        "검색_적합성",
        "근거_기반성",
        "안전법령_판단정확성",
        "실무성",
        "총점",
        "판정",
        "메모",
        "검색정확도(0~5)",
        "답변정확도(0~5)",
        "근거성(0~5)",
        "환각억제(0~5)",
        "실무성(0~5)",
    ]

    normalized_rows = [normalize_evaluation_row(row) for row in rows]
    EVALUATION_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(EVALUATION_PATH, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(normalized_rows)


def update_evaluation_result(
    scenario: dict[str, str],
    results: list[dict[str, Any]],
    answer: str,
    answer_mode: str,
    gemini_status: str,
    scores: dict[str, int],
    memo: str,
) -> tuple[bool, str]:
    try:
        rows = read_evaluation_rows()
        if not rows:
            scenarios, scenario_error = load_question_scenarios(selected_scenario_path)
            if scenario_error:
                return False, scenario_error
            rows = [
                {
                    "번호": r.get("번호", ""),
                    "분류": r.get("분류", ""),
                    "난이도": r.get("난이도", ""),
                    "질문": r.get("질문 시나리오", ""),
                    "기대 검색 문서": r.get("기대 검색 문서", ""),
                    "검색된 주요 근거 문서": "",
                    "Gemini 답변 요약": "",
                    "답변 생성 방식": "",
                    "Gemini 상태": "",
                    "검색_적합성": "0",
                    "근거_기반성": "0",
                    "안전법령_판단정확성": "0",
                    "실무성": "0",
                    "총점": "",
                    "판정": "",
                    "메모": "",
                }
                for r in scenarios
            ]

        normalized_scores = {
            "검색_적합성": int(scores.get("검색_적합성", scores.get("검색정확도(0~5)", 0))),
            "근거_기반성": int(scores.get("근거_기반성", scores.get("근거성(0~5)", 0))),
            "안전법령_판단정확성": int(scores.get("안전법령_판단정확성", scores.get("답변정확도(0~5)", 0))),
            "실무성": int(scores.get("실무성", scores.get("실무성(0~5)", 0))),
        }
        total_score = sum(normalized_scores.values())
        judgment = calculate_judgment(total_score)
        target_no = str(scenario.get("번호", "")).strip()
        updated = False

        for row in rows:
            row.setdefault("답변 생성 방식", "")
            row.setdefault("Gemini 상태", "")

        for row in rows:
            if str(row.get("번호", "")).strip() == target_no:
                row["검색된 주요 근거 문서"] = make_evidence_summary(results)
                row["Gemini 답변 요약"] = make_answer_summary(answer)
                row["답변 생성 방식"] = answer_mode
                row["Gemini 상태"] = gemini_status
                row["검색_적합성"] = str(normalized_scores["검색_적합성"])
                row["근거_기반성"] = str(normalized_scores["근거_기반성"])
                row["안전법령_판단정확성"] = str(normalized_scores["안전법령_판단정확성"])
                row["실무성"] = str(normalized_scores["실무성"])
                row["총점"] = str(total_score)
                row["판정"] = judgment
                row["메모"] = memo
                updated = True
                break

        if not updated:
            return False, f"evaluation_template.tsv에서 번호 {target_no} 행을 찾을 수 없습니다."

        write_evaluation_rows(rows)
        return True, f"평가 결과 저장 완료: Q{int(target_no):02d}, 총점 {total_score}, 판정 {judgment}"
    except Exception as e:
        return False, str(e)


def generate_local_fallback_answer(
    question: str,
    results: list[dict[str, Any]],
    reason: str = "",
) -> str:
    situation_type = classify_safety_situation(question)
    immediate_judgment = build_immediate_judgment(situation_type)
    priority_actions = [f"- {item}" for item in build_priority_actions(situation_type)]
    check_items = [f"- {item}" for item in build_check_items(situation_type)]

    unique_sources = []
    for r in results:
        source_label = f"{r['source']} (chunk_id: {r.get('chunk_id', '정보 없음')})"
        if source_label not in unique_sources:
            unique_sources.append(source_label)

    evidence_lines = []
    for r in results[:5]:
        evidence_lines.append(
            f"- 근거 {r['rank']} | {r['source']} | chunk_id {r.get('chunk_id', '정보 없음')} | 거리값 {format_distance(r['distance'])}\n"
            f"  - {make_preview(r['text'])}"
        )

    source_lines = [f"- {source}" for source in unique_sources[:8]]
    reason_text = reason.strip() or "안정 모드: Gemini API 호출 안 함"
    kras_section = build_kras_risk_assessment_section(
        question,
        results,
        situation_type,
    )

    return "\n".join(
        [
            "## 검색 근거 기반 안전 답변",
            "",
            "외부 LLM API 호출 없이, 검색된 법령·지침 근거를 바탕으로 안전 답변을 제공합니다.",
            "아래 내용은 검색 결과 기반 요약이며, 실제 현장 조치와 법령 해석은 담당 안전관리자, 관계 기관, 전문가 확인이 필요합니다.",
            "",
            "## 질문",
            question,
            "",
            "## 즉시 판단",
            f"- 상황 유형: {situation_type}",
            f"- {immediate_judgment}",
            "",
            "## 우선 조치",
            *priority_actions,
            "",
            "## 확인해야 할 사항",
            *check_items,
            "",
            "## 관련 근거 문서",
            *source_lines,
            "",
            "### 검색된 주요 근거 요약",
            *evidence_lines,
            "",
            kras_section,
            "",
            "## 현장 조치 체크리스트",
            "- 현장 위험이 제거되기 전에는 작업 재개를 서두르지 않습니다.",
            "- 위험 구역, 작업자 위치, 대피 경로, 비상 연락 체계를 확인합니다.",
            "- 검색된 근거 문서의 점검 기준, 작업중지 기준, 보고 절차를 대조합니다.",
            "- 조치 결과와 확인 내용을 기록하고 필요 시 안전관리자 또는 책임자에게 보고합니다.",
            "- 근거가 부족한 부분은 내부 안전관리 기준, 관계 기관, 전문가 확인 절차로 넘깁니다.",
            "",
            "## 주의사항",
            "- 이 답변은 검색된 Vector DB 근거를 바탕으로 한 발표/평가용 안전 답변입니다.",
            "- 문서에 없는 세부 법령 해석, 처벌 수위, 현장별 최종 조치는 단정하지 않습니다.",
            "- 실제 조치 전 최신 법령과 현장 조건을 반드시 확인하세요.",
            "",
            "---",
            f"생성 상태: {reason_text}",
        ]
    )


def call_gemini_with_timeout(
    client,
    prompt: str,
    model_name: str,
    *,
    timeout_seconds: int = GEMINI_RESPONSE_TIMEOUT_SECONDS,
    max_output_tokens: int = GEMINI_MAX_OUTPUT_TOKENS,
) -> str:
    config = types.GenerateContentConfig(
        temperature=0.2,
        max_output_tokens=max_output_tokens,
    )

    executor = ThreadPoolExecutor(max_workers=1)
    future = executor.submit(
        client.models.generate_content,
        model=model_name,
        contents=prompt,
        config=config,
    )

    try:
        response = future.result(timeout=timeout_seconds)
    except FutureTimeoutError as e:
        future.cancel()
        raise TimeoutError(
            f"Gemini 응답 제한 시간 {timeout_seconds}초 초과"
        ) from e
    finally:
        executor.shutdown(wait=False, cancel_futures=True)

    text = getattr(response, "text", None)
    if not isinstance(text, str):
        raise RuntimeError("Gemini 응답에 text가 없습니다.")

    cleaned_text = text.strip()
    if not cleaned_text:
        raise RuntimeError("Gemini 응답이 비어 있습니다.")
    return cleaned_text


def execute_gemini_request(
    prompt: str,
    model_name: str,
    *,
    max_output_tokens: int = GEMINI_MAX_OUTPUT_TOKENS,
) -> dict[str, Any]:
    client, error = load_gemini_client()
    if error:
        error_summary = summarize_gemini_error(error)
        return {
            "called": False,
            "success": False,
            "status": "error",
            "message": (
                "Gemini API를 호출할 수 없습니다. "
                f"오류 요약: {error_summary}"
            ),
            "answer": "",
            "attempts": 0,
            "model": model_name,
            "used_fallback": False,
            "error": error_summary,
            "elapsed": 0.0,
        }

    start_time = time.monotonic()
    last_error = ""
    last_status = "error"
    attempts_used = 0

    for attempt in range(1, GEMINI_MAX_ATTEMPTS + 1):
        attempts_used = attempt
        try:
            answer = call_gemini_with_timeout(
                client,
                prompt,
                model_name,
                timeout_seconds=GEMINI_RESPONSE_TIMEOUT_SECONDS,
                max_output_tokens=max_output_tokens,
            )
            if not answer or not answer.strip():
                raise RuntimeError("Gemini 응답이 비어 있습니다.")
            return {
                "called": True,
                "success": True,
                "status": "success",
                "message": "Gemini 응답 생성에 성공했습니다.",
                "answer": answer,
                "attempts": attempt,
                "model": model_name,
                "used_fallback": False,
                "error": "",
                "elapsed": time.monotonic() - start_time,
            }
        except Exception as e:
            last_error = str(e)
            last_status = classify_gemini_error(last_error)
            should_retry = (
                last_status in {"503", "429", "timeout"}
                or is_temporary_gemini_error(last_error)
            )
            if attempt < GEMINI_MAX_ATTEMPTS and should_retry:
                time.sleep(GEMINI_RETRY_SLEEP_SECONDS)
                continue
            break

    return {
        "called": True,
        "success": False,
        "status": last_status,
        "message": get_gemini_failure_message(last_status, last_error),
        "answer": "",
        "attempts": attempts_used,
        "model": model_name,
        "used_fallback": False,
        "error": summarize_gemini_error(last_error),
        "elapsed": time.monotonic() - start_time,
    }


def test_gemini_connection(model_name: str) -> dict[str, Any]:
    return execute_gemini_request(
        "테스트입니다. 한 문장으로 응답해 주세요.",
        model_name,
        max_output_tokens=80,
    )


def build_legacy_gemini_status(
    result: dict[str, Any],
    *,
    fallback_used: bool,
) -> dict[str, Any]:
    success = bool(result.get("success", False))
    return {
        **result,
        "mode": "gemini" if success else "fallback",
        "reason": str(result.get("error", "")),
        "selected_model": str(result.get("model", GEMINI_MODEL_NAME)),
        "gemini_called": bool(result.get("called", False)),
        "fallback_used": fallback_used,
        "used_fallback": fallback_used,
        "gemini_status": classify_gemini_status(result),
    }


def generate_gemini_answer(
    question: str,
    results: list[dict[str, Any]],
    model_name: str = GEMINI_MODEL_NAME,
):
    prompt = build_prompt(question, results)
    result = execute_gemini_request(prompt, model_name)

    if result.get("success"):
        answer = str(result.get("answer", ""))
        if "KRAS식 위험성평가 기록 초안" not in answer:
            answer = "\n\n".join(
                [
                    answer,
                    build_kras_risk_assessment_section(question, results),
                ]
            )
        status = build_legacy_gemini_status(result, fallback_used=False)
        return answer, status

    fallback_answer = generate_local_fallback_answer(
        question,
        results,
        str(result.get("message") or result.get("error") or "Gemini 호출 실패"),
    )
    status = build_legacy_gemini_status(result, fallback_used=True)
    return fallback_answer, status


# ==============================
# 법령 체크리스트·증빙자료·위험성평가·대화이력 기능
# ==============================
LEGAL_CHECKLIST_ITEMS = [
    ("안전보건 경영방침 수립 여부", "경영방침을 문서화하고 현장 구성원에게 공유했는지 확인"),
    ("안전보건 전담조직 지정 여부", "안전보건 업무 담당 조직과 역할이 지정되어 있는지 확인"),
    ("안전보건 예산 편성 및 집행 여부", "안전보건 예산 편성, 집행 내역, 개선 투자 기록 확인"),
    ("위험성평가 실시 여부", "정기·수시 위험성평가 실시 및 개선대책 수립 여부 확인"),
    ("개선조치 이행 여부", "위험성평가, 점검, 사고조사 후 개선조치 완료 여부 확인"),
    ("작업 전 TBM 기록 여부", "작업 전 위험요인 공유, 보호구, 작업중지 기준 안내 기록 확인"),
    ("보호구 지급대장 보유 여부", "보호구 지급, 교체, 착용 점검 기록 보유 여부 확인"),
    ("교육 기록 보유 여부", "정기교육, 특별교육, TBM 참석자 명부 및 교육자료 확인"),
    ("가스/분진/환기 측정 기록 보유 여부", "측정값, 측정자, 측정장비, 조치 내역 기록 확인"),
    ("사고 및 아차사고 보고서 작성 여부", "사고·아차사고 발생 보고, 원인, 조치 기록 확인"),
    ("재발방지대책 수립 여부", "원인조사 후 재발방지대책과 이행 결과 확인"),
    ("협력업체 안전관리 기록 여부", "협력업체 작업 전 교육, 위험성평가, 작업허가 기록 확인"),
]


def ensure_feature_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    FEATURE_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def default_legal_checklist_rows() -> list[dict[str, Any]]:
    return [
        {
            "항목명": item,
            "설명": description,
            "상태": "미작성",
            "담당자": "",
            "최근 작성일": "",
            "비고": "",
        }
        for item, description in LEGAL_CHECKLIST_ITEMS
    ]


def load_legal_checklist_status() -> list[dict[str, Any]]:
    ensure_feature_dirs()
    if not LEGAL_CHECKLIST_STATUS_PATH.exists():
        rows = default_legal_checklist_rows()
        save_legal_checklist_status(rows)
        return rows
    try:
        data = json.loads(LEGAL_CHECKLIST_STATUS_PATH.read_text(encoding="utf-8"))
        if isinstance(data, list) and data:
            return data
    except Exception:
        pass
    return default_legal_checklist_rows()


def save_legal_checklist_status(rows: list[dict[str, Any]]) -> None:
    ensure_feature_dirs()
    LEGAL_CHECKLIST_STATUS_PATH.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def normalize_excel_value(value: Any) -> Any:
    if isinstance(value, (list, tuple, set)):
        return " | ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    if value is None:
        return ""
    return value


def build_simple_xlsx_workbook(sheet_name: str, rows: list[dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    headers = list(rows[0].keys()) if rows else ["내용"]
    ws.append(headers)
    for row in rows:
        ws.append([normalize_excel_value(row.get(header, "")) for header in headers])
    style_excel_sheet(ws)
    return wb


def build_simple_xlsx_bytes(sheet_name: str, rows: list[dict[str, Any]]) -> bytes:
    wb = build_simple_xlsx_workbook(sheet_name, rows)
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer.getvalue()


def write_simple_xlsx(path: Path, sheet_name: str, rows: list[dict[str, Any]]) -> None:
    ensure_feature_dirs()
    wb = build_simple_xlsx_workbook(sheet_name, rows)
    wb.save(path)
    wb.close()


def style_excel_sheet(ws) -> None:
    thin = Side(style="thin", color="D9E2EC")
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx > 1:
            ws.row_dimensions[row_idx].height = 42
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=10)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx == 1:
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="17324D")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, idx).value or "")
        width = 16
        if "답변" in header or "answer" in header or "비고" in header or "메모" in header:
            width = 48
        elif "설명" in header or "증빙" in header or "근거" in header or "질문" in header:
            width = 36
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def export_legal_checklist_xlsx(rows: list[dict[str, Any]]) -> Path:
    write_simple_xlsx(LEGAL_CHECKLIST_EXPORT_PATH, "legal_checklist", rows)
    return LEGAL_CHECKLIST_EXPORT_PATH


def recommend_evidence_records(question: str, answer: str = "") -> list[str]:
    text = f"{question} {answer}".lower()
    rules = [
        (["메탄", "가스", "환기", "산소"], ["가스 측정 기록지", "환기설비 점검표", "작업중지 지시 기록", "작업재개 승인 기록", "TBM 회의록", "개선조치 완료 사진"]),
        (["발파", "불발", "장약", "화약"], ["발파작업 허가서", "발파 전 점검표", "화약류 사용 기록", "불발 장약 처리 기록", "대피 확인 기록", "작업재개 승인 기록"]),
        (["낙반", "붕락", "막장", "지보", "천반"], ["막장 점검표", "지보 상태 점검 기록", "작업중지 기록", "보강조치 완료 사진", "작업재개 승인 기록"]),
        (["분진", "방진", "마스크", "보호구", "ppe"], ["분진 측정 기록", "방진마스크 지급대장", "보호구 착용 점검표", "작업환경측정 결과", "교육 참석자 명부"]),
        (["중대재해", "사고", "응급", "경영책임자"], ["사고 발생 보고서", "응급조치 기록", "관계기관 보고 기록", "재발방지대책", "개선조치 이행 결과", "경영책임자 보고 기록"]),
        (["전기", "감전", "접지", "절연"], ["전기설비 점검표", "전원 차단 및 잠금표지 기록", "절연저항 측정 기록", "접지 상태 점검 기록", "작업허가서"]),
        (["운반", "장비", "협착", "통로", "차량"], ["장비 일상점검표", "작업동선 분리 계획", "신호수 배치 기록", "후진경보장치 점검표", "작업자 교육 기록"]),
    ]
    recommended: list[str] = []
    for keywords, records in rules:
        if any(keyword in text for keyword in keywords):
            for record in records:
                if record not in recommended:
                    recommended.append(record)
    if not recommended:
        recommended = ["작업 전 TBM 회의록", "위험성평가표", "작업중지·재개 판단 기록", "현장 점검 사진", "교육 참석자 명부"]
    return recommended


def render_recommended_evidence_records(records: list[str]) -> None:
    with st.container(border=True):
        st.markdown("#### 필요 증빙자료")
        st.caption("질문과 답변의 키워드를 기준으로 현장 이행자료 후보를 자동 추천합니다.")
        cols = st.columns(2)
        for idx, record in enumerate(records):
            with cols[idx % 2]:
                st.markdown(f"- {record}")
        st.info("AI 답변만으로 법적 책임이 면제되는 것은 아니며, 실제 점검표·사진·교육기록·작업중지기록 등 현장 이행자료와 함께 관리해야 합니다.")


def source_chunk_labels(results: list[dict[str, Any]]) -> list[str]:
    labels: list[str] = []
    for item in results:
        source = str(item.get("source", "출처 정보 없음"))
        chunk_id = str(item.get("chunk_id", "정보 없음"))
        label = f"{source} / {chunk_id}"
        if label not in labels:
            labels.append(label)
    return labels


def append_conversation_history(row: dict[str, Any]) -> str:
    ensure_feature_dirs()
    history_id = row.get("history_id") or str(uuid.uuid4())
    row["history_id"] = history_id
    with CONVERSATION_HISTORY_PATH.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\n")
    return history_id


def load_conversation_history() -> list[dict[str, Any]]:
    ensure_feature_dirs()
    if not CONVERSATION_HISTORY_PATH.exists():
        CONVERSATION_HISTORY_PATH.write_text("", encoding="utf-8")
        return []
    rows = []
    for line in CONVERSATION_HISTORY_PATH.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            rows.append(json.loads(line))
        except Exception:
            continue
    return rows


def rewrite_conversation_history(rows: list[dict[str, Any]]) -> None:
    ensure_feature_dirs()
    with CONVERSATION_HISTORY_PATH.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def update_conversation_history(history_id: str, updates: dict[str, Any]) -> bool:
    rows = load_conversation_history()
    updated = False
    for row in rows:
        if str(row.get("history_id")) == str(history_id):
            row.update(updates)
            updated = True
            break
    if updated:
        rewrite_conversation_history(rows)
    return updated


def normalize_conversation_history_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        rows = [{"history_id": "", "created_at": "", "question": "", "answer": ""}]
    normalized_rows = []
    for row in rows:
        normalized_rows.append(
            {
                "history_id": row.get("history_id", ""),
                "created_at": row.get("created_at", ""),
                "question": row.get("question", ""),
                "answer": row.get("answer", ""),
                "situation_type": row.get("situation_type", ""),
                "risk_level": row.get("risk_level", ""),
                "evidence_documents": " | ".join(row.get("evidence_documents", [])),
                "source_chunks": " | ".join(row.get("source_chunks", [])),
                "recommended_evidence_records": " | ".join(row.get("recommended_evidence_records", [])),
                "user_action_status": row.get("user_action_status", ""),
                "manager": row.get("manager", ""),
                "action_due_date": row.get("action_due_date", ""),
                "memo": row.get("memo", ""),
            }
        )
    return normalized_rows


def build_conversation_history_xlsx_bytes(rows: list[dict[str, Any]]) -> bytes:
    return build_simple_xlsx_bytes("conversation_history", normalize_conversation_history_rows(rows))


def normalize_history_display_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        items = [str(item).strip() for item in value if str(item).strip()]
    else:
        text = str(value).strip()
        if not text:
            return []
        separator = " | " if " | " in text else "," if "," in text else None
        items = [part.strip() for part in text.split(separator)] if separator else [text]
    return [item for item in items if item]


def short_history_label(row: dict[str, Any], limit: int = 52) -> str:
    question = " ".join(str(row.get("question", "")).split())
    if len(question) > limit:
        question = question[: limit - 3] + "..."
    created_at = str(row.get("created_at", ""))[:16]
    situation = str(row.get("situation_type", "상황 미분류") or "상황 미분류")
    return f"{created_at} | {situation} | {question}"


def render_history_list_card(title: str, items: list[str], empty_message: str = "기록 없음") -> None:
    with st.container(border=True):
        st.markdown(f"#### {title}")
        if not items:
            st.caption(empty_message)
            return
        cols = st.columns(2)
        for idx, item in enumerate(items):
            with cols[idx % 2]:
                st.markdown(
                    f"""
                    <div style="background:#f8fafc;border:1px solid #dbe4ef;border-radius:10px;
                                padding:10px 12px;margin:4px 0;color:#17324d;line-height:1.45;">
                        <strong style="color:#2563eb;">•</strong> {escape(item)}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )


def render_history_report_field(title: str, content: str) -> None:
    st.markdown(
        f"""
        <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:12px;
                    padding:14px 16px;margin-bottom:10px;box-shadow:0 1px 2px rgba(15,23,42,0.04);">
            <div style="font-weight:700;color:#17324d;margin-bottom:6px;">{escape(title)}</div>
            <div style="color:#334155;line-height:1.6;white-space:pre-wrap;">{escape(content or "기록 없음")}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_history_report_card(selected: dict[str, Any], evidence_documents: list[str], recommended_records: list[str]) -> None:
    st.subheader("증빙자료 보고서용 보기")
    render_history_report_field("질문", str(selected.get("question", "")))
    c1, c2 = st.columns(2)
    with c1:
        render_history_report_field("상황 유형", str(selected.get("situation_type", "상황 미분류") or "상황 미분류"))
    with c2:
        render_history_report_field("위험도", str(selected.get("risk_level", "검토 필요") or "검토 필요"))
    render_history_report_field("근거 문서", ", ".join(evidence_documents) if evidence_documents else "기록 없음")
    render_history_report_field("필요 증빙자료", ", ".join(recommended_records) if recommended_records else "기록 없음")
    render_history_report_field("조치 상태", str(selected.get("user_action_status", "미조치") or "미조치"))


def strip_inline_markdown(text: str) -> str:
    return str(text or "").replace("**", "").replace("\\|", "|").strip()


def split_display_items(text: str) -> list[str]:
    cleaned = strip_inline_markdown(text)
    if not cleaned:
        return []
    parts = []
    for raw in cleaned.replace("\r\n", " / ").replace("\n", " / ").split(" / "):
        item = raw.strip(" -•")
        if item:
            parts.append(item)
    return parts or [cleaned]


def kras_risk_badge_html(level: str) -> str:
    styles = {
        "낮음": ("#dcfce7", "#166534", "#86efac"),
        "보통": ("#fef9c3", "#854d0e", "#fde047"),
        "높음": ("#ffedd5", "#9a3412", "#fdba74"),
        "매우 높음": ("#fee2e2", "#991b1b", "#fca5a5"),
    }
    bg, fg, border = styles.get(level, ("#e2e8f0", "#334155", "#cbd5e1"))
    return (
        f'<span style="display:inline-block;background:{bg};color:{fg};border:1px solid {border};'
        'border-radius:999px;padding:2px 9px;margin:0 3px;font-weight:700;white-space:nowrap;">'
        f'{escape(level)}</span>'
    )


def emphasize_risk_terms_html(text: str) -> str:
    escaped = escape(strip_inline_markdown(text))
    for level in ["매우 높음", "높음", "보통", "낮음"]:
        escaped = escaped.replace(escape(level), kras_risk_badge_html(level))
    return escaped


def measure_label_html(label: str) -> str:
    styles = {
        "제거": ("#dbeafe", "#1d4ed8"),
        "대체": ("#e0e7ff", "#4338ca"),
        "공학적 대책": ("#ccfbf1", "#0f766e"),
        "관리적 대책": ("#fef3c7", "#92400e"),
        "보호구/PPE": ("#fce7f3", "#be185d"),
    }
    bg, fg = styles.get(label, ("#e2e8f0", "#334155"))
    return (
        f'<span style="display:inline-block;background:{bg};color:{fg};border-radius:8px;'
        'padding:2px 8px;margin-right:6px;font-weight:800;white-space:nowrap;">'
        f'{escape(label)}</span>'
    )


def emphasize_measure_item_html(item: str) -> str:
    cleaned = strip_inline_markdown(item)
    for label in ["공학적 대책", "관리적 대책", "보호구/PPE", "제거", "대체"]:
        for marker in [f"{label}:", f"{label}："]:
            if cleaned.startswith(marker):
                detail = cleaned[len(marker):].strip()
                return f"{measure_label_html(label)} {emphasize_risk_terms_html(detail)}"
    return emphasize_risk_terms_html(cleaned)


def render_kras_value_card(label: str, value: str) -> None:
    clean_label = strip_inline_markdown(label)
    items = split_display_items(value)
    if len(items) > 1:
        if clean_label == "위험성 감소대책":
            body = "".join(f"<li>{emphasize_measure_item_html(item)}</li>" for item in items)
        else:
            body = "".join(f"<li>{emphasize_risk_terms_html(item)}</li>" for item in items)
        value_html = f"<ul style='margin:0;padding-left:18px;line-height:1.8;'>{body}</ul>"
    else:
        item = items[0] if items else "기록 필요"
        if clean_label == "위험성 감소대책":
            value_html = f"<div style='line-height:1.8;white-space:pre-wrap;'>{emphasize_measure_item_html(item)}</div>"
        else:
            value_html = f"<div style='line-height:1.8;white-space:pre-wrap;'>{emphasize_risk_terms_html(item)}</div>"
    st.markdown(
        f"""
        <div style="background:#ffffff;border:1px solid #dbe4ef;border-radius:12px;
                    padding:13px 15px;margin-bottom:10px;">
            <div style="font-weight:800;color:#17324d;margin-bottom:8px;">• {escape(clean_label)}</div>
            <div style="color:#334155;">{value_html}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def parse_markdown_table_lines(lines: list[str]) -> tuple[list[str], list[list[str]]]:
    rows = []
    for line in lines:
        stripped = line.strip()
        if not stripped.startswith("|"):
            continue
        cells = [cell.strip() for cell in stripped.strip("|").split("|")]
        if cells and all(set(cell) <= {"-"} for cell in cells if cell):
            continue
        rows.append(cells)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def render_generic_kras_table(headers: list[str], rows: list[list[str]]) -> None:
    if not headers or not rows:
        return
    header_html = "".join(f"<th>{escape(strip_inline_markdown(header))}</th>" for header in headers)
    row_html = []
    for row in rows:
        cells = []
        for idx, cell in enumerate(row):
            value = strip_inline_markdown(cell)
            if " / " in value:
                items = split_display_items(value)
                cell_html = "<ul style='margin:0;padding-left:16px;'>" + "".join(f"<li>{escape(item)}</li>" for item in items) + "</ul>"
            else:
                cell_html = escape(value)
            cells.append(f"<td>{cell_html}</td>")
        row_html.append("<tr>" + "".join(cells) + "</tr>")
    st.markdown(
        (
            "<div class='portal-table-scroll table-wrap'>"
            "<table class='portal-data-table'>"
            f"<thead><tr>{header_html}</tr></thead>"
            f"<tbody>{''.join(row_html)}</tbody></table></div>"
        ),
        unsafe_allow_html=True,
    )


def render_kras_readable_markdown(kras_answer: str) -> None:
    content = strip_kras_title(kras_answer)
    lines = content.splitlines()
    idx = 0
    while idx < len(lines):
        line = lines[idx].strip()
        if not line:
            idx += 1
            continue
        if line.startswith(">"):
            st.info(line.lstrip("> ").strip())
            idx += 1
            continue
        if line.startswith("###"):
            st.markdown(f"#### {escape(line.lstrip('#').strip())}")
            idx += 1
            continue
        if line.startswith("|"):
            table_lines = []
            while idx < len(lines) and lines[idx].strip().startswith("|"):
                table_lines.append(lines[idx])
                idx += 1
            headers, rows = parse_markdown_table_lines(table_lines)
            if headers[:2] == ["항목", "기입 초안"]:
                for row in rows:
                    if len(row) >= 2:
                        render_kras_value_card(row[0], row[1])
            else:
                render_generic_kras_table(headers, rows)
            continue
        st.markdown(line)
        idx += 1


def render_export_report_page() -> None:
    st.header("Excel 내보내기 및 보고서")
    st.info("현장 점검 보조자료를 Excel로 내려받아 교수님 보고나 내부 기록 정리에 활용할 수 있습니다.")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("#### 중대재해처벌법 체크리스트")
        checklist_excel_bytes = build_simple_xlsx_bytes("legal_checklist", load_legal_checklist_status())
        st.download_button(
            label="체크리스트 Excel 다운로드",
            data=checklist_excel_bytes,
            file_name="legal_checklist_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="field_download_legal_checklist_excel",
            use_container_width=True,
        )
    with c2:
        st.markdown("#### 위험성평가 초안")
        rows = st.session_state.get(
            "risk_assessment_draft_rows",
            build_risk_assessment_draft("갱내 작업", "막장 또는 갱내 작업장", "현장 위험요인", "작업 전 위험성평가 필요"),
        )
        risk_excel_bytes = build_simple_xlsx_bytes("risk_assessment_draft", rows)
        st.download_button(
            label="위험성평가 초안 Excel 다운로드",
            data=risk_excel_bytes,
            file_name="risk_assessment_draft_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="field_download_risk_assessment_excel",
            use_container_width=True,
        )
    with c3:
        st.markdown("#### 대화 이력")
        history_excel_bytes = build_conversation_history_xlsx_bytes(load_conversation_history())
        st.download_button(
            label="대화 이력 Excel 다운로드",
            data=history_excel_bytes,
            file_name="conversation_history_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="field_download_conversation_history_excel",
            use_container_width=True,
        )
    st.caption("AI 답변과 내보낸 파일은 법적 책임 면제 자료가 아니며, 현장 점검표·사진·교육기록·작업중지 기록과 함께 관리해야 합니다.")


def export_conversation_history_xlsx(rows: list[dict[str, Any]]) -> Path:
    write_simple_xlsx(CONVERSATION_HISTORY_EXPORT_PATH, "conversation_history", normalize_conversation_history_rows(rows))
    return CONVERSATION_HISTORY_EXPORT_PATH


def auto_save_conversation_history(
    question: str,
    answer: str,
    situation_type: str,
    risk_level: str,
    results: list[dict[str, Any]],
    recommended_records: list[str],
    result_key: str,
) -> str | None:
    if not question.strip() or not answer.strip():
        return None
    history_marker = f"history_saved_{result_key}_{abs(hash(question + answer))}"
    if st.session_state.get(history_marker):
        return st.session_state[history_marker]
    history_id = append_conversation_history(
        {
            "history_id": str(uuid.uuid4()),
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "question": question,
            "answer": answer,
            "situation_type": situation_type,
            "risk_level": risk_level,
            "evidence_documents": list(dict.fromkeys(str(item.get("source", "출처 정보 없음")) for item in results)),
            "source_chunks": source_chunk_labels(results),
            "recommended_evidence_records": recommended_records,
            "user_action_status": "미조치",
            "manager": "",
            "action_due_date": "",
            "memo": "",
        }
    )
    st.session_state[history_marker] = history_id
    return history_id


def infer_risk_keywords(text: str) -> list[str]:
    lowered = text.lower()
    mapping = {
        "발파": ["대피", "출입통제", "발파 전 점검", "불발 확인"],
        "불발": ["대피", "출입통제", "불발 장약 처리", "작업재개 승인"],
        "메탄/가스": ["작업중지", "대피", "환기", "재측정", "작업재개 승인"],
        "낙반": ["지보 점검", "천반 확인", "작업중지", "보강 후 재개"],
        "분진": ["살수", "집진", "방진마스크", "작업환경측정"],
        "전기": ["전원 차단", "절연", "잠금표지", "점검자 지정"],
        "장비/운반": ["동선 분리", "신호수 배치", "속도 제한", "후진 경보"],
    }
    selected = []
    for key, actions in mapping.items():
        key_parts = key.lower().replace("/", " ").split()
        if any(part in lowered for part in key_parts):
            selected.extend(actions)
    return selected or ["작업중지 기준 확인", "위험성평가", "보호구 확인", "책임자 승인"]


def build_risk_assessment_draft(
    work_name: str,
    work_place: str,
    hazards: str,
    situation: str,
) -> list[dict[str, Any]]:
    combined = f"{work_name} {work_place} {hazards} {situation}"
    actions = infer_risk_keywords(combined)
    records = recommend_evidence_records(combined, situation)
    priority = "높음" if any(word in combined for word in ["메탄", "가스", "불발", "낙반", "중대재해", "감전"]) else "보통"
    return [
        {
            "작업명": work_name,
            "작업 장소": work_place,
            "위험요인": hazards or "현장 위험요인 확인 필요",
            "위험상황": situation or "작업 전 위험상황 설명 필요",
            "현재 안전조치": "작업 전 점검, TBM, 보호구 착용, 작업구역 확인",
            "추가 개선대책": ", ".join(actions),
            "필요 증빙자료": ", ".join(records),
            "조치 우선순위": priority,
            "담당자": "",
            "조치 기한": "",
        }
    ]


def export_risk_assessment_xlsx(rows: list[dict[str, Any]]) -> Path:
    write_simple_xlsx(RISK_ASSESSMENT_EXPORT_PATH, "risk_assessment_draft", rows)
    return RISK_ASSESSMENT_EXPORT_PATH


def render_legal_checklist_page() -> None:
    st.header("중대재해처벌법 대응 체크리스트")
    st.warning("본 기능은 법적 면책 자료가 아니라 안전보건 확보의무 이행 상태 점검 보조 기능입니다.")
    rows = load_legal_checklist_status()
    status_counts = Counter(row.get("상태", "미작성") for row in rows)
    completed = status_counts.get("완료", 0)
    insufficient = status_counts.get("미흡", 0)
    missing = status_counts.get("미작성", 0)
    total = len(rows) or 1
    response_rate = round(completed / total * 100, 1)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("전체 대응률", f"{response_rate}%")
    col2.metric("완료", completed)
    col3.metric("미흡", insufficient)
    col4.metric("미작성", missing)

    edited_rows = []
    with st.form("legal_checklist_form"):
        for idx, row in enumerate(rows):
            with st.expander(f"{idx + 1}. {row.get('항목명', '')}", expanded=False):
                st.caption(row.get("설명", ""))
                c1, c2, c3 = st.columns([1, 1, 2])
                status = c1.selectbox("상태", ["완료", "미흡", "미작성"], index=["완료", "미흡", "미작성"].index(row.get("상태", "미작성") if row.get("상태") in ["완료", "미흡", "미작성"] else "미작성"), key=f"legal_status_{idx}")
                manager = c2.text_input("담당자", value=row.get("담당자", ""), key=f"legal_manager_{idx}")
                recent_date = c3.text_input("최근 작성일", value=row.get("최근 작성일", ""), placeholder="YYYY-MM-DD", key=f"legal_date_{idx}")
                memo = st.text_area("비고", value=row.get("비고", ""), height=70, key=f"legal_memo_{idx}")
                edited_rows.append({**row, "상태": status, "담당자": manager, "최근 작성일": recent_date, "비고": memo})
        save_clicked = st.form_submit_button("체크리스트 저장", type="primary")
    if save_clicked:
        save_legal_checklist_status(edited_rows)
        st.success(f"체크리스트 저장 완료: {LEGAL_CHECKLIST_STATUS_PATH}")
        st.rerun()

    checklist_excel_bytes = build_simple_xlsx_bytes("legal_checklist", load_legal_checklist_status())
    st.download_button(
        label="체크리스트 Excel 다운로드",
        data=checklist_excel_bytes,
        file_name="legal_checklist_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_legal_checklist_excel",
        use_container_width=True,
    )


def render_risk_assessment_draft_page() -> None:
    st.header("위험성평가 초안")
    st.warning("본 초안은 안전관리자가 검토·수정해야 하며, 최종 법적 문서로 바로 사용할 수 없습니다.")
    with st.form("risk_assessment_form"):
        work_name = st.text_input("작업명", value="갱내 작업")
        work_place = st.text_input("작업 장소", value="막장 또는 갱내 작업장")
        hazards = st.text_area("주요 위험요인", value="메탄가스, 환기 불량, 장비 이동")
        situation = st.text_area("작업 상황 설명", value="작업 전 TBM에서 위험요인을 공유하고 작업 가능 여부를 판단해야 함")
        submitted = st.form_submit_button("위험성평가 초안 생성", type="primary")
    if submitted or "risk_assessment_draft_rows" not in st.session_state:
        st.session_state["risk_assessment_draft_rows"] = build_risk_assessment_draft(work_name, work_place, hazards, situation)
    rows = st.session_state.get("risk_assessment_draft_rows", [])
    st.dataframe(rows, use_container_width=True, hide_index=True)
    risk_excel_bytes = build_simple_xlsx_bytes("risk_assessment_draft", rows)
    st.download_button(
        label="위험성평가 초안 Excel 다운로드",
        data=risk_excel_bytes,
        file_name="risk_assessment_draft_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_risk_assessment_excel",
        use_container_width=True,
    )


def render_conversation_history_page() -> None:
    st.header("대화 이력")
    st.info("대화 이력은 AI 답변과 근거 문서, 추천 증빙자료를 함께 저장하여 안전관리 기록 보조자료로 활용할 수 있습니다. 단, 실제 법적 증빙은 현장 점검표, 사진, 교육기록, 작업중지 기록 등과 함께 관리해야 합니다.")
    rows = load_conversation_history()
    st.caption(f"저장 위치: {CONVERSATION_HISTORY_PATH}")
    if not rows:
        st.info("아직 저장된 대화 이력이 없습니다. 안전 질의 및 답변 메뉴에서 질문을 생성하면 자동 저장됩니다.")
        empty_history_excel_bytes = build_conversation_history_xlsx_bytes(rows)
        st.download_button(
            label="빈 대화 이력 Excel 다운로드",
            data=empty_history_excel_bytes,
            file_name="conversation_history_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_empty_conversation_history_excel",
            use_container_width=True,
        )
        return

    keyword = st.text_input("검색창: 질문 키워드")
    situation_options = ["전체"] + sorted(set(str(row.get("situation_type", "기타")) for row in rows if row.get("situation_type")))
    risk_options = ["전체"] + sorted(set(str(row.get("risk_level", "검토 필요")) for row in rows if row.get("risk_level")))
    f1, f2, f3 = st.columns(3)
    selected_situation = f1.selectbox("상황 유형 필터", situation_options)
    selected_risk = f2.selectbox("위험도 필터", risk_options)
    selected_date = f3.text_input("날짜 필터", placeholder="YYYY-MM-DD")
    filtered = []
    for row in rows:
        if keyword and keyword not in str(row.get("question", "")):
            continue
        if selected_situation != "전체" and row.get("situation_type") != selected_situation:
            continue
        if selected_risk != "전체" and row.get("risk_level") != selected_risk:
            continue
        if selected_date and not str(row.get("created_at", "")).startswith(selected_date):
            continue
        filtered.append(row)
    st.metric("조회 결과", f"{len(filtered)}건")
    if not filtered:
        st.info("조건에 맞는 대화 이력이 없습니다.")
        return
    labels = [short_history_label(row) for row in filtered]
    selected_label = st.selectbox("저장된 질문 목록", labels)
    selected = filtered[labels.index(selected_label)]

    evidence_documents = normalize_history_display_list(selected.get("evidence_documents", []))
    recommended_records = normalize_history_display_list(selected.get("recommended_evidence_records", []))

    with st.container(border=True):
        st.subheader("질문/답변")
        st.markdown("**질문**")
        st.markdown(
            f"""
            <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;
                        padding:14px 16px;color:#1e293b;line-height:1.6;white-space:pre-wrap;">
                {escape(str(selected.get("question", "")))}
            </div>
            """,
            unsafe_allow_html=True,
        )
        with st.expander("당시 답변 보기", expanded=True):
            st.markdown(selected.get("answer", ""))

    col_docs, col_records = st.columns(2)
    with col_docs:
        render_history_list_card("근거 문서", evidence_documents, "저장된 근거 문서가 없습니다.")
    with col_records:
        render_history_list_card("추천 증빙자료", recommended_records, "저장된 추천 증빙자료가 없습니다.")

    render_history_report_card(selected, evidence_documents, recommended_records)

    with st.form(f"history_update_{selected.get('history_id')}"):
        c1, c2, c3 = st.columns(3)
        action_status = c1.selectbox("조치 여부", ["미조치", "진행 중", "조치 완료", "보류"], index=["미조치", "진행 중", "조치 완료", "보류"].index(selected.get("user_action_status", "미조치") if selected.get("user_action_status") in ["미조치", "진행 중", "조치 완료", "보류"] else "미조치"))
        manager = c2.text_input("담당자", value=selected.get("manager", ""))
        due_date = c3.text_input("조치 완료일", value=selected.get("action_due_date", ""), placeholder="YYYY-MM-DD")
        memo = st.text_area("메모", value=selected.get("memo", ""), height=80)
        update_clicked = st.form_submit_button("이력 메타데이터 저장", type="primary")
    if update_clicked:
        ok = update_conversation_history(
            selected.get("history_id", ""),
            {
                "user_action_status": action_status,
                "manager": manager,
                "action_due_date": due_date,
                "memo": memo,
            },
        )
        if ok:
            st.success("대화 이력 메타데이터 저장 완료")
            st.rerun()
        else:
            st.error("대화 이력 저장 실패")
    history_excel_bytes = build_conversation_history_xlsx_bytes(rows)
    st.download_button(
        label="대화 이력 Excel 다운로드",
        data=history_excel_bytes,
        file_name="conversation_history_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_conversation_history_excel",
        use_container_width=True,
    )


def render_notice_guidance_page() -> None:
    st.header("공지 및 지침")
    st.info("MineSafe AI는 공식 광산 안전 문서와 중대재해처벌법 관련 자료를 함께 검색하여 답변합니다.")
    st.markdown(
        """
        - 중대재해처벌법 대응 체크리스트는 안전보건 확보의무 이행 상태를 정리하는 보조 기능입니다.
        - 질문 답변 후 추천되는 증빙자료는 현장 기록 관리 후보이며, 실제 법적 판단은 전문가 검토가 필요합니다.
        - 위험성평가 초안은 안전관리자가 검토·수정해야 하며 최종 법적 문서로 바로 사용할 수 없습니다.
        - 대화 이력은 AI 답변 기록이며, 실제 점검표·사진·교육기록 등 이행자료와 함께 관리해야 합니다.
        """
    )


# ==============================
# 사이드바
# ==============================
audience_mode = st.sidebar.radio(
    "사용자 모드",
    ["현장관리자 모드", "관리자/개발자 모드"],
    index=0,
)
is_admin_mode = audience_mode == "관리자/개발자 모드"

if is_admin_mode:
    st.sidebar.markdown(
        """
        <div class="mscc-sidebar-brand">
            <strong>MineSafe AI</strong><br>
            <div class="portal-sidebar-en">공식 문서 기반 광산 안전관리 AI 시스템</div>
        </div>
        <div class="portal-nav-group">운영 및 답변</div>
        <div class="portal-nav-item active">안전 질의 및 답변</div>
        <div class="portal-nav-item">대화 이력</div>
        <div class="portal-nav-item">공지 및 지침</div>
        <div class="portal-nav-group">지식 관리</div>
        <div class="portal-nav-item">Vector DB 관리</div>
        <div class="portal-nav-item">문서 관리</div>
        <div class="portal-nav-item">모델 관리</div>
        <div class="portal-nav-group">시스템 관리</div>
        <div class="portal-nav-item">사용자 관리</div>
        <div class="portal-nav-item">권한 관리</div>
        <div class="portal-nav-item">설정</div>
        """,
        unsafe_allow_html=True,
    )
else:
    st.sidebar.markdown(
        """
        <div class="mscc-sidebar-brand">
            <strong>MineSafe AI</strong><br>
            <div class="portal-sidebar-en">현장 안전관리자용 화면</div>
        </div>
        <div class="portal-nav-group">실무 메뉴</div>
        <div class="portal-nav-item active">안전 질문 답변</div>
        <div class="portal-nav-item">중대재해처벌법 대응 체크리스트</div>
        <div class="portal-nav-item">위험성평가 초안</div>
        <div class="portal-nav-item">대화 이력</div>
        <div class="portal-nav-item">Excel 내보내기 또는 보고서</div>
        """,
        unsafe_allow_html=True,
    )

chunk_count: int | str = 0
collection, db_error = load_chroma_collection()
if db_error:
    if is_admin_mode:
        st.sidebar.error("Vector DB · 연결 실패")
        st.sidebar.write(db_error)
else:
    try:
        chunk_count = collection.count()
    except Exception as e:
        chunk_count = "확인 실패"
        if is_admin_mode:
            st.sidebar.warning(f"chunk 수 확인 실패: {e}")

    if is_admin_mode:
        st.sidebar.success("Vector DB · 정상")
        st.sidebar.write(f"저장된 chunk 수: **{chunk_count}개**")
        st.sidebar.write(f"DB 폴더: `{VECTOR_DB_DIR.name}`")

law_badges_sidebar = "".join(
    f'<span class="major-law-mini-badge">{escape(label)}</span>'
    for label, _ in MAJOR_ACCIDENT_LAW_DOCS
)
if is_admin_mode:
    st.sidebar.markdown(
        (
            '<div class="major-law-sidebar-card">'
            '<div class="major-law-sidebar-title">법령 자료 반영 현황</div>'
            '<div class="major-law-sidebar-line">중대재해처벌법 자료 반영 완료</div>'
            '<div class="major-law-sidebar-line">반영 자료: FAQ / 질의회시집 / 해설서 / 따라하기 안내서</div>'
            f'<div class="major-law-badge-row">{law_badges_sidebar}</div>'
            f'<div class="major-law-sidebar-line">통합 chunk 수: {MAJOR_ACCIDENT_DOC_TOTAL_CHUNKS}개</div>'
            f'<div class="major-law-sidebar-line">추가 자료 chunk 수: {MAJOR_ACCIDENT_DOC_ADDED_CHUNKS}개</div>'
            f'<div class="major-law-sidebar-line">현재 Vector DB: {escape(VECTOR_DB_DIR.name)}</div>'
            "</div>"
        ),
        unsafe_allow_html=True,
    )
else:
    st.sidebar.markdown(
        (
            '<div class="major-law-sidebar-card">'
            '<div class="major-law-sidebar-title">법령 자료 반영</div>'
            '<div class="major-law-sidebar-line">중대재해처벌법 공식 자료를 함께 활용합니다.</div>'
            f'<div class="major-law-badge-row">{law_badges_sidebar}</div>'
            "</div>"
        ),
        unsafe_allow_html=True,
    )

if is_admin_mode:
    selected_gemini_model = st.sidebar.selectbox(
        "Gemini 모델",
        GEMINI_MODEL_OPTIONS,
        index=0,
    )

    _, gemini_error = load_gemini_client()
    if gemini_error:
        st.sidebar.warning("Gemini API · 키 미감지")
        st.sidebar.caption("로컬 .env 또는 Streamlit Cloud Secrets의 GEMINI_API_KEY를 확인하세요.")
        st.sidebar.caption(gemini_error)
    else:
        st.sidebar.info("Gemini API · API 키 감지됨")
        st.sidebar.caption("실제 연결 상태는 연결 테스트 또는 답변 호출 결과로 확인합니다.")
        st.sidebar.caption(f"선택 모델: {selected_gemini_model}")
        st.sidebar.write(f"호출 제한: 시도당 약 {GEMINI_RESPONSE_TIMEOUT_SECONDS}초")
        st.sidebar.write(f"최대 시도: {GEMINI_MAX_ATTEMPTS}회")

    if st.sidebar.button("Gemini 연결 테스트", use_container_width=True):
        with st.sidebar:
            with st.spinner("선택한 모델로 연결을 확인하는 중입니다..."):
                test_result = test_gemini_connection(
                    selected_gemini_model
                )
        test_state = classify_gemini_status(test_result)
        if test_result.get("success"):
            st.sidebar.success(
                f"연결 테스트 성공 · {test_result.get('model')} · "
                f"{test_result.get('elapsed', 0.0):.1f}초"
            )
            st.sidebar.caption(f"Gemini 응답: {test_result.get('answer', '')}")
        else:
            st.sidebar.warning(
                f"연결 테스트 실패 · {test_result.get('model')} · "
                f"{format_gemini_status(test_state)}"
            )
            st.sidebar.caption(test_result.get("message", "Gemini 연결 테스트 실패"))
            with st.sidebar.expander("오류 상세"):
                st.write(test_result.get("error") or "오류 메시지 없음")
                st.write(f"시도 횟수: {test_result.get('attempts', 0)}")

    st.sidebar.divider()

    answer_mode = st.sidebar.selectbox(
        "답변 생성 방식",
        [STABLE_MODE, GEMINI_MODE, HYBRID_MODE],
        index=0,
    )

    if answer_mode == STABLE_MODE:
        st.sidebar.info("안정 모드 · 외부 LLM 미호출")
    elif answer_mode == GEMINI_MODE:
        st.sidebar.warning("Gemini 모드 · 실패 시 근거 답변 전환")
    else:
        st.sidebar.info(
            "하이브리드 모드\n\n"
            "- 1차: 검색 근거 기반 안정형 답변\n"
            "- 2차: Gemini 기반 보조 답변\n"
            "- Gemini 실패 시에도 1차 답변 유지"
        )

    with st.sidebar.expander("시연 권장 설정"):
        st.markdown(
            "- **답변 모드:** 하이브리드 모드 권장\n"
            "- **Gemini 모델:** `gemini-2.5-flash-lite` 권장\n"
            "- **이유:** 안정형 답변을 먼저 제공하고 Gemini 응답을 보조적으로 확인\n"
            "- Gemini가 실패해도 검색 근거 기반 답변은 계속 제공됩니다."
        )

    top_k = st.sidebar.slider(
        "검색할 근거 문서 수",
        min_value=3,
        max_value=5,
        value=3,
        step=1,
    )

    with st.sidebar.expander("시연 상태 체크"):
        st.markdown(
            "- Vector DB 로드 확인\n"
            "- Gemini 모델 선택 확인\n"
            "- 답변 모드는 하이브리드 모드 권장\n"
            "- 답변과 함께 근거 문서가 표시되는지 확인"
        )

    selected_scenario_label = st.sidebar.selectbox(
        "질문 시나리오 세트",
        list(SCENARIO_SET_OPTIONS.keys()),
        index=list(SCENARIO_SET_OPTIONS.keys()).index("110개 분진·보호구 보강 평가 세트"),
    )
    selected_scenario_path = SCENARIO_SET_OPTIONS[selected_scenario_label]

    st.sidebar.markdown(
        """
        <div class="mscc-sidebar-note">
            테스트 모드에서는 선택한 시나리오 세트의 질문 검색, 답변 검토, 평가 저장과 진행률 확인을 지원합니다.
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.sidebar.caption("시연 시에는 110개 분진·보호구 보강 평가 세트를 선택하는 것을 권장합니다.")
    feature_page = st.sidebar.radio(
        "기능 메뉴",
        ["안전 질의 및 답변", "중대재해처벌법 대응", "위험성평가 초안", "대화 이력", "공지 및 지침"],
        index=0,
    )

    st.sidebar.markdown(
        f"""
        <div class="portal-system-state">
            <div class="portal-system-state-title">시스템 상태</div>
            <div class="portal-system-state-value">정상 운영 중</div>
            <div class="portal-system-state-time">
                마지막 동기화: {escape(time.strftime("%Y-%m-%d %H:%M"))}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
else:
    selected_gemini_model = GEMINI_MODEL_OPTIONS[0]
    gemini_error = "현장관리자 모드에서는 Gemini 상태를 숨깁니다."
    answer_mode = STABLE_MODE
    top_k = 3
    selected_scenario_label = "110개 분진·보호구 보강 평가 세트"
    selected_scenario_path = SCENARIO_SET_OPTIONS[selected_scenario_label]
    feature_page = st.sidebar.radio(
        "기능 메뉴",
        ["안전 질문 답변", "중대재해처벌법 대응 체크리스트", "위험성평가 초안", "대화 이력", "Excel 내보내기 또는 보고서"],
        index=0,
    )


# ==============================
# 상단 운영 상태
# ==============================
scenario_rows, scenario_status_error = load_question_scenarios(selected_scenario_path)
scenario_count = len(scenario_rows) if not scenario_status_error else 0
db_status_text = "정상" if not db_error else "점검 필요"
db_status_accent = "#10b981" if not db_error else "#b45309"
gemini_status_text = (
    "관리자 모드에서 확인"
    if not is_admin_mode
    else ("API 키 감지됨" if not gemini_error else "키 미감지")
)

status_col1, status_col2, status_col3, status_col4 = st.columns(4)
with status_col1:
    render_status_card(
        "Vector DB",
        db_status_text,
        VECTOR_DB_DIR.name,
        db_status_accent,
    )
with status_col2:
    render_status_card(
        "문서 Chunk",
        f"{chunk_count}개",
        COLLECTION_NAME,
        "#64748b",
    )
with status_col3:
    render_status_card(
        "답변 모드",
        short_answer_mode(answer_mode),
        f"{selected_gemini_model} · Gemini API: {gemini_status_text}",
        "#64748b",
    )
with status_col4:
    render_status_card(
        "평가 시나리오",
        f"{scenario_count}개",
        f"{selected_scenario_label}",
        "#64748b",
    )


# ==============================
# 메뉴별 화면 전환
# ==============================
if feature_page in {"중대재해처벌법 대응", "중대재해처벌법 대응 체크리스트"}:
    render_legal_checklist_page()
    st.stop()
if feature_page == "위험성평가 초안":
    render_risk_assessment_draft_page()
    st.stop()
if feature_page == "대화 이력":
    render_conversation_history_page()
    st.stop()
if feature_page == "공지 및 지침":
    render_notice_guidance_page()
    st.stop()
if feature_page == "Excel 내보내기 또는 보고서":
    render_export_report_page()
    st.stop()

# ==============================
# 질문 입력 + 테스트 모드
# ==============================
def run_rag_flow(
    question_text: str,
    top_k_value: int,
    answer_mode: str,
    selected_model: str = GEMINI_MODEL_NAME,
):
    with st.spinner("1단계: Vector DB에서 관련 근거 문서를 검색 중입니다..."):
        rag_results, search_error = search_vector_db(question_text, top_k=top_k_value)

    if search_error:
        return [], "", {}, search_error
    if not rag_results:
        return [], "", {}, "검색 결과가 없습니다."

    fallback_answer = generate_local_fallback_answer(
        question_text,
        rag_results,
        "안정 모드: Gemini API 호출 안 함",
    )

    if answer_mode == STABLE_MODE:
        rag_status = {
            "called": False,
            "success": False,
            "status": "not_called",
            "message": "안정 모드에서는 Gemini API를 호출하지 않습니다.",
            "answer": "",
            "model": selected_model,
            "used_fallback": False,
            "mode": "stable",
            "answer_mode": short_answer_mode(answer_mode),
            "gemini_status": "호출 안 함",
            "reason": "안정 모드에서는 Gemini API를 호출하지 않습니다.",
            "attempts": 0,
            "elapsed": 0.0,
            "selected_model": selected_model,
            "gemini_called": False,
            "fallback_used": False,
        }
        return rag_results, fallback_answer, rag_status, None

    if answer_mode == HYBRID_MODE:
        with st.spinner(
            f"하이브리드 모드: 검색 근거 기반 답변을 우선 사용하고, Gemini 추가 답변을 최대 {GEMINI_RESPONSE_TIMEOUT_SECONDS}초 동안 시도합니다..."
        ):
            gemini_answer, gemini_status = generate_gemini_answer(
                question_text,
                rag_results,
                selected_model,
            )

        gemini_state = classify_gemini_status(gemini_status)
        if gemini_status.get("mode") == "gemini":
            combined_answer = "\n\n".join(
                [
                    fallback_answer,
                    "---",
                    "## Gemini 추가 답변",
                    gemini_answer,
                ]
            )
        else:
            combined_answer = "\n\n".join(
                [
                    fallback_answer,
                    "---",
                    "## Gemini 응답 실패",
                    "Gemini 응답 실패: 검색 근거 기반 답변을 사용합니다.",
                ]
            )

        gemini_status.update(
            {
                "mode": "hybrid",
                "answer_mode": short_answer_mode(answer_mode),
                "gemini_status": gemini_state,
                "fallback_used": gemini_state != "성공",
                "used_fallback": gemini_state != "성공",
            }
        )
        return rag_results, combined_answer, gemini_status, None

    with st.spinner(
        f"Gemini 모드: Gemini 답변 생성 중입니다. {GEMINI_RESPONSE_TIMEOUT_SECONDS}초 내 응답이 없으면 검색 근거 기반 답변으로 전환합니다..."
    ):
        rag_answer, rag_status = generate_gemini_answer(
            question_text,
            rag_results,
            selected_model,
        )

    rag_status.update(
        {
            "answer_mode": short_answer_mode(answer_mode),
            "gemini_status": classify_gemini_status(rag_status),
        }
    )

    return rag_results, rag_answer, rag_status, None


def render_evidence_card(result: dict[str, Any]) -> None:
    distance_text = format_distance(result["distance"])
    source = str(result.get("source", "출처 정보 없음"))
    chunk_id = str(result.get("chunk_id", "정보 없음"))
    rank = result.get("rank", "-")
    preview = make_preview(str(result.get("text", "")), limit=520)

    with st.container(border=True):
        st.markdown(
            (
                f'<div class="mscc-evidence-title">'
                f'근거 {escape(str(rank))} | {escape(source)}</div>'
                f'<div class="mscc-evidence-meta">'
                f'CHUNK {escape(chunk_id)} · DISTANCE {escape(distance_text)}</div>'
                f'<div class="mscc-evidence-preview">{escape(preview)}</div>'
            ),
            unsafe_allow_html=True,
        )
        with st.expander("전체 본문 및 메타데이터"):
            st.write(f"문서명: {source}")
            st.write(f"chunk_id: {chunk_id}")
            st.write(f"distance 또는 유사도: {distance_text}")
            st.markdown("**전체 본문**")
            st.write(result.get("text", ""))
            st.markdown("**메타데이터**")
            st.json(result.get("metadata", {}))


def split_answer_for_dashboard(answer: str) -> tuple[str, str, str]:
    kras_markers = [
        "### KRAS식 위험성평가 기록 초안",
        "## 2. KRAS식 위험성평가 기록 초안",
        "## KRAS식 위험성평가 기록 초안",
    ]
    kras_start = -1
    for marker in kras_markers:
        marker_index = answer.find(marker)
        if marker_index >= 0 and (kras_start < 0 or marker_index < kras_start):
            kras_start = marker_index

    supplement_markers = [
        "\n---\n## Gemini 추가 답변",
        "\n## Gemini 추가 답변",
        "\n---\n## Gemini 응답 실패",
        "\n## Gemini 응답 실패",
    ]
    supplement_start = -1
    for marker in supplement_markers:
        marker_index = answer.find(marker)
        if marker_index >= 0 and (
            supplement_start < 0 or marker_index < supplement_start
        ):
            supplement_start = marker_index

    if kras_start < 0:
        core_end = supplement_start if supplement_start >= 0 else len(answer)
        core_answer = answer[:core_end].strip()
        evidence_marker = core_answer.find("\n## 관련 근거 문서")
        if evidence_marker >= 0:
            core_answer = core_answer[:evidence_marker].strip()
        return core_answer, "", answer[core_end:].strip()

    kras_end = (
        supplement_start
        if supplement_start >= 0 and supplement_start > kras_start
        else len(answer)
    )
    core_answer = answer[:kras_start].strip()
    evidence_marker = core_answer.find("\n## 관련 근거 문서")
    if evidence_marker >= 0:
        core_answer = core_answer[:evidence_marker].strip()
    return (
        core_answer,
        answer[kras_start:kras_end].strip(),
        answer[kras_end:].strip(),
    )


def extract_markdown_section(markdown_text: str, heading: str) -> str:
    marker = f"## {heading}"
    start = markdown_text.find(marker)
    if start < 0:
        return ""
    content_start = start + len(marker)
    next_heading = markdown_text.find("\n## ", content_start)
    content_end = next_heading if next_heading >= 0 else len(markdown_text)
    return markdown_text[content_start:content_end].strip()


def strip_markdown_prefix(text: str) -> str:
    cleaned_lines = []
    for line in text.splitlines():
        cleaned = line.strip()
        if not cleaned:
            continue
        while cleaned.startswith(("- ", "* ", "• ")):
            cleaned = cleaned[2:].strip()
        cleaned_lines.append(cleaned)
    return " ".join(cleaned_lines)


def build_dashboard_summary(
    core_answer: str,
    situation_type: str,
) -> tuple[str, list[str]]:
    immediate_text = strip_markdown_prefix(
        extract_markdown_section(core_answer, "즉시 판단")
    )
    if not immediate_text:
        immediate_text = build_immediate_judgment(situation_type)

    action_items: list[str] = []
    for item in build_priority_actions(situation_type) + build_check_items(situation_type):
        normalized = clean_text(str(item))
        if normalized and normalized not in action_items:
            action_items.append(normalized)
        if len(action_items) >= 6:
            break
    return immediate_text, action_items


def strip_kras_title(kras_answer: str) -> str:
    lines = kras_answer.splitlines()
    while lines and (
        not lines[0].strip()
        or "KRAS식 위험성평가 기록 초안" in lines[0]
    ):
        lines.pop(0)
    return "\n".join(lines).strip()


def decorate_kras_markdown(kras_answer: str) -> str:
    decorated_lines = []
    for line in kras_answer.splitlines():
        if "현재 위험성" in line:
            line = line.replace(
                "매우 높음",
                '<span class="status-badge badge-danger">매우 높음</span>',
            ).replace(
                "Level 5",
                '<span class="status-badge badge-danger">Level 5</span>',
            )
        elif "조치 후 잔여위험성" in line:
            line = line.replace(
                "낮음",
                '<span class="status-badge badge-success">낮음</span>',
            ).replace(
                "Level 2",
                '<span class="status-badge badge-success">Level 2</span>',
            )
        decorated_lines.append(line)
    return "\n".join(decorated_lines)


def render_info_card(
    title: str,
    value: str,
    subtitle: str,
    icon: str,
    tone: str,
) -> None:
    st.markdown(
        (
            f'<div class="info-card info-card-{escape(tone)}">'
            f'<div class="info-card-icon">{escape(icon)}</div>'
            '<div class="info-card-body">'
            f'<div class="info-card-label">{escape(title)}</div>'
            f'<div class="info-card-value">{escape(value)}</div>'
            f'<div class="info-card-subtitle">{escape(subtitle)}</div>'
            "</div></div>"
        ),
        unsafe_allow_html=True,
    )


def classify_evidence_document(source: str) -> tuple[str, str]:
    normalized = source.replace(" ", "").lower()
    if "중대재해처벌법" in normalized or "중대산업재해" in normalized:
        if "faq" in normalized:
            return "FAQ", "law-badge law-badge-faq"
        if "질의회시" in normalized:
            return "질의회시", "law-badge law-badge-reply"
        if "해설서" in normalized:
            return "해설서", "law-badge law-badge-commentary"
        if "따라하기" in normalized or "안내서" in normalized:
            return "따라하기", "law-badge law-badge-major"
        return "중대재해법", "law-badge law-badge-major"
    if "법" in normalized and "지침" not in normalized:
        return "법령", "law-badge"
    if "지침" in normalized or "가이드" in normalized or "안내서" in normalized:
        return "지침", "law-badge law-badge-guideline"
    if "기준" in normalized or "고시" in normalized:
        return "고시", "law-badge law-badge-notice"
    return "참고", "law-badge law-badge-reference"


def render_evidence_table(results: list[dict[str, Any]]) -> None:
    row_parts = []
    for index, item in enumerate(results[:5], start=1):
        source = str(item.get("source", "출처 정보 없음"))
        document_type, badge_class = classify_evidence_document(source)
        status_text = "적용" if index <= 2 else "참고"
        status_class = "badge-success" if index <= 2 else "badge-muted"
        row_parts.append(
            (
                "<tr>"
                f"<td>{escape(str(item.get('rank', index)))}</td>"
                f'<td><span class="{badge_class}">{escape(document_type)}</span></td>'
                f"<td>{escape(source)}</td>"
                f"<td>{escape(str(item.get('chunk_id', '정보 없음')))}</td>"
                f"<td>{escape(format_distance(item.get('distance')))}</td>"
                f'<td><span class="status-badge {status_class}">{status_text}</span></td>'
                "</tr>"
            )
        )
    rows_html = "".join(row_parts)
    st.markdown(
        (
            '<div class="portal-table-scroll table-wrap">'
            '<table class="portal-data-table portal-evidence-table">'
            "<colgroup>"
            '<col style="width:7%"><col style="width:11%">'
            '<col style="width:39%"><col style="width:23%">'
            '<col style="width:11%"><col style="width:9%">'
            "</colgroup>"
            "<thead><tr><th>근거</th><th>구분</th><th>문서명</th>"
            "<th>chunk_id</th><th>거리값</th><th>상태</th></tr></thead>"
            f"<tbody>{rows_html}</tbody></table></div>"
        ),
        unsafe_allow_html=True,
    )


def render_major_accident_law_evidence_panel() -> None:
    badge_html = "".join(
        f'<span class="law-badge law-badge-major">{escape(label)}</span>'
        for label, _ in MAJOR_ACCIDENT_LAW_DOCS
    )
    docs_html = "".join(
        (
            '<div class="major-law-doc-item">'
            f'<span class="law-badge law-badge-major">{escape(label)}</span> '
            f"{escape(doc_name)}</div>"
        )
        for label, doc_name in MAJOR_ACCIDENT_LAW_DOCS
    )
    st.markdown(
        (
            '<div class="major-law-evidence-box">'
            '<div class="major-law-evidence-title">중대재해처벌법 활용 근거</div>'
            '<div class="major-law-evidence-text">'
            "본 시스템은 광산 안전 지침뿐 아니라 중대재해처벌법 관련 공식 자료를 "
            "함께 검색하여 답변 생성에 활용합니다."
            "</div>"
            f'<div class="major-law-badge-row">{badge_html}'
            '<span class="law-badge law-badge-major">중대재해처벌법</span></div>'
            f'<div class="major-law-doc-list">{docs_html}</div>'
            "</div>"
        ),
        unsafe_allow_html=True,
    )


def render_checklist_table(situation_type: str) -> None:
    rows_html = "".join(
        (
            "<tr>"
            f"<td>{index}</td>"
            f"<td>{escape(clean_text(str(item)))}</td>"
            '<td><span class="portal-badge portal-badge-wait">대기</span></td>'
            "<td>-</td>"
            "<td>현장 책임자</td>"
            "</tr>"
        )
        for index, item in enumerate(
            build_check_items(situation_type)[:7],
            start=1,
        )
    )
    st.markdown(
        (
            '<div class="portal-table-scroll table-wrap">'
            '<table class="portal-data-table portal-checklist-table">'
            "<colgroup>"
            '<col style="width:6%"><col style="width:55%">'
            '<col style="width:12%"><col style="width:12%"><col style="width:15%">'
            "</colgroup>"
            "<thead><tr><th>번호</th><th>조치 항목</th><th>이행 상태</th>"
            "<th>이행일</th><th>담당자</th></tr></thead>"
            f"<tbody>{rows_html}</tbody></table></div>"
        ),
        unsafe_allow_html=True,
    )


def render_gemini_runtime_status(
    answer_status: dict[str, Any],
    selected_model: str,
    mode_name: str,
) -> None:
    gemini_state = classify_gemini_status(answer_status)
    gemini_called = bool(
        answer_status.get(
            "called",
            answer_status.get("gemini_called", False),
        )
    )
    fallback_used = bool(
        answer_status.get(
            "used_fallback",
            answer_status.get("fallback_used", False),
        )
    )
    display_gemini_state = format_gemini_status(gemini_state)

    with st.expander("답변 생성 상태 및 외부 LLM 호출 정보"):
        status_col1, status_col2, status_col3, status_col4, status_col5 = st.columns(5)
        status_col1.metric("모드", mode_name)
        status_col2.metric("모델", selected_model)
        status_col3.metric("호출 여부", "호출함" if gemini_called else "호출 안 함")
        status_col4.metric("상태", display_gemini_state)
        status_col5.metric("Fallback", "사용함" if fallback_used else "사용 안 함")

        if answer_status.get("mode") == "stable":
            st.info(
                "안정 모드입니다. 외부 LLM을 호출하지 않고 Vector DB 검색 근거로 답변했습니다."
            )
        elif gemini_state == "성공":
            st.success(
                f"Gemini 답변 생성 완료 · {answer_status.get('attempts', 0)}회 시도 · "
                f"{answer_status.get('elapsed', 0.0):.1f}초"
            )
        else:
            st.warning(
                answer_status.get("message")
                or get_gemini_failure_message(
                    gemini_state,
                    str(answer_status.get("reason", "")),
                )
            )
            st.info(
                "Gemini 응답 생성은 실패했지만 검색 근거 기반 안정형 답변으로 "
                "전환되었습니다. 근거 문서 검색과 안정형 답변 생성 기능은 정상 작동 중입니다."
            )
            st.caption(
                f"시도 횟수: {answer_status.get('attempts', 0)} · "
                f"소요 시간: {answer_status.get('elapsed', 0.0):.1f}초"
            )
            if answer_status.get("reason"):
                st.caption(f"오류 정보: {answer_status.get('reason')}")


def render_rag_result(
    answer: str,
    answer_status: dict[str, Any],
    results: list[dict[str, Any]],
    result_key: str = "rag",
    question_text: str = "",
    auto_save_history: bool = True,
) -> None:
    st.subheader("답변 상세")
    mode_name = answer_status.get("answer_mode", "알 수 없음")
    question_type = (
        results[0].get("question_type", "일반")
        if results
        else "일반"
    )
    reranking_applied = bool(
        results and results[0].get("reranking_applied")
    )
    situation_type = resolve_display_situation_type(question_type, answer)
    risk_level, risk_accent = get_risk_level(situation_type)
    priority_action = get_priority_action(situation_type)
    selected_model = str(
        answer_status.get(
            "model",
            answer_status.get("selected_model", GEMINI_MODEL_NAME),
        )
    )
    elapsed = float(answer_status.get("elapsed", 0.0) or 0.0)
    unique_sources = list(
        dict.fromkeys(str(item.get("source", "출처 정보 없음")) for item in results)
    )
    confidence = "높음" if len(results) >= 3 else "검토 필요"
    generation_time = f"{elapsed:.1f}초" if elapsed > 0 else "즉시 생성"
    core_answer, kras_answer, supplement_answer = split_answer_for_dashboard(answer)
    effective_question = question_text.strip() or extract_markdown_section(answer, "질문") or "질문 정보 없음"
    recommended_records = recommend_evidence_records(effective_question, answer)

    info_cols = st.columns(5)
    info_cards = [
        ("답변 신뢰도", confidence, f"근거 {len(results)}개 확보", "✓", "success"),
        ("위험 등급", risk_level, situation_type, "!", "warning"),
        ("관련 법령", f"{len(unique_sources)}개 문서", "Vector DB 검색 기준", "§", "navy"),
        ("답변 생성 시간", generation_time, mode_name, "⏱", "teal"),
        ("모델 정보", selected_model, "선택된 답변 모델", "M", "blue"),
    ]
    for info_col, (title, value, description, icon, tone) in zip(info_cols, info_cards):
        with info_col:
            render_info_card(title, value, description, icon, tone)

    st.markdown(
        (
            '<div class="mscc-priority-strip">'
            f'<strong>우선 조치</strong> · {escape(priority_action)}'
            "</div>"
        ),
        unsafe_allow_html=True,
    )
    if reranking_applied:
        st.caption(
            results[0].get(
                "reranking_label",
                "광산 특화 문서 우선 정렬 적용",
            )
        )
    else:
        st.caption("기본 벡터 유사도 정렬 적용")

    render_gemini_runtime_status(answer_status, selected_model, mode_name)

    summary_lead, summary_actions = build_dashboard_summary(
        core_answer,
        situation_type,
    )
    summary_items_html = "".join(
        f"<li>{escape(item)}</li>" for item in summary_actions
    )

    answer_col, evidence_col = st.columns([3, 2], gap="large")
    with answer_col:
        with st.container(border=True):
            st.markdown(
                '<div class="card-title-row">'
                '<span class="card-title-icon card-title-icon-success">✓</span>'
                '<div class="portal-card-title">답변 요약</div></div>'
                '<div class="portal-card-subtitle">Evidence-based safety response</div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                (
                    f'<div class="portal-summary-lead">{escape(summary_lead)}</div>'
                    f'<ul class="portal-summary-list">{summary_items_html}</ul>'
                ),
                unsafe_allow_html=True,
            )

    with evidence_col:
        with st.container(border=True):
            st.markdown(
                '<div class="card-title-row">'
                '<span class="card-title-icon card-title-icon-navy">§</span>'
                '<div class="portal-card-title">관련 법령 및 문서 근거</div></div>'
                '<div class="portal-card-subtitle">검색 상위 3~5개 공식 문서</div>',
                unsafe_allow_html=True,
            )
            render_evidence_table(results)

    with st.container(border=True):
        st.markdown(
            '<div class="card-title-row">'
            '<span class="card-title-icon card-title-icon-warning">✓</span>'
            '<div class="portal-card-title">조치 체크리스트</div></div>'
            '<div class="portal-card-subtitle">'
            "사업주·경영책임자 및 현장 관리자 확인용"
            "</div>",
            unsafe_allow_html=True,
        )
        render_checklist_table(situation_type)
        st.caption("이행 상태·이행일·담당자는 현장 확인 후 기록합니다.")

    with st.container(border=True):
        st.markdown(
            '<div class="card-title-row">'
            '<span class="card-title-icon card-title-icon-navy">▣</span>'
            '<div class="portal-card-title">KRAS식 위험성평가 기록 초안</div></div>'
            '<div class="portal-kras-note">'
            "검색 근거와 질문 상황을 바탕으로 작성한 기입 초안이며, "
            "최종 가능성·중대성·위험등급은 현장 기준에 따라 재평가해야 합니다."
            "</div>",
            unsafe_allow_html=True,
        )
        if kras_answer:
            render_kras_readable_markdown(kras_answer)
        else:
            st.info("KRAS 초안이 포함되지 않은 답변입니다. 검색 근거를 확인해 주세요.")

    if supplement_answer:
        with st.expander("Gemini 보조 답변 보기", expanded=False):
            st.markdown(supplement_answer)

    with st.expander("전체 답변 원문 보기", expanded=False):
        st.markdown(answer)

    render_major_accident_law_evidence_panel()

    render_recommended_evidence_records(recommended_records)
    if auto_save_history:
        saved_history_id = auto_save_conversation_history(
            effective_question,
            answer,
            situation_type,
            risk_level,
            results,
            recommended_records,
            result_key,
        )
        if saved_history_id:
            st.caption(f"대화 이력 자동 저장 완료: {saved_history_id}")

    ribbon_items = "".join(
        (
            '<span class="portal-source-chip">'
            f'{escape(str(item.get("source", "출처 정보 없음")))} · '
            f'{escape(str(item.get("chunk_id", "정보 없음")))}'
            "</span>"
        )
        for item in results[:5]
    )
    st.markdown(
        (
            '<div class="portal-evidence-ribbon">'
            '<span class="portal-ribbon-label">문서 근거 · Evidence Ribbon</span>'
            f"{ribbon_items}</div>"
        ),
        unsafe_allow_html=True,
    )

    action_col1, action_col2, action_col3, action_spacer = st.columns([1, 1, 1.2, 3])
    if action_col1.button("인쇄", key=f"print_{result_key}", use_container_width=True):
        st.info("브라우저 인쇄 기능에서 현재 답변 화면을 인쇄할 수 있습니다.")
    if action_col2.button(
        "PDF 내보내기",
        key=f"pdf_{result_key}",
        use_container_width=True,
    ):
        st.info("브라우저 인쇄 메뉴에서 'PDF로 저장'을 선택해 주세요.")
    if action_col3.button(
        "즐겨찾기 추가",
        key=f"favorite_{result_key}",
        use_container_width=True,
    ):
        st.session_state["favorite_answer"] = {
            "answer": answer,
            "sources": unique_sources,
            "saved_at": time.strftime("%Y-%m-%d %H:%M"),
        }
        st.success("현재 답변을 세션 즐겨찾기에 추가했습니다.")

    with st.expander(f"근거 문서 상세 보기 ({len(results)}개)"):
        for result in results:
            render_evidence_card(result)


direct_tab, scenario_tab = st.tabs(["직접 질문", "질문 시나리오 테스트 모드"])

with direct_tab:
    st.subheader("질문 입력")

    demo_questions = [
        "갱내에서 메탄가스 농도가 높게 감지되면 어떻게 조치해야 해?",
        "발파 후 불발이 의심될 때 작업자는 어떻게 대응해야 해?",
        "분진이 많은 굴진 작업에서 방진마스크를 어떤 기준으로 지급하고 착용 상태를 확인해야 해?",
        "광산 현장에서 위험성평가를 실시할 때 어떤 절차로 진행해야 해?",
        "전기설비 점검 중 감전 위험이 있을 때 작업 전 확인해야 할 사항은 뭐야?",
    ]

    with st.expander("시연용 질문 예시"):
        for index, demo_question in enumerate(demo_questions, start=1):
            st.markdown(f"{index}. {demo_question}")

    example_questions = [
        "갱내 메탄가스 기준은 어떻게 확인해야 해?",
        "발파 작업 전에 현장 관리자가 확인해야 할 안전 체크리스트 알려줘",
        "갱내 작업 전에 환기 상태는 어떻게 점검해야 해?",
        "낙반 위험이 있을 때 작업을 중지해야 하는 기준은 뭐야?",
        "중대재해가 발생하면 현장 관리자는 먼저 무엇을 해야 해?",
        "중대재해처벌법상 경영책임자의 안전보건 확보의무를 요약해줘",
    ]

    selected_example = st.selectbox(
        "예시 질문 선택",
        ["직접 입력"] + example_questions,
    )

    default_question = (
        "갱내 메탄가스 기준은 어떻게 확인해야 해?"
        if selected_example == "직접 입력"
        else selected_example
    )

    question = st.text_area(
        "현장 관리자 질문",
        value=default_question,
        height=100,
        key=f"direct_question_{st.session_state['new_question_token']}",
    )

    col1, col2 = st.columns([1, 3])
    with col1:
        search_button = st.button("RAG 답변 생성", type="primary")
    with col2:
        st.caption(
            "Vector DB에서 근거 문서를 찾고 Gemini 답변을 시도합니다. "
            "Gemini가 지연되면 검색 근거 기반 fallback 답변을 바로 표시합니다."
        )

    if search_button:
        if not question.strip():
            st.warning("질문을 입력해 주세요.")
        elif db_error:
            st.error("Vector DB를 로드할 수 없어 검색을 진행할 수 없습니다.")
            st.write(db_error)
        else:
            results, answer, answer_status, error = run_rag_flow(
                question,
                top_k,
                answer_mode,
                selected_gemini_model,
            )
            if error:
                st.error(error)
            else:
                st.success(f"Vector DB 검색 완료: 관련 근거 {len(results)}개")
                render_rag_result(
                    answer,
                    answer_status,
                    results,
                    result_key="direct",
                    question_text=question,
                )

with scenario_tab:
    st.subheader("질문 시나리오 테스트 모드")
    st.caption(
        f"현재 선택된 질문 세트: {selected_scenario_label}"
    )
    st.caption(SCENARIO_SET_DESCRIPTIONS.get(selected_scenario_label, ""))
    scenarios, scenario_error = load_question_scenarios(selected_scenario_path)
    progress_info = render_evaluation_progress(
        scenario_rows=scenarios if not scenario_error else None,
        include_auto_eval=bool(not scenario_error and len(scenarios) > 30),
    )
    render_auto_eval_summary()

    if scenario_error:
        st.error("질문 시나리오를 불러오지 못했습니다.")
        st.write(scenario_error)
    elif not scenarios:
        st.warning("질문 시나리오가 없습니다.")
    else:
        st.success(f"질문 시나리오 {len(scenarios)}개 로드 완료")

        status_by_no = progress_info.get("status_by_no", {}) if progress_info else {}
        scenario_options = {}
        for row in scenarios:
            no_text = str(row.get("번호", "")).strip()
            status_text = status_by_no.get(no_text, "미평가")
            label = f"Q{int(row.get('번호', 0)):02d} | {row.get('분류', '')} | {status_text}"
            scenario_options[label] = row

        selected_label = st.selectbox("테스트 질문 선택", list(scenario_options.keys()))
        selected_scenario = scenario_options[selected_label]

        info_col1, info_col2, info_col3 = st.columns(3)
        info_col1.metric("번호", f"Q{int(selected_scenario.get('번호', 0)):02d}")
        info_col2.metric("분류", selected_scenario.get("분류", ""))
        info_col3.metric("난이도", selected_scenario.get("난이도", ""))

        st.markdown("**질문 시나리오**")
        st.write(selected_scenario.get("질문 시나리오", ""))
        st.markdown("**기대 검색 문서**")
        st.write(selected_scenario.get("기대 검색 문서", ""))
        st.markdown("**정답에 포함되어야 할 핵심 요소**")
        st.write(selected_scenario.get("정답에 포함되어야 할 핵심 요소", ""))

        if st.button("선택한 질문으로 테스트 실행", type="primary"):
            if db_error:
                st.error("Vector DB를 로드할 수 없어 테스트를 진행할 수 없습니다.")
                st.write(db_error)
            else:
                scenario_question = selected_scenario.get("질문 시나리오", "")
                results, answer, answer_status, error = run_rag_flow(
                    scenario_question,
                    top_k,
                    answer_mode,
                    selected_gemini_model,
                )
                if error:
                    st.error(error)
                else:
                    st.session_state["scenario_test_result"] = {
                        "scenario_no": selected_scenario.get("번호", ""),
                        "scenario": selected_scenario,
                        "results": results,
                        "answer": answer,
                        "answer_status": answer_status,
                        "answer_mode": short_answer_mode(answer_mode),
                        "gemini_status": classify_gemini_status(answer_status),
                    }
                    st.success(f"선택 질문 RAG 테스트 완료: 관련 근거 {len(results)}개")

        stored_result = st.session_state.get("scenario_test_result")
        if stored_result:
            stored_scenario = stored_result["scenario"]
            stored_no = int(stored_scenario.get("번호", 0))
            previous_evaluation, previous_evaluation_error = get_evaluation_row(
                stored_no
            )
            has_previous_evaluation = has_saved_evaluation_result(
                previous_evaluation
            )
            st.divider()
            st.caption(f"현재 평가 대상: Q{stored_no:02d} | {stored_scenario.get('분류', '')}")

            render_rag_result(
                stored_result["answer"],
                stored_result["answer_status"],
                stored_result["results"],
                result_key=f"scenario_{stored_no}",
                question_text=stored_scenario.get("질문 시나리오", ""),
            )

            st.subheader("평가 입력")
            with st.expander("평가 기준 보기"):
                st.info(
                    "평가 기준은 4개 항목, 총 100점 만점입니다. "
                    "환각억제는 별도 항목으로 분리하지 않고, 근거 기반성과 안전·법령 판단 정확성에서 함께 감점합니다."
                )
                st.markdown("**검색 적합성**")
                st.caption("질문과 관련된 공식 문서 및 chunk를 제대로 검색했는지 평가한다.")
                st.markdown("**근거 기반성**")
                st.caption("답변이 검색된 근거 문서와 명확히 연결되는지 평가한다.")
                st.markdown("**안전·법령 판단 정확성**")
                st.caption("위험 상황 판단, 작업 중지, 대피, 보고, 보호구, 법령·지침 적용이 정확한지 평가한다.")
                st.markdown("**실무성**")
                st.caption("현장 관리자가 답변을 보고 바로 조치할 수 있을 만큼 구체적인지 평가한다.")
                criteria_text = load_evaluation_criteria()
                if criteria_text:
                    st.markdown(criteria_text)
                else:
                    st.warning("평가 기준 파일을 찾을 수 없습니다.")

            with st.expander("Q01 평가 예시 보기"):
                q01_example_text = load_q01_evaluation_example()
                if q01_example_text:
                    st.markdown(q01_example_text)
                else:
                    st.warning("Q01 평가 예시 파일을 찾을 수 없습니다.")

            score_col1, score_col2, score_col3, score_col4 = st.columns(4)
            scores = {
                "검색_적합성": score_col1.number_input(
                    "검색 적합성 (0~25)", min_value=0, max_value=25, value=0, step=1
                ),
                "근거_기반성": score_col2.number_input(
                    "근거 기반성 (0~25)", min_value=0, max_value=25, value=0, step=1
                ),
                "안전법령_판단정확성": score_col3.number_input(
                    "안전·법령 판단 정확성 (0~25)", min_value=0, max_value=25, value=0, step=1
                ),
                "실무성": score_col4.number_input(
                    "실무성 (0~25)", min_value=0, max_value=25, value=0, step=1
                ),
            }
            memo = st.text_area("메모", height=100)
            total_score = sum(scores.values())
            judgment = calculate_judgment(total_score)

            eval_col1, eval_col2 = st.columns(2)
            eval_col1.metric("총점", total_score)
            eval_col2.metric("판정", judgment)

            st.subheader("재테스트 비교")
            current_evidence = make_evidence_summary(stored_result["results"])
            current_profile = classify_evidence_profile(current_evidence)

            if previous_evaluation_error:
                st.warning(previous_evaluation_error)

            if has_previous_evaluation and previous_evaluation:
                previous_evidence = previous_evaluation.get(
                    "검색된 주요 근거 문서",
                    "",
                )
                previous_profile = classify_evidence_profile(previous_evidence)
                previous_col, current_col = st.columns(2)

                with previous_col:
                    st.markdown("**이전 저장 결과**")
                    st.write(f"이전 상태: {previous_profile}")
                    st.metric(
                        "이전 총점",
                        format_comparison_score(
                            previous_evaluation.get("총점")
                        ),
                    )
                    st.markdown("**이전 검색된 주요 근거 문서**")
                    st.write(previous_evidence or "저장된 검색 결과가 없습니다.")

                with current_col:
                    st.markdown("**현재 재테스트 결과**")
                    st.write(f"현재 상태: {current_profile}")
                    st.metric("현재 입력 예정 총점", f"{total_score}점")
                    st.markdown("**현재 검색된 주요 근거 문서**")
                    st.write(current_evidence or "현재 검색 결과가 없습니다.")

                save_policy = st.radio(
                    "기존 평가 처리 방식",
                    ["기존 평가 유지", "기존 평가 덮어쓰기"],
                    index=0,
                    horizontal=True,
                    key=f"evaluation_save_policy_{stored_no}",
                )
                if save_policy == "기존 평가 유지":
                    st.info(
                        "기존 평가 유지가 선택되어 있습니다. "
                        "저장 버튼을 눌러도 TSV의 기존 평가 결과는 변경되지 않습니다."
                    )
                else:
                    st.warning(
                        "기존 평가 덮어쓰기가 선택되어 있습니다. "
                        "저장하면 현재 근거 문서, 점수와 메모로 기존 행이 갱신됩니다."
                    )
            else:
                st.info("이 질문에는 비교할 기존 저장 평가 결과가 없습니다.")
                st.write(f"현재 상태: {current_profile}")
                st.metric("현재 입력 예정 총점", f"{total_score}점")
                st.markdown("**현재 검색된 주요 근거 문서**")
                st.write(current_evidence or "현재 검색 결과가 없습니다.")
                save_policy = "새 평가 저장"

            if st.button("평가 결과 저장"):
                if (
                    has_previous_evaluation
                    and save_policy == "기존 평가 유지"
                ):
                    st.success(
                        f"Q{stored_no:02d} 기존 평가를 유지했습니다. "
                        "evaluation_template.tsv는 변경하지 않았습니다."
                    )
                else:
                    ok, message = update_evaluation_result(
                        stored_scenario,
                        stored_result["results"],
                        stored_result["answer"],
                        stored_result.get("answer_mode", stored_result.get("answer_status", {}).get("answer_mode", "")),
                        stored_result.get("gemini_status", classify_gemini_status(stored_result.get("answer_status", {}))),
                        scores,
                        memo,
                    )
                    if ok:
                        st.success(message)
                        st.caption(f"저장 파일: {EVALUATION_PATH}")
                        st.rerun()
                    else:
                        st.error("평가 결과 저장 실패")
                        st.write(message)


# ==============================
# 하단 설명
# ==============================
st.divider()

st.markdown(
    """
<div class="mscc-footer">
    MineSafe AI · 공식 안전 문서 검색 기반 현장 의사결정 지원
</div>
"""
    ,
    unsafe_allow_html=True,
)
