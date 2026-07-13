from openpyxl import load_workbook
from pathlib import Path

base = Path("15_fair_eval_v2_50")
target = base / "model_answer_collection_template_50.xlsx"
output = base / "model_answer_collection_template_50_with_minesafe.xlsx"

if not target.exists():
    raise FileNotFoundError(f"파일 없음: {target}")

wb = load_workbook(target)
ws = wb.active

headers = [cell.value for cell in ws[1]]
if "minesafe_ai_answer" not in headers:
    raise ValueError("minesafe_ai_answer 열을 찾지 못했습니다.")

col = headers.index("minesafe_ai_answer") + 1

cleared = 0
for row in range(2, ws.max_row + 1):
    value = ws.cell(row=row, column=col).value
    if value is not None and str(value).strip() == "MineSafe AI 답변":
        ws.cell(row=row, column=col).value = None
        cleared += 1

wb.save(target)

# 기존 잘못 생성된 결과 파일은 삭제
if output.exists():
    output.unlink()

print(f"[OK] placeholder 삭제 완료: {cleared}개")
print(f"[OK] 저장 파일: {target}")
print("[OK] 기존 with_minesafe 결과 파일 삭제 완료")