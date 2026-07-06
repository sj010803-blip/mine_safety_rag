from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "12_compare_experiment"
OUT_PATH = OUT_DIR / "llm_rag_comparison_template.xlsx"


QUESTIONS = [
    ("Q001", "낙반/중대재해", "광산에서 작업자가 낙반으로 사망한 경우 사업주나 경영책임자가 확인해야 할 의무는 무엇인가?"),
    ("Q004", "발파/불발", "발파 후 불발이 의심되는 상황에서 작업자가 바로 접근하려고 하면 관리자는 어떤 조치를 해야 하나?"),
    ("Q010", "환기/유해가스", "갱내에서 메탄가스 농도가 기준 이상으로 측정되면 관리자는 무엇을 해야 하나?"),
    ("Q012", "환기/유해가스", "정전으로 주 환기설비가 멈췄을 때 갱내 작업자를 어떻게 대피시키고 작업 재개는 언제 판단해야 하나?"),
    ("Q066", "전기안전", "전기설비 점검 전 차단·잠금·표지 절차를 꼭 확인해야 하는 이유는 무엇인가?"),
    ("Q067", "전기안전", "감전 위험이 있는 습기 많은 구간에서 작업 전 관리자가 확인해야 할 항목은 무엇인가?"),
    ("Q031", "TBM(툴박스 미팅)", "작업 시작 전에 TBM 공정별 위험요인을 공유할 때 관리자가 먼저 확인해야 할 항목은 무엇인가?"),
    ("Q032", "TBM(툴박스 미팅)", "TBM 작업 전 안전회의에서 작업자에게 무엇을 우선적으로 안내해야 하나?"),
    ("Q101", "분진 관리", "천공 작업 전에 분진 발생을 줄이기 위해 살수와 집진 상태를 어떤 기준으로 확인해야 하나?"),
    ("Q102", "분진 관리", "굴진 작업 중 분진이 평소보다 많이 발생하면 관리자가 우선 확인하고 조치해야 할 사항은 무엇인가?"),
    ("Q103", "분진 관리", "파쇄 작업장에서 비산먼지가 주변 통로까지 퍼질 때 작업관리자는 어떤 통제조치를 해야 하나?"),
    ("Q105", "분진 관리", "작업환경측정 결과 분진 관리가 필요하다고 판단되면 작업 전 회의에서 어떤 내용을 공유해야 하나?"),
    ("Q106", "보호구/PPE", "작업 시작 전 방진마스크를 지급할 때 관리자가 착용 적합성을 어떻게 확인해야 하나?"),
    ("Q107", "보호구/PPE", "작업자가 안전모를 착용했지만 턱끈을 하지 않은 경우 관리자는 어떻게 조치해야 하나?"),
    ("Q108", "보호구/PPE", "보호장갑이나 보안경이 파손된 상태로 작업하려는 작업자가 있을 때 관리자는 무엇을 확인해야 하나?"),
    ("Q109", "보호구/PPE", "유해가스 측정 작업 전에 작업자에게 필요한 보호구와 착용 상태를 어떻게 확인해야 하나?"),
    ("Q110", "보호구/PPE", "보호구를 지급했는데 작업자가 반복적으로 미착용하는 경우 관리자는 어떤 절차로 조치해야 하나?"),
    ("Q020", "위험성평가", "위험성평가 결과 중대한 위험이 확인되었는데 개선조치 전 작업을 계속하려는 경우 관리자는 어떻게 해야 하나?"),
    ("Q025", "사고보고/응급조치", "작업자가 갱내에서 의식을 잃고 쓰러졌을 때 관리자는 어떤 순서로 대응해야 하나?"),
    ("Q030", "작업중지/재개", "위험요인을 제거한 뒤 작업을 재개하려면 어떤 확인 절차가 필요한가?"),
]

TARGETS = ["ChatGPT", "Gemini", "내 사이트 RAG"]


def clean_text(value):
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", " ", text)
    return text.strip()


def set_title(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True, color="1F4E79")
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name="맑은 고딕", size=10, color="666666")


def style_header(ws, row=1, start_col=1, end_col=None):
    if end_col is None:
        end_col = ws.max_column

    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_common_style(ws):
    thin = Side(style="thin", color="D9DEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_score_color(ws, cell_range):
    green = PatternFill("solid", fgColor="D9EAD3")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    red = PatternFill("solid", fgColor="F4CCCC")

    for row in ws[cell_range]:
        for cell in row:
            try:
                value = float(cell.value)
            except Exception:
                continue
            if value >= 85:
                cell.fill = green
            elif value >= 70:
                cell.fill = yellow
            else:
                cell.fill = red


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    ws_overview = wb.active
    ws_overview.title = "1_실험개요"
    ws_questions = wb.create_sheet("2_질문목록_20개")
    ws_answers = wb.create_sheet("3_답변원문_붙여넣기")
    ws_eval = wb.create_sheet("4_비교평가")
    ws_summary = wb.create_sheet("5_평균요약")
    ws_criteria = wb.create_sheet("6_채점기준")

    # =====================================================
    # 1_실험개요
    # =====================================================
    set_title(
        ws_overview,
        "ChatGPT vs Gemini vs 내 사이트 RAG 비교 실험",
        "평가 기준: 검색 적합성, 근거 기반성, 안전·법령 판단 정확성, 실무성 / 각 25점, 총 100점",
    )

    overview_rows = [
        ["항목", "내용"],
        ["비교 목적", "광산 안전관리 업무에서 일반 LLM과 RAG 기반 시스템의 답변 품질 차이를 비교"],
        ["비교 대상", "ChatGPT / Gemini / 내 사이트 RAG"],
        ["질문 수", "대표 질문 20개"],
        ["평가 기준", "검색 적합성, 근거 기반성, 안전·법령 판단 정확성, 실무성"],
        ["총점", "100점"],
        ["내 사이트 특징", "Vector DB 검색 근거, 문서명, chunk_id, KRAS식 위험성평가 초안, 조치 체크리스트 제공"],
        ["주의사항", "일반 LLM을 비판하기보다 광산 안전관리 실무 목적에서의 적합성을 비교"],
    ]

    for r, row in enumerate(overview_rows, start=4):
        for c, value in enumerate(row, start=1):
            ws_overview.cell(r, c, clean_text(value))

    style_header(ws_overview, 4, 1, 2)
    set_widths(ws_overview, {"A": 24, "B": 110})

    # =====================================================
    # 2_질문목록_20개
    # =====================================================
    headers = ["번호", "카테고리", "질문"]
    ws_questions.append(headers)

    for qid, category, question in QUESTIONS:
        ws_questions.append([qid, category, question])

    style_header(ws_questions, 1, 1, 3)
    set_widths(ws_questions, {"A": 12, "B": 20, "C": 90})
    ws_questions.freeze_panes = "A2"

    # =====================================================
    # 3_답변원문_붙여넣기
    # =====================================================
    answer_headers = ["번호", "카테고리", "질문", "비교대상", "답변 원문 붙여넣기", "답변 생성일", "모델명/버전", "비고"]
    ws_answers.append(answer_headers)

    for qid, category, question in QUESTIONS:
        for target in TARGETS:
            ws_answers.append([qid, category, question, target, "", "", "", ""])

    style_header(ws_answers, 1, 1, len(answer_headers))
    set_widths(
        ws_answers,
        {
            "A": 10,
            "B": 18,
            "C": 55,
            "D": 16,
            "E": 95,
            "F": 18,
            "G": 22,
            "H": 30,
        },
    )
    ws_answers.freeze_panes = "A2"

    for row in range(2, ws_answers.max_row + 1):
        ws_answers.row_dimensions[row].height = 70

    # =====================================================
    # 4_비교평가
    # =====================================================
    eval_headers = [
        "번호",
        "카테고리",
        "질문",
        "비교대상",
        "검색 적합성(25)",
        "근거 기반성(25)",
        "안전·법령 판단 정확성(25)",
        "실무성(25)",
        "총점(100)",
        "장점",
        "한계",
        "비고",
    ]

    ws_eval.append(eval_headers)

    row_idx = 2
    for qid, category, question in QUESTIONS:
        for target in TARGETS:
            ws_eval.cell(row_idx, 1, qid)
            ws_eval.cell(row_idx, 2, category)
            ws_eval.cell(row_idx, 3, question)
            ws_eval.cell(row_idx, 4, target)
            # 점수는 직접 입력
            ws_eval.cell(row_idx, 9, f"=SUM(E{row_idx}:H{row_idx})")
            row_idx += 1

    style_header(ws_eval, 1, 1, len(eval_headers))
    set_widths(
        ws_eval,
        {
            "A": 10,
            "B": 18,
            "C": 55,
            "D": 16,
            "E": 16,
            "F": 16,
            "G": 24,
            "H": 13,
            "I": 12,
            "J": 38,
            "K": 38,
            "L": 25,
        },
    )
    ws_eval.freeze_panes = "A2"

    for row in range(2, ws_eval.max_row + 1):
        ws_eval.row_dimensions[row].height = 48

    # 점수 셀 배경
    score_fill = PatternFill("solid", fgColor="EAF2F8")
    for row in range(2, ws_eval.max_row + 1):
        for col in range(5, 9):
            ws_eval.cell(row, col).fill = score_fill
            ws_eval.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws_eval.cell(row, 9).alignment = Alignment(horizontal="center", vertical="center")

    # =====================================================
    # 5_평균요약
    # =====================================================
    set_title(ws_summary, "비교 결과 평균 요약", "4개 평가 기준 및 총점 평균 비교")

    summary_headers = [
        "비교대상",
        "검색 적합성 평균",
        "근거 기반성 평균",
        "안전·법령 판단 정확성 평균",
        "실무성 평균",
        "총점 평균",
    ]
    for c, h in enumerate(summary_headers, start=1):
        ws_summary.cell(4, c, h)

    for r, target in enumerate(TARGETS, start=5):
        ws_summary.cell(r, 1, target)
        ws_summary.cell(r, 2, f'=AVERAGEIF(\'4_비교평가\'!$D:$D,A{r},\'4_비교평가\'!$E:$E)')
        ws_summary.cell(r, 3, f'=AVERAGEIF(\'4_비교평가\'!$D:$D,A{r},\'4_비교평가\'!$F:$F)')
        ws_summary.cell(r, 4, f'=AVERAGEIF(\'4_비교평가\'!$D:$D,A{r},\'4_비교평가\'!$G:$G)')
        ws_summary.cell(r, 5, f'=AVERAGEIF(\'4_비교평가\'!$D:$D,A{r},\'4_비교평가\'!$H:$H)')
        ws_summary.cell(r, 6, f'=AVERAGEIF(\'4_비교평가\'!$D:$D,A{r},\'4_비교평가\'!$I:$I)')

    style_header(ws_summary, 4, 1, 6)
    set_widths(ws_summary, {"A": 18, "B": 20, "C": 20, "D": 28, "E": 16, "F": 16})

    for row in range(5, 8):
        for col in range(2, 7):
            ws_summary.cell(row, col).number_format = "0.00"

    chart = BarChart()
    chart.title = "비교대상별 총점 평균"
    chart.y_axis.title = "점수"
    chart.x_axis.title = "비교대상"
    data = Reference(ws_summary, min_col=6, min_row=4, max_row=7)
    cats = Reference(ws_summary, min_col=1, min_row=5, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 14
    ws_summary.add_chart(chart, "H4")

    # =====================================================
    # 6_채점기준
    # =====================================================
    set_title(ws_criteria, "채점 기준", "각 항목 25점, 총점 100점")

    criteria_rows = [
        ["평가 항목", "점수 기준"],
        ["검색 적합성", "25점: 질문과 직접 관련된 공식 법령·지침·문서 제시 / 20점: 대체로 관련 / 15점: 관련성 있으나 핵심 부족 / 10점 이하: 일반론 또는 부적합"],
        ["근거 기반성", "25점: 문서명·근거 요약·chunk_id 등 출처 명확 / 20점: 법령·지침 언급 / 15점: 일반 근거 중심 / 10점 이하: 출처 불명확"],
        ["안전·법령 판단 정확성", "25점: 작업중지·대피·보고·조사·재발방지 등 판단 적절 / 20점: 주요 판단 적절 / 15점: 광산 특화성 부족 / 10점 이하: 핵심 조치 누락"],
        ["실무성", "25점: KRAS 양식·체크리스트에 바로 활용 가능 / 20점: 조치 구체적 / 15점: 기록용으로 추가 정리 필요 / 10점 이하: 원칙적 조언 중심"],
    ]

    for r, row in enumerate(criteria_rows, start=4):
        for c, value in enumerate(row, start=1):
            ws_criteria.cell(r, c, clean_text(value))

    style_header(ws_criteria, 4, 1, 2)
    set_widths(ws_criteria, {"A": 26, "B": 115})

    # =====================================================
    # 공통 스타일 적용
    # =====================================================
    for ws in wb.worksheets:
        apply_common_style(ws)
        ws.sheet_view.showGridLines = False

    # 헤더 스타일 다시 적용
    style_header(ws_questions, 1, 1, 3)
    style_header(ws_answers, 1, 1, len(answer_headers))
    style_header(ws_eval, 1, 1, len(eval_headers))
    style_header(ws_summary, 4, 1, 6)
    style_header(ws_criteria, 4, 1, 2)

    # 점수 총점 조건 색상
    add_score_color(ws_eval, f"I2:I{ws_eval.max_row}")

    wb.save(OUT_PATH)

    print("[OK] 비교평가 엑셀 템플릿 생성 완료")
    print(f"저장 파일: {OUT_PATH}")
    print(f"질문 수: {len(QUESTIONS)}개")
    print(f"비교 대상: {', '.join(TARGETS)}")
    print(f"평가 행 수: {len(QUESTIONS) * len(TARGETS)}행")


if __name__ == "__main__":
    main()